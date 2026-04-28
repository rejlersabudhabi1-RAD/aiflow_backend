"""
Export Service — SLD Verification
==================================
Generates Excel and PDF quality reports from stored findings.
"""
import io
import logging
from typing import Optional

logger = logging.getLogger(__name__)

HIDDEN_CATEGORIES = set()

FINDINGS_COLUMNS = [
    'SL No', 'Drawing ID', 'Category', 'Rule ID',
    'Issue Observed', 'Action Required', 'Evidence',
    'Direction', 'Severity', 'Status',
]


def generate_excel(document) -> Optional[bytes]:
    """Build an Excel workbook from all findings for `document`."""
    try:
        import pandas as pd
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        logger.error('[SLDExport] pandas/openpyxl not installed')
        return None

    rows = []
    for drawing in document.drawings.order_by('page_index'):
        for finding in drawing.findings.order_by('sl_no'):
            if finding.category in HIDDEN_CATEGORIES:
                continue
            rows.append({
                'SL No':           finding.sl_no,
                'Drawing ID':      drawing.drawing_id,
                'Category':        finding.category,
                'Rule ID':         finding.rule_id,
                'Issue Observed':  finding.issue_observed,
                'Action Required': finding.action_required,
                'Evidence':        finding.evidence,
                'Direction':       finding.direction,
                'Severity':        finding.severity.upper(),
                'Status':          finding.status.capitalize(),
            })

    df = pd.DataFrame(rows, columns=FINDINGS_COLUMNS) if rows else pd.DataFrame(columns=FINDINGS_COLUMNS)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='SLD QC Findings')
        wb = writer.book
        ws = writer.sheets['SLD QC Findings']

        header_fill = PatternFill(fill_type='solid', fgColor='2F5496')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, col_name in enumerate(FINDINGS_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 40

    buf.seek(0)
    return buf.read()


def generate_pdf(document) -> Optional[bytes]:
    """Generate a PDF report from findings (placeholder)."""
    logger.warning('[SLDExport] PDF generation not yet implemented')
    return None


def upload_to_s3(data: bytes, key: str, content_type: str) -> str:
    """Upload bytes to S3 and return the URL (placeholder)."""
    logger.warning('[SLDExport] S3 upload not yet implemented')
    return ''
