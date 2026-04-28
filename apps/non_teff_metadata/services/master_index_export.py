"""
Master Index Excel Exporter
---------------------------

Builds the two-sheet workbook defined in ``master_index_template.json``:

* Primary sheet - one row per batch item, 37 columns, styled header.
  Sheet name is derived from the ``sheets.primary`` token (``{plant}``
  substituted from the batch).

* DOCUMENT_COUNT sheet - aggregate counts by ``document_type``.

The column order, labels and widths come from the template; nothing is
hard-coded here beyond styling defaults.
"""

from __future__ import annotations

import io
import logging
import re
from collections import Counter
from typing import Iterable

from .master_index_service import get_columns, load_template

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Styling constants (soft-coded)
# ---------------------------------------------------------------------------

HEADER_FILL_COLOR = '1F4E78'        # dark blue
HEADER_FONT_COLOR = 'FFFFFF'
HEADER_FONT_NAME  = 'Calibri'
HEADER_FONT_SIZE  = 11
BODY_FONT_NAME    = 'Calibri'
BODY_FONT_SIZE    = 10
ROW_HEIGHT        = 18
HEADER_ROW_HEIGHT = 32
FREEZE_CELL       = 'A2'


def _slugify_plant(plant: str) -> str:
    return re.sub(r'\s+', '_', (plant or 'MASTER').strip()).upper() or 'MASTER'


def _resolve_sheet_name(template: dict, plant: str) -> str:
    name = template.get('sheets', {}).get('primary', '{plant}_METADATA')
    return name.replace('{plant}', _slugify_plant(plant))[:31]  # openpyxl limit


def build_workbook(*, batch_name: str, plant: str, items: Iterable[dict]) -> bytes:
    """
    Build and return the workbook as bytes.

    Parameters
    ----------
    batch_name : used for the filename — caller is responsible for that.
    plant : used to resolve the ``{plant}`` token in the primary sheet name.
    items : iterable of ``NonTeffBatchItem.fields`` dicts.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    template = load_template()
    columns = get_columns()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _resolve_sheet_name(template, plant)

    header_font = Font(name=HEADER_FONT_NAME, size=HEADER_FONT_SIZE,
                       bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill('solid', fgColor=HEADER_FILL_COLOR)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    body_font = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = col.get('width', 14)

    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    ws.freeze_panes = FREEZE_CELL
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Body rows - preserve template column order
    type_counter: Counter = Counter()
    for row_idx, item in enumerate(items, start=2):
        for col_idx, col in enumerate(columns, start=1):
            value = item.get(col['key'], '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = border
        ws.row_dimensions[row_idx].height = ROW_HEIGHT
        type_counter[(item.get('document_type') or 'NA')] += 1

    # DOCUMENT_COUNT sheet
    count_sheet_name = template.get('sheets', {}).get('count', 'DOCUMENT_COUNT')[:31]
    ws2 = wb.create_sheet(title=count_sheet_name)
    ws2.cell(row=1, column=1, value='Document Type').font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=1).alignment = center
    ws2.cell(row=1, column=2, value='Count').font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=2).alignment = center
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 10
    for i, (doc_type, count) in enumerate(type_counter.most_common(), start=2):
        ws2.cell(row=i, column=1, value=doc_type).font = body_font
        ws2.cell(row=i, column=2, value=count).font = body_font
    ws2.cell(row=len(type_counter) + 2, column=1, value='TOTAL').font = Font(bold=True)
    ws2.cell(row=len(type_counter) + 2, column=2,
             value=sum(type_counter.values())).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
