"""
Non-TEFF Metadata — DRF views.

Endpoints
---------
POST   /api/v1/non-teff/upload/              upload_non_teff_file
GET    /api/v1/non-teff/status/<job_id>/     get_non_teff_status
GET    /api/v1/non-teff/results/<job_id>/    get_non_teff_results
GET    /api/v1/non-teff/export/<job_id>/     export_non_teff_excel
"""

import io
import json
import logging
import os
import tempfile
import uuid

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from .models import NonTeffExtractionJob
from .services.extractor import run_extraction_async
from .services import history_archive

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# SOFT-CODED configuration constants
# ---------------------------------------------------------------------------

MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB

# Extension → internal format key
EXTENSION_FORMAT_MAP = {
    '.pdf':  'pdf',
    '.xlsx': 'excel',
    '.xls':  'excel',
    '.docx': 'word',
    '.doc':  'word',
    '.dwg':  'autocad',
    '.dxf':  'autocad',
}

ALLOWED_EXTENSIONS = set(EXTENSION_FORMAT_MAP.keys())


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_file_format(filename):
    """Return the internal format key for a given filename, or None."""
    ext = os.path.splitext(filename.lower())[1]
    return EXTENSION_FORMAT_MAP.get(ext)


def _build_excel_response(job):
    """Build an in-memory Excel file from job results and return an HttpResponse."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, 'openpyxl not installed'

    # Load field definitions
    config_path = os.path.join(os.path.dirname(__file__), 'config', 'non_teff_fields.json')
    with open(config_path, 'r') as f:
        field_config = json.load(f)
    fields = field_config['fields']

    items = (job.result_json or {}).get('items', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Non-TEFF Metadata'

    # Header style
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    # Write headers
    for col_idx, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field['label'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = field['width']

    ws.row_dimensions[1].height = 30

    # Write data rows
    for row_idx, item in enumerate(items, start=2):
        for col_idx, field in enumerate(fields, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(field['key'], ''))

    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, None


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_non_teff_file(request):
    """
    Upload a document for Non-TEFF metadata extraction.

    Request: multipart/form-data with field `file`.
    Response: { job_id, status }
    """
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    if uploaded_file.size > MAX_FILE_SIZE_BYTES:
        return Response(
            {'error': f'File exceeds maximum size of {MAX_FILE_SIZE_BYTES // (1024 * 1024)} MB.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    file_format = _get_file_format(uploaded_file.name)
    if not file_format:
        return Response(
            {'error': f'Unsupported file type. Allowed: {", ".join(ALLOWED_EXTENSIONS)}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Persist file to a temp location for background processing
    suffix = os.path.splitext(uploaded_file.name)[1].lower()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    for chunk in uploaded_file.chunks():
        tmp.write(chunk)
    tmp.flush()
    tmp_path = tmp.name
    tmp.close()

    # Create DB job record
    job = NonTeffExtractionJob.objects.create(
        file_name=uploaded_file.name,
        file_format=file_format,
        status=NonTeffExtractionJob.STATUS_PENDING,
        status_message='Queued for extraction',
        created_by=request.user,
    )

    # Kick off async extraction
    run_extraction_async(str(job.job_id), tmp_path, file_format)

    # Best-effort: archive the source file to S3 under a role-based prefix.
    # Failures here MUST NOT impact the upload response.
    try:
        history_archive.archive_source(job, tmp_path, request.user)
    except Exception:
        logger.warning('Non-TEFF history archive_source failed', exc_info=True)

    return Response({'job_id': str(job.job_id), 'status': job.status}, status=status.HTTP_202_ACCEPTED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_non_teff_status(request, job_id):
    """
    Poll extraction status.

    Response: { job_id, status, progress, message }
    """
    try:
        job = NonTeffExtractionJob.objects.get(pk=job_id)
    except NonTeffExtractionJob.DoesNotExist:
        return Response({'error': 'Job not found.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception:
        return Response({'error': 'Invalid job ID.'}, status=status.HTTP_400_BAD_REQUEST)

    return Response({
        'job_id': str(job.job_id),
        'status': job.status,
        'progress': job.progress,
        'message': job.status_message,
        'error': job.error_message,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_non_teff_results(request, job_id):
    """
    Fetch completed extraction results.

    Response: { job_id, total, items: [...] }
    """
    try:
        job = NonTeffExtractionJob.objects.get(pk=job_id)
    except NonTeffExtractionJob.DoesNotExist:
        return Response({'error': 'Job not found.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception:
        return Response({'error': 'Invalid job ID.'}, status=status.HTTP_400_BAD_REQUEST)

    if job.status != NonTeffExtractionJob.STATUS_COMPLETED:
        return Response(
            {'error': f'Results not ready. Current status: {job.status}'},
            status=status.HTTP_202_ACCEPTED,
        )

    result = job.result_json or {}

    # Best-effort: archive completed result to S3 once. Idempotent (overwrite OK).
    try:
        history_archive.archive_result(job, request.user)
    except Exception:
        logger.warning('Non-TEFF history archive_result failed', exc_info=True)

    return Response({
        'job_id': str(job.job_id),
        'file_name': job.file_name,
        'total': result.get('total', 0),
        'items': result.get('items', []),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_non_teff_excel(request, job_id):
    """
    Export extraction results as an Excel file.
    """
    try:
        job = NonTeffExtractionJob.objects.get(pk=job_id)
    except NonTeffExtractionJob.DoesNotExist:
        return Response({'error': 'Job not found.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception:
        return Response({'error': 'Invalid job ID.'}, status=status.HTTP_400_BAD_REQUEST)

    if job.status != NonTeffExtractionJob.STATUS_COMPLETED:
        return Response(
            {'error': 'Export not available — extraction not yet complete.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    buffer, error = _build_excel_response(job)
    if error:
        return Response({'error': error}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    safe_name = os.path.splitext(job.file_name)[0].replace(' ', '_')
    filename = f'NonTEFF_Metadata_{safe_name}.xlsx'

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# History endpoints (additive — no change to core extraction logic)
# ---------------------------------------------------------------------------

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_non_teff_history(request):
    """
    List past Non-TEFF extractions visible to the current user.
    Role-based: regular users see their own jobs; admins see all.
    Query params:
        ?limit=<int>   default per HISTORY_CONFIG['list_default_limit']
    Response: { role, total, items: [...] }
    """
    try:
        limit = int(request.query_params.get('limit', 0)) or None
    except (TypeError, ValueError):
        limit = None

    items = history_archive.list_history(request.user, limit=limit)
    return Response({
        'role':  history_archive.resolve_user_role(request.user),
        'total': len(items),
        'items': items,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def load_non_teff_history(request, job_id):
    """
    Load full result payload for a past extraction so the user can re-open
    it in the canvas without re-uploading the document.
    """
    payload = history_archive.load_history(request.user, job_id)
    if not payload:
        return Response(
            {'error': 'History entry not found or access denied.'},
            status=status.HTTP_404_NOT_FOUND,
        )
    return Response(payload)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_non_teff_history(request, job_id):
    """
    Delete a past extraction (single-file job or bulk batch) with RBAC.
    """
    ok, err = history_archive.delete_history(request.user, job_id)
    if not ok:
        code = status.HTTP_404_NOT_FOUND if err == 'not_found' else status.HTTP_403_FORBIDDEN
        return Response({'error': err or 'Could not delete entry.'}, status=code)
    return Response({'deleted': True, 'entry_id': job_id})


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_non_teff_history(request, job_id):
    """
    Modify metadata of a past extraction. Currently supported fields
    (soft-coded in the service layer): ``name`` / ``file_name``, ``plant``.
    """
    ok, err, data = history_archive.update_history(request.user, job_id, request.data or {})
    if not ok:
        code = status.HTTP_404_NOT_FOUND if err == 'not_found' else status.HTTP_400_BAD_REQUEST
        return Response({'error': err or 'Could not update entry.'}, status=code)
    return Response({'updated': True, 'entry': data})
