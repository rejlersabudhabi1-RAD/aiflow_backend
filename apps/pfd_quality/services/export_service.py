"""
Export Service — PFD Quality Checker
=====================================
Generates Excel and PDF quality reports from a PFDQDocument's findings.
Also provides upload_to_s3() helper.

All functions return bytes (or None on failure).
Views always call generate_*() directly — no S3 redirects.
"""
import io
import logging
import os
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------
FINDINGS_COLUMNS = [
    'SL No',
    'Drawing ID',
    'Category',
    'Rule ID',
    'Issue Observed',
    'Action Required',
    'Evidence',
    'Direction',
    'Severity',
    'Status',
]

SEVERITY_COLORS = {
    'critical': 'C00000',  # dark red
    'major':    'FF6600',  # orange
    'minor':    'FFC000',  # amber
    'info':     '4472C4',  # blue
}

HEADER_FILL = '2F5496'  # dark blue (company style)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def generate_excel(document) -> Optional[bytes]:
    """Return .xlsx bytes for all findings in the document."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'PFD Quality Findings'

        # ── Title row
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f'PFD Quality Report — {document.file_name}'
        title_cell.font  = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill  = PatternFill('solid', fgColor=HEADER_FILL)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # ── Header row
        header_fill = PatternFill('solid', fgColor=HEADER_FILL)
        thin = Side(style='thin', color='FFFFFF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, col_name in enumerate(FINDINGS_COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = border
        ws.row_dimensions[2].height = 22

        # ── Data rows
        row_num = 3
        for drawing in document.drawings.all().prefetch_related('findings'):
            for finding in drawing.findings.all().order_by('sl_no'):
                row = [
                    finding.sl_no,
                    drawing.drawing_id,
                    finding.category,
                    finding.rule_id,
                    finding.issue_observed,
                    finding.action_required,
                    finding.evidence,
                    finding.direction,
                    finding.severity,
                    finding.status,
                ]
                for col_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border    = border

                # Colour-code Severity cell
                sev_cell = ws.cell(row=row_num, column=9)
                color    = SEVERITY_COLORS.get(finding.severity, '000000')
                sev_cell.fill = PatternFill('solid', fgColor=color)
                sev_cell.font = Font(bold=True, color='FFFFFF')

                ws.row_dimensions[row_num].height = 40
                row_num += 1

        # ── Column widths
        col_widths = [8, 18, 14, 12, 45, 40, 35, 12, 12, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A3'

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as exc:
        logger.exception('[PFDQ Export] Excel generation failed: %s', exc)
        return None


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def generate_pdf(document) -> Optional[bytes]:
    """Return PDF bytes for all findings in the document."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )

        buf    = io.BytesIO()
        doc_rl = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm,
        )

        styles  = getSampleStyleSheet()
        wrap_st = ParagraphStyle('wrap', parent=styles['Normal'], fontSize=7, leading=9)
        hdr_st  = ParagraphStyle('hdr',  parent=styles['Normal'], fontSize=7, leading=9,
                                 textColor=colors.white)

        elements = []

        # Title
        title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=14,
                                     spaceAfter=6)
        elements.append(Paragraph(f'PFD Quality Report — {document.file_name}', title_style))
        elements.append(Spacer(1, 0.3*cm))

        # Table data
        table_data = [[Paragraph(c, hdr_st) for c in FINDINGS_COLUMNS]]

        for drawing in document.drawings.all().prefetch_related('findings'):
            for finding in drawing.findings.all().order_by('sl_no'):
                row = [
                    Paragraph(str(finding.sl_no),             wrap_st),
                    Paragraph(drawing.drawing_id or '',        wrap_st),
                    Paragraph(finding.category or '',          wrap_st),
                    Paragraph(finding.rule_id or '',           wrap_st),
                    Paragraph(finding.issue_observed or '',    wrap_st),
                    Paragraph(finding.action_required or '',   wrap_st),
                    Paragraph(finding.evidence or '',          wrap_st),
                    Paragraph(finding.direction or '',         wrap_st),
                    Paragraph(finding.severity or '',          wrap_st),
                    Paragraph(finding.status or '',            wrap_st),
                ]
                table_data.append(row)

        col_widths_pdf = [1.2*cm, 3.2*cm, 2.4*cm, 2.0*cm, 6.0*cm, 5.5*cm, 4.5*cm,
                          1.8*cm, 1.8*cm, 1.8*cm]
        tbl = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)

        header_bg = colors.HexColor(f'#{HEADER_FILL}')
        tbl_style = TableStyle([
            ('BACKGROUND',  (0, 0), (-1, 0),   header_bg),
            ('TEXTCOLOR',   (0, 0), (-1, 0),   colors.white),
            ('FONTNAME',    (0, 0), (-1, 0),   'Helvetica-Bold'),
            ('FONTSIZE',    (0, 0), (-1, 0),   7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
            ('VALIGN',      (0, 0), (-1, -1),  'TOP'),
            ('FONTSIZE',    (0, 1), (-1, -1),  7),
        ])

        # Severity column colour
        sev_col = 8  # 0-indexed
        sev_map = {
            'critical': '#C00000',
            'major':    '#FF6600',
            'minor':    '#FFC000',
            'info':     '#4472C4',
        }
        r = 1
        for drawing in document.drawings.all().prefetch_related('findings'):
            for finding in drawing.findings.all().order_by('sl_no'):
                hex_c = sev_map.get(finding.severity)
                if hex_c:
                    tbl_style.add('BACKGROUND', (sev_col, r), (sev_col, r), colors.HexColor(hex_c))
                    tbl_style.add('TEXTCOLOR',  (sev_col, r), (sev_col, r), colors.white)
                    tbl_style.add('FONTNAME',   (sev_col, r), (sev_col, r), 'Helvetica-Bold')
                r += 1

        tbl.setStyle(tbl_style)
        elements.append(tbl)

        doc_rl.build(elements)
        return buf.getvalue()

    except Exception as exc:
        logger.exception('[PFDQ Export] PDF generation failed: %s', exc)
        return None


# ---------------------------------------------------------------------------
# S3 upload helper (optional — used by the Celery task only)
# ---------------------------------------------------------------------------

def upload_to_s3(data: bytes, key: str, content_type: str) -> Optional[str]:
    """Upload bytes to S3 and return the public or presigned URL, or None on error."""
    try:
        import boto3

        bucket = os.environ.get('AWS_STORAGE_BUCKET_NAME', '')
        region = os.environ.get('AWS_S3_REGION_NAME', 'us-east-1')
        if not bucket:
            return None

        s3 = boto3.client('s3', region_name=region)
        s3.put_object(
            Bucket=bucket,
            Key=key,
            Body=data,
            ContentType=content_type,
        )
        return f'https://{bucket}.s3.{region}.amazonaws.com/{key}'
    except Exception as exc:
        logger.warning('[PFDQ Export] S3 upload failed (non-fatal): %s', exc)
        return None
