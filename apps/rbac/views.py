"""
RBAC Views - DRF ViewSets for Super Admin Dashboard
"""
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.utils import timezone
from django.db.models import Q, Count
from django_filters.rest_framework import DjangoFilterBackend

from .models import (
    Organization, Module, Permission, Role, RolePermission, RoleModule,
    UserProfile, UserRole, UserStorage, AuditLog
)
from .serializers import (
    OrganizationSerializer, ModuleSerializer, PermissionSerializer,
    RoleSerializer, RoleListSerializer, RolePermissionSerializer, RoleModuleSerializer,
    UserProfileSerializer, UserProfileListSerializer, UserRoleSerializer,
    UserStorageSerializer, AuditLogSerializer,
    UserPermissionCheckSerializer, UserModuleCheckSerializer
)
from .permissions import (
    IsSuperAdmin, IsAdmin, CanManageUsers, CanManageRoles, SameOrganization
)
from .utils import create_audit_log
from .s3_service import S3Service
from .pagination import FlexiblePageNumberPagination


class OrganizationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing organizations
    Admins can view organizations, only super admin can create/edit/delete
    """
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'code', 'primary_contact_email']
    ordering_fields = ['name', 'created_at']
    filterset_fields = ['is_active']
    
    def get_permissions(self):
        """
        Allow admins to read organizations, only super admin can modify
        """
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated, IsAdmin]
        else:
            permission_classes = [IsAuthenticated, IsSuperAdmin]
        return [permission() for permission in permission_classes]
    
    def perform_create(self, serializer):
        org = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='create',
            resource_type='Organization',
            resource_id=org.id,
            resource_repr=str(org),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
    
    def perform_update(self, serializer):
        old_data = {
            'name': serializer.instance.name,
            'is_active': serializer.instance.is_active
        }
        org = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='update',
            resource_type='Organization',
            resource_id=org.id,
            resource_repr=str(org),
            changes={'old': old_data, 'new': serializer.data},
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )


class ModuleViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing application modules
    Admins can view modules, only super admin can create/edit/delete
    """
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'code']
    ordering_fields = ['order', 'name']
    filterset_fields = ['is_active']
    pagination_class = None  # Disable pagination - modules are a small dataset
    
    def get_permissions(self):
        """
        Allow admins to read modules, only super admin can modify
        """
        if self.action in ['list', 'retrieve', 'active']:
            permission_classes = [IsAuthenticated, IsAdmin]
        else:
            permission_classes = [IsAuthenticated, IsSuperAdmin]
        return [permission() for permission in permission_classes]
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated, IsAdmin])
    def active(self, request):
        """Get all active modules"""
        modules = Module.objects.filter(is_active=True).order_by('order', 'name')
        serializer = self.get_serializer(modules, many=True)
        return Response(serializer.data)


class PermissionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing permissions
    Only super admin can create/edit permissions
    """
    queryset = Permission.objects.select_related('module').all()
    serializer_class = PermissionSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'code']
    ordering_fields = ['module__name', 'action', 'name']
    filterset_fields = ['module', 'action', 'is_active']
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def by_module(self, request):
        """Get permissions grouped by module"""
        module_id = request.query_params.get('module_id')
        if module_id:
            permissions = Permission.objects.filter(
                module_id=module_id,
                is_active=True
            )
        else:
            permissions = Permission.objects.filter(is_active=True)
        
        serializer = self.get_serializer(permissions, many=True)
        return Response(serializer.data)


class RoleViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing roles
    Only super admin can create/edit roles
    """
    queryset = Role.objects.prefetch_related('permissions', 'modules').all()
    permission_classes = [IsAuthenticated, CanManageRoles]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'code']
    ordering_fields = ['level', 'name']
    filterset_fields = ['level', 'is_active']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return RoleListSerializer
        return RoleSerializer
    
    def perform_create(self, serializer):
        role = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='create',
            resource_type='Role',
            resource_id=role.id,
            resource_repr=str(role),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
    
    def perform_update(self, serializer):
        role = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='update',
            resource_type='Role',
            resource_id=role.id,
            resource_repr=str(role),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
    
    def perform_destroy(self, instance):
        if instance.is_system_role:
            raise serializers.ValidationError("Cannot delete system roles")
        
        create_audit_log(
            user=self.request.user,
            action='delete',
            resource_type='Role',
            resource_id=instance.id,
            resource_repr=str(instance),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
        instance.delete()
    
    @action(detail=True, methods=['post'])
    def assign_permission(self, request, pk=None):
        """Assign permission to role"""
        role = self.get_object()
        permission_id = request.data.get('permission_id')
        
        if not permission_id:
            return Response(
                {'error': 'permission_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            permission = Permission.objects.get(id=permission_id)
            RolePermission.objects.get_or_create(
                role=role,
                permission=permission,
                defaults={'granted_by': request.user}
            )
            
            create_audit_log(
                user=request.user,
                action='permission_grant',
                resource_type='Role',
                resource_id=role.id,
                resource_repr=str(role),
                metadata={'permission': permission.code},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            return Response({'status': 'permission assigned'})
        except Permission.DoesNotExist:
            return Response(
                {'error': 'Permission not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def revoke_permission(self, request, pk=None):
        """Revoke permission from role"""
        role = self.get_object()
        permission_id = request.data.get('permission_id')
        
        if not permission_id:
            return Response(
                {'error': 'permission_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count = RolePermission.objects.filter(
            role=role,
            permission_id=permission_id
        ).delete()[0]
        
        if deleted_count > 0:
            create_audit_log(
                user=request.user,
                action='permission_revoke',
                resource_type='Role',
                resource_id=role.id,
                resource_repr=str(role),
                metadata={'permission_id': str(permission_id)},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
        
        return Response({'status': 'permission revoked', 'count': deleted_count})
    
    @action(detail=True, methods=['post'])
    def assign_module(self, request, pk=None):
        """Assign module to role"""
        role = self.get_object()
        module_id = request.data.get('module_id')
        
        if not module_id:
            return Response(
                {'error': 'module_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            module = Module.objects.get(id=module_id)
            RoleModule.objects.get_or_create(
                role=role,
                module=module,
                defaults={'granted_by': request.user}
            )
            
            return Response({'status': 'module assigned'})
        except Module.DoesNotExist:
            return Response(
                {'error': 'Module not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def revoke_module(self, request, pk=None):
        """Revoke module from role"""
        role = self.get_object()
        module_id = request.data.get('module_id')
        
        if not module_id:
            return Response(
                {'error': 'module_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count = RoleModule.objects.filter(
            role=role,
            module_id=module_id
        ).delete()[0]
        
        return Response({'status': 'module revoked', 'count': deleted_count})


class UserProfileViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing user profiles with performance optimizations
    Admin can manage users in their organization
    Super admin can manage all users
    
    Performance Features:
    - Optimized queries (66% fewer DB queries)
    - Redis caching for large list requests (5min TTL)
    - Response time: <2s for 276 users (cached)
    
    Flexible Pagination:
    - GET /api/v1/rbac/users/ - Returns 10 users (default)
    - GET /api/v1/rbac/users/?page_size=25 - Returns 25 users
    - GET /api/v1/rbac/users/?page_size=100 - Returns 100 users
    - GET /api/v1/rbac/users/?page_size=1000 - Returns all users (cached)
    """
    permission_classes = [IsAuthenticated, CanManageUsers]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['user__email', 'user__first_name', 'user__last_name', 'employee_id']
    ordering_fields = ['created_at', 'user__email', 'last_login_at']
    filterset_fields = ['organization', 'status', 'is_deleted']
    
    # Use custom pagination for flexible page sizes
    pagination_class = FlexiblePageNumberPagination
    
    def get_permissions(self):
        """
        Custom permissions:
        - 'me' and 'change_password' actions only require authentication
        - Other actions require user management permissions
        """
        if self.action in ['me', 'change_password', 'engineers']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), CanManageUsers()]
    
    def get_queryset(self):
        """
        Filter users based on role with optimized query performance
        
        Optimization Strategy:
        - select_related: Fetch user and organization in single query (JOIN)
        - prefetch_related: Fetch roles and userrole_set efficiently
        - Reduces N+1 query problem from 276+ queries to ~3 queries
        """
        user = self.request.user
        queryset = UserProfile.objects.select_related(
            'user', 'organization'
        ).prefetch_related(
            'roles',
            'userrole_set__role'  # Fix N+1 for primary_role lookup
        ).filter(is_deleted=False)
        
        # Super admin sees all
        try:
            profile = user.rbac_profile
            if not profile.roles.filter(code='super_admin', is_active=True).exists():
                # Other admins see only their organization
                queryset = queryset.filter(organization=profile.organization)
        except UserProfile.DoesNotExist:
            return UserProfile.objects.none()
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return UserProfileListSerializer
        return UserProfileSerializer
    
    def perform_create(self, serializer):
        profile = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='create',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
    
    def perform_update(self, serializer):
        profile = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='update',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
    
    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """Deactivate user"""
        profile = self.get_object()
        profile.status = 'inactive'
        profile.user.is_active = False
        profile.save()
        profile.user.save()
        
        create_audit_log(
            user=request.user,
            action='update',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            metadata={'status_change': 'active -> inactive'},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({'status': 'user deactivated'})
    
    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Activate user"""
        profile = self.get_object()
        profile.status = 'active'
        profile.user.is_active = True
        profile.save()
        profile.user.save()
        
        create_audit_log(
            user=request.user,
            action='update',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            metadata={'status_change': 'inactive -> active'},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({'status': 'user activated'})
    
    @action(detail=True, methods=['delete'])
    def soft_delete(self, request, pk=None):
        """Soft delete user"""
        profile = self.get_object()
        profile.is_deleted = True
        profile.deleted_at = timezone.now()
        profile.deleted_by = request.user
        profile.user.is_active = False
        profile.save()
        profile.user.save()
        
        create_audit_log(
            user=request.user,
            action='delete',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({'status': 'user soft deleted'})
    
    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        """
        Reset user password to default password
        Admin-only action for security
        """
        from django.contrib.auth.hashers import make_password
        from django.utils import timezone
        from django.conf import settings
        
        profile = self.get_object()
        user = profile.user
        
        # Default password (soft-coded in settings)
        default_password = getattr(settings, 'DEFAULT_USER_PASSWORD', 'Rejlers@123')
        
        # Set the password
        user.password = make_password(default_password)
        user.last_password_change = timezone.now()
        user.must_reset_password = True
        user.is_first_login = False
        user.save()
        
        # Set must_change_password flag
        profile.must_change_password = True
        profile.save()
        
        # Log the action
        create_audit_log(
            user=request.user,
            action='reset_password',
            resource_type='User',
            resource_id=user.id,
            resource_repr=f'{user.email}',
            changes={'reset_by': request.user.email, 'must_change_password': True},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({
            'status': 'password reset successfully',
            'message': f'Password has been reset to default. User must change it on next login.',
            'default_password': default_password
        })
    
    @action(detail=False, methods=['post'], url_path='change-password')
    def change_password(self, request):
        """
        Change user's own password
        Required fields: old_password, new_password
        Clears must_change_password flag on success
        """
        from django.contrib.auth.hashers import check_password, make_password
        from django.utils import timezone
        
        user = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        
        # Validation
        if not old_password or not new_password:
            return Response(
                {'error': 'old_password and new_password are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verify old password
        if not check_password(old_password, user.password):
            return Response(
                {'error': 'Invalid old password'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Password strength validation (basic)
        if len(new_password) < 8:
            return Response(
                {'error': 'Password must be at least 8 characters long'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update password
        user.password = make_password(new_password)
        user.last_password_change = timezone.now()
        user.must_reset_password = False
        user.is_first_login = False
        user.save()
        
        # Clear must_change_password flag
        try:
            profile = user.rbac_profile  # Fixed: use rbac_profile instead of profile
            if profile.must_change_password:
                profile.must_change_password = False
                profile.save()
        except Exception as e:
            print(f"Warning: Could not clear must_change_password flag: {e}")
        
        # Log the action
        create_audit_log(
            user=user,
            action='change_password',
            resource_type='User',
            resource_id=user.id,
            resource_repr=f'{user.email}',
            changes={'password_changed': True, 'must_change_password_cleared': True},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({
            'status': 'password changed successfully',
            'message': 'Your password has been updated'
        })
    
    @action(detail=True, methods=['post'])
    def assign_role(self, request, pk=None):
        """Assign role to user"""
        profile = self.get_object()
        role_id = request.data.get('role_id')
        is_primary = request.data.get('is_primary', False)
        
        if not role_id:
            return Response(
                {'error': 'role_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            role = Role.objects.get(id=role_id)
            user_role, created = UserRole.objects.get_or_create(
                user_profile=profile,
                role=role,
                defaults={'assigned_by': request.user, 'is_primary': is_primary}
            )
            
            create_audit_log(
                user=request.user,
                action='role_assign',
                resource_type='UserProfile',
                resource_id=profile.id,
                resource_repr=str(profile),
                metadata={'role': role.name},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            return Response({
                'status': 'created' if created else 'already_exists',
                'role': role.name
            })
        except Role.DoesNotExist:
            return Response(
                {'error': 'Role not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def revoke_role(self, request, pk=None):
        """Revoke role from user"""
        profile = self.get_object()
        role_id = request.data.get('role_id')
        
        if not role_id:
            return Response(
                {'error': 'role_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count = UserRole.objects.filter(
            user_profile=profile,
            role_id=role_id
        ).delete()[0]
        
        if deleted_count > 0:
            create_audit_log(
                user=request.user,
                action='role_revoke',
                resource_type='UserProfile',
                resource_id=profile.id,
                resource_repr=str(profile),
                metadata={'role_id': str(role_id)},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
        
        return Response({'status': 'role revoked', 'count': deleted_count})
    
    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """
        Bulk upload users from CSV/Excel with Email Notifications
        Expected CSV format: email,first_name,last_name,password,department,job_title,phone,role_codes,module_codes
        role_codes and module_codes should be comma-separated (e.g., "admin,engineer" or "PID,PFD")
        
        New Features:
        - Sends welcome email with credentials to each user
        - Aligned with registration form fields
        - Better error handling and reporting
        """
        import csv
        import io
        from django.contrib.auth import get_user_model
        from django.db import transaction
        from apps.users.email_service import send_email
        from apps.users.email_templates import get_email_template
        from django.conf import settings
        import secrets
        import string
        
        User = get_user_model()
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        organization_id = request.data.get('organization_id')
        
        # Validate file extension
        if not file.name.endswith(('.csv', '.txt')):
            return Response(
                {'error': 'Invalid file format. Please upload a CSV file.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get organization
        try:
            organization = Organization.objects.get(id=organization_id) if organization_id else None
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Parse CSV
        try:
            decoded_file = file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            results = {
                'success': [],
                'failed': [],
                'skipped': []
            }
            
            with transaction.atomic():
                for row_num, row in enumerate(reader, start=2):  # Start at 2 (1 is header)
                    try:
                        # Validate required fields
                        email = row.get('email', '').strip()
                        if not email:
                            results['failed'].append({
                                'row': row_num,
                                'email': email or 'N/A',
                                'error': 'Email is required'
                            })
                            continue
                        
                        # Check if user already exists (exclude soft-deleted users)
                        existing_profile = UserProfile.objects.filter(
                            user__email=email, 
                            is_deleted=False
                        ).first()
                        
                        if existing_profile:
                            results['skipped'].append({
                                'row': row_num,
                                'email': email,
                                'reason': 'User already exists'
                            })
                            continue
                        
                        # Generate unique username from email
                        base_username = email.split('@')[0]
                        username = base_username
                        counter = 1
                        while User.objects.filter(username=username).exists():
                            username = f"{base_username}{counter}"
                            counter += 1
                        
                        # Get or generate password
                        password = row.get('password', '').strip()
                        if not password:
                            # Generate secure random password
                            alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
                            password = ''.join(secrets.choice(alphabet) for i in range(12))
                            # Ensure at least one of each type
                            password = secrets.choice(string.ascii_uppercase) + \
                                      secrets.choice(string.ascii_lowercase) + \
                                      secrets.choice(string.digits) + \
                                      secrets.choice("!@#$%^&*") + \
                                      password[4:]
                        
                        # Store original password for email
                        original_password = password
                        
                        # Create user
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            first_name=row.get('first_name', '').strip(),
                            last_name=row.get('last_name', '').strip(),
                            password=password
                        )
                        
                        # Create user profile (using 'phone' not 'phone_number' to align with form)
                        profile = UserProfile.objects.create(
                            user=user,
                            organization=organization,
                            department=row.get('department', '').strip(),
                            job_title=row.get('job_title', '').strip(),
                            phone_number=row.get('phone', '').strip() or row.get('phone_number', '').strip(),
                            status='active'
                        )
                        
                        # Assign roles
                        role_codes = row.get('role_codes', '').strip()
                        if role_codes:
                            role_code_list = [r.strip() for r in role_codes.split(',')]
                            roles = Role.objects.filter(code__in=role_code_list, is_active=True)
                            for idx, role in enumerate(roles):
                                UserRole.objects.create(
                                    user_profile=profile,
                                    role=role,
                                    assigned_by=request.user,
                                    is_primary=(idx == 0)
                                )
                        
                        # Assign modules
                        module_codes = row.get('module_codes', '').strip()
                        if module_codes:
                            module_code_list = [m.strip() for m in module_codes.split(',')]
                            modules = Module.objects.filter(code__in=module_code_list, is_active=True)
                            for module in modules:
                                profile.modules.add(module)
                        
                        results['success'].append({
                            'row': row_num,
                            'email': email,
                            'name': f"{user.first_name} {user.last_name}".strip(),
                            'username': username
                        })
                        
                        # Send welcome email with credentials
                        try:
                            login_url = f"{settings.FRONTEND_URL}/login" if hasattr(settings, 'FRONTEND_URL') else 'https://radai.ae/login'
                            
                            email_context = {
                                'first_name': user.first_name or 'User',
                                'last_name': user.last_name or '',
                                'email': user.email,
                                'username': username,
                                'temp_password': original_password,
                                'login_url': login_url
                            }
                            
                            email_template = get_email_template('welcome', email_context)
                            
                            send_email(
                                to_email=user.email,
                                subject=email_template['subject'],
                                html_body=email_template['html_body'],
                                text_body=email_template['text_body']
                            )
                            
                            # Update result to indicate email sent
                            results['success'][-1]['email_sent'] = True
                            
                        except Exception as email_error:
                            # Log email error but don't fail user creation
                            print(f"⚠️ Failed to send welcome email to {email}: {str(email_error)}")
                            results['success'][-1]['email_sent'] = False
                            results['success'][-1]['email_error'] = str(email_error)
                        
                        # Create audit log
                        create_audit_log(
                            user=request.user,
                            action='bulk_create',
                            resource_type='UserProfile',
                            resource_id=profile.id,
                            resource_repr=str(profile),
                            metadata={'source': 'bulk_upload', 'email_sent': results['success'][-1].get('email_sent', False)},
                            ip_address=request.META.get('REMOTE_ADDR'),
                            user_agent=request.META.get('HTTP_USER_AGENT', '')
                        )
                        
                    except Exception as e:
                        results['failed'].append({
                            'row': row_num,
                            'email': row.get('email', 'N/A'),
                            'error': str(e)
                        })
            
            # Calculate email statistics
            emails_sent = sum(1 for item in results['success'] if item.get('email_sent', False))
            emails_failed = sum(1 for item in results['success'] if not item.get('email_sent', False))
            
            # Create summary audit log
            create_audit_log(
                user=request.user,
                action='bulk_upload',
                resource_type='UserProfile',
                resource_id=None,
                resource_repr='Bulk User Upload',
                metadata={
                    'success_count': len(results['success']),
                    'failed_count': len(results['failed']),
                    'skipped_count': len(results['skipped']),
                    'emails_sent': emails_sent,
                    'emails_failed': emails_failed
                },
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            return Response({
                'message': 'Bulk upload completed successfully!',
                'summary': {
                    'total_processed': len(results['success']) + len(results['failed']) + len(results['skipped']),
                    'successful': len(results['success']),
                    'failed': len(results['failed']),
                    'skipped': len(results['skipped']),
                    'emails_sent': emails_sent,
                    'emails_failed': emails_failed
                },
                'details': results
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download CSV template for bulk upload
        Template aligned with registration form fields
        """
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="user_bulk_upload_template.csv"'
        
        writer = csv.writer(response)
        
        # Header row - aligned with registration form
        writer.writerow([
            'email',           # Required - User's email address
            'first_name',      # Required - User's first name
            'last_name',       # Required - User's last name
            'password',        # Optional - Leave empty for auto-generated password
            'department',      # Optional - e.g., Engineering, Finance
            'job_title',       # Optional - e.g., Senior Engineer
            'phone',           # Optional - Phone number (aligned with form field name)
            'role_codes',      # Optional - Comma-separated role codes (e.g., engineer,reviewer)
            'module_codes'     # Optional - Comma-separated module codes (e.g., PID,PFD,CRS)
        ])
        
        # Example row 1 - Engineer with specific password
        writer.writerow([
            'john.doe@company.com', 
            'John', 
            'Doe', 
            'SecurePass@123',
            'Engineering', 
            'Senior Engineer', 
            '+971501234567',
            'engineer,reviewer', 
            'PID,PFD,CRS'
        ])
        
        # Example row 2 - Manager with auto-generated password
        writer.writerow([
            'jane.smith@company.com', 
            'Jane', 
            'Smith', 
            '',  # Empty password = auto-generated
            'Management', 
            'Project Manager', 
            '+971507654321',
            'manager', 
            'PID,PFD,CRS,PROJECT_CONTROL'
        ])
        
        # Example row 3 - Test user with email xerxez.in@gmail.com
        writer.writerow([
            'xerxez.in@gmail.com', 
            'Test', 
            'User', 
            '',  # Auto-generated password
            'Testing', 
            'Test Engineer', 
            '+971501112233',
            'engineer', 
            'PID,PFD'
        ])
        
        return response

    @action(detail=False, methods=['get', 'patch'], url_path='me')
    def me(self, request):
        """Get or update current user's profile"""
        import traceback
        
        try:
            # PATCH - Update profile
            if request.method == 'PATCH':
                return self.update_my_profile(request)
            
            # GET - Retrieve profile
            # Log the request for debugging
            print(f"\n[DEBUG /rbac/users/me/] User: {request.user}")
            print(f"[DEBUG /rbac/users/me/] User authenticated: {request.user.is_authenticated}")
            print(f"[DEBUG /rbac/users/me/] User email: {getattr(request.user, 'email', 'N/A')}")
            
            # Try to get existing profile with optimized query
            # Use select_related and prefetch_related to avoid N+1 queries
            profile = UserProfile.objects.select_related(
                'user', 
                'organization'
            ).prefetch_related(
                'roles',
                'roles__permissions',  # Prefetch permissions through roles
                'userrole_set__role'
            ).filter(
                user=request.user,
                is_deleted=False
            ).first()
            
            print(f"[DEBUG /rbac/users/me/] Profile found: {profile is not None}")
            
            # If no profile exists, return user info without RBAC data
            if not profile:
                print(f"[DEBUG /rbac/users/me/] No RBAC profile for {request.user.email}")
                return Response({
                    'id': str(request.user.id),
                    'email': request.user.email,
                    'username': request.user.username,
                    'first_name': request.user.first_name,
                    'last_name': request.user.last_name,
                    'roles': [],
                    'organization': None,
                    'status': 'pending',
                    'message': 'RBAC profile not configured. Please contact administrator.'
                })
            
            # Try to serialize the profile
            print(f"[DEBUG /rbac/users/me/] Serializing profile...")
            serializer = self.get_serializer(profile)
            data = serializer.data
            print(f"[DEBUG /rbac/users/me/] Serialization successful")
            print(f"[DEBUG /rbac/users/me/] Roles count: {len(data.get('roles', []))}")
            print(f"[DEBUG /rbac/users/me/] Phone: {data.get('phone')}, Profile Photo: {data.get('profile_photo')}")
            
            return Response(data)
            
        except Exception as e:
            # Log the full error for debugging
            print(f"\n[ERROR /rbac/users/me/] Exception occurred: {str(e)}")
            print(f"[ERROR /rbac/users/me/] Exception type: {type(e).__name__}")
            print(f"[ERROR /rbac/users/me/] Traceback:")
            traceback.print_exc()
            
            # Return safe fallback data to avoid breaking the UI
            return Response({
                'id': str(getattr(request.user, 'id', '')),
                'email': getattr(request.user, 'email', ''),
                'username': getattr(request.user, 'username', ''),
                'first_name': getattr(request.user, 'first_name', ''),
                'last_name': getattr(request.user, 'last_name', ''),
                'roles': [],
                'organization': None,
                'status': 'pending',
                'message': 'Profile temporarily unavailable'
            })
    
    def update_my_profile(self, request):
        """Update current user's profile"""
        try:
            # Get user profile
            profile = UserProfile.objects.filter(
                user=request.user,
                is_deleted=False
            ).first()
            
            if not profile:
                return Response(
                    {'error': 'Profile not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Track changes for audit log
            changes = {}
            
            # Update User model fields
            user = request.user
            if 'first_name' in request.data:
                changes['first_name'] = request.data['first_name']
                user.first_name = request.data['first_name']
            if 'last_name' in request.data:
                changes['last_name'] = request.data['last_name']
                user.last_name = request.data['last_name']
            user.save()
            
            # Update UserProfile fields
            if 'phone' in request.data:
                changes['phone'] = request.data['phone']
                profile.phone = request.data['phone']
            if 'bio' in request.data:
                changes['bio'] = request.data['bio']
                profile.bio = request.data['bio']
            if 'location' in request.data:
                changes['location'] = request.data['location']
                profile.location = request.data['location']
            if 'department' in request.data:
                changes['department'] = request.data['department']
                profile.department = request.data['department']
            if 'job_title' in request.data:
                changes['job_title'] = request.data['job_title']
                profile.job_title = request.data['job_title']
            
            # Handle profile photo upload
            if 'profile_photo' in request.FILES:
                profile.profile_photo = request.FILES['profile_photo']
                changes['profile_photo'] = 'uploaded'

            # Handle engineer_profile JSON — persisted to rbac_engineer_profiles table
            # Accepts both JSON body (dict) and FormData (JSON string)
            ep_raw = request.data.get('engineer_profile')
            if ep_raw is not None:
                import json as _json
                if isinstance(ep_raw, str):
                    try:
                        ep_raw = _json.loads(ep_raw)
                    except Exception:
                        ep_raw = None
                if isinstance(ep_raw, dict):
                    from apps.rbac.models import EngineerProfile
                    ep_obj, _ = EngineerProfile.objects.get_or_create(user_profile=profile)
                    ep_obj.expertise_level          = ep_raw.get('expertise_level', ep_obj.expertise_level)
                    ep_obj.years_experience         = int(ep_raw.get('years_experience') or ep_obj.years_experience or 0)
                    ep_obj.engineering_disciplines  = ep_raw.get('engineering_disciplines', ep_obj.engineering_disciplines)
                    ep_obj.technical_skills         = ep_raw.get('technical_skills', ep_obj.technical_skills)
                    ep_obj.languages                = ep_raw.get('languages', ep_obj.languages)
                    ep_obj.certifications           = ep_raw.get('certifications', ep_obj.certifications)
                    ep_obj.availability_status      = ep_raw.get('availability_status', ep_obj.availability_status)
                    ep_obj.availability_percentage  = int(ep_raw.get('availability_percentage') or ep_obj.availability_percentage or 100)
                    ep_obj.next_available_date      = ep_raw.get('next_available_date') or None
                    ep_obj.max_concurrent_projects  = int(ep_raw.get('max_concurrent_projects') or ep_obj.max_concurrent_projects or 2)
                    ep_obj.preferred_project_types  = ep_raw.get('preferred_project_types', ep_obj.preferred_project_types)
                    ep_obj.current_projects         = ep_raw.get('current_projects', ep_obj.current_projects)
                    ep_obj.save()
                    changes['engineer_profile'] = 'updated'

            profile.save()
            
            # Create audit log (only with serializable data)
            create_audit_log(
                user=request.user,
                action='update_profile',
                resource_type='UserProfile',
                resource_id=profile.id,
                resource_repr=f'{user.email}',
                changes=changes,
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            # Return updated profile
            serializer = self.get_serializer(profile)
            response_data = serializer.data
            print(f"[DEBUG] Profile response - phone: {response_data.get('phone')}, profile_photo: {response_data.get('profile_photo')}")
            return Response(response_data)
            
        except Exception as e:
            import traceback
            print(f"[ERROR] Failed to update profile: {str(e)}")
            print(traceback.format_exc())
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get user statistics"""
        queryset = self.get_queryset()
        
        stats = {
            'total_users': queryset.count(),
            'active_users': queryset.filter(status='active').count(),
            'inactive_users': queryset.filter(status='inactive').count(),
            'suspended_users': queryset.filter(status='suspended').count(),
            'by_organization': list(queryset.values('organization__name').annotate(count=Count('id'))),
        }
        
        return Response(stats)
    
    @action(detail=False, methods=['get'])
    def my_features(self, request):
        """Get list of features current user has access to"""
        from .utils import get_user_accessible_features
        
        features = get_user_accessible_features(request.user)
        return Response({
            'features': list(features.values()),
            'accessible_count': sum(1 for f in features.values() if f['accessible'])
        })

    @action(detail=False, methods=['get'])
    def engineers(self, request):
        """
        List engineers in the same organisation with their competency profiles.
        Used for project team building and assignment matching.

        Query params (all optional):
          discipline     — filter by engineering discipline (case-insensitive contains)
          expertise_level — filter by level code e.g. senior, lead
          available_only  — 'true' to show only available / partial engineers
          skill           — filter by technical skill name (contains)
        """
        try:
            profile = request.user.rbac_profile
        except UserProfile.DoesNotExist:
            return Response({'engineers': [], 'count': 0})

        queryset = UserProfile.objects.select_related(
            'user', 'engineer_profile'
        ).filter(
            is_deleted=False,
            status='active',
            organization=profile.organization,
        )

        # ── Query-param filters ────────────────────────────────────────────
        discipline     = (request.query_params.get('discipline') or '').strip().lower()
        expertise_lvl  = (request.query_params.get('expertise_level') or '').strip().lower()
        available_only = request.query_params.get('available_only', '').lower() == 'true'
        skill_filter   = (request.query_params.get('skill') or '').strip().lower()

        engineers_out = []
        for up in queryset:
            try:
                ep_obj = up.engineer_profile
                ep = ep_obj.to_dict()
            except Exception:
                ep = {}

            # Availability filter
            if available_only and ep.get('availability_status', 'available') not in ('available', 'partial'):
                continue

            # Expertise level filter
            if expertise_lvl and ep.get('expertise_level', '').lower() != expertise_lvl:
                continue

            # Discipline filter (any discipline contains the search string)
            if discipline:
                disciplines_lower = [d.lower() for d in ep.get('engineering_disciplines', [])]
                if not any(discipline in d for d in disciplines_lower):
                    continue

            # Skill filter
            if skill_filter:
                skills_lower = [s.get('name', '').lower() for s in ep.get('technical_skills', [])]
                if not any(skill_filter in s for s in skills_lower):
                    continue

            # Build profile photo URL
            photo_url = None
            if up.profile_photo:
                try:
                    url = up.profile_photo.url
                    if not url.startswith('http'):
                        url = request.build_absolute_uri(url)
                    photo_url = url
                except Exception:
                    pass

            engineers_out.append({
                'id':                    str(up.id),
                'name':                  f"{up.user.first_name} {up.user.last_name}".strip() or up.user.email,
                'email':                 up.user.email,
                'job_title':             up.job_title,
                'department':            up.department,
                'location':              up.location,
                'profile_photo':         photo_url,
                'expertise_level':       ep.get('expertise_level', ''),
                'years_experience':      ep.get('years_experience', ''),
                'engineering_disciplines': ep.get('engineering_disciplines', []),
                'technical_skills':      ep.get('technical_skills', []),
                'certifications':        ep.get('certifications', []),
                'availability_status':   ep.get('availability_status', 'available'),
                'availability_percentage': ep.get('availability_percentage', 100),
                'preferred_project_types': ep.get('preferred_project_types', []),
                'languages':             ep.get('languages', []),
            })

        return Response({'engineers': engineers_out, 'count': len(engineers_out)})

    @action(detail=True, methods=['post'], url_path='assign-modules')
    def assign_modules(self, request, pk=None):
        """
        Assign modules to a user by updating their role's module access
        Body: { "module_codes": ["pid_analysis", "pfd", "qhse"] }
        """
        profile = self.get_object()
        module_codes = request.data.get('module_codes', [])

        if not module_codes:
            return Response(
                {'error': 'module_codes is required (array of module codes)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get modules
        modules = Module.objects.filter(code__in=module_codes, is_active=True)
        if modules.count() != len(module_codes):
            found_codes = set(modules.values_list('code', flat=True))
            missing_codes = set(module_codes) - found_codes
            return Response(
                {'error': f'Some modules not found: {list(missing_codes)}'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get user's primary role (or create custom role)
        user_roles = UserRole.objects.filter(user_profile=profile, is_primary=True)

        if not user_roles.exists():
            user_roles = UserRole.objects.filter(user_profile=profile)

        if not user_roles.exists():
            return Response(
                {'error': 'User has no roles assigned. Please assign a role first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        assigned_count = 0
        for user_role in user_roles:
            role = user_role.role
            for module in modules:
                _, created = RoleModule.objects.get_or_create(
                    role=role,
                    module=module,
                    defaults={'granted_by': request.user}
                )
                if created:
                    assigned_count += 1

        create_audit_log(
            user=request.user,
            action='modules_assign',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            metadata={'module_codes': module_codes, 'assigned_count': assigned_count},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )

        return Response({
            'status': 'modules assigned',
            'user': profile.user.email,
            'modules': [m.name for m in modules],
            'assigned_count': assigned_count
        })

    @action(detail=False, methods=['post'], url_path='bulk-assign-modules')
    def bulk_assign_modules(self, request):
        """
        Bulk assign modules to multiple users
        Body: {
            "user_ids": ["uuid1", "uuid2"],  // or "user_emails": ["email1", "email2"]
            "module_codes": ["pid_analysis", "pfd", "qhse"],
            "all_users": true  // Optional: assign to ALL users in system
        }
        """
        user_ids = request.data.get('user_ids', [])
        user_emails = request.data.get('user_emails', [])
        module_codes = request.data.get('module_codes', [])
        all_users = request.data.get('all_users', False)
        
        if not module_codes:
            return Response(
                {'error': 'module_codes is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # If all_users is True, ignore user_ids and user_emails
        if not all_users and not user_ids and not user_emails:
            return Response(
                {'error': 'Either user_ids, user_emails, or all_users=true is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get modules
        modules = Module.objects.filter(code__in=module_codes, is_active=True)
        if modules.count() != len(module_codes):
            found_codes = set(modules.values_list('code', flat=True))
            missing_codes = set(module_codes) - found_codes
            return Response(
                {'error': f'Some modules not found: {list(missing_codes)}'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get user profiles
        if all_users:
            # Assign to ALL active users in the system
            profiles = UserProfile.objects.filter(is_deleted=False)
        elif user_ids:
            profiles = UserProfile.objects.filter(id__in=user_ids, is_deleted=False)
        else:
            profiles = UserProfile.objects.filter(user__email__in=user_emails, is_deleted=False)
        
        if not profiles.exists():
            return Response(
                {'error': 'No users found matching criteria'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Assign modules to each user's roles
        results = {
            'success': [],
            'failed': [],
            'total_assignments': 0
        }
        
        for profile in profiles:
            try:
                user_roles = UserRole.objects.filter(user_profile=profile)
                
                if not user_roles.exists():
                    results['failed'].append({
                        'user': profile.user.email,
                        'reason': 'No roles assigned'
                    })
                    continue
                
                assigned_count = 0
                for user_role in user_roles:
                    role = user_role.role
                    for module in modules:
                        role_module, created = RoleModule.objects.get_or_create(
                            role=role,
                            module=module,
                            defaults={'granted_by': request.user}
                        )
                        if created:
                            assigned_count += 1
                
                results['success'].append({
                    'user': profile.user.email,
                    'user_id': str(profile.id),
                    'modules_assigned': assigned_count
                })
                results['total_assignments'] += assigned_count
                
            except Exception as e:
                results['failed'].append({
                    'user': profile.user.email,
                    'reason': str(e)
                })
        
        # Create audit log
        create_audit_log(
            user=request.user,
            action='bulk_modules_assign',
            resource_type='UserProfile',
            resource_id=None,
            resource_repr='Bulk Module Assignment',
            metadata={
                'module_codes': module_codes,
                'success_count': len(results['success']),
                'failed_count': len(results['failed']),
                'total_assignments': results['total_assignments']
            },
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({
            'message': 'Bulk module assignment completed',
            'summary': {
                'total_users_processed': len(results['success']) + len(results['failed']),
                'successful': len(results['success']),
                'failed': len(results['failed']),
                'total_module_assignments': results['total_assignments']
            },
            'details': results
        })
    
    @action(detail=False, methods=['get'], url_path='departments')
    def get_departments(self, request):
        """
        Get unique list of departments from UserProfile
        Returns: {
            "departments": ["Engineering", "Sales", ...]
        }
        """
        departments = UserProfile.objects.filter(
            is_deleted=False,
            department__isnull=False
        ).exclude(
            department__exact=''
        ).values_list('department', flat=True).distinct().order_by('department')
        
        return Response({
            'departments': list(departments),
            'count': len(departments)
        })
    
    @action(detail=False, methods=['get'], url_path='job-titles')
    def get_job_titles(self, request):
        """
        Get unique list of job titles from UserProfile
        Returns: {
            "job_titles": ["Engineer", "Manager", ...]
        }
        """
        job_titles = UserProfile.objects.filter(
            is_deleted=False,
            job_title__isnull=False
        ).exclude(
            job_title__exact=''
        ).values_list('job_title', flat=True).distinct().order_by('job_title')
        
        return Response({
            'job_titles': list(job_titles),
            'count': len(job_titles)
        })


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing audit logs
    Read-only access for admins
    """
    queryset = AuditLog.objects.select_related('user').all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['user_email', 'action', 'resource_type']
    ordering_fields = ['-timestamp']
    filterset_fields = ['action', 'resource_type', 'success']
    
    def get_queryset(self):
        """Filter logs based on organization for non-super-admins"""
        user = self.request.user
        queryset = AuditLog.objects.all()
        
        try:
            profile = user.rbac_profile
            if not profile.roles.filter(code='super_admin', is_active=True).exists():
                # Filter to organization users
                org_user_ids = UserProfile.objects.filter(
                    organization=profile.organization
                ).values_list('user_id', flat=True)
                queryset = queryset.filter(user_id__in=org_user_ids)
        except UserProfile.DoesNotExist:
            return AuditLog.objects.none()
        
        return queryset.order_by('-timestamp')
    
    @action(detail=False, methods=['get'])
    def user_activity(self, request):
        """Get activity logs for specific user"""
        user_id = request.query_params.get('user_id')
        if not user_id:
            return Response(
                {'error': 'user_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        logs = self.get_queryset().filter(user_id=user_id)[:50]
        serializer = self.get_serializer(logs, many=True)
        return Response(serializer.data)


class StorageViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing file storage with S3 integration
    """
    queryset = UserStorage.objects.filter(is_deleted=False)
    serializer_class = UserStorageSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['filename', 'file_type']
    ordering_fields = ['-created_at', 'file_size', 'download_count']
    filterset_fields = ['file_type', 'mime_type']
    
    def get_queryset(self):
        """Filter files based on user and organization"""
        user = self.request.user
        queryset = UserStorage.objects.filter(is_deleted=False)
        
        try:
            profile = user.userprofile
            # Super admins see all files
            if not profile.has_permission('file_view_all'):
                # Regular users see only their files and organization files
                queryset = queryset.filter(
                    Q(user_profile=profile) | Q(organization=profile.organization)
                )
        except UserProfile.DoesNotExist:
            return UserStorage.objects.none()
        
        return queryset.order_by('-created_at')
    
    @action(detail=False, methods=['post'])
    def generate_upload_url(self, request):
        """
        Generate pre-signed URL for uploading a file to S3
        
        POST /api/v1/rbac/storage/generate_upload_url/
        {
            "file_name": "drawing.pdf",
            "file_size": 1024000,
            "content_type": "application/pdf",
            "category": "pid_analysis",
            "tags": {"project": "ABC-123"}
        }
        """
        file_name = request.data.get('file_name')
        file_size = request.data.get('file_size')
        content_type = request.data.get('content_type')
        category = request.data.get('category', 'general')
        tags = request.data.get('tags', {})
        
        if not file_name or not file_size:
            return Response(
                {'error': 'file_name and file_size are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            s3_service = S3Service(organization=request.user.userprofile.organization)
            result = s3_service.generate_upload_url(
                user=request.user,
                file_name=file_name,
                file_size=file_size,
                content_type=content_type,
                category=category,
                tags=tags
            )
            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """
        Generate pre-signed URL for downloading a file from S3
        
        GET /api/v1/rbac/storage/{id}/download/
        """
        try:
            s3_service = S3Service(organization=request.user.userprofile.organization)
            result = s3_service.generate_download_url(
                storage_id=pk,
                user=request.user
            )
            return Response(result, status=status.HTTP_200_OK)
        except PermissionError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_403_FORBIDDEN
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def verify_upload(self, request, pk=None):
        """
        Verify that file was successfully uploaded to S3
        
        POST /api/v1/rbac/storage/{id}/verify_upload/
        {
            "checksum": "md5hash"
        }
        """
        checksum = request.data.get('checksum')
        
        try:
            s3_service = S3Service(organization=request.user.userprofile.organization)
            result = s3_service.verify_upload(
                storage_id=pk,
                user=request.user,
                checksum=checksum
            )
            return Response({'verified': result}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get storage statistics for current user
        
        GET /api/v1/rbac/storage/stats/
        """
        try:
            s3_service = S3Service(organization=request.user.userprofile.organization)
            stats = s3_service.get_storage_stats(user=request.user)
            return Response(stats, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """Soft delete file from storage"""
        try:
            storage = self.get_object()
            s3_service = S3Service(organization=request.user.userprofile.organization)
            s3_service.delete_file(storage_id=storage.id, user=request.user)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except PermissionError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_403_FORBIDDEN
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ============================================================================
# ANALYTICS VIEWSETS - AI-Powered Admin Features
# ============================================================================

from .analytics_models import (
    SystemMetrics, UserActivityAnalytics, SecurityAlert, PredictiveInsight,
    FeatureUsageAnalytics, ErrorLogAnalytics, SystemHealthCheck
)
from .analytics_serializers import (
    SystemMetricsSerializer, UserActivityAnalyticsSerializer, SecurityAlertSerializer,
    PredictiveInsightSerializer, FeatureUsageAnalyticsSerializer, ErrorLogAnalyticsSerializer,
    SystemHealthCheckSerializer, DashboardStatsSerializer, RealTimeActivitySerializer
)
from datetime import timedelta, datetime
from django.db.models import Avg, Sum, Count, Max, Min, F


class AnalyticsDashboardViewSet(viewsets.ViewSet):
    """
    AI-Powered Analytics Dashboard
    Comprehensive admin overview with real-time insights
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    
    @action(detail=False, methods=['get'])
    def overview(self, request):
        """
        Get comprehensive dashboard overview
        Includes system health, user stats, security alerts, and AI insights
        """
        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        
        # User Statistics
        total_users = UserProfile.objects.filter(is_deleted=False).count()
        active_today = UserActivityAnalytics.objects.filter(
            date=today, 
            login_count__gt=0
        ).count()
        
        # System Metrics (Latest)
        latest_metrics = SystemMetrics.objects.first()
        
        # Security Alerts
        active_alerts = SecurityAlert.objects.filter(status='new').count()
        critical_alerts = SecurityAlert.objects.filter(
            status='new',
            severity='critical'
        ).count()
        
        # AI Predictions
        active_predictions = PredictiveInsight.objects.filter(
            is_active=True,
            is_acknowledged=False
        ).count()
        high_impact = PredictiveInsight.objects.filter(
            is_active=True,
            is_acknowledged=False,
            impact_level='high'
        ).count()
        
        # Error Statistics
        errors_today = ErrorLogAnalytics.objects.filter(
            last_occurrence__date=today,
            status='open'
        ).count()
        critical_errors = ErrorLogAnalytics.objects.filter(
            status='open',
            severity='critical'
        ).count()
        
        # User Growth
        users_yesterday = UserProfile.objects.filter(
            created_at__date__lte=yesterday,
            is_deleted=False
        ).count()
        growth_rate = ((total_users - users_yesterday) / users_yesterday * 100) if users_yesterday > 0 else 0
        
        # System Health
        latest_health = SystemHealthCheck.objects.first()
        health_score = latest_health.health_score if latest_health else 100.0
        
        data = {
            'total_users': total_users,
            'active_users_today': active_today,
            'total_api_requests_today': latest_metrics.api_requests_count if latest_metrics else 0,
            'system_health_score': health_score,
            'avg_response_time_ms': latest_metrics.avg_response_time_ms if latest_metrics else 0,
            'success_rate_percentage': latest_metrics.success_rate_percentage if latest_metrics else 100,
            'active_connections': latest_metrics.active_connections if latest_metrics else 0,
            'active_alerts_count': active_alerts,
            'critical_alerts_count': critical_alerts,
            'active_predictions_count': active_predictions,
            'high_impact_insights_count': high_impact,
            'errors_today': errors_today,
            'critical_errors_count': critical_errors,
            'user_growth_percentage': round(growth_rate, 2),
            'engagement_trend': 'growing' if growth_rate > 0 else 'stable',
        }
        
        serializer = DashboardStatsSerializer(data)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def real_time_activity(self, request):
        """
        Get real-time activity feed
        Recent user actions, alerts, and system events
        """
        limit = int(request.query_params.get('limit', 20))
        
        # Get recent audit logs
        recent_audits = AuditLog.objects.select_related('user').order_by('-timestamp')[:limit]
        
        activities = []
        for audit in recent_audits:
            activities.append({
                'activity_type': audit.action,
                'user_email': audit.user_email,
                'description': f"{audit.action.title()} {audit.resource_type}",
                'timestamp': audit.timestamp,
                'severity': 'high' if not audit.success else 'normal',
                'metadata': {
                    'resource_id': audit.resource_id,
                    'success': audit.success,
                    'changes': audit.changes
                }
            })
        
        serializer = RealTimeActivitySerializer(activities, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def system_performance(self, request):
        """
        Get system performance metrics over time
        """
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now() - timedelta(days=days)
        
        metrics = SystemMetrics.objects.filter(
            timestamp__gte=start_date
        ).order_by('timestamp').values(
            'timestamp', 'avg_response_time_ms', 'success_rate_percentage',
            'cpu_usage_percentage', 'memory_usage_mb', 'active_connections',
            'api_requests_count', 'failed_requests_count'
        )
        
        return Response(list(metrics))
    
    @action(detail=False, methods=['get'])
    def user_engagement_trends(self, request):
        """
        Get user engagement trends and patterns
        """
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now().date() - timedelta(days=days)
        
        analytics = UserActivityAnalytics.objects.filter(
            date__gte=start_date
        ).values('date').annotate(
            total_logins=Sum('login_count'),
            avg_engagement=Avg('engagement_score'),
            avg_productivity=Avg('productivity_score'),
            users_with_anomalies=Count('id', filter=Q(anomaly_detected=True))
        ).order_by('date')
        
        return Response(list(analytics))


class SystemMetricsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for system performance metrics
    Read-only for admins to monitor system health
    """
    queryset = SystemMetrics.objects.all()
    serializer_class = SystemMetricsSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['timestamp']
    ordering = ['-timestamp']
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        """Get the most recent metrics"""
        latest = self.queryset.first()
        if latest:
            serializer = self.get_serializer(latest)
            return Response(serializer.data)
        return Response({})
    
    @action(detail=False, methods=['get'])
    def averages(self, request):
        """Get average metrics over a time period"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now() - timedelta(days=days)
        
        averages = self.queryset.filter(timestamp__gte=start_date).aggregate(
            avg_response_time=Avg('avg_response_time_ms'),
            avg_success_rate=Avg('success_rate_percentage'),
            avg_cpu=Avg('cpu_usage_percentage'),
            avg_memory=Avg('memory_usage_mb'),
            total_requests=Sum('api_requests_count'),
            total_failed=Sum('failed_requests_count')
        )
        
        return Response(averages)


class UserActivityAnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for user behavior analytics
    AI-powered insights into user patterns and engagement
    """
    queryset = UserActivityAnalytics.objects.select_related('user').all()
    serializer_class = UserActivityAnalyticsSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['date', 'anomaly_detected', 'usage_pattern']
    search_fields = ['user__email', 'user__first_name', 'user__last_name']
    ordering = ['-date']
    
    @action(detail=False, methods=['get'])
    def top_engaged_users(self, request):
        """Get users with highest engagement scores"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now().date() - timedelta(days=days)
        limit = int(request.query_params.get('limit', 10))
        
        top_users = self.queryset.filter(date__gte=start_date).values(
            'user__email', 'user__first_name', 'user__last_name'
        ).annotate(
            avg_engagement=Avg('engagement_score'),
            total_logins=Sum('login_count'),
            total_actions=Sum('drawings_uploaded') + Sum('analyses_completed')
        ).order_by('-avg_engagement')[:limit]
        
        return Response(list(top_users))
    
    @action(detail=False, methods=['get'])
    def anomalies(self, request):
        """Get users with detected anomalies"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now().date() - timedelta(days=days)
        
        anomalies = self.queryset.filter(
            date__gte=start_date,
            anomaly_detected=True
        ).select_related('user').order_by('-risk_score')
        
        serializer = self.get_serializer(anomalies, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def user_timeline(self, request, pk=None):
        """Get activity timeline for a specific user"""
        user_id = pk
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now().date() - timedelta(days=days)
        
        timeline = self.queryset.filter(
            user_id=user_id,
            date__gte=start_date
        ).order_by('date')
        
        serializer = self.get_serializer(timeline, many=True)
        return Response(serializer.data)


class SecurityAlertViewSet(viewsets.ModelViewSet):
    """
    ViewSet for security alerts
    AI-powered threat detection and management
    """
    queryset = SecurityAlert.objects.select_related('user', 'resolved_by').all()
    serializer_class = SecurityAlertSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['severity', 'status', 'alert_type']
    search_fields = ['title', 'description', 'user__email']
    ordering = ['-detection_time']
    
    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """Mark alert as resolved"""
        alert = self.get_object()
        alert.status = 'resolved'
        alert.resolved_at = timezone.now()
        alert.resolved_by = request.user
        alert.resolution_notes = request.data.get('notes', '')
        alert.save()
        
        serializer = self.get_serializer(alert)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def investigate(self, request, pk=None):
        """Mark alert as under investigation"""
        alert = self.get_object()
        alert.status = 'investigating'
        alert.save()
        
        serializer = self.get_serializer(alert)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def critical(self, request):
        """Get all critical unresolved alerts"""
        critical_alerts = self.queryset.filter(
            severity='critical',
            status__in=['new', 'investigating']
        )
        
        serializer = self.get_serializer(critical_alerts, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get security alert statistics"""
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)
        
        stats = {
            'total_alerts': self.queryset.filter(detection_time__gte=start_date).count(),
            'by_severity': dict(self.queryset.filter(
                detection_time__gte=start_date
            ).values('severity').annotate(count=Count('id')).values_list('severity', 'count')),
            'by_status': dict(self.queryset.filter(
                detection_time__gte=start_date
            ).values('status').annotate(count=Count('id')).values_list('status', 'count')),
            'resolution_time_avg_hours': 0,  # Calculate from resolved alerts
        }
        
        return Response(stats)


class PredictiveInsightViewSet(viewsets.ModelViewSet):
    """
    ViewSet for AI-generated predictions and insights
    Machine learning powered recommendations
    """
    queryset = PredictiveInsight.objects.select_related('acknowledged_by').all()
    serializer_class = PredictiveInsightSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['insight_type', 'impact_level', 'is_active', 'is_acknowledged']
    ordering = ['-created_at']
    
    @action(detail=True, methods=['post'])
    def acknowledge(self, request, pk=None):
        """Acknowledge an insight"""
        insight = self.get_object()
        insight.is_acknowledged = True
        insight.acknowledged_by = request.user
        insight.acknowledged_at = timezone.now()
        insight.save()
        
        serializer = self.get_serializer(insight)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get unacknowledged insights"""
        pending = self.queryset.filter(
            is_active=True,
            is_acknowledged=False
        ).order_by('-confidence_score')
        
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def high_priority(self, request):
        """Get high-impact unacknowledged insights"""
        high_priority = self.queryset.filter(
            is_active=True,
            is_acknowledged=False,
            impact_level='high'
        ).order_by('-confidence_score')
        
        serializer = self.get_serializer(high_priority, many=True)
        return Response(serializer.data)


class FeatureUsageAnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for feature usage analytics
    Track feature adoption and health
    """
    queryset = FeatureUsageAnalytics.objects.all()
    serializer_class = FeatureUsageAnalyticsSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['feature_name', 'date', 'trend']
    ordering = ['-date']
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get summary of all features"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now().date() - timedelta(days=days)
        
        summary = self.queryset.filter(date__gte=start_date).values(
            'feature_name'
        ).annotate(
            avg_adoption_rate=Avg('adoption_rate_percentage'),
            avg_health_score=Avg('health_score'),
            total_users=Sum('active_users'),
            total_usage=Sum('total_usage_count')
        ).order_by('-total_usage')
        
        return Response(list(summary))
    
    @action(detail=False, methods=['get'])
    def trending(self, request):
        """Get trending features"""
        trending = self.queryset.filter(
            trend='growing'
        ).order_by('-growth_rate_percentage')[:10]
        
        serializer = self.get_serializer(trending, many=True)
        return Response(serializer.data)


class ErrorLogAnalyticsViewSet(viewsets.ModelViewSet):
    """
    ViewSet for error analytics
    AI-powered error tracking and root cause analysis
    """
    queryset = ErrorLogAnalytics.objects.all()
    serializer_class = ErrorLogAnalyticsSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['severity', 'status', 'error_type']
    search_fields = ['error_type', 'error_message']
    ordering = ['-last_occurrence']
    
    @action(detail=True, methods=['post'])
    def mark_resolved(self, request, pk=None):
        """Mark error as resolved"""
        error = self.get_object()
        error.status = 'resolved'
        error.resolution_notes = request.data.get('notes', '')
        error.resolved_at = timezone.now()
        error.save()
        
        serializer = self.get_serializer(error)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def critical_errors(self, request):
        """Get all critical unresolved errors"""
        critical = self.queryset.filter(
            severity='critical',
            status='open'
        ).order_by('-occurrence_count')
        
        serializer = self.get_serializer(critical, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get error statistics"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now() - timedelta(days=days)
        
        stats = {
            'total_errors': self.queryset.filter(last_occurrence__gte=start_date).count(),
            'total_occurrences': self.queryset.filter(
                last_occurrence__gte=start_date
            ).aggregate(total=Sum('occurrence_count'))['total'] or 0,
            'by_severity': dict(self.queryset.filter(
                last_occurrence__gte=start_date
            ).values('severity').annotate(count=Count('id')).values_list('severity', 'count')),
            'affected_users': self.queryset.filter(
                last_occurrence__gte=start_date
            ).aggregate(total=Sum('affected_users_count'))['total'] or 0,
        }
        
        return Response(stats)


class SystemHealthCheckViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for system health monitoring
    Real-time system status and diagnostics
    """
    queryset = SystemHealthCheck.objects.all()
    serializer_class = SystemHealthCheckSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    ordering = ['-check_time']
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        """Get latest health check"""
        latest = self.queryset.first()
        if latest:
            serializer = self.get_serializer(latest)
            return Response(serializer.data)
        return Response({})
    
    @action(detail=False, methods=['get'])
    def history(self, request):
        """Get health check history"""
        hours = int(request.query_params.get('hours', 24))
        start_time = timezone.now() - timedelta(hours=hours)
        
        history = self.queryset.filter(check_time__gte=start_time).order_by('check_time')
        serializer = self.get_serializer(history, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def component_status(self, request):
        """Get current status of all system components"""
        latest = self.queryset.first()
        if latest:
            return Response({
                'database': latest.database_status,
                'redis': latest.redis_status,
                'celery': latest.celery_status,
                'storage': latest.storage_status,
                'api': latest.api_status,
                'overall': latest.overall_status,
                'health_score': latest.health_score,
                'check_time': latest.check_time,
            })
        return Response({})


# ===========================================================================
# STANDALONE: User Export View — no ViewSet inheritance, no router dependency
# GET /api/v1/rbac/users/export/?file_format=csv   → CSV download
# GET /api/v1/rbac/users/export/?file_format=xlsx  → Excel download
# NOTE: param is 'file_format' (not 'format') to avoid DRF content-negotiation
#       interception — DRF treats ?format=xxx as a renderer override and raises
#       Http404 when no renderer matches (e.g. 'csv' is not a registered renderer).
# ===========================================================================
class UserExportView(APIView):
    permission_classes = [IsAuthenticated, CanManageUsers]

    # Soft-coded export configuration
    EXPORT_HEADERS = [
        'First Name', 'Last Name', 'Email', 'Username',
        'Department', 'Job Title', 'Phone', 'Employee ID',
        'Location', 'Status', 'Roles', 'Organization',
        'Created At', 'Last Login',
    ]
    HEADER_COLOR = '1E40AF'   # Blue header for Excel
    MAX_COL_WIDTH = 40

    def get(self, request):
        import csv
        import datetime
        from io import BytesIO
        from django.http import HttpResponse

        # 'file_format' param — intentionally NOT 'format' to avoid DRF content-negotiation
        # interception: DRF treats ?format=xxx as a renderer format override and raises
        # Http404 when no renderer with that format exists (e.g. 'csv' has no renderer).
        export_format = request.query_params.get('file_format', 'csv').lower()
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

        # Build queryset scoped to the requester's organization (super_admin sees all)
        queryset = UserProfile.objects.select_related(
            'user', 'organization'
        ).prefetch_related(
            'roles'
        ).filter(is_deleted=False)

        try:
            profile = request.user.rbac_profile
            if not profile.roles.filter(code='super_admin', is_active=True).exists():
                queryset = queryset.filter(organization=profile.organization)
        except UserProfile.DoesNotExist:
            queryset = UserProfile.objects.none()

        queryset = queryset.order_by('created_at')

        def build_row(p):
            usr = p.user
            roles = ', '.join(sorted(set(r.name for r in p.roles.all())))
            created = p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else ''
            last_login = p.last_login_at.strftime('%Y-%m-%d %H:%M') if p.last_login_at else ''
            return [
                usr.first_name or '', usr.last_name or '',
                usr.email or '', usr.username or '',
                p.department or '', p.job_title or '',
                p.phone or '', p.employee_id or '',
                p.location or '', p.status or '',
                roles,
                p.organization.name if p.organization else '',
                created, last_login,
            ]

        if export_format == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Users'

            header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            ws.append(self.EXPORT_HEADERS)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for p in queryset:
                ws.append(build_row(p))

            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, self.MAX_COL_WIDTH)

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="users_export_{timestamp}.xlsx"'
        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="users_export_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(self.EXPORT_HEADERS)
            for p in queryset:
                writer.writerow(build_row(p))

        return response
