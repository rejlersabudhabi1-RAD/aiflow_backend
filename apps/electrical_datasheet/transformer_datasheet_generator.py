"""
Power/Distribution Transformer Datasheet Generator
Extracts equipment data from Transformer Sizing Calculation documents
and generates comprehensive datasheets matching the standard ADNOC form.

Columns: SI No. | DESCRIPTION | UNIT | SPECIFIED DESIGN DATA | VENDOR DATA | Rev
"""
import logging
import json
from typing import Dict, List, Optional
from django.conf import settings
from openai import OpenAI
import PyPDF2

logger = logging.getLogger(__name__)


class TransformerDatasheetGenerator:
    """Generate Power/Distribution Transformer datasheets from sizing calculation documents."""

    def __init__(self):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)

    # ──────────────────────────────────────────────────────────────────────────
    # PDF Extraction
    # ──────────────────────────────────────────────────────────────────────────
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extract all text from an uploaded PDF file."""
        try:
            pdf_file.seek(0)
            reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            logger.info(f"[TransformerDatasheet] PDF has {len(reader.pages)} pages")
            for i, page in enumerate(reader.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                    logger.info(f"[TransformerDatasheet] Page {i}: {len(page_text)} chars")
                else:
                    logger.warning(f"[TransformerDatasheet] Page {i}: no text (image-based?)")
            logger.info(f"[TransformerDatasheet] Total: {len(text)} chars")
            return text
        except Exception as e:
            logger.error(f"[TransformerDatasheet] PDF extraction error: {e}", exc_info=True)
            return ""

    # ──────────────────────────────────────────────────────────────────────────
    # Main entry point
    # ──────────────────────────────────────────────────────────────────────────
    def generate_datasheet_from_sizing_calc(self, pdf_file, project_info: Dict = None) -> Dict:
        """
        Generate a transformer datasheet from a sizing calculation PDF.

        Returns:
            {
                'success': bool,
                'datasheet_rows': List[Dict],   # sr_no, description, unit, required_data, vendor_data, rev
                'summary': Dict,
                'extraction_metadata': Dict
            }
        """
        try:
            logger.info("[TransformerDatasheet] Extracting text from sizing calculation PDF…")
            doc_text = self.extract_text_from_pdf(pdf_file)

            if not doc_text or len(doc_text) < 20:
                logger.error(f"[TransformerDatasheet] Insufficient text: {len(doc_text) if doc_text else 0} chars")
                return {
                    'success': False,
                    'error': (
                        'Could not extract text from the PDF. '
                        'The file may be image-based or empty. '
                        'Please provide a text-based transformer sizing calculation document.'
                    )
                }

            logger.info("[TransformerDatasheet] Analysing with AI…")
            datasheet_rows = self._extract_datasheet_with_ai(doc_text, project_info)

            if not datasheet_rows:
                logger.warning("[TransformerDatasheet] AI returned no data – falling back to template")
                datasheet_rows = self._get_default_datasheet_template()

            summary = {
                'total_rows': len(datasheet_rows),
                'equipment_count': sum(1 for r in datasheet_rows if r.get('description', '').strip()),
                'completed_fields': sum(1 for r in datasheet_rows if r.get('vendor_data', '').strip()),
                'missing_fields': sum(1 for r in datasheet_rows if not r.get('vendor_data', '').strip()),
            }

            logger.info(f"[TransformerDatasheet] ✅ Generated {summary['total_rows']} rows")
            return {
                'success': True,
                'datasheet_rows': datasheet_rows,
                'summary': summary,
                'extraction_metadata': {
                    'document_length': len(doc_text),
                    'project_info': project_info or {},
                }
            }

        except Exception as e:
            logger.error(f"[TransformerDatasheet] Error: {e}", exc_info=True)
            return {'success': False, 'error': str(e)}

    # ──────────────────────────────────────────────────────────────────────────
    # AI extraction
    # ──────────────────────────────────────────────────────────────────────────
    def _extract_datasheet_with_ai(self, doc_text: str, project_info: Dict = None) -> List[Dict]:
        """Use GPT-4o to extract transformer datasheet data from a sizing calculation document."""

        prompt = f"""You are a senior electrical engineer specialising in power and distribution transformers.
Analyse the provided Transformer Sizing Calculation document and extract comprehensive datasheet information.

PROJECT INFORMATION:
{json.dumps(project_info or {}, indent=2)}

DOCUMENT CONTENT:
{doc_text[:7000]}

TASK:
Populate a standard transformer datasheet with EXACTLY 6 fields per row, matching the ADNOC/IEC datasheet form:
- SI_NO      : Sequential item number (e.g. 1, 2, A, A.1, A.2 …). Blank for section-header rows.
- DESCRIPTION: Parameter name or section heading.
- UNIT       : Engineering unit (MVA, kV, A, Hz, ℃, %, mm, kg, dB, etc.). Blank if not applicable.
- REQUIRED_DATA : Specified design data – the required/design value (engineer-filled column).
- VENDOR_DATA   : Value extracted from the uploaded sizing-calculation document; empty string "" if not found.
- REV        : Revision marker – empty string "" unless explicitly noted in the document.

Cover ALL sections listed below in order:

A – GENERAL PARTICULARS
  Tag No., Title, Manufacturer / Country of Origin, Year of Manufacture, Quantity,
  Rating, Project Specification, Standards (IEC 60076 series), Design Life, Criticality Rating,
  Inspection Class, Material Certification

B – ENVIRONMENTAL CONDITIONS
  Type of Installation, Altitude, Max Ambient Temperature, Min Ambient Temperature,
  Max Relative Humidity (at 45 °C / at 54 °C), Degree of Protection (IP), Special Conditions

C – GENERAL CHARACTERISTICS
  Rated Power, Rated Secondary Voltage at No Load,
  Rated Voltage at Rated Power and P.F. 0.8, Rated Frequency, Vector Group,
  Connection Symbol and Vector Group, Maximum Flux Density, Number of Windings,
  With Separate Windings, Type of Cooling, Type of Tap Changer / Tapping / NER

D – INSULATION SYSTEMS
  Isolation Quality, Uniform Insulation,
  Power Frequency Withstand Voltage – Primary, Power Frequency Withstand Voltage – Secondary,
  Unearthed Transformer, Zero Sequence Impedance, Positive Sequence Voltage, Zero End Ratio

E – MODE OF OPERATION

F – PRIMARY WINDING
  Voltage, Material (Copper), Max Current Density in Winding, Rated Primary Current

G – SECONDARY WINDING
  High Voltage, Material (Copper), Additional Neutral in Separate Box,
  Earthing System, Max Current Density in Winding, Rated Primary Current

H – ELECTRICAL AND MECHANICAL CHARACTERISTICS
  No-Load Current (Primary), Magnetising Inrush Current & Duration,
  Short Circuit Impedance at Principal Tap, Short Circuit Impedance at Maximum Tap,
  Tolerance on Short Circuit Impedance, Zero Sequence Impedance, Positive Sequence Ratio,
  Primary System Apparent Short Circuit Rating, Max Short Circuit Duration,
  Top Oil Temperature Rise, Average Winding Temperature Rise, Hot Spot Temperature,
  Iron Losses (No Load), Copper Losses (Full Load), Total Losses,
  Efficiency at 0.9 PF – 50% Load / 75% Load / 100% Load,
  Voltage Regulation at 0.9 PF, Max Efficiency at % Load

I – TAP CHANGERS
  Series Parallel, On-Load, No. of Steps, Tapping Step, Tapping Range,
  Voltage Regulator & Parallel Control System

TANK
  Main Material, Thickness of Tank – Sides / Bottom / Radiators,
  Type of Tank (Sealed / Conservator), Radiator Mounting

J – TANK COVER TYPE
  Bolted, Welded, Bell Type, Thickness, Dimensions (L × W × H)

M – WEIGHTS
  Core & Winding, Oil, Tank & Fittings, Volume of Oil, Make of Oil

N – NOISE LEVEL
  Without Cooling, With Cooling

O – CONNECTIONS
  Primary Voltage Side – Cable Connection, Cable Size, Qty of Bushings,
    Plug-in CT, Pull & Test Facility, Cable Box with Oil, Air Cooled,
    Thermal Image Window, Pressure Relief Dampeners
  Secondary Voltage Side – Cable Connection, Cable Size
  Neutral End – Cable Terminal, Cable Size, Thermal Image Window, Pressure Relief Dampeners
  Cooling System – Thermal, Fans & Associated Contactors, Rated Power, Rated Frequency

P – CONTROL AND PROTECTION DEVICES
  Buchholz Relay (no trip contact form C), Buchholz Relay (2 alarm + 2 trip contacts),
  Oil Temperature Indicator, Thermal Image Winding Temperature (2 alarm + 2 trip),
  Oil Temp Indicator with Contacts (2 alarm + 2 trip),
  Winding Temp Indicator with Contacts (2 alarm + 2 trip),
  Thermometer Pockets, Thermowell, Thermistors,
  Liquid Level Gauge (2 contacts alarm/trip), Pressure Relief Valve (2-stage contacts),
  Magnetic Oil Level Indicator, Pressure Vacuum Gauge (4 contacts),
  Primary Phase CT for Transformer & Line Differential Protection,
  Current Transformer for Restricted EF (BREF)

R – ACCESSORIES
  Surge Arrester, Surge Suppression at Primary Side, Air Dryer,
  Filling Eyes & Jacking Lugs, Pulling Eyes, Tank Access Ladder,
  Safety Valve on Tank & Radiators, Filling Valve, Sampling / Drain Valve,
  Pre-filter Isolating Valve, Earth Connection, Marshalling Box

S – INSPECTION & TESTING
  Inspection, Routine Tests, Type Tests & Acoustic Sound Tests, Special Tests

PAINT / COLOUR SPECIFICATION
  Painting, Colour, Painting Thickness – Tank / Radiator, Oil Saturation Thickness

Return ONLY a JSON array. Each element must have exactly these keys:
  "sr_no", "description", "unit", "required_data", "vendor_data", "rev"

Rules:
- Section header rows: sr_no = "", unit = "", required_data = "", vendor_data = "", rev = ""
- Extract ACTUAL values from the document for vendor_data; use "" when not found
- required_data = standard/typical requirement value for a power transformer per IEC 60076 / ADNOC specs
- rev = "" always (unless document specifies a revision letter)
- Include every parameter listed above even if vendor_data is empty
- Return ONLY the JSON array – no markdown, no explanation"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are an expert electrical engineer specialising in "
                            "power transformer datasheets per IEC 60076 and ADNOC standards. "
                            "Return only valid JSON arrays."
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=6000,
            )

            ai_response = response.choices[0].message.content.strip()

            # Strip markdown fences if present
            if "```json" in ai_response:
                ai_response = ai_response.split("```json")[1].split("```")[0]
            elif "```" in ai_response:
                ai_response = ai_response.split("```")[1].split("```")[0]

            datasheet_rows = json.loads(ai_response.strip())

            if isinstance(datasheet_rows, list) and len(datasheet_rows) > 0:
                for i, row in enumerate(datasheet_rows):
                    row.setdefault("sr_no", "")
                    row.setdefault("description", "")
                    row.setdefault("unit", "")
                    row.setdefault("required_data", "")
                    row.setdefault("vendor_data", "")
                    row.setdefault("rev", "")
                    row.pop("remarks", None)  # remove legacy key if present
                logger.info(f"[TransformerDatasheet] AI returned {len(datasheet_rows)} rows")
                return datasheet_rows
            else:
                logger.error("[TransformerDatasheet] Invalid AI response structure")
                return self._get_default_datasheet_template()

        except json.JSONDecodeError as e:
            logger.error(f"[TransformerDatasheet] JSON decode error: {e}")
            return self._get_default_datasheet_template()
        except Exception as e:
            logger.error(f"[TransformerDatasheet] AI extraction error: {e}")
            return self._get_default_datasheet_template()

    # ──────────────────────────────────────────────────────────────────────────
    # Default template  (all sections from the ADNOC transformer datasheet form)
    # ──────────────────────────────────────────────────────────────────────────
    def _get_default_datasheet_template(self) -> List[Dict]:
        """Return the full standard transformer datasheet template (6-column)."""
        R = lambda sr, desc, unit="", req="": {
            "sr_no": sr, "description": desc, "unit": unit,
            "required_data": req, "vendor_data": "", "rev": ""
        }
        H = lambda desc: {
            "sr_no": "", "description": desc, "unit": "",
            "required_data": "", "vendor_data": "", "rev": ""
        }
        return [
            # ── A. GENERAL PARTICULARS ──────────────────────────────────────
            H("A   GENERAL PARTICULARS"),
            R("1",  "TAG NO."),
            R("2",  "TITLE",                          "",   "POWER TRANSFORMER"),
            R("3",  "MANUFACTURER / COUNTRY OF ORIGIN"),
            R("4",  "YEAR OF MANUFACTURE"),
            R("5",  "QUANTITY"),
            R("6",  "RATING",                         "MVA"),
            R("7",  "PROJECT SPECIFICATION"),
            R("8",  "STANDARDS",                      "",   "IEC 60076 SERIES"),
            R("9",  "DESIGN LIFE",                    "YR",  "25"),
            R("10", "CRITICALITY RATING"),
            R("11", "INSPECTION CLASS"),
            R("12", "MATERIAL CERTIFICATION"),

            # ── B. ENVIRONMENTAL CONDITIONS ────────────────────────────────
            H("B   ENVIRONMENTAL CONDITIONS"),
            R("1",  "TYPE OF INSTALLATION",           "",   "OUTDOOR IN SHADED AREA"),
            R("2",  "ALTITUDE",                       "m",  "< 100"),
            R("3",  "MAX AMBIENT TEMPERATURE",        "℃",  "50"),
            R("4",  "MINIMUM AMBIENT TEMPERATURE",    "℃",  "-5"),
            R("5",  "MAX RELATIVE HUMIDITY AT 45 ℃",  "%",  "95"),
            R("5a", "MAX RELATIVE HUMIDITY AT 54 ℃",  "%"),
            R("6",  "DEGREE OF PROTECTION (IP)",      "",   "IP55"),
            R("7",  "SPECIAL CONDITIONS",             "",   "TROPICALIZED"),

            # ── C. GENERAL CHARACTERISTICS ─────────────────────────────────
            H("C   GENERAL CHARACTERISTICS"),
            R("1",  "RATED POWER",                    "MVA"),
            R("2",  "RATED SECONDARY VOLTAGE AT NO LOAD",  "kV",  "11.5"),
            R("3",  "RATED VOLTAGE AT RATED POWER AND P.F. 0.8",  "kV"),
            R("4",  "RATED FREQUENCY",                "Hz",  "50"),
            R("5",  "VECTOR GROUP"),
            R("6",  "CONNECTION SYMBOL AND VECTOR GROUP"),
            R("7",  "MAXIMUM FLUX DENSITY"),
            R("8",  "NUMBER OF WINDINGS"),
            R("9",  "WITH SEPARATE WINDINGS"),
            R("10", "TYPE OF COOLING",                "",   "CORE TYPE"),
            R("11", "TYPE OF TAP CHANGER / TAPPING"),

            # ── D. INSULATION SYSTEMS ───────────────────────────────────────
            H("D   INSULATION SYSTEMS"),
            R("1",  "ISOLATION QUALITY"),
            R("2",  "UNIFORM INSULATION"),
            H("3   POWER FREQUENCY WITHSTAND VOLTAGE"),
            R("3.1","PRIMARY",                        "kV",  "70"),
            R("3.2","",                               "kV",  "28"),
            H("4   IMPULSE WITHSTAND VOLTAGE"),
            R("4.1","SECONDARY",                      "kV",  "28"),
            R("4.2","",                               "kV",  "75"),
            R("5",  "UNEARTHED TRANSFORMER"),
            R("6",  "ZERO SEQUENCE IMPEDANCE"),
            R("7",  "POSITIVE SEQUENCE VOLTAGE"),
            R("8",  "ZERO END RATIO"),

            # ── E. MODE OF OPERATION ────────────────────────────────────────
            H("E   MODE OF OPERATION"),
            R("1",  "OPERATION MODE",                 "",   "PARALLEL (REFER NOTE 10)"),

            # ── F. PRIMARY WINDING ──────────────────────────────────────────
            H("F   PRIMARY WINDING"),
            R("1",  "VOLTAGE",                        "kV"),
            R("2",  "MATERIAL",                       "",   "COPPER"),
            R("3",  "MAXIMUM CURRENT DENSITY IN THE WINDING"),
            R("4",  "RATED PRIMARY CURRENT",          "A"),

            # ── G. SECONDARY WINDING ────────────────────────────────────────
            H("G   SECONDARY WINDING"),
            R("1",  "HIGH VOLTAGE",                   "kV"),
            R("2",  "MATERIAL",                       "",   "COPPER"),
            R("3",  "ADDITIONAL NEUTRAL BROUGHT IN A SEPARATED BOX", "", "YES"),
            R("4",  "EARTHING SYSTEM",                "",   "RESISTIVE"),
            R("5",  "MAXIMUM CURRENT DENSITY IN THE WINDING"),
            R("6",  "RATED PRIMARY CURRENT",          "A"),

            # ── H. ELECTRICAL AND MECHANICAL CHARACTERISTICS ───────────────
            H("H   ELECTRICAL AND MECHANICAL CHARACTERISTICS"),
            R("1",  "NO LOAD CURRENT (PRIMARY)",      "%"),
            R("2",  "MAGNETISING INRUSH CURRENT AND DURATION"),
            H("3   SHORT CIRCUIT IMPEDANCE"),
            R("3.1","TRANSFORMER IMPEDANCE AT PRINCIPAL TAP", "%"),
            R("3.2","TRANSFORMER IMPEDANCE AT MAXIMUM TAP",   "%"),
            R("3.3","TOLERANCE ON SHORT CIRCUIT IMPEDANCE",   "%", "7.5"),
            R("4",  "ZERO SEQUENCE IMPEDANCE"),
            R("5",  "POSITIVE SEQUENCE VS RATIO"),
            R("6",  "ZERO END RATIO",                 "kA",  "40"),
            R("7",  "PRIMARY SYSTEM APPARENT SHORT CIRCUIT RATING", "kA"),
            R("8",  "MAX SHORT CIRCUIT DURATION",     "Sec"),
            R("9",  "TOP OIL TEMPERATURE RISE",       "℃",   "50"),
            R("10", "AVERAGE WINDING TEMPERATURE RISE","℃",  "50"),
            R("11", "HOT SPOT TEMPERATURE",           "℃"),
            R("12", "IRON LOSSES (NO LOAD)"),
            R("13", "COPPER LOSSES (FULL LOAD)"),
            R("14", "STRAY LOSSES"),
            R("15", "TOTAL LOSSES"),
            H("16  EFFICIENCY AT 0.9 POWER FACTOR"),
            R("16.1","50% LOAD",                      "%"),
            R("16.2","75% LOAD",                      "%"),
            R("16.3","100% LOAD",                     "%"),
            H("17  VOLTAGE REGULATION"),
            R("17.1","AT UNITY POWER FACTOR",         "%"),
            R("17.2","AT 0.9 POWER FACTOR",           "%"),
            R("18", "MAX EFFICIENCY AT % LOAD",       "%"),

            # ── I. TAP CHANGERS ─────────────────────────────────────────────
            H("I   TAP CHANGERS"),
            R("1",  "SERIES PARALLEL",                "",   "YES"),
            R("2",  "ON-LOAD",                        "Y/N"),
            R("3",  "NO. OF STEPS",                   "No."),
            R("4",  "TAPPING STEP",                   "%",   "2.5"),
            R("5",  "TAPPING RANGE",                  "%",   "± 10% (STEP OF 2.5%)"),
            R("6",  "VOLTAGE REGULATOR & PARALLEL CONTROL SYSTEM", "", "N/A"),

            # ── TANK ────────────────────────────────────────────────────────
            H("TANK"),
            R("1",  "MAIN MATERIAL (UNDER BASE)",     "mm",  "THICKNESS MIN. 12 MM"),
            R("2",  "THICKNESS OF TANK",              "mm"),
            R("3.1","SIDES",                          "mm"),
            R("3.2","BOTTOM",                         "mm"),
            R("4",  "RADIATORS",                      "mm"),
            R("5",  "TYPE OF TANK (SEALED / CONSERVATOR)", "", "CONSERVATOR"),
            R("6",  "RADIATOR MOUNTING",              "",   "BI-DIRECTIONAL ROLLERS"),

            # ── J. TANK COVER TYPE ──────────────────────────────────────────
            H("J   TANK COVER TYPE"),
            R("1",  "BOLTED",                         "",   "YES"),
            R("2",  "WELDED",                         "",   "NA"),
            R("3",  "BELL TYPE"),
            R("4",  "THICKNESS",                      "mm",  "AS PER BGS-EE-003"),
            H("DIMENSIONS"),
            R("5",  "OVERALL WITH ACCESSORIES (LENGTH × WIDTH × HEIGHT)", "mm"),

            # ── M. WEIGHTS ──────────────────────────────────────────────────
            H("M   WEIGHTS"),
            R("1",  "CORE AND WINDING",               "kg"),
            R("2",  "OIL",                            "LITER"),
            R("3",  "TANK AND FITTINGS",              "kg"),
            R("4",  "VOLUME OF OIL",                  "kg"),
            R("5",  "MAKE OF OIL"),

            # ── N. NOISE LEVEL ──────────────────────────────────────────────
            H("N   NOISE LEVEL"),
            R("1",  "WITHOUT COOLING",                "dB"),
            R("2",  "WITH COOLING",                   "dB"),

            # ── O. CONNECTIONS ──────────────────────────────────────────────
            H("O   CONNECTIONS"),
            H("PRIMARY VOLTAGE SIDE"),
            R("1.1","CABLE CONNECTION",               "",   "YES"),
            R("1.2","CABLE SIZE",                     "",   "4 × 1C × 500 Sqmm / Phase"),
            R("1.3","QUANTITY OF BUSHING AND RATING"),
            R("1.4","QUANTITY OF BUSHING"),
            R("1.5","PLUG IN TYPE OF CURRENT TRANSFORMER"),
            R("1.6","PULL & TEST FACILITY"),
            R("1.7","CABLE BOX WITH OIL",             "",   "NO"),
            R("1.8","AIR COOLED",                     "",   "HEAT INSULATED CABLE BOX"),
            R("1.9","THERMAL IMAGE WINDOW FOR CABLE INVESTIGATION", "", "REQUIRED"),
            R("1.10","PRESSURE RELIEF DAMPENERS",     "",   "REQUIRED"),
            R("1.11","PRESSURE RELIEF DAMPENERS - LINKS", "", "DETACHABLE"),
            H("SECONDARY VOLTAGE SIDE"),
            R("2.1","CABLE CONNECTION"),
            R("2.2","CABLE SIZE",                     "",   "4R × 1C × 500 Sqmm"),
            H("NEUTRAL END"),
            R("3.1","CABLE TERMINAL, IN A SEPARATE NEUTRAL TERMINAL BOX"),
            R("3.2","CABLE SIZE",                     "",   "4 × 1C × 500 Sqmm"),
            R("3.3","THERMAL IMAGE WINDOW FOR CABLE INVESTIGATION", "", "REQUIRED"),
            R("3.4","PRESSURE RELIEF DAMPENERS",      "",   "REQUIRED"),
            H("COOLING SYSTEM"),
            R("1",  "THERMAL"),
            R("2",  "FANS AND ASSOCIATED CONTACTORS"),
            R("3",  "RATED POWER"),
            R("4",  "RATED POWER – FREQUENCY"),

            # ── P. CONTROL AND PROTECTION DEVICES ──────────────────────────
            H("P   CONTROL AND PROTECTION DEVICES"),
            R("1",  "BUCHHOLZ RELAY – NO TRIP CONTACT FORM 'C'",               "", "YES"),
            R("2",  "BUCHHOLZ RELAY WITH TWO ALARM AND TWO TRIP CONTACTS",     "", "YES"),
            R("3",  "OIL TEMPERATURE INDICATOR",                               "", "YES"),
            R("4",  "THERMAL IMAGE TYPE WINDING TEMPERATURE WITH CONTACTS (TWO ALARM + TWO TRIP)", "", "YES"),
            R("5",  "OIL TEMP. INDICATOR WITH CONTACTS (TWO ALARM & TWO TRIP)", "", "YES"),
            R("6",  "WINDING TEMP. INDICATOR WITH CONTACTS (TWO ALARM & TWO TRIP)", "", "YES"),
            R("7",  "THERMO METER POCKETS",                                    "", "YES"),
            R("8",  "THERMOWELL",                                              "", "YES"),
            R("9",  "THERMISTORS"),
            R("10", "LIQUID LEVEL GAUGE WITH 2 CONTACTS (ALARM / TRIP)",       "", "YES"),
            R("11", "PRESSURE RELIEF VALVE WITH OPERATING 2 STAGE CONTACTS",  "", "YES"),
            R("12", "MAGNETIC OIL LEVEL INDICATOR – LEVEL GAUGE WITH TWO CONTACTS (ALARM / TRIP)"),
            R("13", "PRESSURE VACUUM GAUGE WITH OPERATING 4 CONTACTS (ALARM / TRIP)", "", "YES"),
            R("14", "PRIMARY PHASE CT FOR TRANSFORMER & LINE DIFFERENTIAL PROTECTION"),
            R("15", "CURRENT TRANSFORMER FOR RESTRICTED EF (BREF)"),

            # ── R. ACCESSORIES ──────────────────────────────────────────────
            H("R   ACCESSORIES"),
            R("1",  "SURGE ARRESTER",                 "",   "YES"),
            R("2",  "SURGE SUPPRESSING CAPABILITY AT PRIMARY SIDE"),
            R("3",  "AIR DRYER",                      "",   "NA"),
            R("4",  "FILLING EYES & JACKING LUGS"),
            R("5",  "PULLING EYES FOR MOVING TRANSFORMER IN ALL DIRECTIONS"),
            R("6",  "TANK ACCESS LADDER"),
            R("7",  "SAFETY VALVE ON TANK AND RADIATORS"),
            R("8",  "FILLING VALVE ON TANK AND RADIATORS"),
            R("9",  "SAMPLING VALVE (DRAIN VALVE) ON TANK AND RADIATORS"),
            R("10", "PRE-FILTER ISOLATING VALVE",     "",   "YES"),
            R("11", "ORIENTABLE",                     "",   "YES"),
            R("12", "EARTH CONNECTION"),
            R("13", "MARSHALLING BOX",                "",   "YES"),

            # ── S. INSPECTION & TESTING ─────────────────────────────────────
            H("S   INSPECTION & TESTING"),
            R("1",  "INSPECTION",                     "",   "AS PER SPEC. BGS-EE-003"),
            R("2",  "ROUTINE TESTS",                  "",   "AS PER APPENDIX-3 OF BGS-EE-003"),
            R("3",  "TYPE TESTS & ACOUSTIC SOUND TESTS", "", "AS PER APPENDIX-3 OF BGS-EE-003"),
            R("4",  "SPECIAL TESTS",                  "",   "AS PER APPENDIX-3 OF BGS-EE-003"),

            # ── PAINT / COLOUR SPECIFICATION ────────────────────────────────
            H("PAINT / COLOUR SPECIFICATION"),
            R("5",  "PAINTING",                       "",   "RAL 7035 AS PER BGS-EE-003"),
            R("6",  "COLOUR"),
            R("7",  "PAINTING THICKNESS TANK"),
            R("8",  "PAINTING THICKNESS RADIATOR"),
            R("9",  "OIL SATURATION THICKNESS MAGNETIC RADIATOR"),
        ]

    # ──────────────────────────────────────────────────────────────────────────
    # Excel Export
    # ──────────────────────────────────────────────────────────────────────────
    def export_to_excel(self, datasheet_rows: List[Dict], project_info: Dict = None):
        """Export transformer datasheet to formatted Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "Transformer Datasheet"

        header_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font   = Font(color="FFFFFF", bold=True, size=10)
        section_fill  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        section_font  = Font(bold=True, size=10)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row_idx = 1

        # Title row
        ws.merge_cells("A1:F1")
        ws["A1"] = "POWER / DISTRIBUTION TRANSFORMER – DATASHEET"
        ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        row_idx = 3

        # Project info
        if project_info:
            for key, val in project_info.items():
                ws.cell(row=row_idx, column=1, value=key.replace("_", " ").title())
                ws.cell(row=row_idx, column=2, value=val)
                row_idx += 1
            row_idx += 1

        # Column headers – 6 columns
        col_headers = ["SI NO.", "DESCRIPTION", "UNIT", "SPECIFIED DESIGN DATA", "VENDOR DATA", "Rev"]
        col_widths  = [9, 52, 10, 35, 30, 8]
        for col_idx, (header, width) in enumerate(zip(col_headers, col_widths), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[chr(64 + col_idx)].width = width

        row_idx += 1

        # Data rows
        for row_data in datasheet_rows:
            sr_no        = row_data.get("sr_no", "")
            description  = row_data.get("description", "")
            unit         = row_data.get("unit", "")
            req_data     = row_data.get("required_data", "")
            vendor_data  = row_data.get("vendor_data", "")
            rev          = row_data.get("rev", row_data.get("remarks", ""))

            is_section = (not sr_no) and description and not req_data and not vendor_data

            cells = [sr_no, description, unit, req_data, vendor_data, rev]
            aligns = ["center", "left", "center", "left", "left", "center"]

            for col_idx, (val, align) in enumerate(zip(cells, aligns), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
                if is_section:
                    cell.fill = section_fill
                    cell.font = section_font

            row_idx += 1

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
