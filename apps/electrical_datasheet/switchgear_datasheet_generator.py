"""
11KV Switchgear Datasheet Generator from SLD Documents
Extracts equipment data and generates comprehensive datasheets
"""
import logging
import re
import json
from typing import Dict, List, Optional
from django.conf import settings
from openai import OpenAI
import PyPDF2

logger = logging.getLogger(__name__)


class SwitchgearDatasheetGenerator:
    """Generate 11KV switchgear datasheets from SLD documents"""
    
    def __init__(self):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)
    
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extract text from uploaded PDF file"""
        try:
            # Reset file pointer to beginning
            pdf_file.seek(0)
            
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            
            logger.info(f"[SwitchgearDatasheet] PDF has {len(pdf_reader.pages)} pages")
            
            for page_num, page in enumerate(pdf_reader.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                    logger.info(f"[SwitchgearDatasheet] Page {page_num}: Extracted {len(page_text)} chars")
                else:
                    logger.warning(f"[SwitchgearDatasheet] Page {page_num}: No text extracted (might be image-based)")
            
            logger.info(f"[SwitchgearDatasheet] Total extracted text: {len(text)} characters")
            return text
        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] PDF extraction error: {e}", exc_info=True)
            return ""
    
    def generate_datasheet_from_sld(self, pdf_file, project_info: Dict = None) -> Dict:
        """
        Generate comprehensive 11KV switchgear datasheet from SLD PDF
        
        Args:
            pdf_file: Uploaded PDF file
            project_info: Optional project metadata
        
        Returns:
            {
                'success': bool,
                'datasheet_rows': List[Dict],
                'summary': Dict,
                'extraction_metadata': Dict
            }
        """
        try:
            # Extract text from PDF
            logger.info("[SwitchgearDatasheet] Extracting text from SLD PDF...")
            sld_text = self.extract_text_from_pdf(pdf_file)
            
            # More lenient text extraction check - if we have ANY text, try to process it
            if not sld_text or len(sld_text) < 20:
                logger.error(f"[SwitchgearDatasheet] Insufficient text: {len(sld_text) if sld_text else 0} chars")
                return {
                    'success': False,
                    'error': 'Could not extract text from PDF. The PDF might be image-based or empty. Please provide a text-based SLD document.'
                }
            
            logger.info(f"[SwitchgearDatasheet] Extracted {len(sld_text)} characters from PDF")
            
            # Use AI to extract structured datasheet information
            logger.info("[SwitchgearDatasheet] Analyzing SLD with AI...")
            datasheet_rows = self._extract_datasheet_with_ai(sld_text, project_info)
            
            if not datasheet_rows:
                logger.warning("[SwitchgearDatasheet] AI extraction returned no data, using template")
                # Fall back to template with extracted text hints
                datasheet_rows = self._get_default_datasheet_template()
            
            # Calculate summary statistics
            summary = {
                'total_rows': len(datasheet_rows),
                'equipment_count': sum(1 for row in datasheet_rows if row.get('description', '').strip()),
                'completed_fields': sum(1 for row in datasheet_rows if row.get('vendor_data', '').strip()),
                'missing_fields': sum(1 for row in datasheet_rows if not row.get('vendor_data', '').strip())
            }
            
            logger.info(f"[SwitchgearDatasheet] ✅ Generated {summary['total_rows']} datasheet rows")
            
            return {
                'success': True,
                'datasheet_rows': datasheet_rows,
                'summary': summary,
                'extraction_metadata': {
                    'document_length': len(sld_text),
                    'project_info': project_info or {}
                }
            }
            
        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] Error: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }
    
    def _extract_datasheet_with_ai(self, sld_text: str, project_info: Dict = None) -> List[Dict]:
        """Use AI to extract structured datasheet data from SLD text"""
        
        extraction_prompt = f"""You are an expert electrical engineer specializing in 11KV switchgear systems. 
Analyze the provided Single Line Diagram (SLD) document and extract comprehensive equipment datasheet information.

PROJECT INFORMATION:
{json.dumps(project_info or {}, indent=2)}

SLD DOCUMENT CONTENT:
{sld_text[:6000]}

TASK: Extract and structure ALL equipment data into a comprehensive datasheet format with EXACTLY 6 fields per row.

The datasheet MUST follow this exact column structure (same as the physical MV switchgear datasheet form):
- SR_NO: Sequential item number (e.g. 1.0, 1.1, 2, 2.1 ...) — blank for section header rows
- DESCRIPTION: Parameter name or section header
- UNIT: Engineering unit for the parameter (e.g. kV, A, kA, Hz, ℃, mm, kg, %) — blank if not applicable
- REQUIRED_DATA: Specification/standard requirement value filled by the engineer
- VENDOR_DATA: Actual value extracted from the uploaded SLD/document — blank string if not found
- REV: Revision marker — leave as empty string unless a specific revision is noted in the document

Cover ALL the following sections:
1. GENERAL — Equipment tag, service description
2. REFERENCE — Applicable international standards (IEC 60298, IEC 60694, IEC 60255, IEC 60529), ADNOC specs
3. SITE DATA — Location, area classification, climate, altitude, min/max ambient temperature, humidity
4. GENERAL CHARACTERISTICS — Type of switchgear, circuit breaker type, standards, system voltage, frequency, phases, earthing
5. RATINGS AND SHORT CIRCUIT DATA — Rated insulation voltage, rated voltage, rated normal current, SC breaking current, peak withstand current, short time withstand current, power frequency withstand voltage, impulse withstand voltage
6. CONSTRUCTION — Type, IP rating, colour, arc fault classification
7. BUSBAR — Material, shape, busbar rating
8. CIRCUIT BREAKER — Type, operating mechanism, auxiliary supply voltage, number of operating cycles
9. CURRENT TRANSFORMER — Type, number of CT cores, CT ratio, CT class
10. VOLTAGE TRANSFORMER — Type, VT ratio, VT class
11. EARTHING — Main earthing bar, earth fault relay
12. PROTECTION & CONTROL — Protection relay type, protection functions, metering
13. AUXILIARY EQUIPMENT — Anti-condensation heater, space heater rating, lighting
14. MANUFACTURER — Name, model/type, country of origin

Return your response as a JSON array where each object has EXACTLY this structure:
{{
    "sr_no": "<sequential number or empty string for section headers>",
    "description": "<parameter name or section header>",
    "unit": "<engineering unit or empty string>",
    "required_data": "<specification requirement value>",
    "vendor_data": "<value extracted from SLD/document, or empty string>",
    "rev": ""
}}

IMPORTANT GUIDELINES:
- ALWAYS include the "unit" field — use the correct SI/engineering unit for every measured quantity
- Common units: voltage → kV, current → A, breaking current → kA, frequency → Hz, temperature → ℃, dimensions → mm, weight → kg, percentage → %
- Section header rows have blank sr_no, blank unit, blank required_data, blank vendor_data, blank rev
- Extract ACTUAL values from the SLD document for vendor_data where available
- Leave vendor_data as empty string "" when value not found in document
- Leave rev as empty string "" unless document contains a specific revision reference
- Be comprehensive — include ALL standard 11KV switchgear parameters

Return ONLY the JSON array, no additional text."""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4-turbo-preview",
                messages=[
                    {"role": "system", "content": "You are an expert electrical engineer specializing in switchgear datasheets. Extract comprehensive equipment data and return only valid JSON."},
                    {"role": "user", "content": extraction_prompt}
                ],
                temperature=0.2,
                max_tokens=4000
            )
            
            ai_response = response.choices[0].message.content
            
            # Parse JSON response
            # Remove markdown code blocks if present
            if '```json' in ai_response:
                ai_response = ai_response.split('```json')[1].split('```')[0]
            elif '```' in ai_response:
                ai_response = ai_response.split('```')[1].split('```')[0]
            
            datasheet_rows = json.loads(ai_response.strip())
            
            # Validate structure
            if isinstance(datasheet_rows, list) and len(datasheet_rows) > 0:
                # Ensure all rows have required fields
                for i, row in enumerate(datasheet_rows):
                    if 'sr_no' not in row:
                        row['sr_no'] = str(i + 1)
                    if 'description' not in row:
                        row['description'] = ''
                    if 'unit' not in row:
                        row['unit'] = ''
                    if 'required_data' not in row:
                        row['required_data'] = ''
                    if 'vendor_data' not in row:
                        row['vendor_data'] = ''
                    if 'rev' not in row:
                        row['rev'] = ''
                    # Remove legacy 'remarks' key if present (replaced by 'rev')
                    row.pop('remarks', None)
                
                return datasheet_rows
            else:
                logger.error("[SwitchgearDatasheet] Invalid AI response structure")
                return self._get_default_datasheet_template()
                
        except json.JSONDecodeError as e:
            logger.error(f"[SwitchgearDatasheet] JSON decode error: {e}")
            logger.error(f"AI Response: {ai_response[:500]}")
            return self._get_default_datasheet_template()
        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] AI extraction error: {e}")
            return self._get_default_datasheet_template()
    
    def _get_default_datasheet_template(self) -> List[Dict]:
        """Return default 11KV switchgear datasheet template with 6 columns: sr_no, description, unit, required_data, vendor_data, rev"""
        return [
            {"sr_no": "",     "description": "GENERAL",                                        "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "1.0",  "description": "EQUIPMENT TAG NO.",                             "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "1.1",  "description": "SERVICE",                                       "unit": "",    "required_data": "11 KV SWITCHGEAR",                          "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "REFERENCE",                                     "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "1.2",  "description": "APPLICABLE INTERNATIONAL STANDARDS",            "unit": "",    "required_data": "IEC 60298, IEC 60694, IEC 60255, IEC 60529", "vendor_data": "", "rev": ""},
            {"sr_no": "1.3",  "description": "APPLICABLE SPEC./ADNOC-AGES",                   "unit": "",    "required_data": "ADNOC-AGES-SP-1031",                       "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "SITE DATA",                                     "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "1.4",  "description": "SITE LOCATION",                                 "unit": "",    "required_data": "ABU DHABI",                                "vendor_data": "", "rev": ""},
            {"sr_no": "1.5",  "description": "AREA CLASSIFICATION",                           "unit": "",    "required_data": "SAFE AREA",                                "vendor_data": "", "rev": ""},
            {"sr_no": "1.6",  "description": "CLIMATE CONDITIONS",                            "unit": "",    "required_data": "TROPICAL",                                 "vendor_data": "", "rev": ""},
            {"sr_no": "1.7",  "description": "SITE ALTITUDE",                                 "unit": "m",   "required_data": "< 100",                                    "vendor_data": "", "rev": ""},
            {"sr_no": "1.8",  "description": "MINIMUM AMBIENT TEMPERATURE",                   "unit": "℃",   "required_data": "-5",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "1.9",  "description": "MAXIMUM AMBIENT TEMPERATURE",                   "unit": "℃",   "required_data": "50",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "1.10", "description": "MAXIMUM RELATIVE HUMIDITY AT 40 ℃",             "unit": "%",   "required_data": "100",                                      "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "GENERAL CHARACTERISTICS",                       "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "2",    "description": "TYPE OF SWITCHGEAR",                             "unit": "",    "required_data": "METAL ENCLOSED",                            "vendor_data": "", "rev": ""},
            {"sr_no": "2.1",  "description": "TYPE OF CIRCUIT BREAKER",                       "unit": "",    "required_data": "VACUUM / SF6",                             "vendor_data": "", "rev": ""},
            {"sr_no": "2.2",  "description": "STANDARDS",                                     "unit": "",    "required_data": "IEC",                                      "vendor_data": "", "rev": ""},
            {"sr_no": "2.3",  "description": "SYSTEM VOLTAGE",                                "unit": "kV",  "required_data": "11",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "2.4",  "description": "SYSTEM FREQUENCY",                              "unit": "Hz",  "required_data": "50",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "2.5",  "description": "NUMBER OF PHASES",                              "unit": "",    "required_data": "3",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "2.6",  "description": "SYSTEM EARTHING",                               "unit": "",    "required_data": "RESISTANCE EARTHED",                       "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "RATINGS AND SHORT CIRCUIT DATA",                "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "3.1",  "description": "RATED INSULATION VOLTAGE",                      "unit": "kV",  "required_data": "12",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "3.2",  "description": "RATED VOLTAGE",                                 "unit": "kV",  "required_data": "11",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "3.3",  "description": "RATED NORMAL CURRENT",                          "unit": "A",   "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "3.4",  "description": "RATED SHORT CIRCUIT BREAKING CURRENT",          "unit": "kA",  "required_data": "25",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "3.5",  "description": "RATED PEAK WITHSTAND CURRENT",                  "unit": "kA",  "required_data": "65",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "3.6",  "description": "RATED SHORT TIME WITHSTAND CURRENT (3 SEC)",    "unit": "kA",  "required_data": "25",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "3.7",  "description": "RATED POWER FREQUENCY WITHSTAND VOLTAGE",       "unit": "kV",  "required_data": "28",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "3.8",  "description": "RATED IMPULSE WITHSTAND VOLTAGE (1.2/50 μs)",  "unit": "kV",  "required_data": "75",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "CONSTRUCTION",                                  "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "4.1",  "description": "TYPE",                                          "unit": "",    "required_data": "METAL CLAD / METAL ENCLOSED",             "vendor_data": "", "rev": ""},
            {"sr_no": "4.2",  "description": "IP RATING",                                     "unit": "",    "required_data": "IP 54 MIN",                               "vendor_data": "", "rev": ""},
            {"sr_no": "4.3",  "description": "COLOUR",                                        "unit": "",    "required_data": "RAL 7035 (LIGHT GREY)",                    "vendor_data": "", "rev": ""},
            {"sr_no": "4.4",  "description": "ARC FAULT CLASSIFICATION",                      "unit": "",    "required_data": "IAC AFL 25 kA 1 SEC",                    "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "BUSBAR",                                        "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "5.1",  "description": "MATERIAL",                                      "unit": "",    "required_data": "COPPER / ALUMINIUM",                        "vendor_data": "", "rev": ""},
            {"sr_no": "5.2",  "description": "SHAPE",                                         "unit": "",    "required_data": "RECTANGULAR",                              "vendor_data": "", "rev": ""},
            {"sr_no": "5.3",  "description": "BUSBAR RATING",                                 "unit": "A",   "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "CIRCUIT BREAKER",                               "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "6.1",  "description": "TYPE",                                          "unit": "",    "required_data": "VACUUM / SF6",                             "vendor_data": "", "rev": ""},
            {"sr_no": "6.2",  "description": "OPERATING MECHANISM",                           "unit": "",    "required_data": "SPRING CHARGED / STORED ENERGY",          "vendor_data": "", "rev": ""},
            {"sr_no": "6.3",  "description": "AUXILIARY SUPPLY VOLTAGE",                      "unit": "VDC", "required_data": "110 / 220",                               "vendor_data": "", "rev": ""},
            {"sr_no": "6.4",  "description": "NUMBER OF OPERATING CYCLES",                    "unit": "",    "required_data": "AS PER IEC 60056",                       "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "CURRENT TRANSFORMER",                           "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "7.1",  "description": "TYPE",                                          "unit": "",    "required_data": "RESIN CAST",                               "vendor_data": "", "rev": ""},
            {"sr_no": "7.2",  "description": "NUMBER OF CT CORES",                            "unit": "",    "required_data": "AS PER SCHEDULE",                          "vendor_data": "", "rev": ""},
            {"sr_no": "7.3",  "description": "CT RATIO",                                      "unit": "",    "required_data": "AS PER SCHEDULE",                          "vendor_data": "", "rev": ""},
            {"sr_no": "7.4",  "description": "CT CLASS",                                      "unit": "",    "required_data": "5P20, 0.5S",                              "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "VOLTAGE TRANSFORMER",                           "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "8.1",  "description": "TYPE",                                          "unit": "",    "required_data": "RESIN CAST",                               "vendor_data": "", "rev": ""},
            {"sr_no": "8.2",  "description": "VT RATIO",                                      "unit": "",    "required_data": "11000/√3 : 110/√3",                      "vendor_data": "", "rev": ""},
            {"sr_no": "8.3",  "description": "VT CLASS",                                      "unit": "",    "required_data": "3P, 0.5",                                 "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "EARTHING",                                      "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "9.1",  "description": "MAIN EARTHING BAR",                             "unit": "",    "required_data": "COPPER",                                   "vendor_data": "", "rev": ""},
            {"sr_no": "9.2",  "description": "EARTH FAULT RELAY",                             "unit": "",    "required_data": "NUMERICAL / MULTIFUNCTION",                "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "PROTECTION & CONTROL",                          "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "10.1", "description": "PROTECTION RELAY TYPE",                         "unit": "",    "required_data": "NUMERICAL / MULTIFUNCTION",                "vendor_data": "", "rev": ""},
            {"sr_no": "10.2", "description": "PROTECTION FUNCTIONS",                          "unit": "",    "required_data": "OVERCURRENT, EARTH FAULT, DIFFERENTIAL",   "vendor_data": "", "rev": ""},
            {"sr_no": "10.3", "description": "METERING",                                      "unit": "",    "required_data": "DIGITAL MULTIFUNCTION METER",               "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "AUXILIARY EQUIPMENT",                           "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "11.1", "description": "ANTI-CONDENSATION HEATER",                      "unit": "",    "required_data": "REQUIRED",                                "vendor_data": "", "rev": ""},
            {"sr_no": "11.2", "description": "SPACE HEATER RATING",                           "unit": "VAC", "required_data": "230",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "11.3", "description": "LIGHTING",                                      "unit": "VAC", "required_data": "230",                                       "vendor_data": "", "rev": ""},
            {"sr_no": "",     "description": "MANUFACTURER",                                  "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "12.1", "description": "NAME",                                          "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "12.2", "description": "MODEL/TYPE",                                    "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
            {"sr_no": "12.3", "description": "COUNTRY OF ORIGIN",                             "unit": "",    "required_data": "",                                        "vendor_data": "", "rev": ""},
        ]
    
    def export_to_excel(self, datasheet_rows: List[Dict], project_info: Dict = None):
        """Export datasheet to Excel with formatting"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "11KV Switchgear Datasheet"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        section_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add project header
        if project_info:
            ws.merge_cells('A1:E1')
            ws['A1'] = "11KV SWITCHGEAR DATASHEET"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            row_idx = 3
            for key, value in project_info.items():
                ws[f'A{row_idx}'] = key.replace('_', ' ').title()
                ws[f'B{row_idx}'] = value
                row_idx += 1
            
            row_idx += 1
        else:
            row_idx = 1
        
        # Add column headers — 6 columns: SR NO, DESCRIPTION, UNIT, REQUIRED DATA, VENDOR DATA, Rev
        headers = ['SR NO', 'DESCRIPTION', 'UNIT', 'REQUIRED DATA', 'VENDOR DATA', 'Rev']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 8
        
        row_idx += 1
        
        # Add data rows — 6 columns: sr_no, description, unit, required_data, vendor_data, rev
        for row_data in datasheet_rows:
            sr_no = row_data.get('sr_no', '')
            description = row_data.get('description', '')
            unit = row_data.get('unit', '')
            required_data = row_data.get('required_data', '')
            vendor_data = row_data.get('vendor_data', '')
            rev = row_data.get('rev', row_data.get('remarks', ''))  # fallback to remarks for legacy data
            
            # Check if this is a section header
            is_section = (sr_no == '' or sr_no is None) and description and not required_data and not vendor_data
            
            # Add cells
            ws.cell(row=row_idx, column=1, value=sr_no).border = border
            cell_desc = ws.cell(row=row_idx, column=2, value=description)
            cell_desc.border = border
            cell_desc.alignment = Alignment(wrap_text=True, vertical='top')
            
            cell_unit = ws.cell(row=row_idx, column=3, value=unit)
            cell_unit.border = border
            cell_unit.alignment = Alignment(horizontal='center', vertical='top')
            
            cell_req = ws.cell(row=row_idx, column=4, value=required_data)
            cell_req.border = border
            cell_req.alignment = Alignment(wrap_text=True, vertical='top')
            
            cell_vendor = ws.cell(row=row_idx, column=5, value=vendor_data)
            cell_vendor.border = border
            cell_vendor.alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.cell(row=row_idx, column=6, value=rev).border = border
            
            # Apply section styling
            if is_section:
                for col in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = section_fill
                    cell.font = section_font
            
            row_idx += 1
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
