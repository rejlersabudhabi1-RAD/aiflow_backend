"""
Emergency Diesel Generator (EDG) Set Datasheet Generator
Extracts equipment data from EDG Sizing Calculation documents
and generates comprehensive datasheets matching the standard ADNOC form.

Columns: SI No. | DESCRIPTION | UNIT | SPECIFIED DESIGN DATA | VENDOR DATA | Rev
"""
import logging
import json
from typing import Dict, List, Optional
from django.conf import settings
from openai import OpenAI
import PyPDF2

logger = logging.getLogger(__name__)


class DGSetDatasheetGenerator:
    """Generate Emergency Diesel Generator (EDG) Set datasheets from sizing calculation documents."""

    def __init__(self):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)

    # ──────────────────────────────────────────────────────────────────────────
    # PDF Extraction
    # ──────────────────────────────────────────────────────────────────────────
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extract all text from an uploaded PDF file."""
        try:
            pdf_file.seek(0)
            reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            logger.info(f"[DGSetDatasheet] PDF has {len(reader.pages)} pages")
            for i, page in enumerate(reader.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                    logger.info(f"[DGSetDatasheet] Page {i}: {len(page_text)} chars")
                else:
                    logger.warning(f"[DGSetDatasheet] Page {i}: no text (image-based?)")
            logger.info(f"[DGSetDatasheet] Total: {len(text)} chars")
            return text
        except Exception as e:
            logger.error(f"[DGSetDatasheet] PDF extraction error: {e}", exc_info=True)
            return ""

    # ──────────────────────────────────────────────────────────────────────────
    # Main entry point
    # ──────────────────────────────────────────────────────────────────────────
    def generate_datasheet_from_sizing_calc(self, pdf_file, project_info: Dict = None) -> Dict:
        """
        Generate a DG set datasheet from a sizing calculation PDF.

        Returns:
            {
                'success': bool,
                'datasheet_rows': List[Dict],   # sr_no, description, unit, required_data, vendor_data, rev
                'summary': Dict,
                'extraction_metadata': Dict
            }
        """
        try:
            logger.info("[DGSetDatasheet] Extracting text from EDG sizing calculation PDF…")
            doc_text = self.extract_text_from_pdf(pdf_file)

            if not doc_text or len(doc_text) < 20:
                logger.error(f"[DGSetDatasheet] Insufficient text: {len(doc_text) if doc_text else 0} chars")
                return {
                    'success': False,
                    'error': (
                        'Could not extract text from the PDF. '
                        'The file may be image-based or empty. '
                        'Please provide a text-based EDG sizing calculation document.'
                    )
                }

            logger.info("[DGSetDatasheet] Analysing with AI…")
            datasheet_rows = self._extract_datasheet_with_ai(doc_text, project_info)

            if not datasheet_rows:
                logger.warning("[DGSetDatasheet] AI returned no data – falling back to template")
                datasheet_rows = self._get_default_datasheet_template()

            summary = {
                'total_rows': len(datasheet_rows),
                'equipment_count': sum(1 for r in datasheet_rows if r.get('description', '').strip()),
                'completed_fields': sum(1 for r in datasheet_rows if r.get('vendor_data', '').strip()),
                'missing_fields': sum(1 for r in datasheet_rows if not r.get('vendor_data', '').strip()),
            }

            logger.info(f"[DGSetDatasheet] ✅ Generated {summary['total_rows']} rows")
            return {
                'success': True,
                'datasheet_rows': datasheet_rows,
                'summary': summary,
                'extraction_metadata': {
                    'document_length': len(doc_text),
                    'project_info': project_info or {},
                }
            }

        except Exception as e:
            logger.error(f"[DGSetDatasheet] Error: {e}", exc_info=True)
            return {'success': False, 'error': str(e)}

    # ──────────────────────────────────────────────────────────────────────────
    # AI extraction
    # ──────────────────────────────────────────────────────────────────────────
    def _extract_datasheet_with_ai(self, doc_text: str, project_info: Dict = None) -> List[Dict]:
        """Use GPT-4o to extract DG set datasheet data from a sizing calculation document."""

        prompt = f"""You are a senior electrical engineer specialising in Emergency Diesel Generator (EDG) sets.
Analyse the provided EDG Sizing Calculation document and extract comprehensive datasheet information.

PROJECT INFORMATION:
{json.dumps(project_info or {}, indent=2)}

DOCUMENT CONTENT:
{doc_text[:7000]}

TASK:
Populate a standard ADNOC EDG set datasheet with EXACTLY 6 fields per row:
- SR_NO       : Sequential item number (e.g. 1, 2, A, B, C …). Blank for section-header rows.
- DESCRIPTION : Parameter name or section heading.
- UNIT        : Engineering unit (kW, kVA, V, A, Hz, RPM, ℃, %, dB, L, kg, etc.). Blank if not applicable.
- SPECIFIED_DESIGN_DATA : Required/design value (engineer-specified column).
- VENDOR_DATA  : Value extracted from the uploaded sizing-calculation document; empty string "" if not found.
- REV         : Revision marker – empty string "" unless explicitly noted.

Cover ALL sections below in order:

HEADER FIELDS (no section letter)
  Tag No., Title (EMERGENCY DIESEL GENERATOR SET), Manufacturer / Model / Country of Origin,
  Year of Manufacture, Quantity

GENERAL INFO
  Design Life, Criticality Rating, Inspection Class

REFERENCE SPECIFICATIONS
  9.1  Emergency Generator (1000kVA and Above)
  9.2  Synchronous AC Generators 1000kVA and Above
  9.3  Diesel Fuelled Compression Ignition Engines
  9.4  Fire Protection Design Philosophy
  9.5  Direct Current UPS System
  9.6  Painting
  9.7  Key Single Line Diagram
  9.8  Power, Control and Earthing Cables
  9.9  Instrument and Control Design Guideline
  9.10 Instrumentation Furnished with Package Units
  9.11 Local Control Panels

B – ENVIRONMENTAL CONDITIONS
  1. Type of Installation
  2. Atmosphere
  3. Design Ambient Temperature
  4. Altitude
  5. Minimum Ambient Temperature
  6. Maximum Relative Humidity
  7. Average Relative Humidity
  8. Degree of Protection (IP)
  9. Solar Background Radiation Heat Flux
  10. Site Class Definition

C – GENERAL TECHNICAL CHARACTERISTICS (ALTERNATOR)
  1.  Rated Voltage
  2.  Phases
  3.  Frequency
  4.  Name Plate kW / kVA
  5.  Power Factor (PF)
  6.  Speed
  7.  Rotor Construction
  8.  Armature (Stator) Insulation Class / Rise
  9.  Field (Rotor) Insulation Class / Rise
  10. Exciter Insulation Class / Rise
  11. Minimum % Overspeed
  12. Maximum Unique Equipment Vertical Thrust
  13. Bearing Type – Sleeve
  14. Bearing Type – Anti-Friction
  15. Type (Synchronous / Induction)
  16. Service
  17. Duty Type

D – AREA CLASSIFICATION
  1. Zone
  2. Group
  3. Area
  4. Temp Class
  5. Outdoor
  6. Roof Over
  7. Max Sound Pressure Level

E – UNUSUAL CONDITION
  1. Abrasive Dust
  2. External Forces & Moments
  3. Seismic Loading
  4. Corrosive Agents

F – ELECTRICAL SYSTEM CONDITION
  1. Type of System Grounding
  2. Neutral Isolation Switch
  3. 3-Phase Symmetrical Fault Current
  4. Earth Fault Ampere
  5. Electrical Phase Rotation (ABC or ACB)

G – ENGINE CHARACTERISTICS
  1. Engine Manufacturer / Model
  2. Engine Type
  3. Number of Cylinders
  4. Bore × Stroke
  5. Rated Engine Speed at Full Load
  6. Maximum Continuous Engine Power (kW)
  7. Standby Engine Power (kW)
  8. Engine Starting System
  9. Number of Starting Attempts
  10. Starting-to-Full-Load Time
  11. Fuel Consumption at Full Load
  12. Fuel Consumption at 75% Load
  13. Fuel Consumption at 50% Load
  14. Fuel Type / Grade
  15. Fuel Tank Capacity (Day Tank)
  16. Fuel Tank Autonomy
  17. Lube Oil Pressure
  18. Lube Oil Temperature
  19. Coolant Temperature (Inlet / Outlet)
  20. Radiator Cooling – Air Flow

H – GENERATOR / ALTERNATOR CHARACTERISTICS
  1. Alternator Manufacturer / Model
  2. Alternator Type (Brushless / Static Excitation)
  3. Rated Output (kVA)
  4. Rated Voltage (V)
  5. Rated Frequency (Hz)
  6. Rated Speed (RPM)
  7. Number of Poles
  8. Power Factor
  9. Efficiency at 100% Load
  10. Efficiency at 75% Load
  11. Voltage Regulation (No Load to Full Load)
  12. Short Circuit Ratio
  13. Subtransient Reactance (Xd'')
  14. Transient Reactance (Xd')
  15. Synchronous Reactance (Xd)
  16. Winding Temperature Rise (Class F Limit)
  17. Excitation System Type
  18. AVR Manufacturer / Model

I – FUEL SYSTEM
  1. Fuel System Type (Gravity / Pump)
  2. Day Tank Capacity
  3. Day Tank Material
  4. Main Fuel Tank Capacity
  5. Fuel Transfer Pump (Duty / Standby)
  6. Fuel Filter Type
  7. Fuel Level Gauge / Alarm

J – COOLING SYSTEM
  1. Cooling Type (Radiator / Remote Radiator)
  2. Cooling Fan Drive
  3. Radiator Fan Motor Rating (kW)
  4. Coolant Type
  5. Coolant Capacity
  6. Radiator Pressure Cap Setting
  7. Expansion Tank

K – LUBRICATION SYSTEM
  1. Lube Oil Pump Type
  2. Lube Oil Filter Type
  3. Lube Oil Capacity
  4. Lube Oil Grade / Specification
  5. Lube Oil Pressure (Normal Operating)
  6. Pre-Lubrication Provision

L – EXHAUST SYSTEM
  1. Exhaust Temperature at Rated Load
  2. Exhaust Back Pressure (Maximum Allowable)
  3. Exhaust Silencer Type (Industrial / Critical / Hospital)
  4. Exhaust Pipe Material
  5. Exhaust Lagging

M – CONTROL AND PROTECTION PANEL
  1. Control Panel Type (AMF / ATS / Paralleling)
  2. Enclosure Protection (IP Rating)
  3. Control Voltage
  4. Battery Charger
  5. Synchronising Facility
  6. Speed Governor Type
  7. Engine Protection: Over-Speed Trip
  8. Engine Protection: Low Oil Pressure Trip
  9. Engine Protection: High Coolant Temperature Trip
  10. Generator Protection: Over-Voltage
  11. Generator Protection: Under-Voltage
  12. Generator Protection: Over-Frequency
  13. Generator Protection: Under-Frequency
  14. Generator Protection: Reverse Power
  15. Generator Protection: Short Circuit (Overcurrent)
  16. Generator Protection: Earth Fault

N – ACOUSTIC AND VIBRATION
  1. Max Sound Pressure Level at 1 m (dB(A))
  2. Max Sound Power Level (dB(A))
  3. Vibration Isolation System
  4. Anti-Vibration Mounts

O – CIVIL / STRUCTURAL
  1. Skid / Baseframe Material
  2. Skid Dimensions (L × W × H)
  3. Total Operating Weight
  4. Anchor Bolt Size / Pattern
  5. Weatherproof Canopy / Enclosure

P – INSPECTION AND TESTING
  1. Factory Acceptance Test (FAT)
  2. Load Test Duration at Full Load
  3. Transient Response Test
  4. Noise Level Test
  5. Vibration Test
  6. Insulation Resistance Test
  7. High Voltage Test
  8. Protection Relay Testing

Return ONLY a JSON array. Each element must have exactly these keys:
  "sr_no", "description", "unit", "required_data", "vendor_data", "rev"

Rules:
- Section header rows: sr_no = "", unit = "", required_data = "", vendor_data = "", rev = ""
- Extract ACTUAL values from the document for vendor_data; use "" when not found
- required_data = standard / typical requirement value for an ADNOC EDG per BGS-MA-004 / IEC 60034
- rev = "" always unless document specifies a revision letter
- Include every parameter listed above even if vendor_data is empty
- Return ONLY the JSON array – no markdown, no explanation"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are an expert electrical engineer specialising in "
                            "Emergency Diesel Generator (EDG) set datasheets per ADNOC BGS-MA-004 / IEC 60034 standards. "
                            "Return only valid JSON arrays."
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=6000,
            )

            ai_response = response.choices[0].message.content.strip()

            # Strip markdown fences if present
            if "```json" in ai_response:
                ai_response = ai_response.split("```json")[1].split("```")[0]
            elif "```" in ai_response:
                ai_response = ai_response.split("```")[1].split("```")[0]

            datasheet_rows = json.loads(ai_response.strip())

            if isinstance(datasheet_rows, list) and len(datasheet_rows) > 0:
                for row in datasheet_rows:
                    row.setdefault("sr_no", "")
                    row.setdefault("description", "")
                    row.setdefault("unit", "")
                    row.setdefault("required_data", "")
                    row.setdefault("vendor_data", "")
                    row.setdefault("rev", "")
                    row.pop("remarks", None)       # remove legacy key
                    row.pop("specified_design_data", None)  # normalise key
                logger.info(f"[DGSetDatasheet] AI returned {len(datasheet_rows)} rows")
                return datasheet_rows
            else:
                logger.error("[DGSetDatasheet] Invalid AI response structure")
                return self._get_default_datasheet_template()

        except json.JSONDecodeError as e:
            logger.error(f"[DGSetDatasheet] JSON decode error: {e}")
            return self._get_default_datasheet_template()
        except Exception as e:
            logger.error(f"[DGSetDatasheet] AI extraction error: {e}")
            return self._get_default_datasheet_template()

    # ──────────────────────────────────────────────────────────────────────────
    # Default template  (full ADNOC EDG datasheet)
    # ──────────────────────────────────────────────────────────────────────────
    def _get_default_datasheet_template(self) -> List[Dict]:
        """Return the full standard ADNOC EDG set datasheet template (6-column)."""
        R = lambda sr, desc, unit="", req="": {
            "sr_no": sr, "description": desc, "unit": unit,
            "required_data": req, "vendor_data": "", "rev": ""
        }
        H = lambda desc: {
            "sr_no": "", "description": desc, "unit": "",
            "required_data": "", "vendor_data": "", "rev": ""
        }
        return [
            # ── HEADER FIELDS ────────────────────────────────────────────────
            R("1",  "TAG NO."),
            R("2",  "TITLE",                                "",   "EMERGENCY DIESEL GENERATOR SET"),
            R("3",  "MANUFACTURER / MODEL / COUNTRY OF ORIGIN"),
            R("4",  "YEAR OF MANUFACTURE"),
            R("5",  "QUANTITY",                            "No",  "1"),

            # ── GENERAL INFO ─────────────────────────────────────────────────
            R("6",  "DESIGN LIFE",                         "",   "MINIMUM SERVICE LIFE 25 YEARS"),
            R("7",  "CRITICALITY RATING",                  "",   "3"),
            R("8",  "INSPECTION CLASS",                    "",   "1"),

            # ── REFERENCE SPECIFICATIONS ─────────────────────────────────────
            H("9   REFERENCE SPECIFICATION"),
            R("9.1",  "EMERGENCY GENERATOR (1000KVA AND ABOVE)",              "", "BGS-MA-004"),
            R("9.2",  "SYNCHRONOUS AC GENERATORS 1000KVA AND ABOVE",          "", "BGS-MV-004"),
            R("9.3",  "DIESEL FUELLED COMPRESSION IGNITION ENGINES",          "", "BGS-MV-003"),
            R("9.4",  "FIRE PROTECTION DESIGN PHILOSOPHY",                    "", "BGS-MU-200"),
            R("9.5",  "DIRECT CURRENT UPS SYSTEM",                            "", "BGS-EE-009"),
            R("9.6",  "PAINTING",                                             "", "BGS-MX-001"),
            R("9.7",  "KEY SINGLE LINE DIAGRAM",                              ""),
            R("9.8",  "POWER, CONTROL AND EARTHING CABLES",                   "", "BGS-EE-011"),
            R("9.9",  "INSTRUMENT AND CONTROL DESIGN GUIDELINE",              "", "BGS-IU-301"),
            R("9.10", "INSTRUMENTATION FURNISHED WITH PACKAGE UNITS",         "", "BGS-IU-007"),
            R("9.11", "LOCAL CONTROL PANELS",                                 "", "BGS-IU-223"),

            # ── B. ENVIRONMENTAL CONDITIONS ──────────────────────────────────
            H("B   ENVIRONMENTAL CONDITIONS"),
            R("1",  "TYPE OF INSTALLATION",        "",
              "OUTDOOR FOR DIESEL GENERATOR CONTAINER; INDOOR FOR REMOTE CONTROL, PROTECTION AND MONITORING PANEL"),
            R("2",  "ATMOSPHERE",                  "",
              "SALTY, SULPHUROUS AND DUSTY WITH HIGH CONCENTRATION OF WINDBORNE SAND"),
            R("3",  "DESIGN AMBIENT TEMPERATURE",  "°C",
              "54°C FOR DIESEL GENERATOR CONTAINER; 40°C FOR REMOTE CONTROL, PROTECTION & MONITORING PANEL"),
            R("4",  "ALTITUDE",                    "M",   "LESS THAN 1000m AMSL"),
            R("5",  "MINIMUM AMBIENT TEMPERATURE", "°C"),
            R("6",  "MAXIMUM RELATIVE HUMIDITY",   "%",   "95% AT 43°C"),
            R("7",  "AVERAGE RELATIVE HUMIDITY",   "%",   "80% AT 54°C"),
            R("8",  "DEGREE OF PROTECTION (IP)",   "",    "IP55 (MINIMUM)"),
            R("9",  "SOLAR BACKGROUND RADIATION HEAT FLUX", "W/M²", "95 W/M2"),
            R("10", "SITE CLASS DEFINITION",       "",    "SITE CLASS C"),

            # ── C. GENERAL TECHNICAL CHARACTERISTICS ─────────────────────────
            H("C   GENERAL TECHNICAL CHARACTERISTICS"),
            R("1",  "RATED VOLTAGE",               "V",   "415 V ±10%"),
            R("2",  "PHASES",                      "",    "3"),
            R("3",  "FREQUENCY",                   "Hz",  "50 Hz ±2%"),
            R("4",  "NAME PLATE KW / KVA",         "kW/kVA"),
            R("5",  "POWER FACTOR (PF)",           "",    "0.8"),
            R("6",  "SPEED",                       "RPM", "1500 rpm"),
            R("7",  "ROTOR CONSTRUCTION",          "",    "CYLINDRICAL"),
            R("8",  "ARMATURE (STATOR) INSULATION CLASS / RISE", "",
              "F/B (86°C ABOVE 54°C AMBIENT)"),
            R("9",  "FIELD (ROTOR) INSULATION CLASS / RISE",    "",
              "F/B (86°C ABOVE 54°C AMBIENT)"),
            R("10", "EXCITER INSULATION CLASS / RISE",          "",
              "F/B (86°C ABOVE 54°C AMBIENT)"),
            R("11", "MINIMUM % OVERSPEED",         "%",
              "120% FOR DURATION OF 2 MIN. (REFER BGS-MA-004, CL 12.3.8)"),
            R("12", "MAXIMUM UNIQUE EQUIPMENT VERTICAL THRUST", ""),
            R("13", "BEARING TYPE – SLEEVE",       ""),
            R("14", "BEARING TYPE – ANTI-FRICTION",""),
            R("15", "TYPE",                        "",    "SYNCHRONOUS"),
            R("16", "SERVICE",                     "",    "CONTINUOUS FULL LOAD"),
            R("17", "DUTY TYPE",                   "",    "S1"),

            # ── D. AREA CLASSIFICATION ───────────────────────────────────────
            H("D   AREA CLASSIFICATION"),
            R("1",  "ZONE",        "",   "NOT APPLICABLE"),
            R("2",  "GROUP",       "",   "NOT APPLICABLE"),
            R("3",  "AREA",        "",   "NOT APPLICABLE"),
            R("4",  "TEMP CLASS",  "",   "NOT APPLICABLE"),
            R("5",  "OUTDOOR",     "",   "YES"),
            R("6",  "ROOF OVER",   "",   "NO ROOF OVER"),
            R("7",  "MAX SOUND PRESSURE LEVEL", "dB(A)",
              "85 dB(A) @ 1 m FOR TOTAL SKID (DIESEL, ENGINE & ALTERNATOR) AS PER BGS-MU-009"),

            # ── E. UNUSUAL CONDITION ─────────────────────────────────────────
            H("E   UNUSUAL CONDITION"),
            R("1",  "ABRASIVE DUST",               "",   "INFORMATION WILL PROVIDE LATER"),
            R("2",  "EXTERNAL FORCES & MOMENTS",   "",   "INFORMATION WILL PROVIDE LATER"),
            R("3",  "SEISMIC LOADING",             "",   "INFORMATION WILL PROVIDE LATER"),
            R("4",  "CORROSIVE AGENTS",            "",   "INFORMATION WILL PROVIDE LATER"),

            # ── F. ELECTRICAL SYSTEM CONDITION ───────────────────────────────
            H("F   ELECTRICAL SYSTEM CONDITION"),
            R("1",  "TYPE OF SYSTEM GROUNDING",                     "",   "REFER NOTE 8"),
            R("2",  "NEUTRAL ISOLATION SWITCH",                      "",   "YES, REQUIRED"),
            R("3",  "3-PHASE SYMMETRICAL FAULT CURRENT",             "kA",
              "65 kA (INCLUDING DG CONTRIBUTION) FOR 3 SEC"),
            R("4",  "EARTH FAULT AMPERE",                           "A"),
            R("5",  "ELECTRICAL PHASE ROTATION (ABC OR ACB)",        "",   "ABC"),

            # ── G. ENGINE CHARACTERISTICS ────────────────────────────────────
            H("G   ENGINE CHARACTERISTICS"),
            R("1",  "ENGINE MANUFACTURER / MODEL"),
            R("2",  "ENGINE TYPE",                    "",   "4-STROKE DIESEL"),
            R("3",  "NUMBER OF CYLINDERS"),
            R("4",  "BORE × STROKE",                  "mm"),
            R("5",  "RATED ENGINE SPEED AT FULL LOAD", "RPM", "1500"),
            R("6",  "MAXIMUM CONTINUOUS ENGINE POWER", "kW"),
            R("7",  "STANDBY ENGINE POWER",           "kW"),
            R("8",  "ENGINE STARTING SYSTEM",         "",   "ELECTRIC START (DUAL BATTERY)"),
            R("9",  "NUMBER OF STARTING ATTEMPTS",    "",   "3"),
            R("10", "STARTING-TO-FULL-LOAD TIME",     "Sec", "10"),
            R("11", "FUEL CONSUMPTION AT 100% LOAD",  "L/hr"),
            R("12", "FUEL CONSUMPTION AT 75% LOAD",   "L/hr"),
            R("13", "FUEL CONSUMPTION AT 50% LOAD",   "L/hr"),
            R("14", "FUEL TYPE / GRADE",              "",   "DIESEL / GAS OIL TO BS EN 590"),
            R("15", "FUEL TANK CAPACITY (DAY TANK)",  "L"),
            R("16", "FUEL TANK AUTONOMY",             "hrs", "MINIMUM 8 hrs AT FULL LOAD"),
            R("17", "LUBE OIL PRESSURE (NORMAL)",     "bar"),
            R("18", "LUBE OIL TEMPERATURE",           "°C"),
            R("19", "COOLANT TEMPERATURE OUTLET",     "°C"),
            R("20", "AIR FLOW THROUGH RADIATOR",      "m³/s"),

            # ── H. ALTERNATOR / GENERATOR CHARACTERISTICS ────────────────────
            H("H   ALTERNATOR / GENERATOR CHARACTERISTICS"),
            R("1",  "ALTERNATOR MANUFACTURER / MODEL"),
            R("2",  "ALTERNATOR TYPE",               "",   "BRUSHLESS SELF-EXCITING"),
            R("3",  "RATED OUTPUT",                  "kVA"),
            R("4",  "RATED VOLTAGE",                 "V",  "415"),
            R("5",  "RATED FREQUENCY",               "Hz", "50"),
            R("6",  "RATED SPEED",                   "RPM","1500"),
            R("7",  "NUMBER OF POLES",               "",   "4"),
            R("8",  "POWER FACTOR",                  "",   "0.8 LAGGING"),
            R("9",  "EFFICIENCY AT 100% LOAD",       "%"),
            R("10", "EFFICIENCY AT 75% LOAD",        "%"),
            R("11", "VOLTAGE REGULATION (NO LOAD TO FULL LOAD)", "%", "±1%"),
            R("12", "SHORT CIRCUIT RATIO",           ""),
            R("13", "SUBTRANSIENT REACTANCE (Xd'')", "%"),
            R("14", "TRANSIENT REACTANCE (Xd')",     "%"),
            R("15", "SYNCHRONOUS REACTANCE (Xd)",    "%"),
            R("16", "WINDING TEMPERATURE RISE (CLASS F LIMIT)", "°C", "105 (CLASS F)"),
            R("17", "EXCITATION SYSTEM TYPE",        "",   "AUTOMATIC VOLTAGE REGULATOR (AVR)"),
            R("18", "AVR MANUFACTURER / MODEL"),

            # ── I. FUEL SYSTEM ───────────────────────────────────────────────
            H("I   FUEL SYSTEM"),
            R("1",  "FUEL SYSTEM TYPE",              "",   "GRAVITY FEED FROM DAY TANK"),
            R("2",  "DAY TANK CAPACITY",             "L"),
            R("3",  "DAY TANK MATERIAL",             "",   "MILD STEEL (EPOXY COATED INTERNALLY)"),
            R("4",  "MAIN FUEL TANK CAPACITY",       "L"),
            R("5",  "FUEL TRANSFER PUMP",            "",   "1 DUTY + 1 STANDBY"),
            R("6",  "FUEL FILTER TYPE",              "",   "DUPLEX WITH CHANGEOVER VALVE"),
            R("7",  "FUEL LEVEL GAUGE / ALARM",      "",   "PROVIDED"),

            # ── J. COOLING SYSTEM ────────────────────────────────────────────
            H("J   COOLING SYSTEM"),
            R("1",  "COOLING TYPE",                  "",   "RADIATOR (ENGINE-MOUNTED)"),
            R("2",  "COOLING FAN DRIVE",             "",   "ENGINE-DRIVEN"),
            R("3",  "RADIATOR FAN MOTOR RATING",     "kW"),
            R("4",  "COOLANT TYPE",                  "",   "ETHYLENE GLYCOL ANTIFREEZE MIXTURE"),
            R("5",  "COOLANT CAPACITY",              "L"),
            R("6",  "RADIATOR PRESSURE CAP SETTING", "bar"),
            R("7",  "EXPANSION TANK",                "",   "PROVIDED"),

            # ── K. LUBRICATION SYSTEM ────────────────────────────────────────
            H("K   LUBRICATION SYSTEM"),
            R("1",  "LUBE OIL PUMP TYPE",            "",   "GEAR PUMP (ENGINE-DRIVEN)"),
            R("2",  "LUBE OIL FILTER TYPE",          "",   "FULL FLOW SPIN-ON"),
            R("3",  "LUBE OIL CAPACITY",             "L"),
            R("4",  "LUBE OIL GRADE / SPECIFICATION","",   "SAE 15W-40 / API CI-4"),
            R("5",  "LUBE OIL PRESSURE (NORMAL)",    "bar", "3.5 – 5.0"),
            R("6",  "PRE-LUBRICATION PROVISION",     "",   "PROVIDED (ELECTRIC PRELUBE PUMP)"),

            # ── L. EXHAUST SYSTEM ────────────────────────────────────────────
            H("L   EXHAUST SYSTEM"),
            R("1",  "EXHAUST TEMPERATURE AT RATED LOAD",        "°C"),
            R("2",  "EXHAUST BACK PRESSURE (MAX ALLOWABLE)",    "kPa"),
            R("3",  "EXHAUST SILENCER TYPE",                    "",   "CRITICAL GRADE"),
            R("4",  "EXHAUST PIPE MATERIAL",                    "",   "STAINLESS STEEL 316L"),
            R("5",  "EXHAUST LAGGING",                          "",   "PROVIDED"),

            # ── M. CONTROL AND PROTECTION PANEL ─────────────────────────────
            H("M   CONTROL AND PROTECTION PANEL"),
            R("1",  "CONTROL PANEL TYPE",            "",   "AMF (AUTOMATIC MAINS FAILURE)"),
            R("2",  "ENCLOSURE PROTECTION",          "",   "IP54 (MINIMUM)"),
            R("3",  "CONTROL VOLTAGE",               "VDC", "24"),
            R("4",  "BATTERY CHARGER",               "",   "PROVIDED (TRICKLE + BOOST)"),
            R("5",  "SYNCHRONISING FACILITY",        "",   "PROVIDED"),
            R("6",  "SPEED GOVERNOR TYPE",           "",   "ELECTRONIC ISOCHRONOUS"),
            R("7",  "ENGINE PROTECTION: OVER-SPEED TRIP",               "", "PROVIDED"),
            R("8",  "ENGINE PROTECTION: LOW OIL PRESSURE TRIP",         "", "PROVIDED"),
            R("9",  "ENGINE PROTECTION: HIGH COOLANT TEMP TRIP",        "", "PROVIDED"),
            R("10", "GENERATOR PROTECTION: OVER-VOLTAGE",               "", "PROVIDED"),
            R("11", "GENERATOR PROTECTION: UNDER-VOLTAGE",              "", "PROVIDED"),
            R("12", "GENERATOR PROTECTION: OVER-FREQUENCY",             "", "PROVIDED"),
            R("13", "GENERATOR PROTECTION: UNDER-FREQUENCY",            "", "PROVIDED"),
            R("14", "GENERATOR PROTECTION: REVERSE POWER",              "", "PROVIDED"),
            R("15", "GENERATOR PROTECTION: SHORT CIRCUIT (OVERCURRENT)","", "PROVIDED"),
            R("16", "GENERATOR PROTECTION: EARTH FAULT",                "", "PROVIDED"),

            # ── N. ACOUSTIC AND VIBRATION ────────────────────────────────────
            H("N   ACOUSTIC AND VIBRATION"),
            R("1",  "MAX SOUND PRESSURE LEVEL AT 1 m",  "dB(A)", "≤ 85"),
            R("2",  "MAX SOUND POWER LEVEL",            "dB(A)"),
            R("3",  "VIBRATION ISOLATION SYSTEM",       "",       "ANTI-VIBRATION MOUNTS"),
            R("4",  "ANTI-VIBRATION MOUNTS",            "",       "PROVIDED"),

            # ── O. CIVIL / STRUCTURAL ────────────────────────────────────────
            H("O   CIVIL / STRUCTURAL"),
            R("1",  "SKID / BASEFRAME MATERIAL",        "",   "STRUCTURAL STEEL (HOT-DIP GALVANISED)"),
            R("2",  "SKID DIMENSIONS (L × W × H)",      "mm"),
            R("3",  "TOTAL OPERATING WEIGHT",           "kg"),
            R("4",  "ANCHOR BOLT SIZE / PATTERN",       ""),
            R("5",  "WEATHERPROOF CANOPY / ENCLOSURE",  "",   "PROVIDED"),

            # ── P. INSPECTION AND TESTING ────────────────────────────────────
            H("P   INSPECTION AND TESTING"),
            R("1",  "FACTORY ACCEPTANCE TEST (FAT)",    "",   "REQUIRED"),
            R("2",  "LOAD TEST DURATION AT FULL LOAD",  "hrs", "2"),
            R("3",  "TRANSIENT RESPONSE TEST",          "",   "REQUIRED"),
            R("4",  "NOISE LEVEL TEST",                 "",   "REQUIRED"),
            R("5",  "VIBRATION TEST",                   "",   "REQUIRED"),
            R("6",  "INSULATION RESISTANCE TEST",       "",   "REQUIRED"),
            R("7",  "HIGH VOLTAGE TEST",                "",   "REQUIRED"),
            R("8",  "PROTECTION RELAY TESTING",         "",   "REQUIRED"),
        ]

    # ──────────────────────────────────────────────────────────────────────────
    # Excel Export
    # ──────────────────────────────────────────────────────────────────────────
    def export_to_excel(self, datasheet_rows: List[Dict], project_info: Dict = None):
        """Export DG set datasheet to formatted Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "DG Set Datasheet"

        # Dark green header to distinguish from transformer (blue)
        header_fill  = PatternFill(start_color="1E5631", end_color="1E5631", fill_type="solid")
        header_font  = Font(color="FFFFFF", bold=True, size=10)
        section_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
        section_font = Font(bold=True, size=10)
        thin   = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row_idx = 1

        # Title row
        ws.merge_cells("A1:F1")
        ws["A1"] = "EMERGENCY DIESEL GENERATOR (EDG) SET – DATASHEET"
        ws["A1"].font = Font(bold=True, size=13, color="1E5631")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        row_idx = 3

        # Project info
        if project_info:
            for key, val in project_info.items():
                ws.cell(row=row_idx, column=1, value=key.replace("_", " ").title())
                ws.cell(row=row_idx, column=2, value=val)
                row_idx += 1
            row_idx += 1

        # Column headers – 6 columns
        col_headers = ["SI NO.", "DESCRIPTION", "UNIT", "SPECIFIED DESIGN DATA", "VENDOR DATA", "Rev"]
        col_widths  = [9, 52, 10, 35, 30, 8]
        for col_idx, (header, width) in enumerate(zip(col_headers, col_widths), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[chr(64 + col_idx)].width = width

        row_idx += 1

        # Data rows
        for row_data in datasheet_rows:
            sr_no       = row_data.get("sr_no", "")
            description = row_data.get("description", "")
            unit        = row_data.get("unit", "")
            req_data    = row_data.get("required_data", "")
            vendor_data = row_data.get("vendor_data", "")
            rev         = row_data.get("rev", row_data.get("remarks", ""))

            is_section = (not sr_no) and description and not req_data and not vendor_data

            cells  = [sr_no, description, unit, req_data, vendor_data, rev]
            aligns = ["center", "left", "center", "left", "left", "center"]

            for col_idx, (val, align) in enumerate(zip(cells, aligns), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
                if is_section:
                    cell.fill = section_fill
                    cell.font = section_font

            row_idx += 1

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
