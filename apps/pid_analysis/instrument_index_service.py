"""
Instrument Index Service — ALL-INSTRUMENT extraction from P&ID drawings
-----------------------------------------------------------------------
Multi-engine extraction waterfall (soft-coded via EXTRACTION_CONFIG):

  Engine 1  — PyMuPDF text layer  (vector PDFs, instant, free)
  Engine 2  — Gemini Vision       (primary AI — free tier, 1M context)
  Engine 3  — OpenAI Vision       (fallback AI — GPT-4o)
  Engine 4  — Tesseract OCR       (scanned PDFs with no text layer)

All pass results are merged and deduplicated by normalised tag number.

SOFT-CODED: all instrument categories live in INSTRUMENT_CATEGORIES dict
below — add/remove types without touching logic.
"""

import io
import os
import re
import json
import base64
import logging
from datetime import datetime

import time
from pdf2image import convert_from_bytes
from PIL import Image
import openpyxl

# Disable PIL's DecompressionBomb limit — large P&ID drawings (A0/A1 at 150 DPI)
# legitimately produce images above the 89 MP default threshold.
Image.MAX_IMAGE_PIXELS = None
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED CONFIGURATION
# All instrument categories / type descriptions live here.
# ────────────────────────────────────────────────────────────────────────────
INSTRUMENT_CATEGORIES = {
    # ── FLOW ──────────────────────────────────────────────────────────────
    "FT":   {"name": "Flow Transmitter",                  "category": "Flow"},
    "FI":   {"name": "Flow Indicator",                    "category": "Flow"},
    "FIC":  {"name": "Flow Indicating Controller",        "category": "Flow"},
    "FIT":  {"name": "Flow Indicating Transmitter",       "category": "Flow"},
    "FCV":  {"name": "Flow Control Valve",                "category": "Flow"},
    "FE":   {"name": "Flow Element / Orifice",            "category": "Flow"},
    "FQI":  {"name": "Flow Quantity Indicator",           "category": "Flow"},
    "FG":   {"name": "Flow Glass / Sight Glass",          "category": "Flow"},
    "FO":   {"name": "Flow Orifice",                      "category": "Flow"},
    "FY":   {"name": "Flow Relay / Computer",             "category": "Flow"},
    # ── PRESSURE ──────────────────────────────────────────────────────────
    "PT":   {"name": "Pressure Transmitter",              "category": "Pressure"},
    "PI":   {"name": "Pressure Indicator",                "category": "Pressure"},
    "PIC":  {"name": "Pressure Indicating Controller",    "category": "Pressure"},
    "PIT":  {"name": "Pressure Indicating Transmitter",   "category": "Pressure"},
    "PS":   {"name": "Pressure Switch",                   "category": "Pressure"},
    "PSH":  {"name": "Pressure Switch High",              "category": "Pressure"},
    "PSL":  {"name": "Pressure Switch Low",               "category": "Pressure"},
    "PSHH": {"name": "Pressure Switch High-High",         "category": "Pressure"},
    "PSLL": {"name": "Pressure Switch Low-Low",           "category": "Pressure"},
    "PSAL": {"name": "Pressure Switch Alarm Low",         "category": "Pressure"},
    "PSAH": {"name": "Pressure Switch Alarm High",        "category": "Pressure"},
    "PSDL": {"name": "Pressure Switch Differential Low",  "category": "Pressure"},
    "PSDH": {"name": "Pressure Switch Differential High", "category": "Pressure"},
    "PCV":  {"name": "Pressure Control Valve",            "category": "Pressure"},
    "PG":   {"name": "Pressure Gauge",                    "category": "Pressure"},
    "PSV":  {"name": "Pressure Safety Valve",             "category": "Safety"},
    "PRV":  {"name": "Pressure Relief Valve",             "category": "Safety"},
    # ── TEMPERATURE ───────────────────────────────────────────────────────
    "TT":   {"name": "Temperature Transmitter",           "category": "Temperature"},
    "TI":   {"name": "Temperature Indicator",             "category": "Temperature"},
    "TIC":  {"name": "Temperature Indicating Controller", "category": "Temperature"},
    "TIT":  {"name": "Temperature Indicating Transmitter","category": "Temperature"},
    "TS":   {"name": "Temperature Switch",                "category": "Temperature"},
    "TSH":  {"name": "Temperature Switch High",           "category": "Temperature"},
    "TSL":  {"name": "Temperature Switch Low",            "category": "Temperature"},
    "TSHH": {"name": "Temperature Switch High-High",      "category": "Temperature"},
    "TSLL": {"name": "Temperature Switch Low-Low",        "category": "Temperature"},
    "TCV":  {"name": "Temperature Control Valve",         "category": "Temperature"},
    "TW":   {"name": "Thermowell",                        "category": "Temperature"},
    "TE":   {"name": "Temperature Element (Thermocouple)","category": "Temperature"},
    # ── LEVEL ─────────────────────────────────────────────────────────────
    "LT":   {"name": "Level Transmitter",                 "category": "Level"},
    "LI":   {"name": "Level Indicator",                   "category": "Level"},
    "LIC":  {"name": "Level Indicating Controller",       "category": "Level"},
    "LIT":  {"name": "Level Indicating Transmitter",      "category": "Level"},
    "LS":   {"name": "Level Switch",                      "category": "Level"},
    "LSH":  {"name": "Level Switch High",                 "category": "Level"},
    "LSL":  {"name": "Level Switch Low",                  "category": "Level"},
    "LSHH": {"name": "Level Switch High-High",            "category": "Level"},
    "LSLL": {"name": "Level Switch Low-Low",              "category": "Level"},
    "LSAL": {"name": "Level Switch Alarm Low",            "category": "Level"},
    "LSAH": {"name": "Level Switch Alarm High",           "category": "Level"},
    "LSDL": {"name": "Level Switch Differential Low",     "category": "Level"},
    "LSDH": {"name": "Level Switch Differential High",    "category": "Level"},
    "LG":   {"name": "Level Gauge",                       "category": "Level"},
    "LCV":  {"name": "Level Control Valve",               "category": "Level"},
    "LY":   {"name": "Level Relay / Computer",            "category": "Level"},
    # ── DIFFERENTIAL PRESSURE ─────────────────────────────────────────────
    "DPI":  {"name": "Differential Pressure Indicator",   "category": "Differential Pressure"},
    "DPIT": {"name": "DP Indicating Transmitter",         "category": "Differential Pressure"},
    "DPT":  {"name": "Differential Pressure Transmitter", "category": "Differential Pressure"},
    "DPAH": {"name": "DP Alarm High",                     "category": "Differential Pressure"},
    "DPAL": {"name": "DP Alarm Low",                      "category": "Differential Pressure"},
    "DPZY": {"name": "DP Position Transmitter",           "category": "Differential Pressure"},
    # ── ANALYSIS ──────────────────────────────────────────────────────────
    "AT":   {"name": "Analyzer Transmitter",              "category": "Analysis"},
    "AI":   {"name": "Analyzer Indicator",                "category": "Analysis"},
    "AIC":  {"name": "Analyzer Indicating Controller",    "category": "Analysis"},
    "AIT":  {"name": "Analyzer Indicating Transmitter",   "category": "Analysis"},
    # ── SHUTDOWN / ESD / CONTROL VALVES ───────────────────────────────────
    "SDV":  {"name": "Shutdown Valve",                    "category": "Shutdown & ESD"},
    "BDV":  {"name": "Blowdown Valve",                    "category": "Shutdown & ESD"},
    "XV":   {"name": "On/Off Valve (ESD)",                "category": "Shutdown & ESD"},
    "EV":   {"name": "Emergency Valve",                   "category": "Shutdown & ESD"},
    "HCV":  {"name": "Hand Control Valve",                "category": "Control Valves"},
    # ── MOTOR / SOLENOID OPERATED ─────────────────────────────────────────
    "MOV":  {"name": "Motor Operated Valve",              "category": "Motor & Solenoid"},
    "SOV":  {"name": "Solenoid Operated Valve",           "category": "Motor & Solenoid"},
    "AOV":  {"name": "Air Operated Valve",                "category": "Motor & Solenoid"},
    # ── POSITION / VALVE POSITION ─────────────────────────────────────────
    "ZI":   {"name": "Position Indicator",                "category": "Position"},
    "ZT":   {"name": "Position Transmitter",              "category": "Position"},
    "ZS":   {"name": "Position Switch",                   "category": "Position"},
    "ZSH":  {"name": "Position Switch High (Open)",       "category": "Position"},
    "ZSL":  {"name": "Position Switch Low (Closed)",      "category": "Position"},
    "ZCV":  {"name": "Position Control Valve",            "category": "Position"},
    "SVZY": {"name": "Solenoid Valve + Position TX",      "category": "Position"},
    "BVZY": {"name": "Ball Valve + Position TX",          "category": "Position"},
    # ── RESTRICTION / SPECIAL ─────────────────────────────────────────────
    "RO":   {"name": "Restriction Orifice",               "category": "Restriction"},
    "XPD":  {"name": "Special / Explosion Proof Device",  "category": "Special"},
    "XY":   {"name": "Relay / Computer (Special)",        "category": "Special"},
    "WI":   {"name": "Weight Indicator",                  "category": "Weight"},
    "WIT":  {"name": "Weight Indicating Transmitter",     "category": "Weight"},
    "SI":   {"name": "Speed Indicator",                   "category": "Speed"},
    "SIT":  {"name": "Speed Indicating Transmitter",      "category": "Speed"},
    "VI":   {"name": "Vibration Indicator",               "category": "Vibration"},
    "VIT":  {"name": "Vibration Indicating Transmitter",  "category": "Vibration"},
    "HI":   {"name": "Hand Indicator",                    "category": "Hand/Manual"},
    "HS":   {"name": "Hand Switch",                       "category": "Hand/Manual"},
}

# Column definitions for Excel output
EXCEL_COLUMNS = [
    {"key": "index_no",              "label": "Index No.",          "width": 10},
    {"key": "tag_number",            "label": "Tag Number",         "width": 18},
    {"key": "control_system_tag",    "label": "CS Tag",             "width": 18},
    {"key": "instrument_type",       "label": "Instrument Type",    "width": 35},
    {"key": "category",              "label": "Category",           "width": 22},
    {"key": "pid_no",                "label": "P&ID No.",           "width": 22},
    {"key": "service_description",   "label": "Service Description","width": 40},
    {"key": "line_number",           "label": "Line Number",        "width": 20},
    {"key": "equipment_number",      "label": "Equipment No.",      "width": 18},
    {"key": "loop_number",           "label": "Loop No.",           "width": 14},
    {"key": "fail_safe",             "label": "Fail Safe",          "width": 12},
    {"key": "signal_type",           "label": "Signal Type",        "width": 16},
    {"key": "set_point",             "label": "Set Point",          "width": 14},
    {"key": "drawing_number",        "label": "Drawing No.",        "width": 22},
    {"key": "revision",              "label": "Rev.",               "width": 8},
    {"key": "notes",                 "label": "Notes",              "width": 40},
]

# Category colour coding for Excel rows
CATEGORY_COLOURS = {
    "Flow":               "DDEEFF",
    "Pressure":           "FFE4CC",
    "Temperature":        "FFE4E4",
    "Level":              "E4F4E4",
    "Differential Pressure": "FFF9CC",
    "Analysis":           "E8E4FF",
    "Safety":             "FFCCCC",
    "Shutdown & ESD":     "FFD9D9",
    "Control Valves":     "CCFFEE",
    "Motor & Solenoid":   "E0E0FF",
    "Position":           "FFFACC",
    "Restriction":        "DDEEDD",
    "Special":            "F0F0F0",
    "Weight":             "E8F4FF",
    "Speed":              "F4E8FF",
    "Vibration":          "FFE8F4",
    "Hand/Manual":        "F0FFE8",
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED EXTRACTION CONFIGURATION
# Tune here — no logic changes required.
# ────────────────────────────────────────────────────────────────────────────
EXTRACTION_CONFIG = {
    # PDF rendering
    "pdf_dpi":           150,     # DPI for PDF→image conversion (150 is sufficient for A0/A1 P&IDs)
    "max_image_size":    4096,    # Max pixel dimension per image sent to Vision API
    "jpeg_quality":       90,     # JPEG compression quality (0-100)

    # AI engine priority for Vision passes (first available + not-quota-exceeded wins)
    # Supported: 'gemini', 'openai'
    # Note: 'tesseract' is handled separately at the orchestration level (not a Vision engine)
    "ai_engines":        ["gemini", "openai"],

    # Full extraction engine order — controls priority waterfall in extract_instruments
    # 1=gemini, 2=tesseract (OCR, free, no quota), 3=openai (last resort)
    "engine_order":      ["gemini", "tesseract", "openai"],

    # Gemini model
    "gemini_model":      "gemini-2.0-flash",

    # Multi-orientation passes (handles vertical / slanted text)
    "enable_rotation":   True,    # Add 90° CW and 90° CCW passes
    "rotation_angles":   [90, 270],

    # Tiled quadrant scanning (handles dense drawings with tiny tags)
    "enable_tiling":     True,
    "tile_grid":         (2, 2),
    "tile_overlap":      0.12,

    # Tesseract OCR settings
    "enable_tesseract":  True,
    "tesseract_on_all":  True,      # Also run Tesseract on vector PDFs (catches circle text)
    "tesseract_dpi":     150,       # DPI for Tesseract rendering

    # Spatial word-proximity grouping (catches tags split across spans inside circles)
    "spatial_grouping":  True,
    "spatial_radius":    80,        # px at 150 DPI — approx instrument circle diameter

    # OpenAI Vision settings
    "max_tokens_primary":  16000,
    "max_tokens_tile":      8000,
    "temperature":           0.1,
    "model":           "gpt-4o",

    # Minimum text-layer tags before skipping Vision passes
    "min_text_tags":         1,

    # Gemini rate-limit retry: sleep this many seconds then retry once before disabling
    "gemini_retry_delay":    5,
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED DRAWING/P&ID NUMBER RESOLUTION
# Extract DWG/P&ID number from title-block style labels in PDF text.
# ────────────────────────────────────────────────────────────────────────────
DRAWING_NUMBER_CONFIG = {
    # Labels commonly used in title blocks
    "label_patterns": [
        r'\b(?:DWG|DRAWING)\s*(?:NO|NUMBER|#)\b',
        r'\bP\s*&\s*ID\s*(?:NO|NUMBER|#)\b',
        r'\bP\s*ID\s*(?:NO|NUMBER|#)\b',
        r'\bDOCUMENT\s*(?:NO|NUMBER|#)\b',
        r'\bDOC\s*(?:NO|NUMBER|#)\b',
    ],
    # Candidate number formats (kept broad but engineering-oriented)
    "value_patterns": [
        r'\b([A-Z0-9]{2,}(?:-[A-Z0-9]{2,}){2,})\b',
        r'\b([A-Z]{2,}[0-9]{1,}(?:-[A-Z0-9]{1,}){2,})\b',
        r'\b([A-Z0-9]{3,}(?:[./][A-Z0-9]{2,}){2,})\b',
    ],
    # Hard filters for candidate sanity
    "min_length": 8,
    "max_length": 64,
    "window_chars": 140,
}

# Optional S3 legend-sheet discovery. Used only to enrich interpretation,
# not to replace instrument extraction from the P&ID itself.
LEGEND_S3_CONFIG = {
    "enabled": True,
    "filename_keywords": [
        "legend", "legends", "symbol", "symbols", "abbreviation", "abbr",
    ],
    "preferred_extensions": [".pdf"],
    "max_list_keys": 400,
    "max_candidate_files": 3,
    "max_text_chars": 12000,
    "max_pages_per_file": 3,
}

# ────────────────────────────────────────────────────────────────────────────
# CONTROL SYSTEM TAG (CS TAG) DETECTION CONFIG
# Identifies whether an instrument is a DCS/CS instrument or a field device
# and, for field devices, derives the expected CS controller tag.
# Tune here — no logic changes required.
# ────────────────────────────────────────────────────────────────────────────
CS_TAG_CONFIG = {
    # Full ISA-5.1 function-code prefixes (the leading letters before the dash)
    # that identify DCS / control-system instruments.
    # For these: control_system_tag = tag_number (the instrument IS a CS tag).
    "dcs_function_codes": {
        # Flow controllers / recorders
        "FIC", "FRC", "FC", "FFC", "FFIC", "FQC",
        # Pressure controllers / recorders
        "PIC", "PRC", "PC", "PDC",
        # Temperature controllers / recorders
        "TIC", "TRC", "TC", "TDIC",
        # Level controllers / recorders
        "LIC", "LRC", "LC",
        # Analysis / composition controllers
        "AIC", "ARC", "AC",
        # Generic / multi-variable controllers
        "IC", "RC", "HC", "HIC", "HRC",
        "SC", "SIC", "SRC",
        "XC", "YC", "ZC",
        "QIC", "QC", "DC", "EC", "GC", "JC", "KC", "MC", "NC", "OC", "UC", "VC", "WC",
    },
    # Mapping: field-instrument function-code suffix → controller suffix to derive.
    # key   = suffix of field instrument (after the first measured-variable letter)
    # value = controller suffix to substitute
    # e.g.  FT-3901-01 (suffix="T") → FIC-3901-01 (suffix="IC")
    "transmitter_to_controller": {
        "T":   "IC",    # Transmitter          → Indicating Controller  (FT→FIC)
        "IT":  "IC",    # Indicating Transmitter → Indicating Controller
        "E":   "C",     # Element / Sensor     → Controller            (FE→FC)
    },
    # Regex patterns that identify explicit CS/DCS tag annotations in PDF text.
    # Group 1 must capture the tag value.
    "label_patterns": [
        r'(?:CS|DCS|PLC|F&?G|ESD|SIS|SCADA)\s*[-:–]?\s*(?:TAG|NO\.?|NUM\.?)?\s*[-:–]?\s*([A-Z]{2,6}[-]\d{3,6}(?:[-][A-Z0-9]{1,4})?)',
        r'\bCONTROL(?:LER)?\s+TAG\s*[-:–]\s*([A-Z]{2,6}[-]\d{3,6}(?:[-][A-Z0-9]{1,4})?)',
    ],
    # Characters of PDF text after the CS label to search for the tag value.
    "label_window_chars": 80,
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED CONTEXTUAL ENRICHMENT CONFIGURATION
# Regex patterns for extracting service context without Vision AI.
# Add/edit patterns here — no logic changes required.
# ────────────────────────────────────────────────────────────────────────────
ENRICHMENT_CONFIG = {
    # ── Piping line number patterns ──────────────────────────────────────
    # ADNOC / oil-gas format:  NPS["-symbol] - FLUID_CODE - AREA - SPEC
    # Examples: 16"-HC-3901-A2A   10-G-3901-A2A   2''-FG-1001-B3A   3/4"-IA-3901-A2A
    #
    # WHY MULTIPLE PATTERNS:
    #   • The NPS inch symbol ("  ″  ''  ʺ) encodes differently across PDF fonts.
    #   • Pattern 1: handles inch symbol between size digits and first dash.
    #   • Pattern 2: no inch symbol (size directly followed by dash).
    #   • Pattern 3: fraction pipe sizes (3/4", 1/2").
    #   • Pattern 4: partial match — SIZE-FLUID-AREA without spec class.
    #   • Pattern 5: area-first format used by SABIC and some FEED contractors.
    "line_no_re": [
        # Line-list style 5-part format: SIZE-FLUID-SEQUENCE-CLASS-INSULATION
        r'(?<!\w)(\d+(?:\.\d+)?["″\u2033\u02BA\'\u2019\uFF02]?\s*[-–—]\s*[A-Z]{1,4}\s*[-–—]\s*\d{4,6}\s*[-–—]\s*[A-Z]\d[A-Z]\d{1,2}\s*[-–—]\s*[A-Z]{1,2})(?!\w)',
        # Line-list style 6-part format: SIZE-AREA-FLUID-SEQUENCE-CLASS-INSULATION(optional)
        r'(?<!\w)(\d{1,2}["″\u2033\u02BA\'\u2019\uFF02]?\s*[-–—]\s*\d{1,2}\s*[-–—]\s*[A-Z]{1,2}\s*[-–—]\s*\d{4}\s*[-–—]\s*[0-9][A-Z0-9]{5}(?:\s*[-–—]\s*[A-Z]{1,2})?)(?!\w)',
        # Full 4-part: SIZE[inch]-FLUID-AREA-SPEC  (inch symbol optional, many variants)
        r'(?<!\w)(\d{1,3}(?:["″\u2033\u02BA\'\u2019\uFF02]{1,2})?[-]\s*[A-Z]{1,6}\s*[-]\s*\d{3,6}\s*[-]\s*[A-Z][0-9][A-Z0-9]{0,4})(?!\w)',
        # 3-part without SPEC: SIZE-FLUID-AREA  (the spec may be on a separate annotation)
        r'(?<!\w)(\d{1,3}["″\u2033\u02BA\'\u2019]{0,2}[-][A-Z]{1,6}[-]\d{3,6})(?![-A-Z0-9])',
        # Fraction pipe sizes: 3/4"-IA-3901-A2A
        r'(?<!\w)(\d(?:[/]\d)["″\u2033]{0,2}[-][A-Z]{1,6}[-]\d{3,6}(?:[-][A-Z][0-9][A-Z0-9]{0,4})?)(?!\w)',
        # Area-first format: AREA-FLUID-SIZE  e.g. 3901-G-16 or 3901-HC-6"-A2A
        r'(?<!\w)(\d{3,6}[-][A-Z]{1,6}[-]\d{1,3}["″\u2033]{0,2}(?:[-][A-Z][0-9][A-Z0-9]{0,4})?)(?!\w)',
        # Flexible join-product (words joined with dashes by block reconstruction)
        r'\b(\d{1,3}-[A-Z]{1,6}-\d{3,6}-[A-Z][0-9][A-Z0-9]{0,4})\b',
    ],

    # Structured line-number rules borrowed from the working line-list parser.
    # These build canonical line numbers from captured components instead of
    # relying only on raw regex matches.
    "line_no_structured_rules": [
        {
            "name": "five_part_line",
            "pattern": r'(?<!\w)(\d+(?:\.\d+)?)\s*["\']?\s*[-–—]\s*([A-Z]{1,4})\s*[-–—]\s*(\d{4,6})\s*[-–—]\s*([A-Z]\d[A-Z]\d{1,2})\s*[-–—]\s*([A-Z]{1,2})(?!\w)',
        },
        {
            "name": "six_part_line",
            "pattern": r'(?<!\w)(\d{1,2})\s*["\']?\s*[-–—]\s*(\d{1,2})\s*[-–—]\s*([A-Z]{1,2})\s*[-–—]\s*(\d{4})\s*[-–—]\s*([0-9][A-Z0-9]{5})(?:\s*[-–—]\s*([A-Z]{1,2}))?(?!\w)',
        },
    ],

    # ── Equipment tag patterns ───────────────────────────────────────────
    # Whitelist: only codes that are NOT in INSTRUMENT_CATEGORIES.
    # Single-letter equipment codes: V (vessel), E (exchanger), P (pump),
    #   C (compressor/column), K (compressor alt), T (tower/tank), R (reactor), D (drum).
    # Two-letter: LP, HP, HE, VV, KO, SD, TK, AC, WH (none are instrument prefixes).
    # Three-letter: SEP, FLR, CMP, PMP, SCR, EXC, KOD, BFW.
    "equipment_re": [
        # Single-letter ISA equipment codes (safe whitelist — none in INSTRUMENT_CATEGORIES)
        r'\b([VEPCKTR][-]\d{3,5}[A-Z]?(?:[/][A-Z])?)\b',
        # Two-letter equipment codes NOT in INSTRUMENT_CATEGORIES
        r'\b((?:LP|HP|HE|VV|KO|SD|TK|AC|WH|UD|CD|FP|MD|SG)[-]\d{3,5}[A-Z]?)\b',
        # Three-letter codes
        r'\b((?:SEP|FLR|CMP|PMP|SCR|EXC|KOD|BFW|SRT|SKD|VES)[-]\d{3,5}[A-Z]?)\b',
        # Drum variant: D-3901, D-3901A
        r'\b(D[-]\d{3,5}[A-Z]?)\b',
    ],

    # ── Fail-safe position annotations ──────────────────────────────────
    "fail_safe_re": [
        r'\b(FC)\b', r'\b(FO)\b', r'\b(FL)\b',
        r'FAIL[-\s]?(CLOSE[D]?|OPEN|LAST|LOCK(?:ED)?)',
    ],

    # ── Signal / communication type ──────────────────────────────────────
    "signal_re": [
        r'(4[-–/]20\s*m[Aa])',
        r'\b(HART)\b',
        r'\b(Fieldbus|Foundation\s*Fieldbus)\b',
        r'\b(Profibus|PROFIBUS)\b',
        r'\b(Wireless|WirelessHART|ISA100)\b',
        r'\b(Pneumatic|Pneum\.?)\b',
        r'(?<![A-Z])(DI|DO|AO)(?![A-Z])',
        r'\b(24\s*VDC?)\b',
        r'\b(On[-/]Off|Discrete|Digital)\b',
    ],

    # ── Set-point value + engineering unit ──────────────────────────────
    "set_point_re": [
        r'(\d+\.?\d*\s*(?:bara|barg|bar[ag]?|kPa[ag]?|MPa[ag]?|psia|psig|psi))\b',
        r'(\d+\.?\d*\s*(?:\xb0C|\xb0F|degC|degF))\b',
        r'(\d+\.?\d*\s*(?:kg\/h|t\/h|m3\/h|Nm3\/h|MMSCFD|SCFD))\b',
        r'(\d+\.?\d*\s*%\s*(?:FS|LRV|URV|SPAN)?)\b',
        r'(\d+\.?\d*\s*(?:rpm|Hz|m\/s))\b',
    ],

    # Spatial search radius for fail-safe/signal/setpoint (pixels at 150 DPI)
    "spatial_radius": 200,
    # Larger radius for line-number and equipment-number search
    # (pipe label can be 300-500px from the instrument circle on large A0 drawings)
    "spatial_radius_context": 350,
    # Fuzzy area-code tolerance — match line/equip tags whose area code is within
    # this many integer units of the instrument's own area code.
    "area_tolerance": 5,
    # Words to exclude from auto-generated service descriptions
    "exclude_desc_words": {
        "p&id", "drawing", "revision", "sheet", "dwg", "doc", "project", "client",
        "date", "appr", "chck", "eng", "rev", "by", "ref", "scale", "nts",
        "title", "north", "south", "tag", "no.", "number", "the", "and", "for",
        "to", "of", "in", "is", "at", "on", "an", "a",
    },
    # Min word count to use proximity text as service description
    "desc_min_words": 2,
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED INSTRUMENT VALIDATION / CLEAN-UP LAYER
# ----------------------------------------------------------------------------
# Addresses real-world P&ID feedback (see feature-instrument.instructions.md)
#   1. Random P&ID no (not title-block)  → per-page title-block scan
#   2. Tags inside line-numbers picked up as instruments (e.g. PG-45011 inside
#      "12"-13-PG-45011-A0JP08-F")                            → line-context filter
#   3. Tank-level labels mistakenly flagged as instruments
#      (e.g. LSH-800-300 = tank level, not an instrument)    → level-label heuristic
#   4. ISA-5.1 universal tag format: Unit-InstrumentTag-Seq  → format validator
#   5. Accessories auto-inference: TT usually has TE + TW    → accessories map
#   6. "Inline" function codes to prioritise                 → PG, PT, TG, TT …
#
# Every behaviour here can be switched off, or its threshold tuned, without
# touching any extraction regex or engine code.  Pure post-processing.
# ────────────────────────────────────────────────────────────────────────────
INSTRUMENT_VALIDATION_CONFIG = {
    "enabled": True,

    # 1 ── Per-page P&ID number extraction (from title block) -----------------
    "per_page_pid_no": {
        "enabled": True,
        # Fallback to drawing_info["pid_no"] when title-block detect fails
        "fallback_to_input": True,
    },

    # 2 ── Line-context filter ----------------------------------------------
    # If a candidate instrument tag appears as a TOKEN INSIDE a known
    # line-number pattern on the page text, drop it — it's a line spec, not
    # an instrument.  Patterns are the same as ENRICHMENT_CONFIG['line_no_re']
    # so tuning one updates both.
    "line_context_filter": {
        "enabled": True,
        # Minimum confidence before dropping: if we can't positively confirm
        # the tag lives inside a line number, keep it.
        "require_containment": True,
    },

    # 3 ── Level-label heuristic --------------------------------------------
    # Tags matching `^L[SI]?H?L?-\d{3}-\d{3}$` where the numbers are "round"
    # (like 800-300, 1000-500) are commonly tank level / nozzle labels, not
    # instruments.  Flag them via notes and expose in `warnings` field.
    "level_label_filter": {
        "enabled": True,
        # Function-code prefixes to check
        "prefixes": ["LSH", "LSL", "LSHH", "LSLL", "LI", "LG"],
        # If the trailing segment is a round multiple of this → warn
        "round_multiple": 50,
        # Action: 'warn'  → add note + warning, keep row
        #         'drop'  → remove from results
        "action": "warn",
    },

    # 4 ── ISA-5.1 universal tag-format validator ---------------------------
    # Canonical shape: <UNIT>-<INSTRUMENT_TAG>-<SEQUENCE>
    # UNIT     = 2–4 digits (loop / area / unit number)
    # TAG      = 2–6 letters (ISA function code)
    # SEQUENCE = 1–4 chars, digits optionally followed by a single letter
    "format_validator": {
        "enabled": True,
        # Regexes ordered most-specific → most-permissive.  Any match = valid.
        "valid_patterns": [
            # Standard 3-part:  FT-3901-01, PIT-2600-12A
            r"^[A-Z]{2,6}-\d{2,4}-\d{1,4}[A-Z]?$",
            # 2-part (older Gulf / ADNOC FEED): FT-3901A
            r"^[A-Z]{2,6}-\d{2,4}[A-Z]?$",
            # Unit-prefixed 4-part: 26-FT-3901-01
            r"^\d{2}-[A-Z]{2,6}-\d{2,4}-\d{1,4}[A-Z]?$",
        ],
        # When invalid → keep row but add warning + mark `format_valid=False`.
        "action": "warn",
    },

    # 5 ── Accessory auto-inference ----------------------------------------
    # Parent function code → list of accessory codes that are implied.
    # Soft-coded per ISA-5.1 best practice.  Accessories inherit the parent's
    # unit/sequence numbers and are marked `inferred=True` so engineers can
    # confirm visually on the P&ID.
    "accessories": {
        "enabled": True,
        "map": {
            # Temperature transmitter → element + thermowell
            "TT":  ["TE", "TW"],
            "TIT": ["TE", "TW"],
            # Flow transmitter → element / orifice
            "FT":  ["FE"],
            "FIT": ["FE"],
            # Pressure transmitter (often) → isolation valve manifold (not auto-created —
            # valves are not in INSTRUMENT_CATEGORIES, so we leave it off).
            # Level transmitter (DP-type) → HP/LP taps are piping, not instruments.
        },
        # Duplicate guard — skip accessory if the same code+unit+seq already exists
        "dup_guard": True,
    },

    # 6 ── Inline-instrument priority list ---------------------------------
    # Per feedback: extractions should start from basic inline instruments.
    # We mark these with `is_inline=True` and move them to the top of the
    # result list (index_no is recomputed).
    "inline_priority": {
        "enabled": True,
        "codes": ["PG", "PI", "PT", "TG", "TI", "TT", "LG", "LI", "LT", "FG", "FI", "FE"],
    },
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED SMART DEFAULTS for Fail-Safe / Signal / Set-Point columns
# ----------------------------------------------------------------------------
# Post-enrichment layer that (1) **validates** the values produced by the
# proximity regex scan against the instrument's function code/category, and
# (2) fills sensible engineering defaults where the drawing does not spell
# the value out explicitly (local gauges → "Local", transmitters → "4-20mA",
# safety valves → "FO", etc.).
#
# Everything here is tunable without touching extraction / regex logic.
# Set `enabled: False` to revert to pure drawing-only values.
# ────────────────────────────────────────────────────────────────────────────
SMART_FIELD_DEFAULTS_CONFIG = {
    "enabled": True,

    # ── FAIL-SAFE rules ─────────────────────────────────────────────────
    "fail_safe": {
        # Function-code prefix → default fail-safe position
        # (ISA-5.1 / industry convention; override per project if needed)
        "by_prefix": {
            # Shutdown / ESD valves default fail-close
            "SDV":  "FC",
            "ESDV": "FC",
            "ESV":  "FC",
            "XV":   "FC",   # on/off block valve — assume FC unless marked
            "BDV":  "FO",   # blow-down valve — fail-open to relieve
            # Pressure-relief / safety
            "PSV":  "FO",   # pressure safety valve — spring-operated, treat as FO
            "PRV":  "FO",
            "PSE":  "FO",   # rupture disc
            # Control valves — conservative industry default = FC
            # (process fluid isolation on air/power failure).  Project-specific
            # services (e.g. anti-surge PCV, cooling-water FCV) may differ and
            # should be corrected by the explicit drawing callout.
            "FCV":  "FC", "FV": "FC",
            "PCV":  "FC", "PV": "FC",
            "TCV":  "FC", "TV": "FC",
            "LCV":  "FC", "LV": "FC",
            "HCV":  "FC", "HV": "FC",
            "CV":   "FC",
            # Solenoid / motor-operated — energise to open → FC on power loss
            "SOV":  "FC", "MOV": "FC",
            "ZV":   "FC",
        },
        # Category-wide fallback (used only if prefix isn't in by_prefix above)
        "by_category": {
            "Shutdown & ESD":   "FC",
            "Pressure Relief":  "FO",
            "Motor & Solenoid": "FC",
        },
        # Clear value if it was wrongly extracted for a non-valve instrument
        "clear_for_non_valve": True,
        # Function codes that qualify as "valve" — others get fail_safe cleared
        "valve_prefixes": {
            "SDV", "ESDV", "ESV", "XV", "BDV", "PSV", "PRV", "PSE",
            "FCV", "PCV", "TCV", "LCV", "HCV", "FV", "PV", "TV", "LV",
            "CV", "PV", "ZV", "MV", "TSV",
        },
    },

    # ── SIGNAL-TYPE rules ───────────────────────────────────────────────
    "signal_type": {
        # Function-code prefix → default signal type
        "by_prefix": {
            # Transmitters — 4-20mA HART is the industry default
            "PT":  "4-20mA HART", "PIT": "4-20mA HART",
            "TT":  "4-20mA HART", "TIT": "4-20mA HART",
            "FT":  "4-20mA HART", "FIT": "4-20mA HART",
            "LT":  "4-20mA HART", "LIT": "4-20mA HART",
            "AT":  "4-20mA HART", "AIT": "4-20mA HART",
            "DPT": "4-20mA HART", "DT":  "4-20mA HART",
            "VT":  "4-20mA HART", "WT":  "4-20mA HART",
            # Switches / discrete devices
            "PSH": "Discrete (DI)", "PSL": "Discrete (DI)",
            "PSHH":"Discrete (DI)", "PSLL":"Discrete (DI)",
            "TSH": "Discrete (DI)", "TSL": "Discrete (DI)",
            "LSH": "Discrete (DI)", "LSL": "Discrete (DI)",
            "LSHH":"Discrete (DI)", "LSLL":"Discrete (DI)",
            "FSH": "Discrete (DI)", "FSL": "Discrete (DI)",
            "ZSH": "Discrete (DI)", "ZSL": "Discrete (DI)",
            # Solenoids / outputs
            "SOV": "Discrete (DO)", "XY": "Discrete (DO)",
            # Valves (on/off vs modulating)
            "SDV": "Discrete (DO)", "ESDV": "Discrete (DO)",
            "XV":  "Discrete (DO)", "BDV": "Discrete (DO)",
            "FCV": "4-20mA", "PCV": "4-20mA",
            "TCV": "4-20mA", "LCV": "4-20mA",
            # Relays / computers
            "FY": "Digital", "PY": "Digital", "TY": "Digital", "LY": "Digital",
            # Local gauges / indicators / sight glasses → NO signal
            "PG":  "Local (Mechanical)", "PI":  "Local (Mechanical)",
            "TG":  "Local (Mechanical)", "TI":  "Local (Mechanical)",
            "LG":  "Local (Mechanical)", "LI":  "Local (Mechanical)",
            "FG":  "Local (Mechanical)", "FI":  "Local (Mechanical)",
            "SG":  "Local (Mechanical)",
            # Elements — upstream of a transmitter, no signal of its own
            "TE":  "RTD / Thermocouple",
            "TW":  "—",
            "FE":  "—",
            "PE":  "—",
            "LE":  "—",
        },
        # Categories that NEVER carry an electronic signal — always force "Local"
        "local_only_categories": set(),
        # If the regex pass put "4-20mA" on a local-gauge code, override it
        "validate_against_prefix": True,
    },

    # ── SET-POINT rules ─────────────────────────────────────────────────
    "set_point": {
        # When no numeric set point found on the drawing, provide a
        # unit hint based on the instrument category so the engineer
        # sees what kind of value is expected.
        "unit_hint_by_category": {
            "Pressure":              "—— bar(g)",
            "Differential Pressure": "—— mbar",
            "Temperature":           "—— °C",
            "Level":                 "—— %",
            "Flow":                  "—— m³/h",
            "Analysis":              "—— ppm",
            "Shutdown & ESD":        "Trip on alarm",
            "Pressure Relief":       "—— bar(g) (set)",
            "Motor & Solenoid":      "Energise / De-energise",
            "Position":              "Open / Closed",
        },
        # Fallback when category is unknown — derive from the first letter
        # of the ISA function code (P = Pressure, T = Temperature, etc.).
        "unit_hint_by_first_letter": {
            "P": "—— bar(g)",
            "T": "—— °C",
            "F": "—— m³/h",
            "L": "—— %",
            "A": "—— ppm",
            "D": "—— mbar",     # D = Differential
            "S": "—— rpm",      # S = Speed
            "V": "—— m/s",      # V = Vibration/Velocity
            "W": "—— kg",       # W = Weight
            "J": "—— A",        # J = Power (current)
            "I": "—— A",        # I = Current
            "E": "—— V",        # E = Voltage
            "Z": "Open / Closed",  # Z = Position
        },
        # Switches always have a set-point; mark "Field-adjustable" when missing
        "switch_prefixes": {
            "PSH", "PSL", "PSHH", "PSLL",
            "TSH", "TSL", "TSHH", "TSLL",
            "LSH", "LSL", "LSHH", "LSLL",
            "FSH", "FSL", "ZSH", "ZSL",
        },
        "switch_default": "Field-adjustable",
        # Control valves get a default when no explicit SP found on drawing
        "control_valve_prefixes": {
            "FCV", "PCV", "TCV", "LCV", "HCV",
            "FV", "PV", "TV", "LV", "HV", "CV",
        },
        "control_valve_default": "Set by DCS loop",
        # Transmitters / indicators / elements generally have no set-point
        "no_setpoint_prefixes": {
            "PT","PIT","TT","TIT","FT","FIT","LT","LIT","AT","AIT",
            "DPT","PG","PI","TG","TI","LG","LI","FG","FI","SG",
            "TE","TW","FE","PE","LE",
        },
        "no_setpoint_marker": "—",
        # Validate current value's unit matches the category;
        # if mismatch (e.g. "100 bar" on a Temperature instrument) → clear.
        "validate_units_by_category": True,
        "category_units": {
            "Pressure":              {"bar", "barg", "bara", "kpa", "mpa", "psi", "psig", "psia"},
            "Differential Pressure": {"mbar", "kpa", "pa", "inh2o", "mmh2o", "bar"},
            "Temperature":           {"°c", "°f", "degc", "degf", "k"},
            "Flow":                  {"m3/h", "nm3/h", "kg/h", "t/h", "mmscfd", "scfd", "gpm", "bpd", "lpm"},
            "Level":                 {"%", "mm", "m", "ft", "in"},
            "Analysis":              {"ppm", "ppb", "%", "mg/l"},
        },
    },
}

# Human-readable verb per instrument category — used to build default service descriptions
_SERVICE_VERB_MAP = {
    "Flow":                  "Flow Measurement",
    "Pressure":              "Pressure Measurement",
    "Temperature":           "Temperature Measurement",
    "Level":                 "Level Measurement",
    "Differential Pressure": "Differential Pressure Measurement",
    "Analysis":              "Process Analyser",
    "Safety":                "Pressure Safety Relief",
    "Shutdown & ESD":        "Shutdown / ESD Control",
    "Control Valves":        "Flow Control Valve",
    "Motor & Solenoid":      "Actuated Valve",
    "Position":              "Valve Position Monitor",
    "Restriction":           "Flow Restriction Orifice",
    "Speed":                 "Speed Measurement",
    "Vibration":             "Vibration Monitor",
    "Weight":                "Weight / Load Measurement",
    "Hand/Manual":           "Manual Indication",
    "Special":               "Special Purpose Device",
}


class InstrumentIndexService:
    """
    Extract ALL instrument tags from a P&ID drawing.
    Engine waterfall (soft-coded via EXTRACTION_CONFIG['ai_engines']):
      1. PyMuPDF text-layer  — free, instant
      2. Gemini Vision       — primary AI (free tier, 1M context)
      3. OpenAI Vision       — fallback AI (GPT-4o)
      4. Tesseract OCR       — last resort for pure scanned PDFs
    """

    def __init__(self):
        self.extraction_config  = EXTRACTION_CONFIG.copy()
        self._quota_exceeded    = False   # set True on OpenAI 429
        self._gemini_quota_exceeded = False
        self.openai_client      = self._init_openai()
        self.gemini_client      = self._init_gemini()
        self.tesseract_available = self._check_tesseract()

    # ────────────────────────────────────────────────────────────────────
    # Initialisation helpers
    # ────────────────────────────────────────────────────────────────────

    def _init_openai(self):
        try:
            import openai
            api_key = os.environ.get("OPENAI_API_KEY")
            if not api_key:
                logger.warning("[InstrumentIndex] OPENAI_API_KEY not set")
                return None
            client = openai.OpenAI(api_key=api_key)
            logger.info("[InstrumentIndex] ✅ OpenAI client initialised")
            return client
        except Exception as e:
            logger.warning(f"[InstrumentIndex] OpenAI init skipped: {e}")
            return None

    def _init_gemini(self):
        try:
            from google import genai
            api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
            if not api_key:
                logger.warning("[InstrumentIndex] GEMINI_API_KEY not set")
                return None
            client = genai.Client(api_key=api_key)
            logger.info("[InstrumentIndex] ✅ Gemini client initialised")
            return client
        except ImportError:
            logger.warning("[InstrumentIndex] google-genai not installed — Gemini disabled")
            return None
        except Exception as e:
            logger.warning(f"[InstrumentIndex] Gemini init skipped: {e}")
            return None

    def _check_tesseract(self):
        try:
            import pytesseract
            pytesseract.get_tesseract_version()
            logger.info("[InstrumentIndex] ✅ Tesseract available")
            return True
        except Exception:
            logger.info("[InstrumentIndex] Tesseract not available (optional)")
            return False

    # ────────────────────────────────────────────────────────────────────
    # Public entry point
    # ────────────────────────────────────────────────────────────────────

    def extract_instruments(self, pid_bytes, drawing_info, legend_context_override=None):
        """
        Extract all instrument tags from a P&ID PDF (or image).

        Extraction strategy (free-first, AI-enrichment):
          Step 1 — PyMuPDF text layer (3-pass: plain text + block reconstruction + spatial grouping)
          Step 2 — Tesseract OCR alongside text layer on ALL PDFs (catches circle-embedded tags)
          Step 3 — AI enrichment (Gemini, then OpenAI) to fill service descriptions & add missed tags

        Args:
            pid_bytes  : raw file bytes
            drawing_info: dict with drawing_number, drawing_title, revision, project_name

        Returns:
            list[dict]: instrument records
        """
        MIN_TEXT_TAGS = self.extraction_config.get("min_text_tags", 1)
        engine_order  = self.extraction_config.get("engine_order", ["gemini", "tesseract", "openai"])
        cfg = self.extraction_config

        try:
            # Resolve drawing number from title-block text when available.
            # This keeps P&ID No aligned with DWG NO per drawing.
            drawing_info = self._resolve_drawing_info_from_pdf(pid_bytes, drawing_info)
            legend_context = self._merge_legend_contexts(
                legend_context_override,
                self._load_legend_context_from_s3(drawing_info),
            )

            # ── Step 1: PyMuPDF text layer (free, always runs) ────────────
            text_instruments = self._extract_with_text_layer(pid_bytes, drawing_info)
            logger.info(f"[InstrumentIndex] Text-layer: {len(text_instruments)} tags")

            # ── Step 2: Tesseract OCR (free, runs on ALL PDFs) ────────────
            # Catches tags in instrument circles that the text layer may split or miss.
            tess_instruments: list = []
            if self.tesseract_available and cfg.get("tesseract_on_all", True):
                logger.info("[InstrumentIndex] Tesseract OCR — running alongside text layer…")
                tess_instruments = self._extract_with_tesseract(pid_bytes, drawing_info)
                logger.info(f"[InstrumentIndex] Tesseract OCR: {len(tess_instruments)} tags")

            # Merge free-tier results
            all_free = self._merge_instruments(text_instruments + tess_instruments)
            logger.info(f"[InstrumentIndex] Free-tier combined: {len(all_free)} unique tags")

            # ── Step 2.5: Contextual enrichment — pattern-based, zero AI quota ──
            # Fills: loop_number, service_description, line_number, equipment_number,
            #        fail_safe, signal_type, set_point using PDF text layer patterns.
            if all_free:
                logger.info("[InstrumentIndex] Running contextual enrichment (pattern-based)…")
                all_free = self._enrich_from_pdf_context(all_free, pid_bytes)

                # Gemini text-only enrichment (much lighter quota than Vision)
                if not self._gemini_quota_exceeded and self.gemini_client:
                    try:
                        import fitz as _fitz
                        _doc = _fitz.open(stream=pid_bytes, filetype="pdf")
                        _pdf_text = "\n".join(
                            _doc[_p].get_text("text") or "" for _p in range(len(_doc))
                        )
                        _doc.close()
                        logger.info("[InstrumentIndex] Running Gemini text enrichment…")
                        all_free = self._enrich_with_gemini_text(
                            all_free,
                            _pdf_text,
                            legend_text=legend_context.get("text", ""),
                            legend_files=legend_context.get("files", []),
                        )
                    except Exception as _etxt:
                        logger.warning(f"[InstrumentIndex] Gemini text enrichment setup error: {_etxt}")

            # ── Step 3: AI Vision (enrich / fill gaps) ────────────────────
            if len(all_free) >= MIN_TEXT_TAGS:
                # Good free-tier results — use AI to enrich service descriptions and find any missed tags
                enriched = []
                for eng in [e for e in engine_order if e != "tesseract"]:
                    if eng == "gemini" and (self._gemini_quota_exceeded or not self.gemini_client):
                        continue
                    if eng == "openai" and (self._quota_exceeded or not self.openai_client):
                        continue
                    try:
                        pages = self._to_jpeg_pages(pid_bytes)
                        if pages:
                            logger.info(f"[InstrumentIndex] AI enrichment via {eng}…")
                            enriched = self._vision_pass(
                                pages[0], drawing_info, 1,
                                extra_hint=(
                                    "Enrich the already-extracted tags with service descriptions, line numbers, "
                                    "signal types, and fail-safe positions. Also add any missed tags. "
                                    "Focus on ALL instrument circles/bubbles."
                                ),
                                mode=f"enrich_{eng}",
                                max_tokens=cfg["max_tokens_primary"],
                                only_engine=eng,
                            )
                    except Exception as ve:
                        logger.warning(f"[InstrumentIndex] {eng} enrichment skipped: {ve}")
                    if enriched:
                        break

                all_instruments = self._merge_instruments(all_free + enriched)

            else:
                # Sparse/no free-tier results — try full AI Vision extraction
                logger.info("[InstrumentIndex] Sparse free-tier — trying full AI Vision extraction…")
                all_instruments = list(all_free)

                for engine in [e for e in engine_order if e != "tesseract"]:
                    if engine == "gemini" and (self._gemini_quota_exceeded or not self.gemini_client):
                        logger.info("[InstrumentIndex] Gemini unavailable — skipping")
                        continue
                    if engine == "openai" and (self._quota_exceeded or not self.openai_client):
                        logger.info("[InstrumentIndex] OpenAI unavailable — skipping")
                        continue

                    logger.info(f"[InstrumentIndex] Full AI scan via {engine}…")
                    try:
                        pages = self._to_jpeg_pages(pid_bytes)
                    except Exception as pe:
                        logger.error(f"[InstrumentIndex] PDF→image failed: {pe}")
                        continue

                    vision_all: list = []
                    seen_vision: set = set()
                    for page_no, jpeg_page in enumerate(pages, start=1):
                        page_insts = self._analyse_page(jpeg_page, drawing_info, page_no, only_engine=engine)
                        for inst in page_insts:
                            tag = (inst.get("tag_number") or "").strip().upper()
                            if tag and tag not in seen_vision:
                                seen_vision.add(tag)
                                vision_all.append(inst)
                            elif not tag:
                                vision_all.append(inst)

                    if vision_all:
                        all_instruments = self._merge_instruments(all_instruments + vision_all)
                        logger.info(f"[InstrumentIndex] {engine} full scan: +{len(vision_all)} tags → total {len(all_instruments)}")
                        break

            # Sequential index numbers
            for i, inst in enumerate(all_instruments, start=1):
                inst["index_no"] = i

            # Ensure drawing number is populated
            dn = drawing_info.get("drawing_number", "")
            pid_no = drawing_info.get("pid_no") or dn
            if dn or pid_no:
                for inst in all_instruments:
                    if inst.get("pid_no") in ("N/A", "", None):
                        inst["pid_no"] = pid_no
                    if inst.get("drawing_number") in ("N/A", "", None):
                        inst["drawing_number"] = dn

            # ── Smart validation / accessory inference / inline priority ──
            # Soft-coded via INSTRUMENT_VALIDATION_CONFIG (see top of module).
            # Purely additive: filters out false positives (tags inside line
            # numbers, tank-level labels), flags non-ISA-5.1 formats,
            # synthesises accessories (TE+TW for TT, FE for FT), and re-orders
            # inline instruments (PG/PT/TG/TT…) first.
            try:
                all_instruments = self._apply_post_validation(
                    all_instruments, pid_bytes, drawing_info
                )
            except Exception as ve:
                logger.warning(f"[InstrumentIndex] Post-validation skipped: {ve}")

            logger.info(f"[InstrumentIndex] ✅ Total unique instruments: {len(all_instruments)}")
            return all_instruments

        except Exception as e:
            logger.error(f"[InstrumentIndex] extract_instruments error: {e}", exc_info=True)
            return []

    # ────────────────────────────────────────────────────────────────────
    # Contextual enrichment — fills service description, line no., etc.
    # 100 % pattern-based, zero AI quota consumed.
    # ────────────────────────────────────────────────────────────────────

    # ────────────────────────────────────────────────────────────────────
    # SMART VALIDATION / CLEAN-UP LAYER
    # All soft-coded via INSTRUMENT_VALIDATION_CONFIG.  Addresses expert
    # feedback: random P&ID no, tags inside line numbers, tank-level labels,
    # tag-format checks, accessory inference, inline-instrument priority.
    # ────────────────────────────────────────────────────────────────────

    def _extract_per_page_pid_numbers(self, pid_bytes):
        """
        Return dict {page_number (1-based) → detected P&ID No} using the same
        title-block heuristics as the global detector, but scanned page by page.
        If a page has no clear candidate, the entry is omitted.
        """
        result = {}
        try:
            import fitz
        except ImportError:
            return result
        try:
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception:
            return result
        try:
            for page_idx in range(len(doc)):
                try:
                    txt = doc[page_idx].get_text("text") or ""
                    cand = self._extract_drawing_number_from_text(txt)
                    if cand:
                        result[page_idx + 1] = cand
                except Exception:
                    continue
        finally:
            doc.close()
        return result

    def _build_page_text_map(self, pid_bytes):
        """Return {page_number → full text} for line-context filtering."""
        out = {}
        try:
            import fitz
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception:
            return out
        try:
            for page_idx in range(len(doc)):
                try:
                    out[page_idx + 1] = doc[page_idx].get_text("text") or ""
                except Exception:
                    continue
        finally:
            doc.close()
        return out

    def _tag_is_inside_line_number(self, tag, page_texts):
        """
        True if `tag` appears as a substring of any line-number token on any
        page.  Uses ENRICHMENT_CONFIG['line_no_re'] so line-list tuning
        automatically applies here.
        """
        if not tag or not page_texts:
            return False
        tag_u = tag.strip().upper()
        line_patterns = [re.compile(p, re.IGNORECASE) for p in ENRICHMENT_CONFIG.get("line_no_re", [])]
        for text in page_texts.values():
            if not text:
                continue
            for lp in line_patterns:
                for m in lp.finditer(text):
                    token = m.group(0).upper()
                    # Strip separators/spaces for robust containment check
                    token_clean = re.sub(r"\s+", "", token)
                    if tag_u in token_clean:
                        # Confirm it's INSIDE the line (not the full line itself)
                        if token_clean != tag_u and len(token_clean) > len(tag_u):
                            return True
        return False

    def _is_level_label_like(self, tag):
        """
        Heuristic: tags shaped like LSH-800-300 where the trailing numbers are
        round (multiples of e.g. 50) are usually tank level / nozzle labels
        rather than instrument tags.
        """
        cfg = INSTRUMENT_VALIDATION_CONFIG.get("level_label_filter", {})
        if not cfg.get("enabled"):
            return False
        prefixes = tuple(cfg.get("prefixes", []))
        round_mult = int(cfg.get("round_multiple", 50)) or 50
        tag_u = (tag or "").upper().strip()
        m = re.match(r"^([A-Z]{1,4})-(\d{2,4})-(\d{2,4})$", tag_u)
        if not m:
            return False
        if not m.group(1) in prefixes:
            return False
        try:
            n1 = int(m.group(2))
            n2 = int(m.group(3))
        except ValueError:
            return False
        # Both numbers round multiples → strong level-label signal
        return (n1 % round_mult == 0) and (n2 % round_mult == 0)

    def _validate_tag_format(self, tag):
        """Return True if the tag matches any of the ISA-5.1 universal shapes."""
        cfg = INSTRUMENT_VALIDATION_CONFIG.get("format_validator", {})
        if not cfg.get("enabled"):
            return True
        tag_u = (tag or "").strip().upper()
        for pat in cfg.get("valid_patterns", []):
            if re.match(pat, tag_u):
                return True
        return False

    def _infer_accessories(self, instruments, dn, rev):
        """
        Given a list of extracted instruments, create accessory stub records
        (TE, TW for TT; FE for FT, etc.) per INSTRUMENT_VALIDATION_CONFIG.
        Accessories inherit unit/sequence numbers and are marked inferred=True.
        """
        cfg = INSTRUMENT_VALIDATION_CONFIG.get("accessories", {})
        if not cfg.get("enabled"):
            return []
        amap = cfg.get("map", {})
        dup_guard = cfg.get("dup_guard", True)

        existing_tags = {
            (i.get("tag_number") or "").strip().upper() for i in instruments
        }
        inferred = []
        for inst in list(instruments):
            tag = (inst.get("tag_number") or "").strip().upper()
            m = re.match(r"^([A-Z]{2,6})(-.+)$", tag)
            if not m:
                continue
            prefix = m.group(1)
            suffix = m.group(2)  # e.g. "-3901-01"
            if prefix not in amap:
                continue
            for acc_code in amap[prefix]:
                new_tag = f"{acc_code}{suffix}"
                if dup_guard and new_tag in existing_tags:
                    continue
                if acc_code not in INSTRUMENT_CATEGORIES:
                    continue
                rec = self._make_instrument_record(
                    new_tag, acc_code, INSTRUMENT_CATEGORIES[acc_code],
                    dn, rev, f"Inferred accessory of {tag}"
                )
                rec["inferred"] = True
                rec["parent_tag"] = tag
                rec["service_description"] = (
                    inst.get("service_description") or ""
                )
                rec["line_number"] = inst.get("line_number") or "N/A"
                rec["equipment_number"] = inst.get("equipment_number") or "N/A"
                rec["pid_no"] = inst.get("pid_no") or "N/A"
                rec["drawing_number"] = inst.get("drawing_number") or dn
                rec["loop_number"] = inst.get("loop_number") or self._derive_loop_number(new_tag)
                inferred.append(rec)
                existing_tags.add(new_tag)
        return inferred

    def _apply_smart_field_defaults(self, instruments):
        """
        Validate + intelligently default the Fail-Safe / Signal / Set-Point
        columns using SMART_FIELD_DEFAULTS_CONFIG.  Runs after regex-based
        enrichment so explicit drawing values always win; we only touch a
        field if it is N/A or clearly mis-assigned (e.g. 4-20mA on a local
        gauge, or a pressure set-point on a temperature instrument).
        """
        cfg = SMART_FIELD_DEFAULTS_CONFIG
        if not cfg.get("enabled", True) or not instruments:
            return instruments

        fs_cfg  = cfg.get("fail_safe", {})
        sig_cfg = cfg.get("signal_type", {})
        sp_cfg  = cfg.get("set_point", {})

        fs_prefix_map = fs_cfg.get("by_prefix", {})
        fs_cat_map    = fs_cfg.get("by_category", {})
        valve_prefixes = set(fs_cfg.get("valve_prefixes", set()))

        sig_prefix_map = sig_cfg.get("by_prefix", {})

        sp_unit_hint   = sp_cfg.get("unit_hint_by_category", {})
        sp_hint_letter = sp_cfg.get("unit_hint_by_first_letter", {})
        sp_switches    = set(sp_cfg.get("switch_prefixes", set()))
        sp_cvalves     = set(sp_cfg.get("control_valve_prefixes", set()))
        sp_no_setpt    = set(sp_cfg.get("no_setpoint_prefixes", set()))
        sp_units_cat   = sp_cfg.get("category_units", {})

        cleared_fs = cleared_sig = cleared_sp = 0
        filled_fs = filled_sig = filled_sp = 0

        for inst in instruments:
            tag = (inst.get("tag_number") or "").strip().upper()
            m = re.match(r"^([A-Z]{2,6})", tag)
            prefix = m.group(1) if m else ""
            category = inst.get("category") or ""

            # ── 1) FAIL-SAFE ─────────────────────────────────────────────
            cur_fs = (inst.get("fail_safe") or "").strip()
            is_valve = prefix in valve_prefixes
            if fs_cfg.get("clear_for_non_valve", True) and not is_valve:
                if cur_fs and cur_fs not in ("N/A", "—", "-"):
                    inst["fail_safe"] = "N/A"
                    cleared_fs += 1
                    cur_fs = "N/A"
            if cur_fs in ("", "N/A", None) and is_valve:
                default_fs = fs_prefix_map.get(prefix) or fs_cat_map.get(category)
                if default_fs:
                    inst["fail_safe"] = default_fs
                    filled_fs += 1

            # ── 2) SIGNAL TYPE ───────────────────────────────────────────
            cur_sig = (inst.get("signal_type") or "").strip()
            prefix_default_sig = sig_prefix_map.get(prefix)

            # Validate: override a clearly-wrong value on local-only devices
            if sig_cfg.get("validate_against_prefix", True) and prefix_default_sig:
                looks_local = prefix_default_sig.startswith("Local") or prefix_default_sig == "—"
                if looks_local and cur_sig and cur_sig not in ("N/A", "", "—"):
                    # Only override if current value looks like an electronic signal
                    if re.search(r"(4-20|HART|FIELD|PROFI|DI|DO|DISCRETE|DIGITAL)", cur_sig, re.IGNORECASE):
                        inst["signal_type"] = prefix_default_sig
                        cleared_sig += 1
                        cur_sig = prefix_default_sig

            if cur_sig in ("", "N/A", None) and prefix_default_sig:
                inst["signal_type"] = prefix_default_sig
                filled_sig += 1

            # ── 3) SET POINT ─────────────────────────────────────────────
            cur_sp = (inst.get("set_point") or "").strip()

            # Validate current set-point's unit matches category
            if sp_cfg.get("validate_units_by_category", True) and cur_sp and cur_sp not in ("N/A", "—"):
                expected_units = sp_units_cat.get(category)
                if expected_units:
                    sp_low = cur_sp.lower()
                    if not any(u in sp_low for u in expected_units):
                        inst["set_point"] = "N/A"
                        cleared_sp += 1
                        cur_sp = "N/A"

            if cur_sp in ("", "N/A", None):
                if prefix in sp_switches:
                    inst["set_point"] = sp_cfg.get("switch_default", "Field-adjustable")
                    filled_sp += 1
                elif prefix in sp_no_setpt:
                    inst["set_point"] = sp_cfg.get("no_setpoint_marker", "—")
                    filled_sp += 1
                elif prefix in sp_cvalves:
                    inst["set_point"] = sp_cfg.get("control_valve_default", "Set by DCS loop")
                    filled_sp += 1
                elif category in sp_unit_hint:
                    inst["set_point"] = sp_unit_hint[category]
                    filled_sp += 1
                elif prefix and prefix[0] in sp_hint_letter:
                    inst["set_point"] = sp_hint_letter[prefix[0]]
                    filled_sp += 1

        logger.info(
            f"[SmartDefaults] fail_safe: cleared={cleared_fs} filled={filled_fs} | "
            f"signal: cleared={cleared_sig} filled={filled_sig} | "
            f"set_point: cleared={cleared_sp} filled={filled_sp}"
        )
        return instruments

    def _apply_post_validation(self, instruments, pid_bytes, drawing_info):
        """
        Orchestrate all post-processing fixes.  Pure add-on — never modifies
        extraction/regex behaviour.  Safe to disable via
        INSTRUMENT_VALIDATION_CONFIG['enabled'] = False.
        """
        if not INSTRUMENT_VALIDATION_CONFIG.get("enabled", True):
            return instruments
        if not instruments:
            return instruments

        dn = drawing_info.get("drawing_number", "")
        rev = drawing_info.get("revision", "0")

        # 1) Per-page P&ID number stamping
        ppg_cfg = INSTRUMENT_VALIDATION_CONFIG.get("per_page_pid_no", {})
        if ppg_cfg.get("enabled", True):
            pid_per_page = self._extract_per_page_pid_numbers(pid_bytes)
            if pid_per_page:
                for inst in instruments:
                    page = inst.get("page") or inst.get("page_number")
                    resolved = pid_per_page.get(page)
                    if resolved:
                        inst["pid_no"] = resolved
                    elif ppg_cfg.get("fallback_to_input", True):
                        inst.setdefault("pid_no", drawing_info.get("pid_no") or dn)

        # Pre-compute the page→text map once for the line-context filter
        page_texts = self._build_page_text_map(pid_bytes)

        # 2) Line-context filter  +  3) Level-label filter  +  4) Format check
        lc_cfg = INSTRUMENT_VALIDATION_CONFIG.get("line_context_filter", {})
        lvl_cfg = INSTRUMENT_VALIDATION_CONFIG.get("level_label_filter", {})
        filtered = []
        dropped = 0
        for inst in instruments:
            tag = (inst.get("tag_number") or "").strip().upper()
            warnings = list(inst.get("warnings") or [])

            # 2) inside line-number?
            if lc_cfg.get("enabled", True) and self._tag_is_inside_line_number(tag, page_texts):
                dropped += 1
                logger.info(
                    f"[Validator] ⛔ drop '{tag}' — appears inside line number token"
                )
                continue

            # 3) tank-level label?
            if self._is_level_label_like(tag):
                if lvl_cfg.get("action", "warn") == "drop":
                    dropped += 1
                    logger.info(
                        f"[Validator] ⛔ drop '{tag}' — tank-level label, not instrument"
                    )
                    continue
                warnings.append("Possible tank-level / nozzle label — verify manually")

            # 4) ISA-5.1 universal format check
            if not self._validate_tag_format(tag):
                warnings.append("Tag does not match ISA-5.1 Unit-Tag-Sequence format")
                inst["format_valid"] = False
            else:
                inst["format_valid"] = True

            if warnings:
                inst["warnings"] = warnings
            filtered.append(inst)

        if dropped:
            logger.info(f"[Validator] Dropped {dropped} tag(s) via line-context / level-label filters")

        # 5) Accessory auto-inference
        inferred = self._infer_accessories(filtered, dn, rev)
        if inferred:
            logger.info(f"[Validator] ➕ {len(inferred)} accessory instruments inferred")
            filtered.extend(inferred)

        # 5b) Smart defaults for Fail-Safe / Signal / Set-Point
        filtered = self._apply_smart_field_defaults(filtered)

        # 6) Inline-instrument priority re-sort
        ip_cfg = INSTRUMENT_VALIDATION_CONFIG.get("inline_priority", {})
        if ip_cfg.get("enabled", True):
            inline_codes = set(c.upper() for c in ip_cfg.get("codes", []))
            def _inline_key(inst):
                tag = (inst.get("tag_number") or "").upper()
                m = re.match(r"^([A-Z]{2,6})", tag)
                code = m.group(1) if m else ""
                is_inline = code in inline_codes
                inst["is_inline"] = is_inline
                # sort: inline first (0), then others (1); stable within group
                return (0 if is_inline else 1,)
            filtered.sort(key=_inline_key)

        # Rebuild sequential index numbers
        for i, inst in enumerate(filtered, start=1):
            inst["index_no"] = i

        return filtered

    def _enrich_from_pdf_context(self, instruments, pid_bytes):
        """
        Post-extraction contextual enrichment using the PDF text layer.

        Fills N/A fields without consuming any AI quota:
          loop_number         — derived from the tag itself (deterministic)
          service_description — category verb + area code
          line_number         — piping line tag matched by area code + spatial proximity
          equipment_number    — equipment tag matched by area code + spatial proximity
          fail_safe           — FC / FO / FL from spatial proximity
          signal_type         — 4-20mA / HART / Discrete from proximity + global scan
          set_point           — value+unit from spatial proximity

        Four-layer strategy:
          L1: Deterministic  — loop_number + default service_description (always).
          L2: Block scan     — joins block lines to reconstruct multi-span line nos.
          L3: Spatial        — finds line nos, equip tags, fail-safe, signal, setpoint
                               near each instrument's position on the page.
          L4: Area-code      — matches globally scanned line nos / equip tags by the
                               shared numeric area code (with fuzzy ±tolerance).
        """
        ec = ENRICHMENT_CONFIG

        # Pre-compile all enrichment patterns (once per call)
        line_no_pats = [re.compile(p, re.IGNORECASE) for p in ec["line_no_re"]]
        line_no_structured_rules = [
            {
                "name": rule["name"],
                "pattern": re.compile(rule["pattern"], re.IGNORECASE),
            }
            for rule in ec.get("line_no_structured_rules", [])
        ]
        equip_pats   = [re.compile(p, re.IGNORECASE) for p in ec["equipment_re"]]
        fail_pats    = [re.compile(p, re.IGNORECASE) for p in ec["fail_safe_re"]]
        signal_pats  = [re.compile(p, re.IGNORECASE) for p in ec["signal_re"]]
        sp_pats      = [re.compile(p, re.IGNORECASE) for p in ec["set_point_re"]]

        radius         = ec["spatial_radius"]
        ctx_radius     = ec.get("spatial_radius_context", 350)
        area_tolerance = ec.get("area_tolerance", 5)
        stopwords      = ec["exclude_desc_words"]
        _area_re       = re.compile(r'(\d{3,6})')

        def _first_match(text, pats):
            text = _normalize_engineering_text(text)
            for p in pats:
                m = p.search(text)
                if m:
                    return _normalize_context_match(m.group(1) if m.lastindex else m.group(0))
            return None

        def _all_matches(text, pats):
            text = _normalize_engineering_text(text)
            out = []
            for p in pats:
                out.extend(
                    _normalize_context_match(m.group(1) if m.lastindex else m.group(0))
                    for m in p.finditer(text)
                )
            return out

        def _normalize_engineering_text(text):
            if not text:
                return ''
            cleaned = str(text)
            cleaned = (cleaned
                .replace('\u2013', '-')
                .replace('\u2014', '-')
                .replace('_', '-')
                .replace('\u201c', '"')
                .replace('\u201d', '"')
                .replace('\u2033', '"')
                .replace('\u02BA', '"')
                .replace('\uFF02', '"')
                .replace('\u2018', "'")
                .replace('\u2019', "'"))
            cleaned = re.sub(r'\s*[-]\s*', '-', cleaned)
            cleaned = re.sub(r'\s+', ' ', cleaned)
            return cleaned.strip()

        def _normalize_context_match(value):
            if not value:
                return ''
            normalized = _normalize_engineering_text(value).upper()
            normalized = re.sub(r'(?<=\d)\s*(?:"|\')', '"', normalized)
            normalized = re.sub(r'"{2,}', '"', normalized)
            return normalized.strip('- ').strip()

        def _canonicalize_line_number(size, *parts):
            size_part = _normalize_context_match(size)
            if size_part and '"' not in size_part:
                size_part = f'{size_part}"'
            ordered_parts = [size_part]
            ordered_parts.extend(_normalize_context_match(part) for part in parts if part)
            return "-".join(part for part in ordered_parts if part)

        def _extract_line_numbers(text):
            candidates = _all_matches(text, line_no_pats)
            normalized_text = _normalize_engineering_text(text)

            for rule in line_no_structured_rules:
                for match in rule["pattern"].finditer(normalized_text):
                    groups = [grp.strip() for grp in match.groups() if grp is not None]
                    if rule["name"] == "five_part_line" and len(groups) == 5:
                        line_no = _canonicalize_line_number(
                            groups[0], groups[1], groups[2], groups[3], groups[4]
                        )
                        if line_no:
                            candidates.append(line_no)
                    elif rule["name"] == "six_part_line" and len(groups) >= 5:
                        line_no = _canonicalize_line_number(*groups)
                        if line_no:
                            candidates.append(line_no)

            return list(dict.fromkeys(c for c in candidates if c))

        def _pick_best_line_number(candidates, area_hint=''):
            if not candidates:
                return None
            if area_hint:
                exact = [cand for cand in candidates if _area(cand) == area_hint]
                if exact:
                    return exact[0]
                fuzzy = _fuzzy_lookup(area_hint, { _area(c): [c] for c in candidates if _area(c) })
                if fuzzy:
                    return fuzzy[0]
            return candidates[0]

        def _area(s):
            m = _area_re.search(s or '')
            return m.group(1) if m else ''

        def _fuzzy_lookup(area_key, by_area_dict):
            """Return candidates for area_key with fuzzy numeric tolerance."""
            if not area_key:
                return []
            if area_key in by_area_dict:
                return by_area_dict[area_key]
            # Numeric proximity search within tolerance
            try:
                target = int(area_key)
                best, best_dist = None, area_tolerance + 1
                for k in by_area_dict:
                    try:
                        d = abs(int(k) - target)
                        if d <= area_tolerance and d < best_dist:
                            best_dist, best = d, k
                    except ValueError:
                        pass
                return by_area_dict[best] if best else []
            except ValueError:
                return []

        def _build_join_variants(tokens, max_span=4):
            clean_tokens = [_normalize_engineering_text(tok) for tok in tokens if _normalize_engineering_text(tok)]
            if not clean_tokens:
                return []

            variants = [
                " ".join(clean_tokens),
                "-".join(clean_tokens),
                "".join(clean_tokens),
            ]

            for span in range(2, min(max_span, len(clean_tokens)) + 1):
                dash_chunks = []
                plain_chunks = []
                spaced_chunks = []
                for idx in range(len(clean_tokens) - span + 1):
                    chunk = clean_tokens[idx:idx + span]
                    dash_chunks.append("-".join(chunk))
                    plain_chunks.append("".join(chunk))
                    spaced_chunks.append(" ".join(chunk))
                variants.append(" ".join(dash_chunks))
                variants.append(" ".join(plain_chunks))
                variants.append(" ".join(spaced_chunks))

            return list(dict.fromkeys(v for v in variants if v))

        def _reconstruct_line_no(word_tokens, area_hint=''):
            """
            Reconstruct piping line number from a sequence of word tokens.

            Piping line format: {size}["]-{fluid}-{area}-{spec}
            The PDF may split "16"" and "HC-3901-A2A" into separate word tokens.
            This function tries to stitch them back together.

            Handles tokens like:
              ["16\"", "HC-3901-A2A"]       → "16-HC-3901-A2A"
              ["16\"", "HC", "3901", "A2A"] → "16-HC-3901-A2A"
              ["16", "HC-3901-A2A"]         → "16-HC-3901-A2A"
            """
            _SIZE  = re.compile(r'^\d{1,3}(?:["″\u2033\u02BA\'\u2019]{1,2})?$')
            _FLUID = re.compile(r'^[A-Z]{1,6}$', re.IGNORECASE)
            _AREA  = re.compile(r'^\d{3,6}$')
            _SPEC  = re.compile(r'^[A-Z][0-9][A-Z0-9]{0,3}$', re.IGNORECASE)
            # A combined token like "HC-3901-A2A" (fluid + rest)
            _FLUID_REST = re.compile(
                r'^([A-Z]{1,6})[-](\d{3,6})[-]([A-Z][0-9][A-Z0-9]{0,3})$', re.IGNORECASE
            )
            _FLUID_AREA = re.compile(r'^([A-Z]{1,6})[-](\d{3,6})$', re.IGNORECASE)

            words = [w.strip() for w in word_tokens if w.strip()]

            for i, w in enumerate(words):
                clean_size = re.sub(r'["″\u2033\u02BA\'\u2019]+', '', w)
                if not _SIZE.match(w) and not re.match(r'^\d{1,3}$', clean_size):
                    continue
                rest = words[i + 1:i + 5]  # look ahead up to 4 tokens

                for j, rw in enumerate(rest):
                    # Case A: next word is "fluid-area-spec"
                    m = _FLUID_REST.match(rw)
                    if m:
                        fluid, area_code, spec = m.group(1), m.group(2), m.group(3)
                        if area_hint and not area_code.startswith(area_hint[:2]):
                            continue
                        return f"{clean_size}-{fluid.upper()}-{area_code}-{spec.upper()}"

                    # Case B: next word is "fluid-area"
                    m2 = _FLUID_AREA.match(rw)
                    if m2:
                        fluid, area_code = m2.group(1), m2.group(2)
                        if area_hint and not area_code.startswith(area_hint[:2]):
                            continue
                        # Look for spec in the word after
                        if j + 1 < len(rest) and _SPEC.match(rest[j + 1]):
                            return f"{clean_size}-{fluid.upper()}-{area_code}-{rest[j+1].upper()}"
                        return f"{clean_size}-{fluid.upper()}-{area_code}"

                    # Case C: fluid/area/spec in separate tokens
                    if not _FLUID.match(rw):
                        continue
                    fluid = rw.upper()
                    remaining = rest[j + 1:]
                    for k_idx, kw in enumerate(remaining):
                        if not _AREA.match(kw):
                            continue
                        if area_hint and not kw.startswith(area_hint[:2]):
                            continue
                        area_code = kw
                        after = remaining[k_idx + 1:]
                        if after and _SPEC.match(after[0]):
                            return f"{clean_size}-{fluid}-{area_code}-{after[0].upper()}"
                        return f"{clean_size}-{fluid}-{area_code}"
            return None

        # ── L1: Deterministic (loop_number + default service description + CS tag ISA) ─
        for inst in instruments:
            tag = inst.get("tag_number", "")
            if not inst.get("loop_number") or inst["loop_number"] == "N/A":
                inst["loop_number"] = self._derive_loop_number(tag)
            if not inst.get("service_description"):
                inst["service_description"] = self._infer_service_description(
                    tag, inst.get("instrument_type", ""), inst.get("category", "")
                )
            # CS Tag L1: if the instrument IS a DCS controller, mark it immediately
            if inst.get("control_system_tag") in ("N/A", "", None):
                _, is_dcs = self._derive_cs_tag_isa(tag)
                if is_dcs:
                    inst["control_system_tag"] = tag.upper()

        # ── Open PDF ──────────────────────────────────────────────────────
        try:
            import fitz
        except ImportError:
            logger.warning("[Enrich] PyMuPDF not installed — skipping contextual enrichment")
            return instruments

        try:
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception as ex:
            logger.warning(f"[Enrich] Cannot open PDF for enrichment: {ex}")
            return instruments

        global_line_nos: list = []
        global_equip:    list = []
        global_signals:  list = []
        global_sps:      list = []
        all_words_pages: list = []

        # ── L2: Block-level scan ─────────────────────────────────────────
        # Plain text alone misses line numbers split across spans (e.g. "16""
        # as one span, "HC-3901-A2A" as the next). Block reconstruction joins
        # them with a dash, matching: 16-HC-3901-A2A → line_no regex.
        for pno in range(len(doc)):
            pg = doc[pno]

            # 2a: Full plain text
            txt = _normalize_engineering_text(pg.get_text("text") or "")
            global_line_nos.extend(_extract_line_numbers(txt))
            global_equip.extend(_all_matches(txt, equip_pats))
            global_signals.extend(_all_matches(txt, signal_pats))
            global_sps.extend(_all_matches(txt, sp_pats))

            # 2b: Word-level pair/triple joining (catches split tokens on same line)
            try:
                words_raw = pg.get_text("words") or []
                all_words_pages.append(words_raw)
                for wi_r in range(len(words_raw)):
                    for span_len in (2, 3, 4):
                        if wi_r + span_len > len(words_raw):
                            break
                        # Only join words on roughly the same text line (y-delta < 12 pts)
                        y_start = words_raw[wi_r][1]
                        y_end   = words_raw[wi_r + span_len - 1][3]
                        if abs(y_end - y_start) > 12:
                            break
                        for joiner in ("-", ""):
                            chunk = _normalize_engineering_text(joiner.join(
                                words_raw[wi_r + s][4].strip()
                                for s in range(span_len)
                            ))
                            global_line_nos.extend(_extract_line_numbers(chunk))
                            global_equip.extend(_all_matches(chunk, equip_pats))
            except Exception:
                if len(all_words_pages) <= pno:
                    all_words_pages.append([])

            # 2c: Block-line reconstruction (joins block's internal lines)
            try:
                page_dict = pg.get_text("dict")
                for block in page_dict.get("blocks", []):
                    if block.get("type") != 0:
                        continue
                    block_lines = []
                    for ln in block.get("lines", []):
                        lt = "".join(s.get("text", "") for s in ln.get("spans", [])).strip()
                        if lt:
                            block_lines.append(lt)
                    for joiner in ("-", " ", ""):
                        combined = _normalize_engineering_text(joiner.join(block_lines))
                        global_line_nos.extend(_extract_line_numbers(combined))
                        global_equip.extend(_all_matches(combined, equip_pats))
                    # Also try _reconstruct_line_no on block tokens
                    reconstructed = _reconstruct_line_no(block_lines)
                    if reconstructed:
                        global_line_nos.append(reconstructed)
            except Exception:
                pass

        # Deduplicate while preserving order
        global_line_nos = list(dict.fromkeys(ln.upper() for ln in global_line_nos if ln))
        global_equip    = list(dict.fromkeys(eq.upper() for eq in global_equip if eq))
        global_signals  = list(dict.fromkeys(global_signals))
        global_sps      = list(dict.fromkeys(global_sps))

        # Filter equipment: remove anything that IS an instrument prefix
        global_equip = [
            eq for eq in global_equip
            if not self._match_instrument_code(eq.split('-')[0].upper())[0]
        ]

        logger.info(
            f"[Enrich] Global scan: {len(global_line_nos)} line nos, "
            f"{len(global_equip)} equip tags, "
            f"{len(global_signals)} signal types, "
            f"{len(global_sps)} set points"
        )

        # ── L2 area-code lookup tables (built from global scan) ──────────
        ln_by_area: dict = {}
        for ln in global_line_nos:
            a = _area(ln)
            if a:
                ln_by_area.setdefault(a, []).append(ln)

        eq_by_area: dict = {}
        for eq in global_equip:
            a = _area(eq)
            if a:
                eq_by_area.setdefault(a, []).append(eq)

        # ── L3: Spatial proximity enrichment ─────────────────────────────
        # Locate each instrument tag in the page word-list, then search
        # within ctx_radius pixels for line numbers, equipment tags,
        # fail-safe annotations, signal types, and set points.
        tag_to_inst: dict = {}
        for inst in instruments:
            n = self._normalize_tag(inst.get("tag_number", ""))
            if n:
                tag_to_inst[n] = inst

        _valve_categories = {
            "Shutdown & ESD", "Control Valves", "Motor & Solenoid",
            "Flow", "Pressure", "Level", "Temperature",
        }

        for pno, words in enumerate(all_words_pages):
            for wi in range(len(words)):
                for span in range(1, 4):
                    if wi + span > len(words):
                        break
                    chunk_dash  = "-".join(words[wi + s][4].strip() for s in range(span))
                    chunk_plain = "".join(words[wi + s][4].strip() for s in range(span))
                    norm = self._normalize_tag(chunk_dash)
                    if norm not in tag_to_inst:
                        norm = self._normalize_tag(chunk_plain)
                    if norm not in tag_to_inst:
                        continue

                    inst = tag_to_inst[norm]
                    x0 = words[wi][0];            y0 = words[wi][1]
                    x1 = words[wi + span - 1][2]; y1 = words[wi + span - 1][3]
                    px = (x0 + x1) / 2;           py = (y0 + y1) / 2
                    tag_area = inst.get("loop_number", "") or _area(inst.get("tag_number", ""))

                    # Collect ALL nearby words with their positions (using larger radius)
                    nearby_words_sorted: list = []  # (dist, text, x, y)
                    for nw in words:
                        nt = nw[4].strip()
                        if not nt or len(nt) < 2:
                            continue
                        nx = (nw[0] + nw[2]) / 2
                        ny = (nw[1] + nw[3]) / 2
                        d  = ((nx - px) ** 2 + (ny - py) ** 2) ** 0.5
                        if d < ctx_radius:
                            nearby_words_sorted.append((d, nt, nx, ny))
                    nearby_words_sorted.sort()
                    nearby = [w[1] for w in nearby_words_sorted]
                    nearby_layout = [
                        w[1] for w in sorted(
                            nearby_words_sorted,
                            key=lambda item: (round(item[3] / 10), item[2])
                        )
                    ]
                    anchor_start = max(0, wi - 12)
                    anchor_end = min(len(words), wi + span + 12)
                    anchor_window = [
                        words[idx][4].strip()
                        for idx in range(anchor_start, anchor_end)
                        if words[idx][4].strip()
                    ]
                    nearby_variants = _build_join_variants(nearby, max_span=4)
                    nearby_layout_variants = _build_join_variants(nearby_layout, max_span=4)
                    anchor_variants = _build_join_variants(anchor_window, max_span=5)

                    # — Line number —
                    if inst.get("line_number") in ("N/A", "", None):
                        ln_found = None
                        for candidate_text in nearby_layout_variants + anchor_variants + nearby_variants:
                            ln_found = _pick_best_line_number(
                                _extract_line_numbers(candidate_text),
                                tag_area,
                            )
                            if ln_found:
                                break
                        if not ln_found:
                            ln_found = (
                                _reconstruct_line_no(nearby_layout, tag_area)
                                or _reconstruct_line_no(anchor_window, tag_area)
                                or _reconstruct_line_no(nearby, tag_area)
                            )
                        if ln_found:
                            ln_found = _normalize_context_match(ln_found)
                            inst["line_number"] = ln_found
                            # Immediately feed back into lookup table
                            la = _area(ln_found)
                            if la:
                                ln_by_area.setdefault(la, [])
                                if ln_found not in ln_by_area[la]:
                                    ln_by_area[la].append(ln_found)

                    # — Equipment number —
                    if inst.get("equipment_number") in ("N/A", "", None):
                        eq_candidates = []
                        for candidate_text in nearby_layout_variants + anchor_variants + nearby_variants:
                            eq_candidates.extend(_all_matches(candidate_text, equip_pats))
                        for eq_c in eq_candidates:
                            code = eq_c.split('-')[0].upper()
                            if not self._match_instrument_code(code)[0]:
                                inst["equipment_number"] = _normalize_context_match(eq_c)
                                ea = _area(eq_c)
                                if ea:
                                    eq_by_area.setdefault(ea, [])
                                    normalized_eq = _normalize_context_match(eq_c)
                                    if normalized_eq not in eq_by_area[ea]:
                                        eq_by_area[ea].append(normalized_eq)
                                break

                    # — Fail-safe (control/shutdown valves only) —
                    if inst.get("fail_safe") in ("N/A", "", None):
                        if inst.get("category") in _valve_categories:
                            fs_str = " ".join(w[1] for w in nearby_words_sorted if w[0] < radius)
                            fs = _first_match(fs_str, fail_pats)
                            if fs:
                                fsu = fs.upper()
                                if "CLOSE" in fsu:
                                    fs = "FC"
                                elif "OPEN" in fsu:
                                    fs = "FO"
                                elif "LAST" in fsu or "LOCK" in fsu:
                                    fs = "FL"
                                inst["fail_safe"] = fs.upper()[:3]

                    # — Signal type —
                    if inst.get("signal_type") in ("N/A", "", None):
                        sig_str = " ".join(w[1] for w in nearby_words_sorted if w[0] < radius)
                        sig = _first_match(sig_str, signal_pats)
                        if sig:
                            sigu = sig.upper()
                            if "4" in sigu and "20" in sigu:
                                sig = "4-20mA"
                            elif "HART" in sigu:
                                sig = "HART"
                            elif "FIELD" in sigu:
                                sig = "Fieldbus"
                            elif "PROFI" in sigu:
                                sig = "Profibus"
                            elif "PNEUM" in sigu:
                                sig = "Pneumatic"
                            elif any(x in sigu for x in ("DISCRETE", "ON/OFF", "ON-OFF")):
                                sig = "Discrete (0/1)"
                            elif sigu in ("AO",):
                                sig = "4-20mA"
                            elif sigu in ("DO",):
                                sig = "Discrete (0/1)"
                            inst["signal_type"] = sig

                    # — Set point —
                    if inst.get("set_point") in ("N/A", "", None):
                        sp_str = " ".join(w[1] for w in nearby_words_sorted if w[0] < radius)
                        sp = _first_match(sp_str, sp_pats)
                        if sp:
                            inst["set_point"] = sp.strip()

                    # — Control System Tag (L3: proximity label scan) —
                    if inst.get("control_system_tag") in ("N/A", "", None):
                        _cs_context = " ".join(nearby_layout[:40])
                        for _cs_pat in [re.compile(p, re.IGNORECASE)
                                        for p in CS_TAG_CONFIG["label_patterns"]]:
                            _cs_m = _cs_pat.search(_cs_context)
                            if _cs_m:
                                inst["control_system_tag"] = _cs_m.group(1).upper()
                                break

                    # — Service description: enrich default with nearby context —
                    current_desc = inst.get("service_description", "")
                    is_default = current_desc.startswith(
                        _SERVICE_VERB_MAP.get(inst.get("category", ""), "~NONE~")
                    )
                    if is_default:
                        desc_candidates = [
                            w for w in nearby[:15]
                            if len(w) > 2
                            and w.lower() not in stopwords
                            and not re.match(r'^\d+$', w)
                            and not self._match_instrument_code(w.upper())[0]
                            and not _first_match(w, line_no_pats + equip_pats + sp_pats)
                        ]
                        if len(desc_candidates) >= ec["desc_min_words"]:
                            inst["service_description"] = " ".join(desc_candidates[:5]).title()

                    break  # tag located — stop trying longer spans

        doc.close()

        # ── L4: Area-code matching (fallback for spatially unlocated tags) ─
        for inst in instruments:
            tag  = inst.get("tag_number", "")
            loop = inst.get("loop_number", "N/A")
            area = loop if loop != "N/A" else _area(tag)

            if inst.get("line_number") in ("N/A", "", None) and area:
                candidates = _fuzzy_lookup(area, ln_by_area)
                if candidates:
                    inst["line_number"] = candidates[0]

            if inst.get("equipment_number") in ("N/A", "", None) and area:
                candidates = _fuzzy_lookup(area, eq_by_area)
                if candidates:
                    inst["equipment_number"] = candidates[0]

            # Category-aware default signal type from global document signals
            if inst.get("signal_type") in ("N/A", "", None) and global_signals:
                cat = inst.get("category", "")
                for sig_raw in global_signals:
                    sigu = sig_raw.upper()
                    if "4" in sigu and "20" in sigu:
                        if cat in ("Flow", "Pressure", "Temperature", "Level",
                                   "Differential Pressure", "Analysis"):
                            inst["signal_type"] = "4-20mA"
                            break
                    if any(x in sigu for x in ("DISCRETE", "DI", "DO")):
                        if cat in ("Shutdown & ESD", "Motor & Solenoid", "Position"):
                            inst["signal_type"] = "Discrete (0/1)"
                            break

        logger.info(f"[Enrich] Complete — {len(instruments)} instruments enriched, "
                    f"{sum(1 for i in instruments if i.get('line_number') not in ('N/A','',None))} have line nos, "
                    f"{sum(1 for i in instruments if i.get('equipment_number') not in ('N/A','',None))} have equip tags")

        # ── L5: Cross-reference CS tags (field transmitters → controllers) ─
        instruments = self._cross_ref_cs_tags(instruments)

        return instruments

    def _enrich_with_gemini_text(self, instruments, pdf_text, legend_text="", legend_files=None):
        """
        Use Gemini's text-only API (NOT Vision) to infer service descriptions
        and contextual fields from the raw PDF text content.

        Why text instead of vision:
          - Text API uses far less quota than vision inference.
          - The raw PDF text already contains line numbers, equipment tags,
            service labels — Gemini just needs to correlate them to each tag.
          - A single text call replaces many expensive multi-angle vision passes.

        Only runs when the Gemini client is available and quota not exceeded.
        Skips instruments where all fields are already populated.
        """
        if not self.gemini_client or self._gemini_quota_exceeded:
            return instruments
        if not pdf_text or len(pdf_text.strip()) < 50:
            return instruments
        legend_files = legend_files or []

        # Only enrich instruments that still have gap fields
        _gap_fields = ("service_description", "line_number", "equipment_number",
                       "fail_safe", "signal_type", "set_point", "control_system_tag")
        needs_enrich = [
            inst for inst in instruments
            if any(
                not inst.get(f) or inst.get(f) == "N/A"
                or (f == "service_description" and inst.get(f, "").startswith(
                    _SERVICE_VERB_MAP.get(inst.get("category", ""), "~NONE~")
                ))
                for f in _gap_fields
            )
        ]
        if not needs_enrich:
            logger.info("[Enrich] Gemini text: all fields already populated — skipping")
            return instruments

        tag_list = [inst["tag_number"] for inst in needs_enrich if inst.get("tag_number")]
        if not tag_list:
            return instruments

        # Limit text to first 10 000 chars to avoid large token usage
        text_chunk = pdf_text[:10000]
        legend_chunk = (legend_text or "")[:6000]
        legend_block = ""
        if legend_chunk:
            legend_block = f"""

    LEGEND / SYMBOL SHEET TEXT FROM AWS S3 (supplemental context only):
    {legend_chunk}

    Use the legend-sheet context to interpret abbreviations, DCS/CS conventions,
    signal notations, fail-safe symbols, and instrument/function-code meanings.
    If legend context helps you infer a field, prefer that interpretation over guesswork.
    """

        prompt = f"""You are a senior P&ID / FEED engineer. The text below was extracted from a P&ID drawing.
Correlate EACH instrument tag to any service context visible in the text, then return a JSON object.

PDF TEXT (raw extraction — up to 10 000 chars):
{text_chunk}
    {legend_block}

Instrument tags to enrich (provide data for as many as you can):
{json.dumps(tag_list[:60])}

Return ONLY a JSON object — no markdown fences, no explanation:
{{
  "PIT-3901-01": {{
    "service_description": "Pig Receiver Inlet Pressure",
    "line_number": "10\\"-G-3901-A2A",
    "equipment_number": "LP-3901",
    "fail_safe": "N/A",
    "signal_type": "4-20mA",
    "set_point": "75 barg (PSHH)",
    "control_system_tag": "PIC-3901-01"
  }}
}}

Rules:
- Use "N/A" for any field you cannot determine from the text.
- fail_safe values: FC (fail closed), FO (fail open), FL (fail last), N/A.
- signal_type: 4-20mA, HART, Fieldbus, Profibus, Discrete (0/1), Pneumatic, N/A.
- set_point: include the number AND unit, e.g. "75 barg", "250 °C", "12000 kg/h".
- control_system_tag: the DCS/CS tag for this instrument if identifiable from context; N/A otherwise.
- Only include tags from the list above.
"""

        try:
            from google.genai import types as _gtypes
            model = self.extraction_config.get("gemini_model", "gemini-2.0-flash")
            response = self.gemini_client.models.generate_content(
                model=model,
                contents=prompt,
                config=_gtypes.GenerateContentConfig(
                    temperature=0.05,
                    max_output_tokens=8000,
                ),
            )
            raw = response.text or ""
            s = raw.find("{");  e = raw.rfind("}") + 1
            if s >= 0 and e > s:
                enrichment_map = json.loads(raw[s:e])
                if isinstance(enrichment_map, dict):
                    updated = 0
                    legend_marked_tags = set()
                    for inst in instruments:
                        tag = inst.get("tag_number", "")
                        enrichment = enrichment_map.get(tag, {})
                        if not enrichment:
                            continue
                        for field in _gap_fields:
                            val = (enrichment.get(field) or "").strip()
                            existing = (inst.get(field) or "").strip()
                            is_gap = (
                                not existing
                                or existing == "N/A"
                                or (field == "service_description"
                                    and existing.startswith(
                                        _SERVICE_VERB_MAP.get(inst.get("category", ""), "~NONE~")
                                    ))
                            )
                            if val and val != "N/A" and is_gap:
                                inst[field] = val
                                updated += 1
                                if legend_chunk:
                                    legend_marked_tags.add(tag)
                    if legend_marked_tags:
                        for inst in instruments:
                            tag = inst.get("tag_number", "")
                            if tag in legend_marked_tags:
                                inst["notes"] = self._append_note_source(
                                    inst.get("notes", ""),
                                    "Legends sheet",
                                )
                    logger.info(f"[Enrich] Gemini text enrichment: updated {updated} fields across {len(instruments)} instruments")
        except Exception as ex:
            err = str(ex)
            if "429" in err or "quota" in err.lower() or "rate" in err.lower():
                self._gemini_quota_exceeded = True
                logger.warning("[Enrich] Gemini text quota exceeded — disabling Gemini for this request")
            else:
                logger.warning(f"[Enrich] Gemini text enrichment error: {ex}")

        return instruments

    # ────────────────────────────────────────────────────────────────────
    # Text-layer extraction (PyMuPDF — free, no API quota)
    # ────────────────────────────────────────────────────────────────────

    # Strict regex: requires dash/underscore separator (e.g. FT-3901-01)
    _TAG_RE = re.compile(
        r'\b([A-Z]{1,5})'
        r'[-_]'
        r'(\d{3,6}[A-Z]?)'
        r'(?:[-_](\d{1,3}[A-Z]?))?'
        r'\b',
        re.IGNORECASE,
    )

    # Flexible regex: allows space as separator (e.g. "FT 3901-01" in a bubble)
    _TAG_RE_FLEX = re.compile(
        r'\b([A-Z]{1,5})'
        r'[\s]'
        r'(\d{3,6}[A-Z]?)'
        r'(?:[-_](\d{1,3}[A-Z]?))?'
        r'\b',
        re.IGNORECASE,
    )

    _TAG_MIN_LEN = 5

    def _match_instrument_code(self, tag_upper):
        """Return (matched_code, cfg_entry) or (None, None) for a tag string."""
        prefix = re.match(r'^([A-Z]+)', tag_upper)
        if not prefix:
            return None, None
        code = prefix.group(1)
        for length in range(len(code), 0, -1):
            candidate = code[:length]
            if candidate in INSTRUMENT_CATEGORIES:
                return candidate, INSTRUMENT_CATEGORIES[candidate]
        return None, None

    def _make_instrument_record(self, tag, matched_code, cfg_entry, dn, rev, note):
        """Build a standardised instrument dict."""
        return {
            "tag_number":          tag.upper(),
            "control_system_tag":  "N/A",
            "instrument_type":     cfg_entry["name"],
            "category":            cfg_entry["category"],
            "pid_no":              dn,
            "service_description": "",
            "line_number":         "N/A",
            "equipment_number":    "N/A",
            "loop_number":         "N/A",
            "fail_safe":           "N/A",
            "signal_type":         "N/A",
            "set_point":           "N/A",
            "drawing_number":      dn,
            "revision":            rev,
            "notes":               note,
        }

    def _derive_loop_number(self, tag):
        """
        Derive loop number from the instrument tag number.
        The numeric area/loop code is the middle segment after the first dash.
          PIT-3901-01  → 3901
          FIT-3901-08A → 3901
          SDV-3901-01  → 3901
          FT-101A      → 101
        """
        m = re.search(r'[-_](\d{3,6})', tag or '')
        return m.group(1) if m else 'N/A'

    def _derive_cs_tag_isa(self, tag):
        """
        Determine the Control System Tag from ISA-5.1 function code analysis.

        Returns (cs_tag, is_dcs_instrument):
          is_dcs_instrument=True  → the instrument IS a DCS/CS device;
                                    cs_tag == tag (it is its own CS tag)
          is_dcs_instrument=False → field instrument; cs_tag is the *expected*
                                    controller tag derived via the function-code
                                    substitution rules (may not exist on this drawing)
          cs_tag == "N/A"         → derivation not possible

        Soft-coded via CS_TAG_CONFIG — no changes to this method needed for tuning.
        """
        cfg = CS_TAG_CONFIG
        tag_upper = (tag or "").strip().upper()
        # Match:  function-code letters  +  dash+numbers+optional-suffix
        m = re.match(r'^([A-Z]{2,6})([-–]\d.*)$', tag_upper)
        if not m:
            return "N/A", False
        func_code, numeric_suffix = m.group(1), m.group(2)

        # Step 1 — Is this already a DCS controller instrument?
        if func_code in cfg["dcs_function_codes"]:
            return tag_upper, True

        # Step 2 — Is it a field transmitter/element? Derive the controller tag.
        first_letter = func_code[0]
        rest         = func_code[1:]
        for field_suffix, ctrl_suffix in cfg["transmitter_to_controller"].items():
            if rest == field_suffix:
                derived = f"{first_letter}{ctrl_suffix}{numeric_suffix}"
                return derived, False

        return "N/A", False

    def _cross_ref_cs_tags(self, instruments):
        """
        Post-processing: cross-reference field instruments against the complete
        list of instruments extracted from this drawing.

        For each field transmitter (e.g. FT-3901-01) whose control_system_tag is
        still N/A, check whether the derived controller (FIC-3901-01 or FIC-3901)
        actually appears in the instruments list.  If found → set the CS tag.
        If not found but derivation still produced a plausible controller tag,
        store the expected tag (prefixed "Expected:") so engineers know what
        controller *should* be on the drawing.

        Soft-coded via CS_TAG_CONFIG.
        """
        # Build lookup: normalised tag → instrument record
        tag_map = {
            (i.get("tag_number") or "").strip().upper(): i
            for i in instruments
        }

        for inst in instruments:
            if inst.get("control_system_tag") not in ("N/A", "", None):
                continue  # already resolved

            tag = (inst.get("tag_number") or "").strip().upper()
            derived_cs, is_dcs = self._derive_cs_tag_isa(tag)

            if is_dcs:
                # This instrument IS a DCS/CS device
                inst["control_system_tag"] = tag
                continue

            if derived_cs == "N/A":
                continue

            # Check exact match in drawing's instrument list
            if derived_cs in tag_map:
                inst["control_system_tag"] = derived_cs
                continue

            # Try abbreviated loop match (drop the last suffix)
            # e.g. try FIC-3901 when FIC-3901-01 not found
            parts = derived_cs.split("-")
            if len(parts) >= 3:
                short_cs = f"{parts[0]}-{parts[1]}"
                if short_cs in tag_map:
                    inst["control_system_tag"] = tag_map[short_cs]["tag_number"]
                    continue

            # Controller not found on this drawing — store expected tag for reference
            # (wrapped in "Expected:" so users know it's inferred, not confirmed)
            inst["control_system_tag"] = f"Expected: {derived_cs}"

        cs_found = sum(
            1 for i in instruments
            if i.get("control_system_tag") not in ("N/A", "", None)
            and not (i.get("control_system_tag") or "").startswith("Expected:")
        )
        logger.info(f"[Enrich] CS tag cross-reference: {cs_found} confirmed CS tags resolved")
        return instruments

    def _infer_service_description(self, tag, instrument_type, category):
        """
        Generate a meaningful base service description from tag + Category.
        Used as a fallback when no contextual text is available from the drawing.
        """
        loop = self._derive_loop_number(tag)
        verb = _SERVICE_VERB_MAP.get(category, instrument_type or "Measurement")
        if loop and loop != 'N/A':
            return f"{verb} — System {loop}"
        return verb

    def _resolve_drawing_info_from_pdf(self, pid_bytes, drawing_info):
        """
        Resolve drawing_number/pid_no from PDF title-block text (DWG NO / P&ID NO).
        Soft-coded via DRAWING_NUMBER_CONFIG and only overrides when a strong
        candidate is found.
        """
        info = dict(drawing_info or {})
        current_dn = (info.get("drawing_number") or "").strip()
        current_pid = (info.get("pid_no") or "").strip()

        detected = self._extract_drawing_number_from_pdf(pid_bytes)
        if not detected:
            return info

        if detected != current_dn or detected != current_pid:
            logger.info(
                f"[InstrumentIndex] Drawing number detected from title block: {detected} "
                f"(input was drawing='{current_dn or 'N/A'}', pid='{current_pid or 'N/A'}')"
            )

        info["drawing_number"] = detected
        info["pid_no"] = detected
        return info

    def _extract_drawing_number_from_pdf(self, pid_bytes):
        """Extract best DWG/P&ID number candidate from PDF text."""
        try:
            import fitz
        except ImportError:
            return ""

        try:
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception:
            return ""

        try:
            page_count = len(doc)
            # Title blocks are usually on first/last pages; scan those first.
            page_indices = list(dict.fromkeys(
                [0, 1, page_count - 2, page_count - 1] if page_count > 2 else list(range(page_count))
            ))

            text_chunks = []
            for idx in page_indices:
                if 0 <= idx < page_count:
                    text_chunks.append(doc[idx].get_text("text") or "")

            text = "\n".join(text_chunks)
            return self._extract_drawing_number_from_text(text)
        except Exception:
            return ""
        finally:
            doc.close()

    def _extract_drawing_number_from_text(self, text):
        """Label-aware extraction of drawing number from raw OCR/PDF text."""
        if not text:
            return ""

        cfg = DRAWING_NUMBER_CONFIG
        normalized = re.sub(r'\s+', ' ', text.upper())
        label_pats = [re.compile(p, re.IGNORECASE) for p in cfg.get("label_patterns", [])]
        value_pats = [re.compile(p, re.IGNORECASE) for p in cfg.get("value_patterns", [])]
        min_len = cfg.get("min_length", 8)
        max_len = cfg.get("max_length", 64)
        window = cfg.get("window_chars", 140)

        def _is_valid_candidate(val):
            if not val:
                return False
            v = val.strip().upper().strip("-:;,. ")
            if len(v) < min_len or len(v) > max_len:
                return False
            if not re.search(r'[A-Z]', v) or not re.search(r'\d', v):
                return False
            if not any(sep in v for sep in ('-', '/', '.')):
                return False
            # Filter obvious non-document tokens
            blocked = {
                "PROCESS", "INSTRUMENT", "DRAWING", "NUMBER", "REVISION",
                "SCALE", "SHEET", "TITLE", "PROJECT", "CLIENT",
            }
            if v in blocked:
                return False
            # Filter instrument-like tags (FT-3901-01 etc.)
            if re.match(r'^[A-Z]{1,5}-\d{3,6}(?:-\d{1,3}[A-Z]?)?$', v):
                return False
            return True

        # 1) Prefer value near explicit title-block labels (DWG NO, P&ID NO, ...)
        for lp in label_pats:
            for lm in lp.finditer(normalized):
                start = max(0, lm.end())
                end = min(len(normalized), lm.end() + window)
                region = normalized[start:end]
                for vp in value_pats:
                    for vm in vp.finditer(region):
                        cand = vm.group(1)
                        if _is_valid_candidate(cand):
                            return cand.strip().upper()

        # 2) Fallback: strongest global candidate in the scanned pages
        all_candidates = []
        for vp in value_pats:
            for vm in vp.finditer(normalized):
                cand = vm.group(1)
                if _is_valid_candidate(cand):
                    all_candidates.append(cand.strip().upper())

        if not all_candidates:
            return ""

        # Prefer richer segmented identifiers (more separators/segments)
        all_candidates = list(dict.fromkeys(all_candidates))
        all_candidates.sort(key=lambda s: (s.count('-') + s.count('/') + s.count('.'), len(s)), reverse=True)
        return all_candidates[0]

    def _load_legend_context_from_s3(self, drawing_info):
        """
        Optionally load related legend/symbol sheets from S3.

        This is a soft enrichment path only: the P&ID remains the primary
        extraction source, while legend sheets help interpret conventions such as
        control-system tags, signal notation, fail-safe symbols, and abbreviations.
        """
        cfg = LEGEND_S3_CONFIG
        context = {"text": "", "files": []}

        if not cfg.get("enabled", True):
            return context

        use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
        s3_ready = os.environ.get("S3_READY", "false").lower() == "true"
        if not (use_s3 and s3_ready):
            return context

        try:
            from apps.core.s3_service import get_s3_service
        except Exception as ex:
            logger.warning(f"[InstrumentIndex] Legend S3 helper unavailable: {ex}")
            return context

        drawing_number = (drawing_info.get("drawing_number") or "").strip().upper()
        project_name = (drawing_info.get("project_name") or "").strip()
        base_path = (os.environ.get("PFD_BASE_PATH") or "").strip().strip("/")

        search_prefixes = []
        if base_path and project_name:
            search_prefixes.append(f"{base_path}/{project_name}/")
        if base_path:
            search_prefixes.append(f"{base_path}/")
        if project_name:
            search_prefixes.append(f"{project_name}/")
        search_prefixes.append("")

        search_prefixes = list(dict.fromkeys(p for p in search_prefixes if p is not None))
        keywords = tuple(k.lower() for k in cfg.get("filename_keywords", []))
        exts = tuple(e.lower() for e in cfg.get("preferred_extensions", [".pdf"]))
        max_keys = cfg.get("max_list_keys", 400)
        max_files = cfg.get("max_candidate_files", 3)

        s3 = get_s3_service()
        candidates = []

        for prefix in search_prefixes:
            listing = s3.list_files(prefix=prefix, max_keys=max_keys)
            if not listing.get("success"):
                continue

            for file_info in listing.get("files", []):
                key = file_info.get("key") or ""
                filename = os.path.basename(key).lower()
                if not filename.endswith(exts):
                    continue
                if not any(word in filename for word in keywords):
                    continue

                score = 0
                if "legend" in filename:
                    score += 5
                if "symbol" in filename:
                    score += 4
                if project_name and project_name.lower() in key.lower():
                    score += 2
                if drawing_number:
                    drawing_tokens = [tok for tok in re.split(r'[-_/ .]+', drawing_number.lower()) if len(tok) >= 3]
                    score += sum(1 for tok in drawing_tokens if tok in key.lower())

                candidates.append((score, key))

            if candidates:
                break

        if not candidates:
            return context

        selected_keys = []
        seen = set()
        for _, key in sorted(candidates, key=lambda item: (item[0], len(item[1])), reverse=True):
            if key in seen:
                continue
            seen.add(key)
            selected_keys.append(key)
            if len(selected_keys) >= max_files:
                break

        if not selected_keys:
            return context

        legend_chunks = []
        for key in selected_keys:
            raw_bytes = self._download_s3_object_bytes(s3, key)
            if not raw_bytes:
                continue
            legend_text = self._extract_pdf_text_bytes(
                raw_bytes,
                max_pages=cfg.get("max_pages_per_file", 3),
            )
            if legend_text:
                legend_chunks.append(f"LEGEND FILE: {os.path.basename(key)}\n{legend_text}")

        if not legend_chunks:
            return context

        context["text"] = "\n\n".join(legend_chunks)[:cfg.get("max_text_chars", 12000)]
        context["files"] = selected_keys
        logger.info(
            f"[InstrumentIndex] Loaded legend context from S3: {len(selected_keys)} file(s)"
        )
        return context

    def build_legend_context_from_uploaded_file(self, legend_bytes, filename):
        """Build legend context from an explicitly uploaded legend/symbol PDF."""
        text = self._extract_pdf_text_bytes(
            legend_bytes,
            max_pages=LEGEND_S3_CONFIG.get("max_pages_per_file", 3),
        )
        if not text:
            return {"text": "", "files": []}
        return {
            "text": text[:LEGEND_S3_CONFIG.get("max_text_chars", 12000)],
            "files": [filename],
        }

    def _merge_legend_contexts(self, primary_context, secondary_context):
        """Merge uploaded and S3 legend contexts, preferring uploaded text first."""
        primary_context = primary_context or {}
        secondary_context = secondary_context or {}

        primary_text = (primary_context.get("text") or "").strip()
        secondary_text = (secondary_context.get("text") or "").strip()
        merged_files = []
        for item in (primary_context.get("files") or []) + (secondary_context.get("files") or []):
            if item and item not in merged_files:
                merged_files.append(item)

        merged_text = "\n\n".join(chunk for chunk in (primary_text, secondary_text) if chunk)
        if merged_text:
            merged_text = merged_text[:LEGEND_S3_CONFIG.get("max_text_chars", 12000)]

        return {
            "text": merged_text,
            "files": merged_files,
        }

    def _download_s3_object_bytes(self, s3_service, s3_key):
        try:
            result = s3_service.download_file(s3_key)
            if not result.get("success"):
                return b""
            body = result.get("body")
            if not body:
                return b""
            return body.read()
        except Exception as ex:
            logger.warning(f"[InstrumentIndex] Legend S3 download failed for {s3_key}: {ex}")
            return b""

    def _extract_pdf_text_bytes(self, pdf_bytes, max_pages=3):
        if not pdf_bytes:
            return ""
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            try:
                return "\n".join(
                    doc[idx].get_text("text") or ""
                    for idx in range(min(len(doc), max_pages))
                )
            finally:
                doc.close()
        except Exception as ex:
            logger.warning(f"[InstrumentIndex] Legend text extraction failed: {ex}")
            return ""

    def _append_note_source(self, note, marker):
        base = (note or "").strip()
        if not marker:
            return base
        if not base:
            return marker
        if marker in base:
            return base
        return f"{base} + {marker}"

    def _scan_for_tags(self, text, seen_tags, dn, rev, instruments, note="PDF text layer"):
        """
        Extract instrument tags from a text string using both strict and flexible regex.
        Appends new (not-yet-seen) records to instruments in-place.
        Internal note labels are mapped to user-friendly strings for the frontend.
        """
        # Map verbose internal labels → concise user-visible source labels
        _note_map = {
            "PDF text (plain)":            "PDF text layer",
            "PDF block reconstruction":    "PDF circle/bubble",
            "PDF block pair":              "PDF circle/bubble",
            "PDF spatial grouping":        "PDF circle (spatial)",
            "PDF spatial triple":          "PDF circle (spatial)",
        }
        # Tesseract notes: simplify PSM details
        if note.startswith("Tesseract"):
            if "spatial triple" in note:
                display_note = "OCR circle (3-part)"
            elif "spatial" in note:
                display_note = "OCR circle (spatial)"
            else:
                display_note = "OCR text"
        else:
            display_note = _note_map.get(note, note)

        for pattern in (self._TAG_RE, self._TAG_RE_FLEX):
            for m in pattern.finditer(text):
                # Normalise: collapse whitespace separators to dash
                raw = re.sub(r'\s+', '-', m.group(0).strip())
                full_tag = raw.upper()
                if len(full_tag) < self._TAG_MIN_LEN:
                    continue
                matched_code, cfg_entry = self._match_instrument_code(full_tag)
                if not matched_code:
                    continue
                norm = self._normalize_tag(full_tag)
                if norm in seen_tags:
                    continue
                seen_tags.add(norm)
                instruments.append(self._make_instrument_record(full_tag, matched_code, cfg_entry, dn, rev, display_note))


    def _extract_with_text_layer(self, pid_bytes, drawing_info):
        """
        Advanced 3-pass text-layer extraction using PyMuPDF.

        Pass A  — Full-page plain text regex
                  Catches complete single-span tags: "FIT-3901-08A"

        Pass B  — Block-level text reconstruction
                  Joins lines within each text block (instrument circles are one block).
                  Catches multi-line tags: ["FIT", "3901", "08A"] → "FIT-3901-08A"

        Pass C  — Spatial word-proximity grouping
                  Finds isolated prefix words (e.g. "FT") and nearby number words
                  ("3901-01") within the instrument circle diameter.
                  Catches split-span tags that no other pass can reconstruct.

        All three passes feed the same deduplication set.
        """
        instruments: list = []
        seen_tags:   set  = set()
        dn  = drawing_info.get("drawing_number", "N/A")
        rev = drawing_info.get("revision", "0")

        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.warning("[InstrumentIndex] PyMuPDF not installed — skipping text-layer extraction")
            return []

        try:
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception as e:
            logger.warning(f"[InstrumentIndex] PyMuPDF open error: {e}")
            return []

        # Pre-build a sorted prefix list (longest-first) + compiled patterns for Pass C
        all_pfx = sorted(INSTRUMENT_CATEGORIES.keys(), key=len, reverse=True)
        _pfx_exact  = re.compile(
            r'^(' + '|'.join(re.escape(p) for p in all_pfx) + r')$', re.IGNORECASE
        )
        # Number part: "3901-01A" / "3901" / "001"
        _num_part = re.compile(r'^\d{2,6}[A-Z]?(?:[-]\d{1,3}[A-Z]?)?$', re.IGNORECASE)
        # Bare suffix: "01" / "08A"
        _suffix   = re.compile(r'^\d{1,3}[A-Z]?$', re.IGNORECASE)

        use_spatial = self.extraction_config.get("spatial_grouping", True)
        radius      = self.extraction_config.get("spatial_radius", 80)

        for page_no in range(len(doc)):
            page = doc[page_no]

            # ── Pass A: Full-page plain text ─────────────────────────────
            plain = page.get_text("text") or ""
            if len(plain.strip()) > 10:
                self._scan_for_tags(plain, seen_tags, dn, rev, instruments, "PDF text (plain)")

            # ── Pass B: Block-level reconstruction ───────────────────────
            # P&ID instrument circles are typically ONE text block whose lines contain:
            #   line 1 → instrument type prefix  "FIT"
            #   line 2 → loop / area number      "3901"
            #   line 3 → suffix                  "08A"
            # Joining with "-" reconstructs the full tag.
            try:
                page_dict = page.get_text("dict")
                for block in page_dict.get("blocks", []):
                    if block.get("type") != 0:
                        continue  # skip image blocks
                    lines_text = []
                    for ln in block.get("lines", []):
                        lt = "".join(s.get("text", "") for s in ln.get("spans", [])).strip()
                        if lt:
                            lines_text.append(lt)
                    if not lines_text:
                        continue
                    # Try joining with dash, no separator, and space
                    for joiner in ("-", "", " "):
                        combined = joiner.join(lines_text)
                        if combined:
                            self._scan_for_tags(
                                combined, seen_tags, dn, rev, instruments,
                                "PDF block reconstruction"
                            )
                    # Also try consecutive-line pairs (handles 2-part split)
                    for i in range(len(lines_text) - 1):
                        for joiner in ("-", ""):
                            pair = joiner.join([lines_text[i], lines_text[i + 1]])
                            self._scan_for_tags(
                                pair, seen_tags, dn, rev, instruments,
                                "PDF block pair"
                            )
            except Exception as be:
                logger.debug(f"[InstrumentIndex] Block pass error p{page_no+1}: {be}")

            # ── Pass C: Spatial word-proximity grouping ───────────────────
            # For each word that is a pure instrument prefix, find the closest
            # number word within `radius` pixels and combine them into a tag.
            if use_spatial:
                try:
                    words = page.get_text("words")
                    # words: (x0, y0, x1, y1, text, block_no, line_no, word_no)
                    for wi, w in enumerate(words):
                        wtext = w[4].strip()
                        if not _pfx_exact.match(wtext):
                            continue
                        # centre of this prefix word
                        px = (w[0] + w[2]) / 2
                        py = (w[1] + w[3]) / 2

                        # Collect all nearby number-like words
                        candidates: list[tuple[float, str, str]] = []  # (dist, raw, joined)
                        for nw in words:
                            ntext = nw[4].strip()
                            if not (_num_part.match(ntext)):
                                continue
                            nx = (nw[0] + nw[2]) / 2
                            ny = (nw[1] + nw[3]) / 2
                            dist = ((nx - px) ** 2 + (ny - py) ** 2) ** 0.5
                            if dist < radius:
                                candidates.append((dist, ntext, f"{wtext}-{ntext}"))

                        # Also look for a bare suffix word near the best number candidate
                        candidates.sort()
                        for _, ntext, joined_tag in candidates:
                            self._scan_for_tags(
                                joined_tag, seen_tags, dn, rev, instruments,
                                "PDF spatial grouping"
                            )
                            # Check if a suffix word is also nearby
                            # (builds "FT-3901-01" from three separate words)
                            for nw2 in words:
                                nx2 = (nw2[0] + nw2[2]) / 2
                                ny2 = (nw2[1] + nw2[3]) / 2
                                if not _suffix.match(nw2[4].strip()):
                                    continue
                                dist2 = ((nx2 - px) ** 2 + (ny2 - py) ** 2) ** 0.5
                                if dist2 < radius and nw2[4].strip() != ntext:
                                    triple = f"{joined_tag}-{nw2[4].strip()}"
                                    self._scan_for_tags(
                                        triple, seen_tags, dn, rev, instruments,
                                        "PDF spatial triple"
                                    )
                except Exception as se:
                    logger.debug(f"[InstrumentIndex] Spatial pass error p{page_no+1}: {se}")

        doc.close()
        logger.info(
            f"[InstrumentIndex] Text-layer 3-pass result: {len(instruments)} unique instrument tags"
        )
        return instruments

    # ────────────────────────────────────────────────────────────────────
    # Tesseract OCR — runs on ALL PDFs (vector + scanned)
    # ────────────────────────────────────────────────────────────────────

    def _extract_with_tesseract(self, pid_bytes, drawing_info):
        """
        Tesseract OCR extraction with spatial word-proximity grouping.

        Runs on ALL PDFs (not just scanned) because even vector P&IDs benefit
        from OCR — instrument circle text is sometimes poorly extracted by
        the PDF text layer due to how AutoCAD writes character spans.

        Uses:
          - Auto-contrast preprocessing for better circle text recognition
          - Multiple PSM modes (11=sparse best for P&IDs, 6=uniform block, 3=auto)
          - Spatial grouping: prefix word + nearby number word → full tag
          - Both strict and flexible regex on OCR output
        """
        if not self.extraction_config.get("enable_tesseract", True):
            return []
        if not self.tesseract_available:
            return []

        try:
            import pytesseract
            from pytesseract import Output as TessOutput
        except ImportError:
            return []

        cfg = self.extraction_config
        dn  = drawing_info.get("drawing_number", "N/A")
        rev = drawing_info.get("revision", "0")
        instruments: list = []
        seen_tags:   set  = set()

        tess_dpi = cfg.get("tesseract_dpi", 150)
        try:
            pil_images = convert_from_bytes(pid_bytes, dpi=tess_dpi)
        except Exception as e:
            logger.warning(f"[InstrumentIndex] Tesseract: pdf2image failed: {e}")
            return []

        # Compiled patterns for spatial grouping
        all_pfx = sorted(INSTRUMENT_CATEGORIES.keys(), key=len, reverse=True)
        _pfx_exact = re.compile(
            r'^(' + '|'.join(re.escape(p) for p in all_pfx) + r')$', re.IGNORECASE
        )
        _num_part = re.compile(r'^\d{2,6}[A-Z]?(?:[-]\d{1,3}[A-Z]?)?$', re.IGNORECASE)
        _suffix   = re.compile(r'^\d{1,3}[A-Z]?$', re.IGNORECASE)
        spatial_radius = cfg.get("spatial_radius", 80) * (tess_dpi / 150)

        # PSM modes: 11=sparse (best for P&IDs),  6=uniform block,  3=fully auto
        psm_modes = [11, 6, 3]

        for page_no, img in enumerate(pil_images, start=1):
            # Pre-process: grayscale + auto-contrast improves OCR on blueprint drawings
            try:
                from PIL import ImageOps
                gray = img.convert("L")
                gray = ImageOps.autocontrast(gray, cutoff=2)
            except Exception:
                gray = img

            for psm in psm_modes:
                try:
                    tess_config = f"--psm {psm} --oem 3"
                    data = pytesseract.image_to_data(
                        gray, config=tess_config, output_type=TessOutput.DICT
                    )
                except Exception as te:
                    logger.debug(f"[InstrumentIndex] Tesseract PSM {psm} p{page_no} error: {te}")
                    continue

                # Build word list with confidence filter (>= 30 %)
                conf_threshold = 30
                word_list = [
                    (
                        data["left"][i], data["top"][i],
                        data["width"][i], data["height"][i],
                        data["text"][i],
                    )
                    for i in range(len(data["text"]))
                    if data["text"][i].strip() and int(data["conf"][i] or 0) >= conf_threshold
                ]

                # Pass A: plain-text scan of all words joined
                full_ocr_text = " ".join(w[4] for w in word_list)
                self._scan_for_tags(
                    full_ocr_text, seen_tags, dn, rev, instruments, f"Tesseract PSM {psm}"
                )

                # Pass B: spatial grouping — find prefix + nearby number words
                for wi, w in enumerate(word_list):
                    wtext = w[4].strip()
                    if not _pfx_exact.match(wtext):
                        continue
                    px = w[0] + w[2] / 2
                    py = w[1] + w[3] / 2

                    nearby: list[tuple[float, str]] = []
                    for nw in word_list:
                        ntext = nw[4].strip()
                        if not _num_part.match(ntext):
                            continue
                        nx = nw[0] + nw[2] / 2
                        ny = nw[1] + nw[3] / 2
                        dist = ((nx - px) ** 2 + (ny - py) ** 2) ** 0.5
                        if dist < spatial_radius:
                            nearby.append((dist, ntext))

                    nearby.sort()
                    for _, ntext in nearby:
                        candidate = f"{wtext.upper()}-{ntext.upper()}"
                        self._scan_for_tags(
                            candidate, seen_tags, dn, rev, instruments,
                            f"Tesseract spatial PSM {psm}"
                        )
                        # Also look for a trailing suffix word
                        for nw2 in word_list:
                            ntext2 = nw2[4].strip()
                            if not _suffix.match(ntext2) or ntext2 == ntext:
                                continue
                            nx2 = nw2[0] + nw2[2] / 2
                            ny2 = nw2[1] + nw2[3] / 2
                            dist2 = ((nx2 - px) ** 2 + (ny2 - py) ** 2) ** 0.5
                            if dist2 < spatial_radius:
                                triple = f"{candidate}-{ntext2.upper()}"
                                self._scan_for_tags(
                                    triple, seen_tags, dn, rev, instruments,
                                    f"Tesseract spatial triple PSM {psm}"
                                )

        logger.info(f"[InstrumentIndex] Tesseract result: {len(instruments)} unique tags")
        return instruments

    # ────────────────────────────────────────────────────────────────────
    # PDF → JPEG conversion
    # ────────────────────────────────────────────────────────────────────

    def _to_jpeg_pages(self, pid_bytes):
        """Convert PDF (or image) to list of JPEG bytes, one entry per page."""
        is_pdf = pid_bytes[:4] == b"%PDF"

        if is_pdf:
            logger.info("[InstrumentIndex] Converting PDF to images…")
            try:
                pil_images = convert_from_bytes(pid_bytes, dpi=self.extraction_config["pdf_dpi"])
            except Exception as e:
                logger.error(f"[InstrumentIndex] pdf2image failed: {e}")
                # Fallback: send raw bytes as a single "page"
                return [pid_bytes]
        else:
            # Already an image
            pil_images = [Image.open(io.BytesIO(pid_bytes))]

        jpeg_pages = []
        for img in pil_images:
            jpeg_pages.append(self._pil_to_jpeg(img))
        return jpeg_pages

    def _pil_to_jpeg(self, img, max_size=None):
        """Resize + convert PIL image to JPEG bytes."""
        if max_size is None:
            max_size = self.extraction_config["max_image_size"]
        # Resize if too large
        if img.width > max_size or img.height > max_size:
            ratio = min(max_size / img.width, max_size / img.height)
            img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.Resampling.LANCZOS)
        # Flatten transparency
        if img.mode in ("RGBA", "LA", "P"):
            bg = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            bg.paste(img, mask=img.split()[-1] if img.mode in ("RGBA", "LA") else None)
            img = bg
        elif img.mode != "RGB":
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=self.extraction_config.get("jpeg_quality", 90), optimize=True)
        return buf.getvalue()

    # ────────────────────────────────────────────────────────────────────
    # Per-page AI analysis — multi-angle, multi-pass
    # ────────────────────────────────────────────────────────────────────

    def _analyse_page(self, jpeg_bytes, drawing_info, page_no, only_engine=None):
        """
        Multi-pass extraction for one P&ID page.
          Pass 1  — Full drawing, 0°
          Pass 2+ — Full drawing rotated (vertical text)
          Tile passes — 2×2 quadrant zoom (dense drawings)

        only_engine: when set, restricts all Vision calls to that engine only.
        """
        cfg = self.extraction_config
        img = Image.open(io.BytesIO(jpeg_bytes))

        all_pass_instruments = []

        # ── Pass 1: Full drawing, normal orientation ─────────────────────
        logger.info(f"[InstrumentIndex] Page {page_no} — Pass 1 (0° full drawing) [{only_engine or 'auto'}]")
        p1 = self._vision_pass(
            jpeg_bytes, drawing_info, page_no,
            extra_hint="Standard orientation. Extract ALL instrument tags visible.",
            mode="primary", max_tokens=cfg["max_tokens_primary"], only_engine=only_engine,
        )
        logger.info(f"[InstrumentIndex] Page {page_no} — Pass 1: {len(p1)} instruments")
        all_pass_instruments.extend(p1)

        # ── Rotation passes: catch vertical / slanted tags ───────────────
        if cfg["enable_rotation"]:
            for angle in cfg["rotation_angles"]:
                rotated_img = img.rotate(-angle, expand=True)
                rot_jpeg = self._pil_to_jpeg(rotated_img)
                label = f"rotated_{angle}cw"
                hint = (
                    f"IMAGE ROTATED {angle}° CLOCKWISE. "
                    "Tags that were printed vertically now appear horizontal. "
                    "Focus on catching instrument tags along pipe runs and diagonal areas. "
                    "Do NOT re-report tags already clearly horizontal in the standard view."
                )
                logger.info(f"[InstrumentIndex] Page {page_no} — Rotation pass {angle}°CW [{only_engine or 'auto'}]")
                pr = self._vision_pass(
                    rot_jpeg, drawing_info, page_no,
                    extra_hint=hint, mode=label, max_tokens=cfg["max_tokens_primary"], only_engine=only_engine,
                )
                logger.info(f"[InstrumentIndex] Page {page_no} — Rotation {angle}°: {len(pr)} instruments")
                all_pass_instruments.extend(pr)

        # ── Tile passes: zoomed quadrant scan ────────────────────────────
        if cfg["enable_tiling"]:
            tiles = self._generate_tiles(img, cfg["tile_grid"], cfg["tile_overlap"])
            for tile_idx, tile_jpeg in enumerate(tiles):
                row = tile_idx // cfg["tile_grid"][1]
                col = tile_idx % cfg["tile_grid"][1]
                hint = (
                    f"ZOOMED TILE — Quadrant row={row+1}, col={col+1} of a {cfg['tile_grid'][0]}×{cfg['tile_grid'][1]} grid. "
                    "This is a high-resolution crop of part of the P&ID. "
                    "Extract EVERY instrument tag visible, including small or partially visible ones."
                )
                logger.info(
                    f"[InstrumentIndex] Page {page_no} — Tile ({row+1},{col+1}) [{only_engine or 'auto'}]"
                )
                pt = self._vision_pass(
                    tile_jpeg, drawing_info, page_no,
                    extra_hint=hint, mode=f"tile_r{row+1}c{col+1}",
                    max_tokens=cfg["max_tokens_tile"], only_engine=only_engine,
                )
                logger.info(f"[InstrumentIndex] Page {page_no} — Tile ({row+1},{col+1}): {len(pt)} instruments")
                all_pass_instruments.extend(pt)

        # ── Merge & deduplicate ──────────────────────────────────────────
        merged = self._merge_instruments(all_pass_instruments)

        # Enrich category from tag prefix
        for inst in merged:
            inst = self._enrich_category(inst)

        logger.info(
            f"[InstrumentIndex] Page {page_no} — "
            f"Total across all passes: {len(all_pass_instruments)}, "
            f"unique after merge: {len(merged)}"
        )
        return merged

    # ────────────────────────────────────────────────────────────────────
    # Vision pass dispatcher
    # ────────────────────────────────────────────────────────────────────

    def _vision_pass(self, jpeg_bytes, drawing_info, page_no,
                     extra_hint="", mode="primary", max_tokens=None, only_engine=None):
        """
        Try AI engines in priority order (EXTRACTION_CONFIG['ai_engines']).
        only_engine: when set, restricts to that single engine (e.g. 'gemini' or 'openai').
        Falls back to next engine on failure / quota exceeded.
        """
        if max_tokens is None:
            max_tokens = self.extraction_config["max_tokens_primary"]

        b64 = base64.b64encode(jpeg_bytes).decode("utf-8")
        if len(b64) > 20 * 1024 * 1024:
            logger.warning(f"[InstrumentIndex] {mode} image >20 MB — skipping")
            return []

        prompt = self._build_prompt(drawing_info, page_no, extra_hint=extra_hint)
        fallback_prompt = self._build_fallback_prompt(drawing_info, page_no, extra_hint=extra_hint)

        engines = self.extraction_config.get("ai_engines", ["gemini", "openai"])
        if only_engine:
            engines = [only_engine]
        for engine in engines:
            instruments = []
            if engine == "gemini":
                if self._gemini_quota_exceeded or not self.gemini_client:
                    continue
                instruments = self._call_gemini_vision(b64, prompt, f"{mode}_gemini", max_tokens)
                if not instruments:
                    instruments = self._call_gemini_vision(b64, fallback_prompt, f"{mode}_gemini_fb", max_tokens)
            elif engine == "openai":
                if self._quota_exceeded or not self.openai_client:
                    continue
                instruments = self._call_openai_vision(b64, prompt, f"{mode}_openai", max_tokens)
                if not instruments:
                    instruments = self._call_openai_vision(b64, fallback_prompt, f"{mode}_openai_fb", max_tokens)

            if instruments:
                logger.info(f"[InstrumentIndex] {mode} — {engine} returned {len(instruments)} instruments")
                return instruments

        logger.warning(f"[InstrumentIndex] {mode} — all AI engines exhausted or quota exceeded")
        return []

    # ────────────────────────────────────────────────────────────────────
    # Tile generator
    # ────────────────────────────────────────────────────────────────────

    def _generate_tiles(self, img, grid=(2, 2), overlap=0.12):
        """
        Crop a PIL image into rows×cols tiles with fractional overlap.
        Tiles are converted to JPEG bytes at full extraction config size.
        """
        rows, cols = grid
        w, h = img.size
        stride_w = w / cols
        stride_h = h / rows
        pad_w = int(w * overlap / 2)
        pad_h = int(h * overlap / 2)

        tiles = []
        for row in range(rows):
            for col in range(cols):
                left  = max(0, int(col * stride_w) - pad_w)
                upper = max(0, int(row * stride_h) - pad_h)
                right = min(w, int((col + 1) * stride_w) + pad_w)
                lower = min(h, int((row + 1) * stride_h) + pad_h)
                tile = img.crop((left, upper, right, lower))
                tiles.append(self._pil_to_jpeg(tile))
        return tiles

    # ────────────────────────────────────────────────────────────────────
    # Instrument merger / deduplicator
    # ────────────────────────────────────────────────────────────────────

    def _normalize_tag(self, tag):
        """Normalise tag number for deduplication (uppercase, stripped of spaces/dashes)."""
        return re.sub(r"[-_\s]+", "", (tag or "").upper().strip())

    def _merge_instruments(self, instruments):
        """
        Deduplicate instrument records from multiple passes.
        Rules:
        1. Same normalised tag → keep richest record (back-fill N/A fields).
        2. Partial-tag absorption: if a new tag's normalised key is a PREFIX of
           an existing longer tag (e.g. PI3700 vs PI370012), keep the longer one.
           This prevents OCR from generating both "PI-3700" and "PI-3700-12" as
           separate entries when they are the same physical instrument read at
           different precision levels.
        3. Records with no tag_number are always included as-is.
        """
        seen: dict[str, dict] = {}   # normalised_key → record
        no_tag: list = []

        def _norm_stripped(tag: str) -> str:
            """Remove ALL non-alphanumeric chars for prefix-match comparison."""
            return re.sub(r"[^A-Z0-9]", "", (tag or "").upper())

        for inst in instruments:
            raw_tag = inst.get("tag_number") or ""
            norm     = self._normalize_tag(raw_tag)
            norm_s   = _norm_stripped(raw_tag)   # for prefix checks

            if not norm:
                no_tag.append(inst)
                continue

            if norm in seen:
                # Same tag — back-fill missing fields
                existing = seen[norm]
                for key, value in inst.items():
                    if value and value != "N/A" and (
                        not existing.get(key) or existing[key] == "N/A"
                    ):
                        existing[key] = value
                continue

            # Check if the current tag is a prefix of an already-stored longer tag
            # e.g. "PI3700" is a prefix of "PI370012"
            is_prefix_of_existing = any(
                existing_norm_s.startswith(norm_s) and existing_norm_s != norm_s
                for existing_norm_s in (_norm_stripped(r.get("tag_number", "")) for r in seen.values())
            )
            if is_prefix_of_existing:
                # This candidate is less specific — skip it
                continue

            # Check if any existing tag is a prefix of this new (longer) tag
            # → replace the shorter existing tag with this more specific one
            to_remove = [
                k for k, r in seen.items()
                if norm_s.startswith(_norm_stripped(r.get("tag_number", "")))
                and _norm_stripped(r.get("tag_number", "")) != norm_s
                and _norm_stripped(r.get("tag_number", ""))  # not empty
            ]
            for k in to_remove:
                old = seen.pop(k)
                # Back-fill fields from the shorter record into the new longer one
                for key, value in old.items():
                    if value and value != "N/A" and (
                        not inst.get(key) or inst.get(key) == "N/A"
                    ):
                        inst[key] = value

            seen[norm] = dict(inst)

        return list(seen.values()) + no_tag

    def _call_vision(self, b64_image, prompt, mode_label, max_tokens=None):
        """Call gpt-4o with vision content; return parsed list or []."""
        cfg = self.extraction_config
        if max_tokens is None:
            max_tokens = cfg["max_tokens_primary"]
    # ────────────────────────────────────────────────────────────────────
    # Gemini Vision call
    # ────────────────────────────────────────────────────────────────────

    def _call_gemini_vision(self, b64_image, prompt, mode_label, max_tokens=None):
        """Call Gemini Vision (gemini-2.0-flash); return parsed list or []."""
        try:
            from google.genai import types as _gtypes
            cfg = self.extraction_config
            model = cfg.get("gemini_model", "gemini-2.0-flash")
            logger.info(f"[InstrumentIndex] Calling Gemini Vision ({mode_label}, model={model})…")

            image_bytes = base64.b64decode(b64_image)
            image_part  = _gtypes.Part.from_bytes(data=image_bytes, mime_type="image/jpeg")

            system_text = (
                "You are an expert P&ID analyst and process instrumentation engineer "
                "with 20+ years of experience reading ADNOC / oil & gas engineering drawings. "
                "You can identify ALL types of instrument symbols regardless of orientation — "
                "horizontal, vertical, rotated at any angle. "
                "Extract EVERY instrument tag you see. Return ONLY a valid JSON array."
            )

            response = self.gemini_client.models.generate_content(
                model=model,
                contents=[
                    _gtypes.Content(
                        role="user",
                        parts=[_gtypes.Part.from_text(text=prompt), image_part],
                    )
                ],
                config=_gtypes.GenerateContentConfig(
                    system_instruction=system_text,
                    temperature=cfg.get("temperature", 0.1),
                    max_output_tokens=max_tokens or cfg["max_tokens_primary"],
                ),
            )
            raw = response.text or ""
            logger.info(f"[InstrumentIndex] Gemini response {len(raw)} chars ({mode_label})")
            return self._parse_response(raw)

        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "quota" in err_str.lower() or "rate" in err_str.lower():
                retry_delay = self.extraction_config.get("gemini_retry_delay", 5)
                logger.warning(
                    f"[InstrumentIndex] Gemini rate-limit ({mode_label}) — "
                    f"waiting {retry_delay}s then retrying once…"
                )
                time.sleep(retry_delay)
                # Single retry attempt
                try:
                    from google.genai import types as _gtypes
                    cfg = self.extraction_config
                    model = cfg.get("gemini_model", "gemini-2.0-flash")
                    image_bytes = base64.b64decode(b64_image)
                    image_part  = _gtypes.Part.from_bytes(data=image_bytes, mime_type="image/jpeg")
                    system_text = (
                        "You are an expert P&ID analyst and process instrumentation engineer "
                        "with 20+ years of experience reading ADNOC / oil & gas engineering drawings. "
                        "Extract EVERY instrument tag you see. Return ONLY a valid JSON array."
                    )
                    response = self.gemini_client.models.generate_content(
                        model=model,
                        contents=[_gtypes.Content(role="user", parts=[_gtypes.Part.from_text(text=prompt), image_part])],
                        config=_gtypes.GenerateContentConfig(
                            system_instruction=system_text,
                            temperature=cfg.get("temperature", 0.1),
                            max_output_tokens=max_tokens or cfg["max_tokens_primary"],
                        ),
                    )
                    raw = response.text or ""
                    logger.info(f"[InstrumentIndex] Gemini retry succeeded {len(raw)} chars ({mode_label})")
                    return self._parse_response(raw)
                except Exception as e2:
                    err2 = str(e2)
                    if "429" in err2 or "quota" in err2.lower() or "rate" in err2.lower():
                        self._gemini_quota_exceeded = True
                        logger.error(f"[InstrumentIndex] Gemini quota exhausted ({mode_label}) — Gemini disabled for this request")
                    else:
                        logger.error(f"[InstrumentIndex] Gemini retry failed ({mode_label}): {e2}")
                    return []
            else:
                logger.error(f"[InstrumentIndex] Gemini Vision error ({mode_label}): {e}")
            return []

    # ────────────────────────────────────────────────────────────────────
    # OpenAI Vision call
    # ────────────────────────────────────────────────────────────────────

    def _call_openai_vision(self, b64_image, prompt, mode_label, max_tokens=None):
        """Call OpenAI gpt-4o Vision; return parsed list or []."""
        cfg = self.extraction_config
        if max_tokens is None:
            max_tokens = cfg["max_tokens_primary"]

        if self._quota_exceeded:
            logger.warning(f"[InstrumentIndex] Skipping OpenAI call ({mode_label}) — quota exceeded")
            return []
        try:
            logger.info(f"[InstrumentIndex] Calling OpenAI Vision ({mode_label}, max_tokens={max_tokens})…")
            resp = self.openai_client.chat.completions.create(
                model=cfg.get("model", "gpt-4o"),
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are an expert P&ID analyst and process instrumentation engineer "
                            "with 20+ years of experience reading ADNOC / oil & gas engineering drawings. "
                            "You can identify ALL types of instrument symbols regardless of orientation. "
                            "Extract EVERY instrument tag you see. Return ONLY a valid JSON array."
                        ),
                    },
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{b64_image}",
                                    "detail": "high",
                                },
                            },
                        ],
                    },
                ],
                max_tokens=max_tokens,
                temperature=cfg.get("temperature", 0.1),
            )
            raw = resp.choices[0].message.content
            logger.info(f"[InstrumentIndex] OpenAI response {len(raw)} chars ({mode_label})")
            return self._parse_response(raw)
        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "insufficient_quota" in err_str or "rate_limit" in err_str.lower():
                self._quota_exceeded = True
                logger.error(f"[InstrumentIndex] OpenAI quota/rate-limit hit ({mode_label}) — OpenAI disabled")
            else:
                logger.error(f"[InstrumentIndex] OpenAI Vision error ({mode_label}): {e}", exc_info=True)
            return []

    # Keep _call_vision as alias for backward compatibility
    def _call_vision(self, b64_image, prompt, mode_label, max_tokens=None):
        return self._call_openai_vision(b64_image, prompt, mode_label, max_tokens)

    # ────────────────────────────────────────────────────────────────────
    # Prompt templates
    # ────────────────────────────────────────────────────────────────────

    def _build_prompt(self, drawing_info, page_no, extra_hint=""):
        type_list = ", ".join(
            f"{k} ({v['name']})" for k, v in INSTRUMENT_CATEGORIES.items()
        )
        hint_block = f"\n⚡ SCAN CONTEXT: {extra_hint}\n" if extra_hint else ""
        return f"""
🎯 MISSION: Extract the COMPLETE Instrument Index from this P&ID drawing.
Page {page_no} — Drawing: {drawing_info.get('drawing_number', 'N/A')} — {drawing_info.get('drawing_title', 'N/A')}
Project: {drawing_info.get('project_name', 'N/A')}   Revision: {drawing_info.get('revision', '0')}
{hint_block}
─────────────────────────────────────────────
WHAT TO EXTRACT
─────────────────────────────────────────────
Extract EVERY instrument tag visible on this drawing.  A P&ID typically has
15–60+ instruments.  Do NOT skip any.

Target instrument tag prefixes (non-exhaustive):
{type_list}

Tag format examples from ADNOC / oil & gas:
  FIT-3901-08A   TI-3901-01   PIT-3901-03   LIT-3601-01   SDV-3901-01
  SVZY-3901-03   MOV-3901-01  PSV-3901-01   RO-3901-01    XPD-3901-01

─────────────────────────────────────────────
WHERE TO LOOK — INSTRUMENT CIRCLES / BUBBLES ARE THE PRIMARY SOURCE
─────────────────────────────────────────────
In P&ID drawings, instrument tags are shown INSIDE CIRCLES (instrument bubbles).
These circles may contain:
  • A single line of text:  "FT-3901-01"
  • Two lines:              "FT" (top) and "3901-01" (bottom)
  • Three lines:            "FIT" + "3901" + "08A"
READ every circle regardless of its size, angle, or position.

Also scan:
1. Circles / bubbles on ALL process lines — horizontal AND vertical pipe runs
2. Tags printed VERTICALLY along pipe runs (rotate reading angle)
3. Circles inside control loops (dashed boxes)
4. Tags connected to equipment nozzles (any angle)
5. Instrument tables in title block or margins
6. Any isolated letter+number combination inside or near a symbol

─────────────────────────────────────────────
FIELDS TO EXTRACT PER INSTRUMENT
─────────────────────────────────────────────
For EACH instrument found, return:

- tag_number          : Full tag e.g. "PIT-3901-01" — REQUIRED
- control_system_tag  : The DCS / Control-System tag for this instrument.
                        • If the instrument bubble is HEXAGONAL or inside a SHARED-DISPLAY / DCS box → it IS a CS tag; use the same tag number.
                        • If a second tag is shown near the instrument (in a box, or labelled "CS TAG", "DCS TAG") → use that tag.
                        • If unknown → "N/A"
- instrument_type     : Full description e.g. "Pressure Indicating Transmitter"
- category            : e.g. Flow / Pressure / Temperature / Level / Safety / Shutdown & ESD / etc.
- pid_no              : P&ID drawing number (default: "{drawing_info.get('drawing_number','N/A')}")
- service_description : What the instrument measures (e.g. "Pig Receiver Inlet Pressure")
- line_number         : Process line tag where instrument is installed
- equipment_number    : Associated equipment tag (vessel, pump, compressor, etc.)
- loop_number         : Control/safety loop number if shown
- fail_safe           : Fail-safe position — "FC" (fail closed), "FO" (fail open), "FL" (fail last), "N/A"
- signal_type         : "4-20mA", "Discrete (0/1)", "HART", "Fieldbus", "Pneumatic", "N/A"
- set_point           : Alarm / trip set point if shown on drawing or in instrument list
- drawing_number      : "{drawing_info.get('drawing_number','N/A')}"
- revision            : "{drawing_info.get('revision','0')}"
- notes               : Any relevant remark, special service (H2S, NACE, SIL), or uncertainty

─────────────────────────────────────────────
OUTPUT
─────────────────────────────────────────────
Return ONLY a JSON array — no markdown fences, no explanation text.
Example single record:
[
  {{
    "tag_number":         "PIT-3901-01",
    "control_system_tag": "PIC-3901-01",
    "instrument_type":    "Pressure Indicating Transmitter",
    "category":           "Pressure",
    "pid_no":             "{drawing_info.get('drawing_number','N/A')}",
    "service_description":"Pig Receiver Inlet Pressure",
    "line_number":        "10\\\"-G-3901-A2A",
    "equipment_number":   "LP-3901",
    "loop_number":        "3901",
    "fail_safe":          "N/A",
    "signal_type":        "4-20mA",
    "set_point":          "75 barg (PSHH)",
    "drawing_number":     "{drawing_info.get('drawing_number','N/A')}",
    "revision":           "{drawing_info.get('revision','0')}",
    "notes":              "SIL-rated loop"
  }}
]

⚠️ CRITICAL: Extract ALL instruments — including vertical/rotated text.
A response of [] or < 5 items for a process P&ID almost certainly means you missed instruments.
Scan in ALL orientations. Start response with [ and end with ].
"""

    def _build_fallback_prompt(self, drawing_info, page_no, extra_hint=""):
        """Simpler, more aggressive fallback prompt."""
        hint_block = f"\n⚡ {extra_hint}\n" if extra_hint else ""
        return f"""
EMERGENCY FALLBACK — Extract ALL instrument tags from this P&ID.
Page {page_no}  |  Drawing: {drawing_info.get('drawing_number', 'N/A')}
{hint_block}
Instructions:
1. Find EVERY circle or bubble containing a text tag on this drawing.
2. READ THE TAG even if it is printed vertically, upside-down, or at an angle.
3. Tag examples: FIT-1234, TI-56, SDV-3901-01, MOV-3901-02, LIT-101A, PT-8001, PSHH-001.
4. For each tag extract as much data as you can see.

Return JSON array only, format:
[
  {{
    "tag_number": "TAG-NO",
    "control_system_tag": "CS-TAG or N/A",
    "instrument_type": "Description",
    "category": "Category",
    "pid_no": "{drawing_info.get('drawing_number','N/A')}",
    "service_description": "what it measures",
    "line_number": "line tag or N/A",
    "equipment_number": "equipment tag or N/A",
    "loop_number": "N/A",
    "fail_safe": "N/A",
    "signal_type": "N/A",
    "set_point": "N/A",
    "drawing_number": "{drawing_info.get('drawing_number','N/A')}",
    "revision": "{drawing_info.get('revision','0')}",
    "notes": ""
  }}
]

Return ONLY the JSON array.
"""

    # ────────────────────────────────────────────────────────────────────
    # Response parsing helpers
    # ────────────────────────────────────────────────────────────────────

    def _parse_response(self, raw):
        """Extract JSON array from raw AI response text."""
        try:
            text = raw.strip()
            # Strip markdown fences
            for fence in ("```json", "```"):
                if fence in text:
                    start = text.find(fence) + len(fence)
                    end = text.find("```", start)
                    if end > start:
                        text = text[start:end].strip()
                        break

            # Find JSON array boundaries
            s = text.find("[")
            e = text.rfind("]") + 1
            if s >= 0 and e > s:
                data = json.loads(text[s:e])
                if isinstance(data, list):
                    return data

            # Try entire payload
            data = json.loads(text)
            if isinstance(data, list):
                return data
            if isinstance(data, dict) and "instruments" in data:
                return data["instruments"]

        except json.JSONDecodeError as exc:
            logger.error(f"[InstrumentIndex] JSON decode error: {exc}")
        except Exception as exc:
            logger.error(f"[InstrumentIndex] Parse error: {exc}", exc_info=True)
        return []

    def _enrich_category(self, inst):
        """Fill in instrument_type + category from INSTRUMENT_CATEGORIES if missing."""
        tag = (inst.get("tag_number") or "").strip().upper()
        if not tag:
            return inst

        # Extract function code (letters before first digit or dash-digit)
        match = re.match(r"^([A-Z]+)", tag)
        if not match:
            return inst
        code = match.group(1)

        cfg = INSTRUMENT_CATEGORIES.get(code)
        if cfg:
            if not inst.get("instrument_type"):
                inst["instrument_type"] = cfg["name"]
            if not inst.get("category"):
                inst["category"] = cfg["category"]
        return inst

    # ────────────────────────────────────────────────────────────────────
    # Excel export
    # ────────────────────────────────────────────────────────────────────

    def generate_excel(self, instruments, drawing_info):
        """
        Build an openpyxl workbook with two sheets:
          1. Instrument Index — row per instrument, category-coloured
          2. Summary          — count per category

        Returns bytes of the .xlsx file.
        """
        wb = openpyxl.Workbook()

        # ── Sheet 1: Instrument Index ────────────────────────────────────
        ws = wb.active
        ws.title = "Instrument Index"
        ws.sheet_view.showGridLines = True

        # Header style
        hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill   = PatternFill("solid", fgColor="1F4E79")
        hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin       = Side(style="thin", color="CCCCCC")
        std_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title row
        ws.row_dimensions[1].height = 22
        title_cell = ws.cell(row=1, column=1, value="INSTRUMENT INDEX")
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Metadata row
        ws.row_dimensions[2].height = 16
        ws.cell(row=2, column=1, value=f"Drawing: {drawing_info.get('drawing_number','N/A')}")
        ws.cell(row=2, column=5, value=f"Title: {drawing_info.get('drawing_title','N/A')}")
        ws.cell(row=2, column=9, value=f"Rev: {drawing_info.get('revision','0')}")
        ws.cell(row=2, column=11, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

        # Header row (row 4)
        ws.row_dimensions[4].height = 30
        for col_idx, col_def in enumerate(EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=4, column=col_idx, value=col_def["label"])
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = std_border
            ws.column_dimensions[cell.column_letter].width = col_def["width"]

        # Data rows
        DATA_START = 5
        for row_offset, inst in enumerate(instruments):
            row_no = DATA_START + row_offset
            ws.row_dimensions[row_no].height = 15

            category = inst.get("category") or "Special"
            fill_hex  = CATEGORY_COLOURS.get(category, "F5F5F5")
            row_fill  = PatternFill("solid", fgColor=fill_hex)
            std_font  = Font(name="Calibri", size=9)
            std_align = Alignment(vertical="center", wrap_text=False)

            for col_idx, col_def in enumerate(EXCEL_COLUMNS, start=1):
                val = inst.get(col_def["key"], "")
                cell = ws.cell(row=row_no, column=col_idx, value=val if val != "N/A" else "")
                cell.font = std_font
                cell.fill = row_fill
                cell.alignment = std_align
                cell.border = std_border

        # Freeze header
        ws.freeze_panes = "A5"

        # Auto-filter on header row
        ws.auto_filter.ref = (
            f"A4:{ws.cell(row=4, column=len(EXCEL_COLUMNS)).column_letter}4"
        )

        # ── Sheet 2: Summary ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 14

        ws2.row_dimensions[1].height = 22
        ws2.cell(row=1, column=1, value="INSTRUMENT INDEX — SUMMARY").font = Font(
            bold=True, size=13, color="1F4E79"
        )
        ws2.cell(row=2, column=1, value=f"Drawing: {drawing_info.get('drawing_number','N/A')}")
        ws2.cell(row=2, column=2, value=f"Total: {len(instruments)}")

        ws2.row_dimensions[4].height = 22
        for col, hdr in [(1, "Category"), (2, "Count")]:
            c = ws2.cell(row=4, column=col, value=hdr)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = std_border

        # Build category counts
        counts: dict[str, int] = {}
        for inst in instruments:
            cat = inst.get("category") or "Unknown"
            counts[cat] = counts.get(cat, 0) + 1

        row_n = 5
        for cat, cnt in sorted(counts.items()):
            fill_hex = CATEGORY_COLOURS.get(cat, "F5F5F5")
            for col, val in [(1, cat), (2, cnt)]:
                c = ws2.cell(row=row_n, column=col, value=val)
                c.font = Font(name="Calibri", size=10)
                c.fill = PatternFill("solid", fgColor=fill_hex)
                c.alignment = Alignment(vertical="center")
                c.border = std_border
            row_n += 1

        # Total row
        total_cell = ws2.cell(row=row_n, column=1, value="TOTAL")
        total_cell.font = Font(bold=True, size=10, color="1F4E79")
        ws2.cell(row=row_n, column=2, value=len(instruments)).font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
