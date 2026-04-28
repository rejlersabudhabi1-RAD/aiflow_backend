"""
Equipment Analysis Views - P&ID Equipment List Extraction
"""

import base64
import json
import logging
import os
import re
import threading
import uuid
from functools import lru_cache

from django.core.cache import cache
from django.http import HttpResponse
from rest_framework import status as drf_status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

# ── Soft-coded async-result cache constants ───────────────────────────────────
# Results are stored in Redis (same broker as Celery) so all gunicorn workers
# can read them regardless of which worker handled the original upload.
EQ_RESULT_CACHE_TTL_S   = 14400    # 4 hours — how long results stay in Redis
EQ_RESULT_CACHE_KEY_FMT = 'eq_analysis:{upload_id}'  # must match tasks.py


def _dispatch_eq_task(task_fn, upload_id: str, config: dict, *task_args) -> None:
    """
    Dispatch a Celery task with automatic thread fallback.

    Async backend is selected via `async_backend` in equipment_type_config.json:
      "celery"  — use Celery broker (.delay). Falls back to thread on failure.
      "thread"  — always use a background thread (no broker needed).

    Either way the task writes its result to Redis cache and the frontend
    polls /status/<upload_id>/ normally — the caller is unaffected.
    """
    backend = config.get('async_backend', 'celery')

    def _run_in_thread():
        try:
            task_fn.apply(args=task_args)
        except Exception as exc:
            logger.error('[EQDispatch] Thread execution failed upload_id=%s: %s',
                         upload_id, exc, exc_info=True)
            cache.set(
                EQ_RESULT_CACHE_KEY_FMT.format(upload_id=upload_id),
                {'status': 'failed', 'error': str(exc)},
                EQ_RESULT_CACHE_TTL_S,
            )

    if backend == 'thread':
        logger.info('[EQDispatch] async_backend=thread  upload_id=%s', upload_id)
        t = threading.Thread(target=_run_in_thread, daemon=True)
        t.start()
        return

    # Default: try Celery, fall back to thread on any broker error
    try:
        task_fn.delay(*task_args)
        logger.info('[EQDispatch] Celery task dispatched  upload_id=%s', upload_id)
    except Exception as dispatch_exc:
        logger.warning(
            '[EQDispatch] Celery broker unavailable (%s) — falling back to thread  upload_id=%s',
            dispatch_exc, upload_id,
        )
        t = threading.Thread(target=_run_in_thread, daemon=True)
        t.start()


def _eq_get_result_entry(upload_id: str) -> dict | None:
    """
    Look up a result entry for upload_id.
    Checks Redis cache first (written by Celery worker), then falls back to
    the in-process _result_store (written by synchronous callers in tests).
    """
    entry = cache.get(EQ_RESULT_CACHE_KEY_FMT.format(upload_id=upload_id))
    if entry is not None:
        return entry
    return _result_store.get(upload_id)

# Lazy import — avoids circular import at module load; models resolved at request time.
def _get_equipment_models():
    from apps.pid_analysis.models import PIDEquipmentType, PIDEquipmentItem  # noqa
    return PIDEquipmentType, PIDEquipmentItem

# Lazy import — avoids circular import at module load; models resolved at request time.
def _get_equipment_models():
    from apps.pid_analysis.models import PIDEquipmentType, PIDEquipmentItem  # noqa
    return PIDEquipmentType, PIDEquipmentItem

logger = logging.getLogger(__name__)

_CONFIG_PATH = os.path.join(
    os.path.dirname(__file__), 'config', 'equipment_type_config.json'
)


@lru_cache(maxsize=1)
def _load_config() -> dict:
    try:
        with open(_CONFIG_PATH, 'r', encoding='utf-8') as fh:
            raw = json.load(fh)
        return {k: v for k, v in raw.items() if not k.startswith('_')}
    except Exception as exc:
        logger.warning('[EquipmentList] Could not load config: %s - using defaults', exc)
        return {
            'extraction': {'context_window_chars': 120, 'description_max_words': 5},
            'type_labels': {
                'V': 'Vessel', 'P': 'Pump', 'E': 'Heat Exchanger', 'T': 'Tank',
                'K': 'Compressor', 'C': 'Column / Tower', 'H': 'Heater / Cooler',
                'D': 'Drum / Separator', 'R': 'Reactor',
            },
            'fluid_keywords': ['crude', 'gas', 'oil', 'water', 'steam'],
            'excel_columns': [
                {'key': 'sl_no',            'label': 'S. No',             'width': 6 },
                {'key': 'tag',              'label': 'Tag Number',         'width': 14},
                {'key': 'type_label',       'label': 'Equipment Type',     'width': 22},
                {'key': 'description',      'label': 'Description',        'width': 30},
                {'key': 'drawing_ref',      'label': 'Drawing Reference',  'width': 22},
                {'key': 'line_connections', 'label': 'Line Connections',   'width': 30},
                {'key': 'service_fluid',    'label': 'Service / Fluid',    'width': 20},
            ],
        }


_LINE_TAG_RE = re.compile(
    r'(?<![A-Za-z0-9])'
    r'(\d+(?:\.\d+)?)\s*["\u201c\u201d\u2019\'`]{1,2}'
    r'[\s\-_]{0,3}([A-Z]{1,4})[\s\-_]+(\d{3,6})[\s\-_]+(\d{4,8})'
    r'(?:[\s\-_]+([A-Z0-9]{1,8}))?'
    r'(?![A-Za-z0-9])',
    re.IGNORECASE,
)


def _normalize_text(text: str) -> str:
    """
    Normalize Unicode variants that CAD tools write in place of ASCII characters.
    Runs on every extracted text string before regex matching.

    Soft-coded by category — add mappings here without touching callers.
    """
    # Non-ASCII hyphens / dashes → ASCII hyphen  (most common CAD encoding issue)
    # U+2010 HYPHEN, U+2011 NON-BREAKING HYPHEN, U+2012/2013/2014 DASHES,
    # U+2212 MINUS SIGN, U+FE63 SMALL HYPHEN-MINUS, U+FF0D FULLWIDTH HYPHEN
    UNICODE_HYPHENS = '\u2010\u2011\u2012\u2013\u2014\u2212\ufe63\uff0d'
    for ch in UNICODE_HYPHENS:
        text = text.replace(ch, '-')
    # Non-breaking space → regular space
    text = text.replace('\u00a0', ' ').replace('\u202f', ' ')
    # Collapse repeated ASCII hyphens (OCR artifact: "V-803--TF" → "V-803-TF")
    import re as _re
    text = _re.sub(r'-{2,}', '-', text)
    return text


def _extract_text_from_pdf(file_obj, config=None, _page_index=None) -> str:
    """
    Extract all text from a PDF with three progressive strategies.

    _page_index (int | None): when specified (0-based), extract only that page.
    When None (default), all pages are processed — original behaviour.

    Strategy 1 — block text  : get_text('text') reading-order blocks (fast).
    Strategy 2 — spatial words: get_text('words') sorted by (y-bucket, x) to
        reconstruct left→right order regardless of CAD stream order.
    Strategy 3 — span proximity: iterate spans/chars to bond fragments that a
        CAD tool stored as separate micro-elements (e.g. "V-308" + "-TF").

    All three are run for EVERY vector PDF and their results concatenated so
    the downstream regex sees the text in every possible form.

    Scanned PDFs: OCR fallback (Tesseract) triggered when the combined vector
    text is shorter than the soft-coded min_vector_chars threshold.

    All config values are soft-coded in equipment_type_config.json.
    """
    cfg        = config or {}
    ext_cfg    = cfg.get('extraction', {})
    ocr_angles = ext_cfg.get('ocr_rotation_angles', [0, 90])
    ocr_psm_modes  = ext_cfg.get('ocr_psm_modes', [11, 6])
    ocr_scale      = float(ext_cfg.get('ocr_render_scale', 3.0))
    # Soft-coded: primary render scale for large-format pages (A0/A1 P&IDs)
    _OCR_SCALE_LARGE      = float(ext_cfg.get('ocr_render_scale_large_format', 4.0))
    # Soft-coded: additional full-page scales pooled on large pages
    _OCR_EXTRA_SCALES     = [float(s) for s in ext_cfg.get('ocr_additional_scales_large_format', [2.0])]
    _LARGE_PAGE_THRESHOLD = float(ext_cfg.get('ocr_large_page_threshold_pts', 900))
    # Soft-coded: tile-based OCR grid for large-format pages
    _TILE_ROWS     = int(ext_cfg.get('ocr_tile_rows', 3))
    _TILE_COLS     = int(ext_cfg.get('ocr_tile_cols', 4))
    _TILE_SCALE    = float(ext_cfg.get('ocr_tile_scale', 3.0))
    _TILE_PSM      = int(ext_cfg.get('ocr_tile_psm', 6))
    _TILE_OVERLAP  = float(ext_cfg.get('ocr_tile_overlap_frac', 0.12))
    # Soft-coded: vertical bucket height (pts) for spatially-sorted word pass
    _Y_BUCKET_PTS  = int(ext_cfg.get('spatial_word_y_bucket_pts', 15))
    # Soft-coded: max x-gap (pts) to bond two horizontally adjacent span fragments
    _SPAN_BOND_GAP = float(ext_cfg.get('span_bond_gap_pts', 20.0))
    # Soft-coded: threshold below which OCR fallback is triggered (chars)
    _MIN_VECTOR    = int(ext_cfg.get('min_vector_chars_for_ocr_skip', 200))
    # Soft-coded: always append OCR results even when vector text is long enough
    _ALWAYS_OCR    = bool(ext_cfg.get('always_include_ocr', True))

    text_parts: list = []
    file_bytes = None

    try:
        import fitz
        file_bytes = file_obj.read()
        doc = fitz.open(stream=file_bytes, filetype='pdf')

        for _pg_idx, page in enumerate(doc):
            if _page_index is not None and _pg_idx != _page_index:
                continue

            # ── Strategy 1: block text ──────────────────────────────────
            # TEXT_DEHYPHENATE excluded: structural hyphens in tags (V-308-TF)
            # must NOT be removed when they span a line boundary in the stream.
            blk_text = _normalize_text(page.get_text('text') or '')
            text_parts.append(blk_text)

            # ── Strategy 2: spatially-sorted word tokens ────────────────
            # get_text('words') → (x0,y0,x1,y1, word, block, line, word_no)
            words_raw = page.get_text('words')
            if words_raw:
                spatial = sorted(
                    words_raw,
                    key=lambda w: (round(w[1] / _Y_BUCKET_PTS) * _Y_BUCKET_PTS, w[0]),
                )
                spatial_text = _normalize_text(' '.join(w[4] for w in spatial))
                text_parts.append(spatial_text)

            # ── Strategy 3: span proximity bonding ─────────────────────
            # CAD tools (AutoCAD, SmartPlant, AVEVA) often write each word or
            # sub-token as an independent text span with a small positional gap.
            # get_text('words') treats a gap as a word boundary, so "V-308-TF"
            # may arrive as ["V-308", "-TF"] → joined with space → "V-308 -TF"
            # which breaks the regex.
            #
            # This pass iterates over character-level spans, sorts them by
            # (y-bucket, x) and bonds adjacent fragments whose right-edge to
            # next-left-edge gap is ≤ _SPAN_BOND_GAP pts, producing the
            # reconstructed token before adding a space.
            try:
                span_tokens: list = []  # (x0, reconstructed_text)
                raw_dict = page.get_text('rawdict')
                for blk in raw_dict.get('blocks', []):
                    for ln in blk.get('lines', []):
                        for sp in ln.get('spans', []):
                            txt = (sp.get('text') or '').strip()
                            if not txt:
                                continue
                            x0  = float(sp['bbox'][0])
                            y0  = float(sp['bbox'][1])
                            x1  = float(sp['bbox'][2])
                            row = round(y0 / _Y_BUCKET_PTS) * _Y_BUCKET_PTS
                            span_tokens.append((row, x0, x1, txt))

                span_tokens.sort(key=lambda t: (t[0], t[1]))

                bonded_parts: list = []
                buf = ''
                last_x1 = None
                last_row = None

                for row, x0, x1, txt in span_tokens:
                    if last_row is None:
                        buf = _normalize_text(txt)
                        last_x1 = x1
                        last_row = row
                    elif row == last_row and last_x1 is not None and (x0 - last_x1) <= _SPAN_BOND_GAP:
                        # Bond: adjacent on same row without visible gap
                        buf += _normalize_text(txt)
                        last_x1 = x1
                    else:
                        if buf:
                            bonded_parts.append(buf)
                        buf = _normalize_text(txt)
                        last_x1 = x1
                        last_row = row

                if buf:
                    bonded_parts.append(buf)

                if bonded_parts:
                    text_parts.append(' '.join(bonded_parts))
            except Exception as exc:
                logger.debug('[EquipmentList] Span bond pass failed: %s', exc)

        doc.close()
    except Exception as exc:
        logger.debug('[EquipmentList] PyMuPDF issue: %s', exc)

    full_text = '\n'.join(text_parts).strip()
    print(f'[EQ-DIAG] Vector text len={len(full_text)}  preview={repr(full_text[:200])}', flush=True)

    if _ALWAYS_OCR or len(full_text) < _MIN_VECTOR:
        # ── OCR fallback ─────────────────────────────────────────────────
        try:
            import fitz
            import pytesseract
            from PIL import Image, ImageEnhance, ImageFilter
            import io

            if file_bytes is None:
                file_obj.seek(0)
                file_bytes = file_obj.read()
            doc = fitz.open(stream=file_bytes, filetype='pdf')
            ocr_parts: list = []
            for _pg_idx, page in enumerate(doc):
                if _page_index is not None and _pg_idx != _page_index:
                    continue
                _r = page.rect
                _page_min_dim = min(abs(_r.width), abs(_r.height))
                _is_large = _page_min_dim > _LARGE_PAGE_THRESHOLD

                # Build list of scales to run for this page:
                # large-format pages run the primary large scale PLUS any
                # additional scales (e.g. 4.0 + 2.0) — different scales produce
                # different word-boundary decisions in Tesseract so pooling them
                # maximises tag coverage.
                if _is_large:
                    _scales_to_run = [_OCR_SCALE_LARGE] + _OCR_EXTRA_SCALES
                else:
                    _scales_to_run = [ocr_scale]

                seen_snippets: set = set()
                for _run_scale in _scales_to_run:
                    print(f'[EQ-DIAG] OCR page min_dim={_page_min_dim:.0f}  scale={_run_scale}', flush=True)
                    mat      = fitz.Matrix(_run_scale, _run_scale)
                    pix      = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
                    base_img = Image.open(io.BytesIO(pix.tobytes('png')))
                    _effective_dpi = int(72 * _run_scale)
                    base_img = ImageEnhance.Contrast(base_img).enhance(1.8)
                    base_img = base_img.filter(ImageFilter.SHARPEN)

                    for angle in ocr_angles:
                        rotated = base_img.rotate(-angle, expand=True) if angle != 0 else base_img
                        for psm in ocr_psm_modes:
                            ocr_text = pytesseract.image_to_string(
                                rotated, config=f'--oem 1 --psm {psm} --dpi {_effective_dpi}'
                            )
                            if not ocr_text.strip():
                                continue
                            fingerprint = ' '.join(ocr_text.split())[:200]
                            if fingerprint not in seen_snippets:
                                seen_snippets.add(fingerprint)
                                ocr_parts.append(_normalize_text(ocr_text))

                # ── Tile-based OCR pass for large-format pages ────────────────
                # Splits the page into a grid of tiles and OCRs each tile
                # independently.  Tesseract processes smaller, focused regions
                # more accurately than a single A0-scale image, so this pass
                # catches equipment tags in dense areas of the drawing that the
                # full-page pass misses.
                if _is_large and _TILE_ROWS > 0 and _TILE_COLS > 0:
                    print(f'[EQ-DIAG] Tiling {_TILE_ROWS}x{_TILE_COLS} scale={_TILE_SCALE} psm={_TILE_PSM}', flush=True)
                    _tile_dpi = int(72 * _TILE_SCALE)
                    _tmat  = fitz.Matrix(_TILE_SCALE, _TILE_SCALE)
                    _tpix  = page.get_pixmap(matrix=_tmat, colorspace=fitz.csGRAY)
                    _tfull = Image.open(io.BytesIO(_tpix.tobytes('png')))
                    _tfull = ImageEnhance.Contrast(_tfull).enhance(2.0)
                    _tfull = _tfull.filter(ImageFilter.SHARPEN)
                    _tw, _th = _tfull.size
                    for _ri in range(_TILE_ROWS):
                        for _ci in range(_TILE_COLS):
                            _x0 = max(0, int(_ci * _tw / _TILE_COLS - _tw * _TILE_OVERLAP / 2))
                            _y0 = max(0, int(_ri * _th / _TILE_ROWS - _th * _TILE_OVERLAP / 2))
                            _x1 = min(_tw, int((_ci + 1) * _tw / _TILE_COLS + _tw * _TILE_OVERLAP / 2))
                            _y1 = min(_th, int((_ri + 1) * _th / _TILE_ROWS + _th * _TILE_OVERLAP / 2))
                            _tile = _tfull.crop((_x0, _y0, _x1, _y1))
                            _tile_text = pytesseract.image_to_string(
                                _tile, config=f'--oem 1 --psm {_TILE_PSM} --dpi {_tile_dpi}'
                            )
                            if not _tile_text.strip():
                                continue
                            _fp = ' '.join(_tile_text.split())[:200]
                            if _fp not in seen_snippets:
                                seen_snippets.add(_fp)
                                ocr_parts.append(_normalize_text(_tile_text))
            doc.close()
            ocr_combined = '\n'.join(ocr_parts)
            print(f'[EQ-DIAG] OCR text len={len(ocr_combined)}  preview={repr(ocr_combined[:200])}', flush=True)
            # When always_include_ocr=true, APPEND to existing vector text.
            # When it's a pure OCR fallback (vector text was too short), REPLACE.
            if _ALWAYS_OCR and full_text:
                full_text = full_text + '\n' + ocr_combined
            else:
                full_text = ocr_combined
        except Exception as exc:
            logger.debug('[EquipmentList] Tesseract fallback issue: %s', exc)
            print(f'[EQ-DIAG] Tesseract fallback error: {exc}', flush=True)

    return full_text




# ---------------------------------------------------------------------------
# Equipment Register (18-field tabular document) extraction
# All thresholds / field-header variants are in equipment_type_config.json
# ---------------------------------------------------------------------------

_PAGE_Y_OFFSET      = 50000   # Vertical offset per PDF page so rows stay distinct
_Y_CLUSTER_TOL     = 12      # px — words within this y-distance are on the same row (vector PDF)
_Y_CLUSTER_TOL_OCR = 22      # px — wider tolerance for OCR; coords can drift more on scanned pages


def _cluster_words_into_rows(word_triples: list, y_tol: int = _Y_CLUSTER_TOL) -> list:
    """
    word_triples: list of (text, x, y) — may span multiple pages.
    Returns sorted list-of-rows, each row = [(text, x, y), ...] sorted by x.
    """
    if not word_triples:
        return []

    word_triples = sorted(word_triples, key=lambda w: (round(w[2] / _PAGE_Y_OFFSET), w[2]))
    rows: list = []
    current: list = [word_triples[0]]
    row_y = word_triples[0][2]

    for item in word_triples[1:]:
        # Treat items on different pages as always new rows
        same_page = abs(item[2] - row_y) < _PAGE_Y_OFFSET // 2
        if same_page and abs(item[2] - row_y) <= y_tol:
            current.append(item)
        else:
            rows.append(sorted(current, key=lambda w: w[1]))
            current = [item]
            row_y = item[2]

    if current:
        rows.append(sorted(current, key=lambda w: w[1]))
    return rows


def _extract_words_with_coords(file_obj, config: dict) -> tuple:
    """
    Returns (word_triples, used_ocr).
    word_triples: [(text, x, y), ...] from the PDF.
    Tries vector (PyMuPDF) first; falls back to pytesseract image_to_data.
    """
    cfg       = config.get('extraction', {})
    ocr_scale = float(cfg.get('ocr_render_scale', 3.0))
    word_list: list = []

    try:
        import fitz
        file_bytes = file_obj.read()
        doc = fitz.open(stream=file_bytes, filetype='pdf')
        for page_num, page in enumerate(doc):
            for entry in page.get_text('words'):
                x0, y0, x1, y1, word = entry[0], entry[1], entry[2], entry[3], entry[4]
                w = word.strip()
                if w:
                    word_list.append((w, x0, y0 + page_num * _PAGE_Y_OFFSET))
        doc.close()
    except Exception as exc:
        logger.debug('[EquipRegister] PyMuPDF words error: %s', exc)

    if len(word_list) > 30:
        return word_list, False  # vector PDF — use directly

    # ------- OCR fallback -------
    try:
        import fitz, pytesseract, io
        from PIL import Image, ImageEnhance, ImageFilter

        file_obj.seek(0)
        file_bytes = file_obj.read()
        doc = fitz.open(stream=file_bytes, filetype='pdf')
        ocr_words: list = []

        for page_num, page in enumerate(doc):
            mat = fitz.Matrix(ocr_scale, ocr_scale)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
            img = Image.open(io.BytesIO(pix.tobytes('png')))
            img = ImageEnhance.Contrast(img).enhance(2.0)
            img = img.filter(ImageFilter.SHARPEN)

            # Try 0° first; also 90° for landscape CAD title blocks
            for angle in cfg.get('ocr_rotation_angles', [0, 90, 180, 270])[:2]:
                rotated = img.rotate(-angle, expand=True) if angle else img
                try:
                    data = pytesseract.image_to_data(
                        rotated,
                        config='--oem 1 --psm 6',
                        output_type=pytesseract.Output.DICT,
                    )
                    for i, word in enumerate(data['text']):
                        w = str(word).strip()
                        raw_conf = data['conf'][i]
                        conf = int(raw_conf) if str(raw_conf).lstrip('-').isdigit() else 0
                        if w and conf > 20:
                            x = float(data['left'][i]) / ocr_scale
                            y = float(data['top'][i]) / ocr_scale + page_num * _PAGE_Y_OFFSET
                            ocr_words.append((w, x, y))
                except Exception:
                    continue

        doc.close()
        return ocr_words, True

    except Exception as exc:
        logger.debug('[EquipRegister] OCR fallback error: %s', exc)
        return [], True


def _norm_header(text: str) -> str:
    """Normalise a column header for fuzzy matching: uppercase, collapse punctuation/spaces.

    Strips & so 'P&ID' and 'P & ID' and 'P ID' all normalise to the same 'P ID' form,
    which means variants only need to cover the stripped form once.
    """
    s = text.upper()
    s = re.sub(r'[.\-/()\[\]&,]', ' ', s)   # & added: P&ID → P ID, A&E → A E
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# ── Soft-coded revision normalisation constants ──────────────────────────────
# Adjust these to cover new drawing conventions without touching logic.
#
# _REVISION_STRIP_PREFIX_RE  — removes leading "Rev"/"Revision"/"Rev." prefix
#                               common in registers that repeat the column label
#                               inside the cell (e.g. "Rev A" → "A").
# _REVISION_VALID_RE         — first match is the cleaned revision mark.
#                               Covers: single letter (A-Z), digit (0-9),
#                               letter+digit (A1, P0, R1), digit+letter (0A),
#                               3-letter IFC codes (IFC, IFD, IFA, IFR, AFD).
# _REVISION_MAX_RAW_LEN      — raw cell text longer than this is unlikely to be
#                               a revision mark; skip to avoid capturing dates or
#                               description bleed-over.
_REVISION_STRIP_PREFIX_RE = re.compile(
    r'^(?:rev(?:ision)?\s*(?:no\.?\s*|mark\s*|code\s*|\.\s*)?)',
    re.IGNORECASE,
)
_REVISION_VALID_RE = re.compile(
    r'\b(IF[CDABR]|AF[CD]|[A-Z]{1,2}[0-9]?|[0-9]{1,2}[A-Z]?)\b'
)
# Rejects date-like strings (e.g. 24/03/2025) that bleed from adjacent date columns.
_REVISION_DATE_RE = re.compile(r'\d[/\-.:]\d')
_REVISION_MAX_RAW_LEN = 20

# ── Soft-coded register extraction constants ──────────────────────────────────
# _REGISTER_HEADER_SCAN_ROWS — number of coordinate-sorted rows to scan when
#                              searching for the equipment register header.
#                              Set high (2000) so multi-page CRS/cover sheets
#                              before the equipment table are always skipped in
#                              favour of the real register header, which scores
#                              much higher (10+ columns vs 2-3 CRS columns).
_REGISTER_HEADER_SCAN_ROWS = 2000

# _REGISTER_HEADER_MIN_SCORE — minimum number of recognised fields the best
#                              header row must match before register mode is
#                              accepted.  Raised above min_cols (4) so that a
#                              CRS table whose header shares only "Rev" and
#                              "Description" with the equipment register schema
#                              is rejected and causes a fall-back to P&ID mode.
#                              Typical equipment register headers match 8-14
#                              fields; CRS headers match 2-4.
_REGISTER_HEADER_MIN_SCORE = 6

# _REGISTER_TAG_FILTER_RE    — after rows are extracted, any row whose 'tag'
#                              value does NOT contain this pattern is discarded
#                              as a footnote, note, or separator row.
#                              Equipment tags always look like "X-NNN" or
#                              "XX-NNNx" (1-2 cap letters + hyphen + 2-5 digits).
#                              This removes garbage like "units are in mm unless",
#                              "Pacakge document for" (note rows), ITEM-001
#                              fallback placeholders, and duplicate-text OCR noise.
#                              Set to None to disable filtering.
_REGISTER_TAG_FILTER_RE = re.compile(r'\b[A-Z]{1,2}-[0-9]{2,5}')

# _REGISTER_REPEATED_HDR_MARGIN — how much the header-match score must exceed
#   min_cols before a data row is treated as a repeated column-header and skipped.
#   Short variants like "R" (revision) and "NO" (sl_no) cause substring false-
#   positives against description text, so data rows routinely score 3-4.
#   Real repeated headers (same words as the header row) score 10+.
_REGISTER_REPEATED_HDR_MARGIN = 3

# _REV_PRE_TAG_WIN_CHARS     — characters to read BEFORE the matched tag in the
#                              raw OCR/vector text, used to find the revision cell
#                              in tabular PDFs where columns are newline-separated.
# _REV_PRE_TAG_TOKENS        — maximum number of newline-split tokens to inspect.
_REV_PRE_TAG_WIN_CHARS = 80
_REV_PRE_TAG_TOKENS    = 3

# ---------------------------------------------------------------------------
# QUANTITY REQUIRED — soft-coded constants.
# QUANTITY_REQUIRED_DEFAULT   : value placed in 'quality_required' when no
#                               explicit count callout is found on the drawing
#                               (virtually all single-unit equipment = 1).
# QUANTITY_REQUIRED_PATTERN   : regex to extract an explicit count callout.
#                               Matches: "QTY: 2", "QUANTITY = 1", "NO. REQD 3"
# QUALITY_SPEC_NACE_FULL_PAT  : sour-service NACE compliance string (full ref).
# QUALITY_SPEC_NACE_SHORT_PAT : abbreviated NACE reference.
# ---------------------------------------------------------------------------
QUANTITY_REQUIRED_DEFAULT  = '1'
QUANTITY_REQUIRED_PATTERN  = r'(?:QTY|QUANTITY|NO\.?\s*REQD?|NO\.?\s*REQUIRED|COUNT)\s*[:=]?\s*(\d+)'
QUALITY_SPEC_NACE_FULL     = 'NACE MR0175'
QUALITY_SPEC_NACE_SHORT    = 'NACE'

# ── Equipment-list Excel download filename (soft-coded) ────────────────────
# Kept short and constant per user request — earlier behaviour produced very
# long filenames like "<sanitised-drawing-ref>_equipment_list.xlsx".
EQUIPMENT_EXCEL_FILENAME   = 'Equipment_list.xlsx'

# ── Dimension extraction minimum-value filters (soft-coded) ────────────────
# Dimension values BELOW these thresholds are rejected as pipe-size / nozzle /
# instrument-level false positives picked up by OCR from context text.
# P&ID pipe annotations (e.g. "2"-FL-…" → OCR "2 M") are the classic source.
# Increase thresholds here only — no logic changes required.
DIMENSION_LENGTH_MIN_M    = 0.5   # metres  — reject vessel length  < 0.5 M
DIMENSION_LENGTH_MIN_MM   = 500   # mm      — reject vessel length  < 500 mm
DIMENSION_DIAMETER_MIN_M  = 0.3   # metres  — reject vessel diameter < 0.3 M
DIMENSION_DIAMETER_MIN_MM = 300   # mm      — reject vessel diameter < 300 mm

# _REVISION_USE_TOPMOST      — when True, the first non-empty revision value
#                              found in the register (topmost row) is applied to
#                              ALL extracted rows.  Equipment registers typically
#                              carry one document revision; individual rows should
#                              all reflect the current (topmost) revision mark.
#                              Set to False to keep per-row revision values.
_REVISION_USE_TOPMOST = True

# _HEADER_MAX_SPAN_ROWS      — maximum number of consecutive rows that can form
#                              a table column header.  CAD equipment registers
#                              commonly split long column labels across 3 lines
#                              (e.g. "Des./Set" / "Press." / "Min (PSIG)").
#                              Raising this from 2 → 3 ensures the qualifier
#                              row ("Min"/"Max") is included when building the
#                              column-x map, fixing extraction of
#                              design_pressure_min and design_pressure_max.
_HEADER_MAX_SPAN_ROWS = 3

# ── Soft-coded title-block revision extraction patterns ───────────────────────
# Used in P&ID drawing mode to extract the DOCUMENT revision from the title
# block, which applies uniformly to all equipment on the drawing.
#
# _TITLEBLOCK_REV_LABEL_RE   — explicit "REV" / "REVISION" label in title block,
#                              followed by the revision mark.  Capture group 1
#                              is the revision value.
# _TITLEBLOCK_DRAWN_CTX_RE   — title-block revision-history row pattern:
#                              a single revision mark that appears in a line
#                              containing DR(AWN)/CH(ECKED)/AP(PROVED) keywords
#                              (e.g. "A  IFR  12/04/2025  MAK  AKR  HJS").
#                              The revision mark is always the first short token.
# _TITLEBLOCK_ISOLATED_RE    — last-resort: single isolated letter/digit on its
#                              own line that appears within a few lines of the
#                              drawing number pattern (project doc-no format).
_TITLEBLOCK_REV_LABEL_RE   = re.compile(
    r'(?:^|\n)\s*REV(?:ISION)?\.?\s*[:\-]\s*([A-Z0-9]{1,3})\s*(?:\n|$)',
    re.IGNORECASE | re.MULTILINE,
)
# _TITLEBLOCK_REVTABLE_ROW_RE — matches a revision-history table row in the
# O&G title block convention:
#   REV_MARK  DD/MM/YYYY  ISSUED FOR .../APPROVED FOR .../RE-APPROVED FOR ...
# Captures groups: (1) REV mark, (2) day, (3) month, (4) year.
# Strategy 0 in _extract_titleblock_revision finds ALL matches and returns
# the mark whose date is the LATEST — correct regardless of OCR read order
# (rows may appear oldest→newest or newest→oldest depending on tile/scale).
# Handles both numeric (0, 1, 2) and alpha (A, B, C, IFC) revision marks.
_TITLEBLOCK_REVTABLE_ROW_RE = re.compile(
    r'(?:^|\s)([0-9]{1,2}[A-Z]?|[A-Z]{1,3})\s+'
    r'(\d{2})[/\-](\d{2})[/\-](\d{4})\s+'
    r'(?:ISSUED|APPROVED|RE[\s\-]?APPROVED|RETURNED|INCORPORATED)',
    re.IGNORECASE,
)
_TITLEBLOCK_DRAWN_CTX_RE   = re.compile(
    r'(?:DR[\'.]?N|DRW|DRAWN|CH[\'.]?D|CHK|CHECKED|APP?[\'.]?D|APPROVED)',
    re.IGNORECASE,
)
# Matches document numbers such as PJ6-EXD-MRI-BQDA-0023 (4–5 hyphen segments,
# last segment is 4–6 digits, each segment is 2–6 alphanumeric chars).
# The first segment may contain digits (e.g. PJ6) so [A-Z0-9]+ is used.
# Anchored with word boundary; minimum total length 10 to avoid short tags.
_TITLEBLOCK_DWG_NO_RE      = re.compile(
    r'\b([A-Z0-9]{2,6}(?:-[A-Z0-9]{2,6}){3,4})\b',
    re.IGNORECASE,
)
# Label that precedes the drawing number in the title block
_TITLEBLOCK_DWG_LABEL_RE   = re.compile(
    r'(?:DWG\.?\s*NO\.?|DRAWING\s*NO\.?|DOCUMENT\s*NO\.?|DOC\.?\s*NO\.?)',
    re.IGNORECASE,
)
# Reference-context words that appear BEFORE a 'DWG NO.' label and indicate
# that the number following it is a reference document, NOT the title-block
# drawing number.  Look-back window is widened to _DWG_LABEL_LOOKBACK_CHARS
# so that longer prefixes (e.g. "REFERENCE DRAWING NO.", "PFD MUBARRAZ ISLAND")
# are captured.
# PFD is included because P&ID reference-document lists typically list the
# associated PFD with its own DWG. NO. entry — that number must be excluded
# so only the actual title-block P&ID number is returned.
_TITLEBLOCK_DWG_REF_CTX_RE = re.compile(
    r'\b(FEED|PFD|P\.F\.D|REF(?:ERENCE)?|RELATED|FROM|VENDOR|CLIENT|PREVIOUS|PARENT|APPLICABLE|ATTACH(?:ED|MENT)?|LIST|TABLE|INDEX)\b',
    re.IGNORECASE,
)
# How many characters to look back before the 'DWG NO.' label when checking
# for reference-context words (soft-coded so it can be tuned without touching logic).
_DWG_LABEL_LOOKBACK_CHARS  = 80

# ── Soft-coded title-block coordinate extraction constants ────────────────────
# Engineering drawings always have the title block in the bottom strip of the
# page.  These constants define how large that strip is (as a fraction of the
# total page height) and how wide the horizontal scan window is (in points)
# after a 'DWG. NO.' label word is found within that strip.
# Adjust these without touching any logic in _extract_titleblock_dwg_no_by_coords.
_TITLEBLOCK_STRIP_FRACTION  = 0.30   # bottom 30 % of page height (wider = safer)
_TITLEBLOCK_SCAN_WINDOW_PT  = 400    # pts to the right of the label word
# PyMuPDF returns individual words; "DWG. NO." is split into two tokens.
# This regex matches the FIRST token (the trigger word: DWG / DRAWING / DOC /
# DOCUMENT).  After finding a trigger, the coord function checks whether the
# very next word on the same row is "NO" / "NO." to confirm the label.
_TITLEBLOCK_DWG_TRIGGER_RE  = re.compile(
    r'^(DWG\.?|DRAW(?:ING)?\.?|DOC(?:UMENT)?\.?)$',
    re.IGNORECASE,
)
# Matches the "NO" / "NO." word that follows the trigger word in the title block.
_TITLEBLOCK_NO_WORD_RE      = re.compile(r'^NO\.?$', re.IGNORECASE)
# Maximum horizontal gap (pts) between the trigger word and the "NO." word
# when they are on the same row (handles varying spacing in different CAD tools).
_TITLEBLOCK_LABEL_GAP_PT    = 80
# Engineering title blocks often use a VERTICALLY-STACKED layout: the label
# "DWG. NO." is on one row and the actual document number is in the cell BELOW
# it on the next row.  This constant controls how many points below the trigger
# word we continue to search for a doc-number pattern.
# NOTE: on rotated A1/A0 sheets PyMuPDF stores the vertical title-block column
# in PDF coordinate space, meaning the value cell can be ~240 pts below the
# label.  300 pts covers that gap while staying within the title-block column.
_TITLEBLOCK_BELOW_SCAN_PT   = 300
# How many characters after a 'DWG. NO.' label to search in the plain text for
# the document number.  Title blocks interleave the label row with other cells
# (REV., DATE, DESCRIPTION, company name …) before the value row appears;
# 600 chars is sufficient to cross that gap without reaching unrelated content.
_TITLEBLOCK_DWG_LABEL_WINDOW_CHARS = 600

# ── Soft-coded operating temperature range normalisation ─────────────────────
# Equipment registers sometimes store two operating temperatures in a single
# cell (e.g. shell/tube, inlet/outlet, or min/max condition) separated by "/".
# e.g. "105/60 °F" → "60 – 105 °F"  (ascending range, engineering convention)
_TEMP_RANGE_SEPARATOR = ' \u2013 '   # en-dash with spaces  (matches frontend constant)
_TEMP_SLASH_RE        = re.compile(
    r'^(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)\s*(°[FC]|DEG\s*[FC]|[FC])?$',
    re.IGNORECASE,
)


def _normalize_oper_temp(raw: str) -> str:
    """
    Normalise an operating temperature string that contains two values
    separated by "/" into a clean ascending range.

    Examples:
        "105/60 °F"  ->  "60 – 105 °F"
        "60 / 105"   ->  "60 – 105 °F"   (°F assumed)
        "175 °F"     ->  "175 °F"         (unchanged)
        ""           ->  ""               (unchanged)
    """
    if not raw or '/' not in raw:
        return raw
    m = _TEMP_SLASH_RE.match(raw.strip())
    if not m:
        return raw
    v1, v2 = float(m.group(1)), float(m.group(2))
    raw_unit = (m.group(3) or '').strip().upper()
    # Normalise unit display
    if raw_unit in ('F', 'DEGF') or raw_unit.endswith('F'):
        unit = '°F'
    elif raw_unit in ('C', 'DEGC') or raw_unit.endswith('C'):
        unit = '°C'
    else:
        unit = '°F'   # default for process equipment
    lo, hi = sorted([v1, v2])
    fmt = lambda v: str(int(v)) if v == int(v) else str(v)   # strip ".0" suffix
    return f'{fmt(lo)}{_TEMP_RANGE_SEPARATOR}{fmt(hi)} {unit}'


# Soft-coded: regex for non-numeric pressure qualifiers appended after "/" such as
# "FV" (Full Vacuum), "FULL VAC", "FULL VACUUM".  These are engineering annotations
# that describe a limit condition and are NOT part of the numeric pressure value.
# Applied as a post-processing strip on design_pressure_min / design_pressure_max.
_PRESS_STRIP_SUFFIX_RE = re.compile(
    r'\s*/\s*(?:FV|FULL\s*VAC(?:UUM)?)\s*$', re.IGNORECASE
)


def _clean_pressure_value(s: str) -> str:
    """Strip non-numeric trailing qualifiers like '/ FV' from a pressure string.

    Examples:
        "195 psig/ FV"   ->  "195 psig"
        "195 psig"       ->  "195 psig"   (unchanged)
        ""               ->  ""           (unchanged)
    """
    if not s or '/' not in s:
        return s
    return _PRESS_STRIP_SUFFIX_RE.sub('', s.strip()).strip()


def _extract_titleblock_dwg_no_by_coords(file_bytes: bytes) -> str:
    """
    Extract the drawing / document number from the title block using page
    coordinates (PyMuPDF).

    Engineering drawings always place the title block in the BOTTOM strip of
    the last (or only) page.  This function restricts its search to that strip
    (controlled by _TITLEBLOCK_STRIP_FRACTION) so it is immune to reference-
    document tables that also contain 'DWG. NO.' labels higher up the page.

    Key fix: PyMuPDF splits "DWG. NO." into two separate word tokens.  This
    function therefore looks for a TRIGGER word ("DWG", "DRAWING", "DOC",
    "DOCUMENT") and then confirms the following same-row word is "NO" / "NO."
    before scanning to the right for the document-number pattern.

    Soft-coded constants:
        _TITLEBLOCK_STRIP_FRACTION, _TITLEBLOCK_SCAN_WINDOW_PT,
        _TITLEBLOCK_DWG_TRIGGER_RE, _TITLEBLOCK_NO_WORD_RE,
        _TITLEBLOCK_LABEL_GAP_PT, _TITLEBLOCK_DWG_NO_RE.
    """
    try:
        import fitz as _fitz
        doc = _fitz.open(stream=file_bytes, filetype='pdf')
        # Title block is on the LAST page of multi-page drawings; try all pages
        pages_to_check = list(range(len(doc) - 1, -1, -1))  # last page first
        for pg_idx in pages_to_check:
            pg = doc[pg_idx]
            pg_height = pg.rect.height
            strip_top  = pg_height * (1.0 - _TITLEBLOCK_STRIP_FRACTION)

            words = pg.get_text('words')  # (x0, y0, x1, y1, word, block, line, span)
            strip_words = [w for w in words if (w[1] + w[3]) / 2 >= strip_top]
            print(f'[EQ-DIAG][DwgNo] Coord: page={pg_idx} height={pg_height:.0f} strip_top={strip_top:.0f} strip_words={len(strip_words)}', flush=True)
            if not strip_words:
                continue

            # Sort left-to-right, top-to-bottom within the strip
            strip_words.sort(key=lambda w: (round(w[1] / 10) * 10, w[0]))

            for i, w in enumerate(strip_words):
                token = w[4].strip()
                if not _TITLEBLOCK_DWG_TRIGGER_RE.match(token):
                    continue

                trigger_x0    = w[0]   # left edge of trigger word
                trigger_x1    = w[2]   # right edge of trigger word
                trigger_y_mid = (w[1] + w[3]) / 2

                # Locate the subsequent words on the same row
                same_row = [
                    nw for nw in strip_words[i + 1:]
                    if abs((nw[1] + nw[3]) / 2 - trigger_y_mid) <= 20
                    and nw[0] >= trigger_x1
                ]

                # Check if next word is "NO" / "NO." to confirm it's a DWG NO. label
                scan_anchor_x1 = trigger_x1
                scan_start_idx = 0
                if same_row:
                    next_w     = same_row[0]
                    is_no_word = _TITLEBLOCK_NO_WORD_RE.match(next_w[4].strip())
                    gap_ok     = (next_w[0] - trigger_x1) <= _TITLEBLOCK_LABEL_GAP_PT
                    if is_no_word and gap_ok:
                        scan_anchor_x1 = next_w[2]
                        scan_start_idx = 1

                # ── Scan to the RIGHT (same row) for doc number ──
                for nw in same_row[scan_start_idx:]:
                    if nw[0] - scan_anchor_x1 > _TITLEBLOCK_SCAN_WINDOW_PT:
                        break
                    cand = nw[4].strip().upper()
                    if _TITLEBLOCK_DWG_NO_RE.fullmatch(cand) and re.search(r'-[0-9]{4,6}$', cand):
                        doc.close()
                        print(f'[EQ-DIAG][DwgNo] Coord (right) page={pg_idx} y={trigger_y_mid:.0f} trigger={token!r}: {cand!r}', flush=True)
                        return cand

                # ── Scan BELOW the trigger (vertically-stacked title block layout) ──
                # Many engineering title blocks place "DWG. NO." label on one row
                # and the actual document number in the cell directly below it.
                # Scan all words within _TITLEBLOCK_BELOW_SCAN_PT pts below the
                # trigger whose x0 is near the trigger's x0 column.
                for nw in strip_words[i + 1:]:
                    nw_y_mid = (nw[1] + nw[3]) / 2
                    if nw_y_mid <= trigger_y_mid:
                        continue   # above or same row — already scanned
                    if nw_y_mid > trigger_y_mid + _TITLEBLOCK_BELOW_SCAN_PT:
                        break      # too far below
                    # x must be roughly in the same column as the trigger
                    if abs(nw[0] - trigger_x0) > _TITLEBLOCK_SCAN_WINDOW_PT:
                        continue
                    cand = nw[4].strip().upper()
                    if _TITLEBLOCK_DWG_NO_RE.fullmatch(cand) and re.search(r'-[0-9]{4,6}$', cand):
                        doc.close()
                        print(f'[EQ-DIAG][DwgNo] Coord (below) page={pg_idx} y={trigger_y_mid:.0f}→{nw_y_mid:.0f} trigger={token!r}: {cand!r}', flush=True)
                        return cand
        doc.close()
    except Exception as _e:
        import logging as _log
        _log.getLogger(__name__).warning('[EquipmentList] Coord DWG NO extraction failed: %s', _e)
    return ''


def _extract_titleblock_dwg_no(text: str) -> str:
    """
    Extract the drawing / document number from a P&ID title block.

    Strategy 1: look for the label 'DWG. NO.' / 'DRAWING NO.' followed by
        a document-number pattern within 120 chars.
    Strategy 2: find the most frequent multi-segment document number
        (4–5 segments, ends in 4-digit sequence) anywhere in the text.
        Filters out equipment tags (≤3 chars prefix + ≤5 digits) and
        change-request numbers.

    Returns the extracted drawing number string, or '' if not found.
    """
    # Strategy 1: label-adjacent.
    # Collect ALL 'DWG NO.'-label adjacent candidates, skip any whose
    # pre-context contains a reference qualifier (FEED, REF, RELATED, etc.).
    # Return the LAST valid candidate — title-block labels are OCR'd last
    # because the title block sits at the bottom of the drawing sheet.
    # Lookback window and reference-context pattern are soft-coded module
    # constants (_DWG_LABEL_LOOKBACK_CHARS, _TITLEBLOCK_DWG_REF_CTX_RE).
    _label_candidates: list = []
    for lbl_m in _TITLEBLOCK_DWG_LABEL_RE.finditer(text):
        pre_ctx = text[max(0, lbl_m.start() - _DWG_LABEL_LOOKBACK_CHARS):lbl_m.start()]
        if _TITLEBLOCK_DWG_REF_CTX_RE.search(pre_ctx):
            continue
        window = text[lbl_m.end():lbl_m.end() + _TITLEBLOCK_DWG_LABEL_WINDOW_CHARS]
        m = _TITLEBLOCK_DWG_NO_RE.search(window)
        if m:
            candidate = m.group(1).upper()
            # Must end with digits — excludes equipment tags like V-308-TF
            if re.search(r'-[0-9]{4,6}$', candidate):
                _label_candidates.append(candidate)
    if _label_candidates:
        # Prefer the last match (title block is at the bottom of the sheet)
        best = _label_candidates[-1]
        print(f'[EQ-DIAG][DwgNo] Found via label strategy (last of {len(_label_candidates)}): {best!r}', flush=True)
        return best

    # Strategy 2: for each 'DWG NO.' label, find the nearest doc-number after it.
    # Return the candidate found after the LAST label occurrence — because the
    # actual title-block label is the last one OCR'd (it is at the bottom of the
    # sheet), while reference-table labels appear earlier.
    # Falls back to the most-frequent candidate if no label-anchored match found.
    _s2_all: list = []
    for m in _TITLEBLOCK_DWG_NO_RE.finditer(text):
        cand = m.group(1).upper()
        if re.search(r'-[0-9]{4,6}$', cand) and len(cand) >= 10:
            _s2_all.append((m.start(), cand))
    if _s2_all:
        _lbl_positions = [lm.end() for lm in _TITLEBLOCK_DWG_LABEL_RE.finditer(text)]
        if _lbl_positions:
            # Try each label from LAST to FIRST; return the nearest candidate after it
            for lbl_end in reversed(_lbl_positions):
                for cpos, cval in sorted(_s2_all, key=lambda x: x[0]):
                    if cpos >= lbl_end:
                        print(f'[EQ-DIAG][DwgNo] Strategy2 last-label nearest: {cval!r}', flush=True)
                        return cval
        # Fallback: most common candidate
        from collections import Counter as _Counter
        freq_best = _Counter(c for _, c in _s2_all).most_common(1)[0][0]
        print(f'[EQ-DIAG][DwgNo] Strategy2 frequency fallback: {freq_best!r}', flush=True)
        return freq_best

    return ''


def _extract_titleblock_revision(text: str) -> str:
    """
    Extract the DOCUMENT revision mark from a P&ID title block.

    Uses three progressive strategies (all soft-coded via module constants):

    Strategy 1 — Explicit label: looks for "REV[.] A" / "REVISION: B" patterns
        anywhere in the extracted text.  Returns the LAST such match because
        the topmost OCR text is often the legend/cover sheet; the title block
        with the current revision appears later.

    Strategy 2 — Drawn/Checked/Approved context: scans each line that contains
        DR'N / CH'D / AP'D keywords (the revision-history table in the title
        block) and extracts the first short token on that line as the revision.
        Takes the LAST such match (= the most recent / lowest-numbered row in
        the revision table, which in O&G conventions is the CURRENT revision).

    Strategy 3 — Isolated token near drawing number: looks for a standalone
        1-3 char alphanumeric token on lines adjacent to a project document
        number (pattern: two-letter-code–section–discipline–doc-no).

    All candidate values are validated by _clean_revision before return.
    Returns '' if no valid revision found.
    """
    if not text:
        return ''

    # Strategy 0: revision table row (most reliable — O&G title block convention)
    # Pattern: REV_MARK  DD/MM/YYYY  ISSUED/APPROVED FOR ...
    # Finds ALL matching rows, parses the date from each, and returns the mark
    # whose date is LATEST.  This is correct regardless of OCR read order —
    # some drawings OCR oldest→newest, others newest→oldest depending on how
    # tiles/spatial passes are merged.
    rev_row_matches = list(_TITLEBLOCK_REVTABLE_ROW_RE.finditer(text))
    if rev_row_matches:
        best_mark = ''
        best_date = (0, 0, 0)   # (year, month, day)
        for rm in rev_row_matches:
            cleaned = _clean_revision(rm.group(1))
            if not cleaned:
                continue
            try:
                day   = int(rm.group(2))
                month = int(rm.group(3))
                year  = int(rm.group(4))
                rev_date = (year, month, day)
            except (IndexError, ValueError):
                # Date parse failed — accept as fallback only
                if not best_mark:
                    best_mark = cleaned
                continue
            if rev_date > best_date:
                best_date = rev_date
                best_mark = cleaned
        if best_mark:
            print(
                f'[EQ-DIAG][TitleBlock] Rev found via table-row strategy: {best_mark!r} '
                f'(date {best_date[2]:02d}/{best_date[1]:02d}/{best_date[0]})',
                flush=True,
            )
            return best_mark

    # Strategy 1: explicit "REV: value" label (strict colon required)
    matches = _TITLEBLOCK_REV_LABEL_RE.findall(text)
    if matches:
        # Validate each match; take the last valid one
        for raw in reversed(matches):
            cleaned = _clean_revision(raw)
            if cleaned:
                print(f'[EQ-DIAG][TitleBlock] Rev found via label strategy: {cleaned!r}', flush=True)
                return cleaned

    # Strategy 2: drawn/checked/approved context rows
    lines = text.split('\n')
    last_rev_from_ctx = ''
    for line in lines:
        if _TITLEBLOCK_DRAWN_CTX_RE.search(line):
            tokens = [t.strip() for t in line.split() if t.strip()]
            for tok in tokens:
                cleaned = _clean_revision(tok)
                if cleaned:
                    last_rev_from_ctx = cleaned
                    break
    if last_rev_from_ctx:
        print(f'[EQ-DIAG][TitleBlock] Rev found via DR/CH/AP context: {last_rev_from_ctx!r}', flush=True)
        return last_rev_from_ctx

    # Strategy 3: isolated short token near project document number
    dwg_matches = list(_TITLEBLOCK_DWG_NO_RE.finditer(text))
    for dwg_m in dwg_matches:
        # Look at a ±300 char window around the drawing number
        win_start = max(0, dwg_m.start() - 300)
        win_end   = min(len(text), dwg_m.end() + 300)
        window    = text[win_start:win_end]
        for ln in window.split('\n'):
            stripped = ln.strip()
            if re.match(r'^[A-Z0-9]{1,3}$', stripped, re.IGNORECASE):
                cleaned = _clean_revision(stripped)
                if cleaned:
                    print(f'[EQ-DIAG][TitleBlock] Rev found via DWG-no proximity: {cleaned!r}', flush=True)
                    return cleaned

    print('[EQ-DIAG][TitleBlock] No document revision found in title block', flush=True)
    return ''


def _clean_revision(raw: str) -> str:
    """Normalise an extracted revision cell value to a short clean mark.

    Steps
    -----
    1. Strip whitespace; bail out if cell is too long to be a revision mark.
    2. Strip leading 'Rev'/'Revision' prefix (some registers duplicate the
       column label inside every cell, e.g. 'Rev A' → 'A').
    3. If the remaining text is already 1-3 chars and alphanumeric, return it.
    4. Otherwise scan for the first token matching _REVISION_VALID_RE.
    5. Return the match in uppercase, or '' if nothing valid found.

    Soft-coded via module-level constants:
      _REVISION_STRIP_PREFIX_RE, _REVISION_DATE_RE, _REVISION_VALID_RE, _REVISION_MAX_RAW_LEN
    """
    if not raw:
        return ''
    s = raw.strip()
    if len(s) > _REVISION_MAX_RAW_LEN:
        # Cell is too long to be a valid revision mark — likely a bleed-over
        # from an adjacent wide column (description, remarks).
        return ''
    # Reject date-like values (e.g. 24/03/2025) that bleed from adjacent columns
    if _REVISION_DATE_RE.search(s):
        return ''
    # Reject "Note N" or "(Note N)" bleed from description/remarks columns
    if re.search(r'\bnote\b', s, re.IGNORECASE):
        return ''
    # Strip 'Rev' / 'Revision' prefix
    s = _REVISION_STRIP_PREFIX_RE.sub('', s).strip()
    if not s:
        return ''
    # If already short and clean, return immediately (fast path)
    if re.match(r'^[A-Za-z0-9]{1,3}$', s):
        return s.upper()
    # Scan for first valid revision token in the (possibly noisy) remainder
    m = _REVISION_VALID_RE.search(s.upper())
    return m.group(1) if m else ''


def _find_header_range(rows: list, field_variants: dict, min_cols: int) -> tuple:
    """
    Scan first _REGISTER_HEADER_SCAN_ROWS rows for the table header row(s).
    Supports single-row and double-row headers (common in CAD documents).
    Returns (start_idx, end_idx_exclusive) or None if not found.

    Uses _REGISTER_HEADER_MIN_SCORE (≥ min_cols) so that low-scoring CRS/cover
    headers that share only "Rev" and "Description" with the equipment schema are
    rejected rather than used as a fallback.
    """
    scan_limit = min(_REGISTER_HEADER_SCAN_ROWS, len(rows))
    best_score = 0
    best_range: tuple = (0, 1)

    all_variants_norm = {
        k: [_norm_header(v) for v in variants]
        for k, variants in field_variants.items()
    }

    for start in range(scan_limit):
        for span in range(1, _HEADER_MAX_SPAN_ROWS + 1):
            end = min(start + span, len(rows))
            combined_norm = _norm_header(
                ' '.join(t for row in rows[start:end] for (t, x, y) in row)
            )
            score = sum(
                1 for variants_norm in all_variants_norm.values()
                if any(v in combined_norm for v in variants_norm)
            )
            if score > best_score:
                best_score = score
                best_range = (start, end)

    # Use the stricter _REGISTER_HEADER_MIN_SCORE threshold so that a CRS table
    # whose header only matches "Rev" + "Description" (score ≤ 3) is rejected.
    required_score = max(min_cols, _REGISTER_HEADER_MIN_SCORE)
    print(f'[EQ-DIAG][Register] header scan: total_rows={len(rows)} scan_limit={scan_limit}'
          f'  best_score={best_score}  best_range={best_range}  required={required_score}', flush=True)
    if best_score < required_score:
        return None
    return best_range


def _build_col_map(header_rows: list, field_variants: dict) -> dict:
    """
    Build mapping: field_key -> x_center from the header row(s).

    Handles multi-line CAD table headers by:
    1. Sorting all header words by (x, y) so same-column words are adjacent.
    2. Grouping into x-column clusters.
    3. Trying left-neighbor merges for short unmatched clusters (handles
       "Des./Set Press. Min" where Min lands in its own cluster).
    4. Greedy conflict resolution to avoid two clusters claiming the same field.
    """
    all_words_y = [(t.strip(), float(x), float(y))
                   for row in header_rows for (t, x, y) in row if t.strip()]
    if not all_words_y:
        return {}

    all_variants_norm = {
        k: [_norm_header(v) for v in variants]
        for k, variants in field_variants.items()
    }

    # ── Sort by x, then y ────────────────────────────────────────────────────
    sorted_by_x = sorted(all_words_y, key=lambda w: (w[1], w[2]))

    # ── Adaptive x-cluster tolerance ─────────────────────────────────────────
    distinct_xs = sorted(set(round(v[1]) for v in sorted_by_x))
    if len(distinct_xs) > 1:
        gaps = [distinct_xs[i + 1] - distinct_xs[i] for i in range(len(distinct_xs) - 1)]
        median_gap = sorted(gaps)[len(gaps) // 2]
        x_col_tol = max(median_gap * 0.8, 8.0)
    else:
        x_col_tol = 15.0

    # ── Form initial clusters ─────────────────────────────────────────────────
    col_clusters: list = []
    current: list = [sorted_by_x[0]]
    for we in sorted_by_x[1:]:
        cm = sum(w[1] for w in current) / len(current)
        if abs(we[1] - cm) <= x_col_tol:
            current.append(we)
        else:
            col_clusters.append(current)
            current = [we]
    if current:
        col_clusters.append(current)

    # ── Helper: build phrase + x-center from a cluster list ──────────────────
    def _cluster_info(cluster: list) -> tuple:
        ro = sorted(cluster, key=lambda w: (w[2], w[1]))
        phrase = _norm_header(' '.join(w[0] for w in ro))
        x_c = sum(w[1] for w in cluster) / len(cluster)
        return phrase, x_c

    # ── Helper: score a phrase against a single field ─────────────────────────
    def _score(phrase: str, field_key: str) -> int:
        best = 0
        for variant in all_variants_norm.get(field_key, []):
            if phrase == variant:
                s = len(variant) * 2
            elif len(variant) >= 3 and variant in phrase:
                s = len(variant)
            elif len(phrase) >= 3 and len(variant) >= 5 and phrase in variant:
                s = len(phrase)
            else:
                continue
            if s > best:
                best = s
        return best

    # ── Step 4: Build (score, x_center, field_key) candidates ────────────────
    # Each cluster produces ALL matching fields (not just best), then we do
    # greedy conflict-free assignment.  Also try LEFT-MERGE for short clusters
    # (catches "Min"/"Max" separated from their prefix by a gap).
    all_matches: list = []  # (score, x_center, field_key)

    for ci, cluster in enumerate(col_clusters):
        phrase, x_c = _cluster_info(cluster)

        # Also try merging with left neighbor (helps "Des./Set Press." + "Min")
        if ci > 0 and len(cluster) <= 2:
            merged = col_clusters[ci - 1] + cluster
            merged_phrase, merged_xc = _cluster_info(merged)
        else:
            merged_phrase, merged_xc = None, None

        for field_key in all_variants_norm:
            s = _score(phrase, field_key)
            use_x = x_c   # always use a local copy — do NOT mutate x_c
            # Prefer merged phrase only if it yields a strictly better score
            if merged_phrase is not None:
                ms = _score(merged_phrase, field_key)
                if ms > s:
                    s, use_x = ms, merged_xc
            if s > 0:
                all_matches.append((s, use_x, field_key))

    # ── Step 5: Greedy conflict-free assignment ───────────────────────────────
    # Sort by score desc, then by x (stable ordering for equal scores).
    all_matches.sort(key=lambda m: (-m[0], m[1]))

    field_best: dict = {}  # field_key -> (score, x_center)
    # Track which physical x-centers have already been "used" (±5pt tolerance)
    used_x: list = []

    for score, x_center, field_key in all_matches:
        if field_key in field_best:
            continue  # field already claimed
        # Check whether a different field already claimed this x_center
        already_used = any(abs(x_center - ux) < 5.0 for ux in used_x)
        if already_used:
            continue
        field_best[field_key] = (score, x_center)
        used_x.append(x_center)

    return {k: v[1] for k, v in field_best.items()}


def _assign_row_to_cols(data_row: list, col_map: dict) -> dict:
    """
    Assign each word in data_row to the nearest column by x-distance.
    Returns dict field_key -> value_string.
    """
    if not data_row or not col_map:
        return {}

    sorted_cols = sorted(col_map.items(), key=lambda c: c[1])   # (key, x)
    n_cols = len(sorted_cols)

    # Midpoints between adjacent columns
    boundaries = [
        (sorted_cols[i][1] + sorted_cols[i + 1][1]) / 2
        for i in range(n_cols - 1)
    ]

    buckets: dict = {k: [] for k in col_map}
    for (text, x, _y) in data_row:
        col_idx = 0
        for bi, bx in enumerate(boundaries):
            if x > bx:
                col_idx = bi + 1
            else:
                break
        assigned = sorted_cols[col_idx][0]
        buckets[assigned].append((x, text))

    return {
        k: ' '.join(txt for _, txt in sorted(items)).strip()
        for k, items in buckets.items()
        if items
    }


def _pid_item_to_register_schema(pid_item: dict) -> dict:
    """Map a P&ID-extraction item to the 18-field register schema."""
    return {
        'sl_no':               str(pid_item.get('sl_no', '')),
        'revision':            _clean_revision(str(pid_item.get('revision', ''))),
        'tag':                 pid_item.get('tag', ''),
        'description':         pid_item.get('description', ''),
        'design_flowrate':     pid_item.get('design_flowrate', ''),
        'oper_pressure':       pid_item.get('oper_pressure', ''),
        'oper_temperature':    pid_item.get('oper_temperature', ''),
        'design_pressure_min': pid_item.get('design_pressure_min', ''),
        'design_pressure_max': pid_item.get('design_pressure_max', ''),
        'design_temp_min':     pid_item.get('design_temp_min', ''),
        'design_temp_max':     pid_item.get('design_temp_max', ''),
        'moc':                 pid_item.get('material_class', ''),
        'insulation':          pid_item.get('insulation', ''),
        'dimension_length':    pid_item.get('dimension_length', ''),
        'dimension_diameter':  pid_item.get('dimension_diameter', ''),
        'motor_rating':        pid_item.get('motor_rating', ''),
        'pid_no':              pid_item.get('drawing_ref', ''),
        'quality_required':    pid_item.get('quality_required', ''),
        'phase':               pid_item.get('service_fluid', ''),
        'remarks':             pid_item.get('remarks', '') or pid_item.get('process_notes', ''),
        # Backward-compat fields kept for status/results endpoints
        'type_label':         pid_item.get('type_label', ''),
        'area':               pid_item.get('area', ''),
        'drawing_ref':        pid_item.get('drawing_ref', ''),
        'line_connections':   pid_item.get('line_connections', []),
        'nozzle_connections': pid_item.get('nozzle_connections', []),
        'service_fluid':      pid_item.get('service_fluid', ''),
        'material_class':     pid_item.get('material_class', ''),
        'process_notes':      pid_item.get('process_notes', ''),
    }


def _extract_equipment_register_rows(file_obj, config: dict):
    """
    Extract 18-field Equipment Register from a tabular CAD/PDF document.

    Uses coordinate-based table detection (PyMuPDF words + pytesseract
    image_to_data as fallback) so it works on both vector and scanned PDFs.

    Returns list of equipment dicts if the document is a register table,
    or None if the document doesn't look like a register (triggers P&ID fallback).
    """
    field_variants = config.get('equip_register_fields', {})
    min_cols       = int(config.get('equip_register_min_columns', 4))
    min_rows       = int(config.get('equip_register_min_rows', 2))
    # Soft-coded: shortest page dimension (pts) above which we treat the doc
    # as a large-format P&ID drawing and skip register detection entirely.
    # A4 landscape smallest dim = 595 pts; A3 = 842 pts; A1/A0 >> 1000 pts.
    # Equipment registers are A4; P&IDs are A1-A0. Threshold = 900 pts.
    max_drawing_min_dim = int(config.get('equip_register_skip_if_page_dim_gt', 900))

    if not field_variants:
        return None  # Config missing — skip register mode

    # ── Page-size guard ──────────────────────────────────────────────────────
    # Large-format drawings (A1/A0 P&IDs) can accidentally match headers from
    # equipment data boxes (DIAMETER, LENGTH, OPERATING PRESS, etc.).
    # Skip register mode when the smallest page dimension exceeds the threshold.
    try:
        import fitz as _fitz
        _fb = file_obj.read()
        _doc = _fitz.open(stream=_fb, filetype='pdf')
        if _doc.page_count > 0:
            _r = _doc[0].rect
            _min_dim = min(abs(_r.width), abs(_r.height))
            print(f'[EQ-DIAG][Register] page_min_dim={_min_dim:.0f}pts  threshold={max_drawing_min_dim}', flush=True)
            if _min_dim > max_drawing_min_dim:
                _doc.close()
                print('[EQ-DIAG][Register] Large-format drawing -> skipping register mode', flush=True)
                return None
        _doc.close()
        file_obj.seek(0)
    except Exception as _exc:
        logger.debug('[EquipRegister] Page-size check failed: %s', _exc)
        try:
            file_obj.seek(0)
        except Exception:
            pass

    logger.info('[EquipRegister] Starting coordinate-based table extraction')

    word_list, used_ocr = _extract_words_with_coords(file_obj, config)
    if not word_list:
        logger.info('[EquipRegister] No words extracted')
        return None

    # Use wider y-tolerance for OCR pages — coordinates are less precise
    y_tol = _Y_CLUSTER_TOL_OCR if used_ocr else _Y_CLUSTER_TOL
    rows = _cluster_words_into_rows(word_list, y_tol=y_tol)
    if len(rows) < 3:
        logger.info('[EquipRegister] Too few rows (%d)', len(rows))
        return None

    header_range = _find_header_range(rows, field_variants, min_cols)
    if header_range is None:
        logger.info('[EquipRegister] No register header detected')
        return None

    h_start, h_end = header_range
    col_map = _build_col_map(rows[h_start:h_end], field_variants)
    if len(col_map) < min_cols:
        logger.info('[EquipRegister] Too few columns mapped (%d)', len(col_map))
        return None

    logger.info('[EquipRegister] Columns detected: %s', list(col_map.keys()))

    # All variants for repeated-header detection
    all_variants_norm = {
        k: [_norm_header(v) for v in variants]
        for k, variants in field_variants.items()
    }

    equipment: list = []
    row_counter = 0

    for row in rows[h_end:]:
        row_text = ' '.join(t for (t, x, y) in row).strip()
        if not row_text or len(row_text) < 2:
            continue

        # Skip repeated header rows (some CAD drawings repeat headers each page)
        combined_norm = _norm_header(row_text)
        hdr_score = sum(
            1 for v_list in all_variants_norm.values()
            if any(v in combined_norm for v in v_list)
        )
        if hdr_score >= min_cols + _REGISTER_REPEATED_HDR_MARGIN:
            continue

        values = _assign_row_to_cols(row, col_map)
        tag_val = values.get('tag', '').strip()
        sl_val  = values.get('sl_no', '').strip()

        # ── Tag clean + rescue ────────────────────────────────────────────────
        # Full equipment tag pattern: 1-3 cap letters, hyphen, 2-5 digits,
        # optional alpha suffix (A, B, A/B/C), optional project suffix (-TF, -1F).
        # Soft-coded via _REGISTER_TAG_FULL_RE.  Applied in order:
        #  1. Strip trailing noise from column-assigned tag (e.g. "PX-851-TF MRD …")
        #  2. If no valid tag in the column value, check whether neighbouring words
        #     in the same row land a valid tag (handles x-coordinate drift that
        #     places the tag token in the revision column instead of the tag column).
        if tag_val:
            _tm = re.search(
                r'\b([A-Z]{1,3}-[0-9]{2,5}[A-Za-z]?(?:[/\-][A-Z0-9]{1,4})*)\b',
                tag_val
            )
            tag_val = _tm.group(1) if _tm else tag_val

        # Check unconditionally whether the revision column contains an equipment
        # tag that landed there due to x-coordinate drift.  When that happens the
        # real tag is in the rev cell and description text bleeds into the tag cell.
        # We accept the swap only when what remains after removing the tag from rev
        # is a plausible revision mark of 0-2 characters (e.g. "1", "A", "1A", "").
        _TAG_FULL_RE = re.compile(
            r'\b([A-Z]{1,3}-[0-9]{2,5}[A-Za-z]?(?:[/\-][A-Z0-9]{1,4})*)\b'
        )
        _rev_raw = values.get('revision', '')
        if _rev_raw:
            _rv_tag = _TAG_FULL_RE.search(_rev_raw)
            if _rv_tag:
                _cleaned_rev = _rev_raw.replace(_rv_tag.group(1), '').strip()
                if re.match(r'^[0-9A-Za-z]{0,2}$', _cleaned_rev):
                    tag_val = _rv_tag.group(1)
                    values['revision'] = _cleaned_rev
                    print(f'[EQ-DIAG][Register] rev-rescue: extracted tag {tag_val!r} '
                          f'from rev col; rev cleaned to {values["revision"]!r}', flush=True)

        if not _REGISTER_TAG_FILTER_RE.search(tag_val):
            # Tag column didn't yield a valid tag — scan the full row text (all
            # words ordered by x-position) and take the leftmost matching token.
            _row_plain = ' '.join(t for (t, x, y) in sorted(row, key=lambda w: w[1]))
            _rescue = re.search(
                r'\b([A-Z]{1,3}-[0-9]{2,5}[A-Za-z]?(?:[/\-][A-Z0-9]{1,4})*)\b',
                _row_plain
            )
            if _rescue:
                tag_val = _rescue.group(1)
                print(f'[EQ-DIAG][Register] tag-rescue: row "{_row_plain[:50]}" → {tag_val}', flush=True)
        # ── End tag clean + rescue ────────────────────────────────────────────

        # Count non-empty fields; skip blank/near-blank rows regardless of tag presence.
        # A row that has ≥ 2 mapped fields is kept even if neither tag nor sl_no are
        # populated — this handles: (a) registers where tag column wasn't matched,
        # (b) multi-line continuation rows caught here before the post-merge pass.
        populated = sum(1 for v in values.values() if v.strip())
        if populated < 2:
            continue

        row_counter += 1
        item: dict = {
            'sl_no':               sl_val or str(row_counter),
            'revision':            _clean_revision(values.get('revision', '')),
            'tag':                 tag_val or f'ITEM-{row_counter:03d}',
            'description':         values.get('description', ''),
            'design_flowrate':     values.get('design_flowrate', ''),
            'oper_pressure':       values.get('oper_pressure', ''),
            'oper_temperature':    _normalize_oper_temp(values.get('oper_temperature', '')),
            'design_pressure_min': _clean_pressure_value(values.get('design_pressure_min', '')),
            'design_pressure_max': _clean_pressure_value(values.get('design_pressure_max', '')),
            'design_temp_min':     values.get('design_temp_min', ''),
            'design_temp_max':     values.get('design_temp_max', ''),
            'moc':                 values.get('moc', ''),
            'insulation':          values.get('insulation', ''),
            'dimension_length':    values.get('dimension_length', ''),
            'dimension_diameter':  values.get('dimension_diameter', ''),
            'motor_rating':        values.get('motor_rating', ''),
            'pid_no':              values.get('pid_no', ''),
            'quality_required':    values.get('quality_required', ''),
            'phase':               values.get('phase', ''),
            'remarks':             values.get('remarks', ''),
            # Backward-compat fields
            'type_label':         '',
            'area':               '',
            'drawing_ref':        '',
            'line_connections':   [],
            'nozzle_connections': [],
            'service_fluid':      values.get('oper_pressure', ''),
            'material_class':     values.get('moc', ''),
            'process_notes':      values.get('remarks', ''),
        }
        equipment.append(item)

    # Post-merge: combine continuation rows (wrapped cells) into the preceding item.
    # A row is a continuation candidate if both tag and sl_no are empty and it has
    # ≤ 3 mapped fields — typically the second line of a wrapped description cell.
    if equipment:
        merged_equip: list = [equipment[0]]
        for item in equipment[1:]:
            is_cont = (
                not item.get('tag') and not item.get('sl_no')
                and sum(1 for v in item.values() if isinstance(v, str) and v.strip()) <= 3
            )
            if is_cont and merged_equip:
                prev = merged_equip[-1]
                for fld in ('description', 'remarks', 'moc', 'insulation'):
                    if item.get(fld):
                        prev[fld] = (prev.get(fld, '') + ' ' + item[fld]).strip()
            else:
                merged_equip.append(item)
        equipment = merged_equip

    # ── Tag-pattern filter ────────────────────────────────────────────────────
    # Remove footnote/separator rows whose 'tag' cell doesn't contain a real
    # equipment tag (e.g. "Note 1: units are in mm", "Water Treatment Package",
    # ITEM-NNN placeholders, duplicated OCR noise lines).
    # Controlled by _REGISTER_TAG_FILTER_RE; set to None to disable.
    if _REGISTER_TAG_FILTER_RE is not None:
        before_filter = len(equipment)
        equipment = [
            _item for _item in equipment
            if _REGISTER_TAG_FILTER_RE.search(_item.get('tag', ''))
        ]
        removed = before_filter - len(equipment)
        if removed:
            print(f'[EQ-DIAG][Register] Tag-filter removed {removed} non-equipment rows '
                  f'(footnotes/notes/separators), {len(equipment)} rows kept', flush=True)

    # Confirm enough populated rows to treat this as a real register
    well_populated = sum(
        1 for item in equipment
        if sum(
            1 for k, v in item.items()
            if k not in ('sl_no', 'tag', 'type_label', 'area', 'drawing_ref',
                         'line_connections', 'nozzle_connections')
            and (v if not isinstance(v, list) else v)
        ) >= 3   # lowered from 5 — sparse registers (tag + desc + one pressure) should qualify
    )

    if len(equipment) < min_rows or well_populated < min_rows:
        logger.info('[EquipRegister] Insufficient populated rows (total=%d, well_pop=%d)',
                    len(equipment), well_populated)
        return None

    # ── Topmost revision override ─────────────────────────────────────────────
    # Apply the topmost (first) non-empty revision value to ALL rows.
    # Equipment registers carry one document revision; individual row cells
    # often read wrongly due to column-coordinate drift or adjacent date bleed.
    # Controlled by _REVISION_USE_TOPMOST constant (True by default).
    if _REVISION_USE_TOPMOST and equipment:
        topmost_rev = next(
            (item['revision'] for item in equipment if item.get('revision')),
            ''
        )
        if topmost_rev:
            for item in equipment:
                item['revision'] = topmost_rev
            print(f'[EQ-DIAG][Register] Topmost-revision applied: "{topmost_rev}" → all {len(equipment)} rows', flush=True)

    logger.info('[EquipRegister] Extracted %d register rows (OCR=%s)', len(equipment), used_ocr)
    return equipment


# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Data-box index — global scanner for equipment data boxes in P&ID drawings
# ---------------------------------------------------------------------------

# Soft-coded: chars to look BACK from the label to find the associated tag.
_DATABOX_TAG_LOOKBACK_CHARS = 1500


def _norm_databox_label(label: str) -> str:
    """Lowercase + collapse punctuation/whitespace for map key comparison."""
    s = label.lower()
    s = re.sub(r'[./()_\-]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# Soft-coded: stop patterns for data-box value cleaning.
# When OCR merges a data-box value with adjacent text (e.g. NOTES on the same
# visual row), these patterns mark where the useful value ends.
# Order matters: more-specific patterns first.
# Soft-coded value-stop patterns for _clean_databox_value.
# Each branch marks where the useful numeric/unit value ends and junk begins.
_DATABOX_VALUE_STOPS = re.compile(
    r'(?:'
    r'\s+\d+\.\s+[A-Z]{3,}'            # note number: "  2. CORROSION..."
    r'|\bSLOPE\b'                      # slope annotation (SLOPE 1:200)
    r'|\bSOUR\s+GAS\b'                 # flow routing annotation
    r'|\bIN\s+SCOPE\b'                 # scope note
    r'|\bWILL\s+BE\b'                  # notes prose
    r'|\bSHALL\b'                      # notes prose
    r'|\bSCOPE\b'                      # scope note
    r'|\bHEADER\b'                     # pipe destination
    r'|\bPRESSURIZED\b'               # ADNOC note prose after design press
    r'|\bINTERLOCK\b'                 # safety interlock note prose
    r'|\bELIMINATOR\b'                # internals note prose (MIST ELIMINATOR)
    r'|\bVORTEX\b'                    # internals note prose (VORTEX BREAKER)
    r'|>\s*FL\b'                      # ADNOC to-flare annotation (>FL)
    r'|\s*>'                          # routing arrow (> FL -ACBN-8128); never valid in a data value
    r'|[;|]'                           # column separator / pipe char
    r'|//'                             # // separating dual-condition values
    r'|\s{3,}'                         # 3+ spaces = OCR column gap
    r'|\s*\\\s'                       # backslash-space OCR artifact (e.g. \ 82106)
    r'|\s+(?:[A-Z]{1,3}\d{4,}[A-Z]?|\d{4,}[A-Z]+)\b'  # ref tags with letter (8210A, Re21084) — NOT bare 1412/1219
    r'|[\"\u201c\u201d\u00a5][A-Z]{2,4}-'   # pipe tag starting with inch mark (incl. mojibake)
    r'|\s+\d+[\"\u201c\u201d\u00a5]\s*-\s*[A-Z]'  # N"-PIPETAG inline (e.g. 6"-SG-)
    r')',
    re.IGNORECASE,
)

# Known data-box label starts — if the value contains one of these after the
# first word, truncate there (OCR ran two rows together).
_DATABOX_LABEL_STARTS_RE = re.compile(
    r'\b(?:NOMINAL|OPERATING|DESIGN|LENGTH|DIAMETER|MATERIAL|MOTOR|HEAT|DUTY'
    r'|CAPACITY|VOLUME|RATED|TRIM|INSULATION|WEIGHT|SHELL|INTERNALS?)\b',
    re.IGNORECASE,
)


def _clean_databox_value(raw: str) -> str:
    """
    Truncate a raw data-box extracted value at the first noise/stop pattern.

    Data-box values on large P&IDs are often followed on the same OCR line by:
    - NOTES text (e.g. "327 M3  2. CORROSION COUPON AND PROBE WILL BE...")
    - A second data-box label (OCR merged two rows)
    - Column separators (spaces, pipes, semicolons)

    Soft-coded via _DATABOX_VALUE_STOPS and _DATABOX_LABEL_STARTS_RE.
    Returns the cleaned, stripped value or '' if nothing useful remains.
    """
    if not raw:
        return ''

    # Truncate at first stop pattern
    stop_m = _DATABOX_VALUE_STOPS.search(raw)
    if stop_m:
        raw = raw[:stop_m.start()]

    # Truncate at the SECOND occurrence of a known label keyword
    # (first occurrence may be part of the value itself e.g. "SHELL DIAMETER")
    label_hits = list(_DATABOX_LABEL_STARTS_RE.finditer(raw))
    if len(label_hits) >= 2:
        raw = raw[:label_hits[1].start()]

    raw = raw.strip(' .,;:()')
    # Reject values that are pure punctuation or very short after cleaning.
    # Minimum 2 chars: allows valid 2-char engineering codes (CS, SS, FV)
    # while rejecting single-character OCR noise ('J', 'X', etc.).
    if len(raw) < 2:
        return ''

    # ── OCR unit correction ──────────────────────────────────────────────
    # Tesseract commonly misreads 'M3' as 'MS', 'M²' as 'M2', 'm³' as 'm3'.
    # Soft-coded replacements applied ONLY to the unit portion (after a digit).
    _ocr_unit_fixes = [
        (re.compile(r'\bft\s*[?³]\s*/\s*h', re.IGNORECASE), 'ft3/H'),  # ft³/hr, ft?/hr → ft3/H
        (re.compile(r'(\d+\.?\d*)\s*MS\b', re.IGNORECASE), r'\1 M3'),  # 327 MS → 327 M3
        (re.compile(r'\bM\s+3\b'),                              'M3'),   # M 3 → M3
        (re.compile(r'\bM3/[Hh]\b'),                            'M3/H'), # normalise
        (re.compile(r'"Fo\b'),                                   '°F'),   # "Fo → °F
        (re.compile(r'\'F\b'),                                   '°F'),   # 'F  → °F
        (re.compile(r'┬░'),                                      '°'),    # mojibake °
    ]
    for pat, repl in _ocr_unit_fixes:
        raw = pat.sub(repl, raw)

    # Remove stray OCR mojibake sequences
    raw = re.sub(r'\xc2\xb0|\u00c2\u00b0|ΓÇ[\x00-\xff]', '', raw)
    raw = raw.strip()

    # ── Dimension value normaliser ────────────────────────────────────────
    # When OCR merges a dimension value with the next label/annotation the
    # result is e.g. "5.0 M iN" or "15.0 M LONG".  Strip trailing alphabetic
    # junk that is not a recognised pressure/temperature/flow unit.
    # Triggered only when the value starts with a digit (not e.g. "CS + LINING").
    if raw and raw[0].isdigit():
        _dim_m = re.match(
            r'^(-?\d+(?:\.\d+)?)\s*(mm|M|NB|DN)\b',
            raw, re.IGNORECASE,
        )
        if _dim_m:
            _remainder = raw[_dim_m.end():].strip()
            # Keep remainder only if it looks like a valid unit/modifier (psig, °F, etc.)
            _is_valid_suffix = bool(re.match(
                r'^(?:psig|psia|psi|barg|bara|kpag|mpa|°[fc]|/[hH]|m3|kw|hp|mw)',
                _remainder, re.IGNORECASE,
            ))
            if _remainder and not _is_valid_suffix:
                raw = f'{_dim_m.group(1)} {_dim_m.group(2).upper()}'.strip()

    # ── Unit boundary truncation ──────────────────────────────────────────
    # When OCR merges junk text after a recognised engineering unit
    # (e.g. "1412 psig Lezi0 Re21084"), truncate at the end of the unit.
    # Cannot use lookbehind (variable length) — use re.match with capture.
    _unit_trunc_m = re.match(
        r'^(-?[\d][\d,.\s]*'
        r'(?:psig|psia|psi|barg|bara|kpag|mpa|'
        r'ft3/h(?:r)?|m3/h(?:r)?|kw|mw|gpm|bbl/d|'
        r'°\s*[fc]|deg[fc]|degf|degc|f\b|c\b))',
        raw, re.IGNORECASE,
    )
    if _unit_trunc_m and len(_unit_trunc_m.group(1)) < len(raw):
        _after_unit = raw[len(_unit_trunc_m.group(1)):].strip()
        # Don't truncate if followed by '/' — that may be a MIN/MAX pair separator
        # (e.g. "-13.2 F / 185 F" or "1219 psig / 1412 psig").
        # Use lstrip() since OCR often leaves a space before the slash.
        if not _after_unit.lstrip().startswith('/'):
            raw = _unit_trunc_m.group(1).strip()

    # ── MOC trailing noise cleanup ─────────────────────────────────────────
    # Strip OCR junk that follows a known material spec, e.g.
    # "CS + LINING ix 3.K" → "CS + LINING".
    _moc_m = re.match(
        r'^((?:CS|SS|316L?|304L?|317L|DSS|SDSS|DUPLEX|A516|INCONEL|MONEL|'
        r'HASTELLOY|GRE|FRP|HDPE|PVC|CARBON\s*STEEL|STAINLESS|ALLOY\s*STEEL)'
        r'(?:\s*[+/&]\s*(?:LINING|CLAD(?:DING)?|LINED?\b[^,;\n]{0,20}?|'
        r'RUBBER|EPOXY|FRP|GRE|HDPE|NEOPRENE))?)',
        raw, re.IGNORECASE,
    )
    if _moc_m and len(_moc_m.group(1)) < len(raw):
        _after_moc = raw[_moc_m.end():].strip()
        # Only truncate if the remainder doesn't look like a meaningful continuation
        if not re.match(r'^(?:ASTM|AISI|ISO|EN|NACE|\+|-)', _after_moc, re.IGNORECASE):
            raw = _moc_m.group(1).strip()

    # ── Full-vacuum code truncation ──────────────────────────────────────
    # 'FV' = full vacuum; any text following it is OCR noise from adjacent
    # annotation (e.g. 'FV SUG CACHER V_803-TE').  Truncate to bare 'FV'.
    if re.match(r'^FV\b', raw, re.IGNORECASE) and len(raw) > 2:
        raw = 'FV'

    return raw


def _build_databox_index(text: str, config: dict) -> dict:
    """
    Global equipment data-box scanner for P&ID drawings.

    Scans the full OCR text for known engineering labels
    (e.g. "OPERATING PRESS.", "NOMINAL CAPACITY :") and associates each
    extracted value with the nearest equipment tag found within
    _DATABOX_TAG_LOOKBACK_CHARS before the label position.

    Returns {TAG: {field_key: value_string}} merged later into per-tag
    results so narrow context-window extraction does not miss data-box
    values that appear far from the tag in OCR order.

    Label to field mappings are soft-coded in equipment_type_config.json
    under 'databox_label_map'.  Scan window is 'databox_scan_window_chars'
    in the 'extraction' section.  A list value in the map indicates a
    MIN/MAX split: the raw value is split on '/' and assigned to
    [0]=first field, [1]=second field.
    """
    ext_cfg   = config.get('extraction', {})
    db_window = int(ext_cfg.get('databox_scan_window_chars', _DATABOX_TAG_LOOKBACK_CHARS))
    label_map = config.get('databox_label_map', {})
    if not label_map:
        return {}

    _DB_TAG_RE = re.compile(
        r'\b([A-Z]{1,3}-[0-9]{2,5}[A-Za-z]?(?:-[A-Z0-9]{1,4})?)\b',
        re.IGNORECASE,
    )

    index: dict = {}

    # Sort longest label first so specific variants (e.g. "design press (min/max)")
    # are matched before shorter ones ("design press") — first-match-wins.
    sorted_variants = sorted(label_map.items(), key=lambda kv: -len(kv[0]))

    for label_variant, field_or_pair in sorted_variants:
        words = label_variant.strip().upper().split()
        if not words:
            continue
        # Allow 0-5 punctuation/space chars between words
        pat_str = r'[\s./()\-]{0,5}'.join(re.escape(w) for w in words)
        # Capture up to 80 chars after the colon (reduced from 120 to limit
        # runaway captures); stop at newline or semicolon.
        # _clean_databox_value() further truncates at noise patterns.
        pat_str = pat_str + r'[^:\n]{0,35}:\s*([^;\n]{1,80})'
        try:
            pat = re.compile(pat_str, re.IGNORECASE)
        except re.error:
            continue

        for m in pat.finditer(text):
            raw_value = _clean_databox_value(m.group(1))
            if not raw_value:
                continue

            win_start   = max(0, m.start() - db_window)
            pre_text    = text[win_start:m.start()]
            tag_matches = list(_DB_TAG_RE.finditer(pre_text))
            if not tag_matches:
                # Wider post-window: 400 chars covers data-box inline formats
                post_text   = text[m.end():min(len(text), m.end() + 400)]
                tag_matches = list(_DB_TAG_RE.finditer(post_text))
            if not tag_matches:
                continue

            tag = tag_matches[-1].group(1).upper()
            if tag not in index:
                index[tag] = {}

            if isinstance(field_or_pair, list) and len(field_or_pair) == 2:
                parts = re.split(r'\s*/\s*', raw_value, maxsplit=1)
                if len(parts) == 2:
                    v0 = _clean_databox_value(parts[0])
                    v1 = _clean_databox_value(parts[1])
                    # ── Smart numerical min/max assignment ──────────────────
                    # When both target fields end with _min / _max, use numeric
                    # ordering rather than position so that a box labelled
                    # (MAX/MIN) with "185 F / -13.2 F" and one labelled (MIN/MAX)
                    # with "-13.2 F / 185 F" both map correctly regardless of
                    # the order in which the values appear in the cell.
                    _f0, _f1 = field_or_pair[0], field_or_pair[1]
                    _is_minmax = (
                        (_f0.endswith('_min') or _f0.endswith('_max')) and
                        (_f1.endswith('_min') or _f1.endswith('_max'))
                    )
                    if _is_minmax and v0 and v1:
                        _n0 = re.search(r'-?\d+(?:\.\d+)?', v0)
                        _n1 = re.search(r'-?\d+(?:\.\d+)?', v1)
                        if _n0 and _n1:
                            _flt0, _flt1 = float(_n0.group()), float(_n1.group())
                            _min_f = _f0 if _f0.endswith('_min') else _f1
                            _max_f = _f0 if _f0.endswith('_max') else _f1
                            _min_v = v0 if _flt0 <= _flt1 else v1
                            _max_v = v0 if _flt0 >= _flt1 else v1
                            if _min_f not in index[tag]:
                                index[tag][_min_f] = _min_v
                            if _max_f not in index[tag]:
                                index[tag][_max_f] = _max_v
                        else:
                            if v0 and _f0 not in index[tag]:
                                index[tag][_f0] = v0
                            if v1 and _f1 not in index[tag]:
                                index[tag][_f1] = v1
                    else:
                        if v0 and field_or_pair[0] not in index[tag]:
                            index[tag][field_or_pair[0]] = v0
                        if v1 and field_or_pair[1] not in index[tag]:
                            index[tag][field_or_pair[1]] = v1
                else:
                    if field_or_pair[0] not in index[tag]:
                        index[tag][field_or_pair[0]] = raw_value
            else:
                if field_or_pair not in index[tag]:
                    index[tag][field_or_pair] = raw_value

    if index:
        print(
            f'[EQ-DIAG][DataBox] Indexed {len(index)} equipment: '
            + str({k: list(v.keys()) for k, v in index.items()}),
            flush=True,
        )
    return index


def _extract_equipment_items(text: str, drawing_ref: str, config: dict) -> list:
    """
    All field extraction patterns are soft-coded in equipment_type_config.json.
    Add / adjust patterns there without touching this function.

    Fields returned per item
    ------------------------
    tag, type_label, description, area, drawing_ref,
    line_connections, nozzle_connections, service_fluid,
    material_class, process_notes
    """
    ext_cfg     = config.get('extraction', {})
    type_labels = config.get('type_labels', {})
    fluid_kws   = [kw for kw in config.get('fluid_keywords', []) if not kw.startswith('_')]
    ctx_win                 = int(ext_cfg.get('context_window_chars', 160))
    desc_words              = int(ext_cfg.get('description_max_words', 6))
    desc_ctx_chars          = int(ext_cfg.get('description_context_chars', 400))
    desc_min_len            = int(ext_cfg.get('description_min_word_length', 3))
    area_ctx_chars          = int(ext_cfg.get('area_context_chars', 600))
    area_from_tag_heuristic = bool(ext_cfg.get('area_from_tag_heuristic', True))
    nozzle_ctx_chars        = int(ext_cfg.get('nozzle_context_chars', 400))
    mat_ctx_chars           = int(ext_cfg.get('material_context_chars', 400))
    service_ctx_chars       = int(ext_cfg.get('service_context_chars', 400))
    note_ctx_chars          = int(ext_cfg.get('note_context_chars', 400))
    # Standards refs, conjunctions and short noise tokens to exclude from description
    _desc_stop_words        = {
        'API','ASME','ANSI','ISO','DIN','NACE','NOTE','REF','SEE','PER',
        'AND','FOR','THE','OR','TO','OF','IN','AT','BY','NO','AS','IS','ON',
        # Drawing / title block words that appear near tags on P&IDs
        'LOCATION','MUBARRAZ','ISLAND','SCALE','NTS','DATE','DESCRIPTION',
        'REFERENCE','DOCUMENTS','DRAWINGS','DOCUMENT','DRAWING','TITLE',
        'COMPANY','PROJECT','SHEET','SIZE','ENGINEERING','CONSULTANT',
        # Company names (may appear in nearby company block)
        'REJLERS','DORSCH','HOLDING','GMBH','ABU','DHABI','UAE','HAMDAN',
        # Revision table words (near tags in OCR order)
        'ISSUED','APPROVED','REVIEW','HAZOP','CONSTRUCTION','INCORPORATED',
        'RETURNED','REAPPROVED','INFORMATION','COMMENTS',
        # P&ID noise
        'ALARM','TRIP','OPEN','HALF','SLOPE','SOUR','FLARE','HEADER',
        'WELL','FLUID','PHASE','LINE','TYPE','NOTE','NOTES','SCOPE',
    }

    # Soft-coded via tag_pattern in equipment_type_config.json.
    # The optional (?:-[A-Za-z0-9]{1,4})? captures project train/unit suffixes such
    # as -TF, -1F, -2A that are common in O&G tag numbering (e.g. V-308-TF,
    # V-805-1F).  Without this suffix, duplicate-deduplication collapses
    # equipment with the same base number but different trains into one row.
    # tag_pattern_ignorecase (default True): compile with IGNORECASE so OCR
    # lowercase variants like C-010c-TF are found; tag is always uppercased.
    _tag_pat_default = r'\b([A-Za-z]{1,2})-([0-9]{3,5}[A-Za-z]?(?:-[A-Za-z0-9]{1,4})?)\b'
    _tag_ic = bool(ext_cfg.get('tag_pattern_ignorecase', True))
    tag_re = re.compile(
        ext_cfg.get('tag_pattern', _tag_pat_default),
        re.IGNORECASE if _tag_ic else 0,
    )

    # --- Soft-coded helper patterns (read once per call) ------------------
    # Used by description strategy 1: identify bare tag lines and pure-noise tokens.
    # Must also match the extended suffix form so lines like "V-308-TF" are
    # not misidentified as description text.  IGNORECASE covers OCR lowercase.
    _tag_like_re  = re.compile(r'^[A-Za-z]{1,2}-\d{3,5}[A-Za-z]?(?:-[A-Za-z0-9]{1,4})?$', re.IGNORECASE)
    _noise_tok_re = re.compile(r'^[\d\.\+\-\/\%\(\)\[\]]{1,6}$')

    # Soft-coded reject patterns for description lines — lines matching any of
    # these are skipped entirely rather than partially filtered.
    # Covers: pipe designations (20"-PL-...), document/drawing numbers
    # (PJ6-EXD-...-0023), instrument/valve tags (FT-1234), fraction-inch
    # size tokens (3/4"), grid refs (A1-H8), and OCR fragments (|[, =£).
    _desc_line_reject_re = re.compile(
        r'(?:'
        r'\d+["\']-[A-Z]{2,4}-'          # pipe designation: 20"-PL-...
        r'|[A-Z]{2,4}-[A-Z]{2,4}-[A-Z]{2,4}-[A-Z]{2,4}-\d{4}'  # doc number
        r'|PJ\d[-_][A-Z]'                # project document prefix PJ6-...
        r'|\b[A-Z]{2,3}-\d{4,6}\b'       # instrument/valve tags FT-1234
        r'|\d+\s*/\s*\d+'                # fractions 3/4
        r'|[|=\[\]£$@#]'                 # OCR junk characters
        r'|^\d{1,3}["\']?\s*[-]\s*[A-Z]{2,4}'  # starts with size then tag type
        r'|\bFROM\s+[A-Z]|\bTO\s+[A-Z]|\bVIA\s+[A-Z]'  # flow routing text
        r'|\bLINE\s*\d|\bNOTE\s*\d|\bSHEET\s*\d'  # line/note/sheet refs
        r'|\bSLOPE\s*1?\s*[:.]'          # SLOPE 1:100 annotations
        r'|\bNTS\b|\bSCALE\b'            # scale annotations
        r')',
        re.IGNORECASE,
    )
    area_re    = re.compile(
        ext_cfg.get('area_pattern',
                    r'(?:AREA|UNIT|TRAIN|BAY|SECTION|BATTERY|MODULE|MOD|ZONE|BLOCK|SKID|PLANT|FIELD|STREAM)\s*[:\-]?\s*([A-Z0-9]{1,8})'),
        re.IGNORECASE,
    )
    nozzle_re         = re.compile(
        ext_cfg.get('nozzle_pattern', r'\bN[-]?[0-9]{1,2}[A-Z]?\b')
    )
    mat_re            = re.compile(
        ext_cfg.get('material_class_pattern',
                    r'\b(A1[A-Z]R?|B1[A-Z]|C1[A-Z]|D1[A-Z]|[A-D]2[A-Z]'
                    r'|CS|SS|316L?|304L?|317L|321|347|2205|254SMO'
                    r'|DSS|SDSS|DUPLEX|INCONEL|HASTELLOY|MONEL'
                    r'|GRE|FRP|HDPE|CPVC|PVC|PVDF|A516|A240|A312|A106)\b'),
        re.IGNORECASE,
    )
    material_label_re = re.compile(
        ext_cfg.get('material_label_pattern',
                    r'(?:MATERIAL|MTL|SHELL|BODY|CASING|LINER'
                    r'|WETTED\s*PARTS?|INTERNALS?)'
                    r'\s*[:\-/]\s*([A-Z0-9][A-Z0-9/\-\s\.]{1,28})'),
        re.IGNORECASE,
    )
    service_label_re  = re.compile(
        ext_cfg.get('service_label_pattern',
                    r'(?:SERVICE|FLUID|MEDIUM|PROCESS\s*FLUID'
                    r'|CONTENTS|PRODUCT|DUTY)'
                    r'\s*[:\.\.\-]\s*([A-Za-z][A-Za-z0-9\s/\-]{1,30})'),
        re.IGNORECASE,
    )
    note_re           = re.compile(
        ext_cfg.get('note_pattern',
                    r'(?:(?:SEE\s+)?NOTE\s*[-\s\(]?[0-9]+[\)\.]*'
                    r'|\bHOLD\b(?:\s*[-]?\s*[0-9]+)?'
                    r'|\bTBD\b|\bTBC\b'
                    r'|\bREF[.\s]+DWG[.\s]+[A-Z0-9/\-]+'
                    r'|SEE\s+(?:DWG|SPEC|DOC)[.]*\s*[A-Z0-9/\-]+)'),
        re.IGNORECASE,
    )
    # -----------------------------------------------------------------------

    instr_valve_prefixes = {
        'FT','FI','FIC','FC','PT','PI','PIC','PC','LT','LI','LIC','LC',
        'TT','TI','TIC','TC','AT','AI','FY','PY','LY','TY',
        'HV','FV','XV','PV','SDV','BDV','PSV','PRV','CV','LV','TV',
        'FE','TE','LE','PE','HS','HIC','HI',
    }

    # Soft-coded: exact tag suffix values that identify non-equipment tokens
    # (e.g. project change-request numbers like PJ-2025-CR-002 which OCR
    # garbles into P-2028-CR — "CR" is a document type, not an equipment suffix).
    _exclude_suffixes = {s.upper() for s in ext_cfg.get('exclude_tag_suffixes', ['CR', 'NCR', 'WO', 'TQ', 'MDR', 'MOM', 'MR'])}

    seen = set()
    results = []

    # ── Global data-box index (built once, merged per-tag below) ─────────
    # Scans the full OCR text for LABEL : VALUE pairs in equipment data boxes
    # so narrow context-window extraction doesn't miss values that are spatially
    # far from the tag in OCR text order (common on large-format P&IDs).
    _databox_idx = _build_databox_index(text, config)

    # ── Slash-variant tag expansion ────────────────────────────────────────
    # OCR on P&IDs sometimes reads multi-unit tags like "P-851A/B/C-TF" as a
    # single token. Expand these into individual variants (P-851A-TF,
    # P-851B-TF, P-851C-TF) and append them to the text so the main loop
    # finds each unit independently.
    #
    # Soft-coded: slash_ocr_substitutions — OCR chars that should be treated
    # as '/' in this context (e.g. '?' → '/' when OCR misreads the slash in
    # P-851A/B/C-TF as P-851A?B/C-TF). Applied locally to a copy of the text
    # so the substitution only affects slash-expansion; the main text is left
    # intact to avoid corrupting description or parameter extraction.
    _slash_ocr_subs = ext_cfg.get('slash_ocr_substitutions', [['?', '/']])
    _slash_text = text
    for _bad, _good in _slash_ocr_subs:
        # Only substitute inside plausible tag tokens: letter-digit?letter pattern
        _slash_text = re.sub(
            r'(?<=[A-Za-z])' + re.escape(_bad) + r'(?=[A-Za-z])',
            _good, _slash_text
        )

    # ── Whitespace-tolerant variant chain normalization (soft-coded) ──────
    # Title blocks routinely render multi-unit tags with a space between the
    # serial number and the first variant letter, e.g. "P-851 A/B/C-TF".
    # The slash-expansion regex below requires the variant letter to sit
    # immediately after the base, so without this step the title token is
    # never expanded and only the explicitly-tagged variants on the drawing
    # body (often missing one of the three) make it into the result list.
    # This step collapses ONE optional whitespace gap, ONLY when the next
    # character pair is "<letter>/<letter>" — keeping the normalization
    # tightly scoped so unrelated text is never altered.
    # Soft-coded by 'slash_collapse_space_before_variant' (default true).
    if bool(ext_cfg.get('slash_collapse_space_before_variant', True)):
        _slash_text = re.sub(
            r'\b([A-Za-z]{1,2}-\d{3,5})\s+([A-Za-z])(?=\s*/\s*[A-Za-z])',
            r'\1\2',
            _slash_text,
        )

    _slash_re = re.compile(
        r'\b([A-Za-z]{1,2}-\d{3,5})([A-Za-z])/([A-Za-z])(?:/([A-Za-z]))?(?:-([A-Za-z0-9]{1,4}))?\b',
        re.IGNORECASE,
    )
    _slash_expanded: list[str] = []
    for _sm in _slash_re.finditer(_slash_text):
        _base = _sm.group(1).upper()
        _sfx  = _sm.group(5).upper() if _sm.group(5) else ''
        for _v in [_sm.group(2).upper(), _sm.group(3).upper()] + ([_sm.group(4).upper()] if _sm.group(4) else []):
            _slash_expanded.append(f'{_base}{_v}' + (f'-{_sfx}' if _sfx else ''))
    if _slash_expanded:
        text = text + '\n' + '\n'.join(_slash_expanded)
        print(f'[EQ-DIAG] Slash expansion added: {_slash_expanded}', flush=True)

    # ── Propagate databox index to slash-expanded sibling tags ────────────
    # When a data box is labelled "P-851A/B/C-TF", the databox scanner indexes
    # only the first letter variant (P-851A) since that is the last tag token
    # it sees before the label in OCR text order.  The slash expansion above
    # creates P-851B / P-851C as synthetic items so each unit gets its own row
    # in the equipment register.  Without propagation Gate2-TierA would drop
    # P-851B and P-851C because they are absent from _databox_idx.
    # Soft-coded: controlled by 'databox_slash_sibling_propagate' (default true).
    if bool(ext_cfg.get('databox_slash_sibling_propagate', True)) and _slash_expanded and _databox_idx:
        _sibling_re = re.compile(r'^([A-Z]+-\d+)[A-Z](?:-[A-Z0-9]+)?$', re.IGNORECASE)
        for _exp_tag in _slash_expanded:
            _exp_upper = _exp_tag.upper()
            if _exp_upper in _databox_idx:
                continue  # already has its own entry
            _sm2 = _sibling_re.match(_exp_upper)
            if not _sm2:
                continue
            _base_prefix = _sm2.group(1).upper()
            # Find a sibling that IS in the index with the same base prefix+number
            for _idx_tag, _idx_data in _databox_idx.items():
                if _idx_tag.upper().startswith(_base_prefix) and _idx_tag.upper() != _exp_upper:
                    _databox_idx[_exp_upper] = dict(_idx_data)  # shallow copy
                    break

    for m in tag_re.finditer(text):
        prefix = m.group(1).upper()
        tag    = m.group(0).upper()   # always uppercase — OCR may emit lowercase letters

        if prefix in instr_valve_prefixes:
            continue
        if type_labels and prefix not in type_labels:
            continue
        # Filter non-equipment project-reference suffixes
        _tag_suffix_m = re.search(r'-([A-Z]{1,4})$', tag)
        if _tag_suffix_m and _tag_suffix_m.group(1) in _exclude_suffixes:
            continue

        # ── Gate 0: tag-context annotation reject ────────────────────
        # Soft-coded via 'tag_context_reject_patterns' in equipment_type_config.json.
        # If ANY pattern matches within 80 chars of the tag, the occurrence is
        # skipped WITHOUT adding to 'seen' (so a legitimate data-box occurrence
        # of the same tag later in the OCR stream is still extracted).
        # Use-case: revision-cloud bubbles annotated directly on P&IDs contain
        # tag text that is NOT equipment on this drawing.
        _ctx_reject_pats = ext_cfg.get('tag_context_reject_patterns', [])
        if _ctx_reject_pats:
            _surrounding = text[max(0, m.start() - 80): min(len(text), m.end() + 80)]
            if any(re.search(_crpat, _surrounding) for _crpat in _ctx_reject_pats):
                print(f'[EQ-DIAG] Gate0-ctx-reject skipped: {tag!r}', flush=True)
                continue

        # ── Gate 1: connector-arrow context check ────────────────────────
        # On ADNOC/O&G P&IDs the sheet-edge continuation arrows use the
        # format: "[KEYWORD] [DESCRIPTION] [TAG]\n[DWG-NO]" (e.g.
        # "SOUR GAS TO MEA INLET SCRUBBER V-804-TF\nPJ6-EXD-MRI-BQDA-0024").
        # These are cross-sheet references to equipment that lives on a
        # different drawing and has no data box here.
        # Detection: a flow-routing keyword within lookback_chars BEFORE the
        # tag AND a multi-segment drawing number within lookahead_chars AFTER.
        # When detected, skip this occurrence WITHOUT adding to 'seen' so
        # the same tag can still be processed if it appears in its own data
        # box later in the OCR text of this very drawing.
        # Soft-coded via connector_context_enabled / connector_context_lookback_chars
        # / connector_context_lookahead_chars / connector_keywords_pattern in
        # equipment_type_config.json extraction section.
        if bool(ext_cfg.get('connector_context_enabled', True)):
            _conn_lookback  = int(ext_cfg.get('connector_context_lookback_chars', 80))
            _conn_lookahead = int(ext_cfg.get('connector_context_lookahead_chars', 120))
            _conn_kws_pat   = ext_cfg.get(
                'connector_keywords_pattern',
                r'\b(?:FROM|TO|VIA|INTO|INLET|OUTLET|SUCTION|DISCHARGE|DEST(?:INATION)?|SOURCE)\b',
            )
            _conn_before = text[max(0, m.start() - _conn_lookback): m.start()]
            _conn_after  = text[m.end(): min(len(text), m.end() + _conn_lookahead)]
            if (re.search(_conn_kws_pat, _conn_before, re.IGNORECASE)
                    and _TITLEBLOCK_DWG_NO_RE.search(_conn_after)):
                print(
                    f'[EQ-DIAG] Gate1-connector-ref skipped: {tag!r} '
                    f'(keyword in lookback + dwg-no in lookahead)',
                    flush=True,
                )
                # Do NOT add to seen — allow a data-box occurrence of the
                # same tag later in the text to be processed normally.
                continue

        if tag in seen:
            continue
        seen.add(tag)

        start = max(0, m.start() - ctx_win)
        end   = min(len(text), m.end() + ctx_win)
        ctx   = text[start:end]

        type_label = type_labels.get(prefix, 'Equipment')

        # ── Description — multi-strategy extraction ───────────────────────
        after       = text[m.end(): m.end() + desc_ctx_chars]
        description = ''

        # Strategy 0: data-box title — on P&IDs the equipment data box has the
        # format: TAG_LINE\nDESCRIPTION_LINE (e.g. "V-803-TF\nMRD OIL SLUG CATCHER").
        # Look for 2-6 consecutive ALL-CAPS words on the first non-blank line
        # after the tag that is NOT a pipe/tag/doc reference.
        # Soft-coded: description_databox_min_words, description_databox_max_words
        _db_min_w = int(ext_cfg.get('description_databox_min_words', 2))
        _db_max_w = int(ext_cfg.get('description_databox_max_words', 6))
        for _dln in (ln.strip() for ln in after.split('\n') if ln.strip()):
            if _tag_like_re.match(_dln) or _desc_line_reject_re.search(_dln):
                continue
            _dln_toks = _dln.split()
            # Require at least _db_min_w tokens, all purely alphabetic (or
            # common hyphenated words like "THREE-PHASE"), no digits
            _alpha_toks = [
                t.strip('.,;:/()"\'[]')
                for t in _dln_toks
                if re.match(r'^[A-Za-z][A-Za-z\-]{1,}$', t.strip('.,;:/()"\'[]'))
                and len(t) >= desc_min_len
                and t.upper() not in _desc_stop_words
            ]
            if len(_alpha_toks) >= _db_min_w:
                description = ' '.join(_alpha_toks[:_db_max_w]).title()
                break

        # Strategy 1: newline-segmented lines right after the tag.
        # Each line is checked for "description-likeness":
        # skip bare tag IDs, pipe designations and pure digit/symbol noise.
        if not description:
            desc_lines = []
            for _ln in (ln.strip() for ln in after.split('\n') if ln.strip()):
                if _tag_like_re.match(_ln):
                    continue
                if _desc_line_reject_re.search(_ln):
                    continue
                _toks = [t.strip('.,;:/()"\'[]') for t in _ln.split()]
                _valid = [
                    t for t in _toks
                    if len(t) >= desc_min_len
                    and not t.isdigit()
                    and not _tag_like_re.match(t)
                    and not _noise_tok_re.match(t)
                    and t.upper() not in _desc_stop_words
                    and not re.search(r'\d{2,}', t)       # skip tokens with 2+ digits
                    and not re.match(r'^[A-Z]{1,3}-\d', t)  # skip tag-like tokens
                ]
                if _valid:
                    desc_lines.append(' '.join(_valid[:5]))
                if len(desc_lines) >= 1:
                    break
            if desc_lines:
                description = ' '.join(desc_lines).title()

        # Strategy 2: ALL-CAPS word scan in narrower ctx_win (improved filter)
        if not description:
            _cap_words = re.findall(r'\b[A-Z][A-Z]{2,19}\b', after[:ctx_win])
            _filtered_caps = [
                w for w in _cap_words
                if not re.match(r'^[A-Z]{1,2}-\d', w)
                and w not in _desc_stop_words
                and len(w) >= desc_min_len
            ][:desc_words]
            if _filtered_caps:
                description = ' '.join(w.capitalize() for w in _filtered_caps[:3])

        # Strategy 3: fall back to the equipment TypeLabel
        if not description:
            description = type_label

        # ── Line connections (piping designation tokens) ───────────────────
        lc_tokens = []
        for lm in _LINE_TAG_RE.finditer(ctx):
            token = lm.group(0).strip()
            if token and token not in lc_tokens:
                lc_tokens.append(token)

        # ── Service / fluid — multi-strategy extraction ───────────────────
        _svc_start    = max(0, m.start() - service_ctx_chars)
        _svc_end      = min(len(text), m.end() + service_ctx_chars)
        _svc_ctx      = text[_svc_start:_svc_end]
        service_fluid = ''
        # Strategy 1: label-based — SERVICE: CRUDE OIL, FLUID: NITROGEN, MEDIUM: GAS
        _svc_lm = service_label_re.search(_svc_ctx)
        if _svc_lm:
            _raw_svc = _svc_lm.group(1).split('\n')[0].strip().rstrip('.,;')
            if len(_raw_svc) >= 2:
                service_fluid = _raw_svc[:35].title()
        # Strategy 2: keyword scan in wider context
        if not service_fluid:
            _svc_lower = _svc_ctx.lower()
            found_fluids = [kw for kw in fluid_kws if kw in _svc_lower]
            service_fluid = ', '.join(found_fluids[:2]).title() if found_fluids else ''
        # Strategy 3: derive from fluid code embedded in already-found line connection tags.
        # e.g. "4"-HO-5665-033842-X" → fluid code "HO" → "Hydrocarbon Oil".
        # Guaranteed to find something whenever line connections were extracted.
        if not service_fluid and lc_tokens:
            _lf_map = {k: v for k, v in config.get('line_fluid_code_map', {}).items()
                       if not str(k).startswith('_')}
            for _lc in lc_tokens:
                _fc_m = re.match(r'^[\d½¾¼]+\s*["\'?]\s*[-_]\s*([A-Z]{1,4})\s*[-_]', _lc)
                if _fc_m:
                    _fc = _fc_m.group(1).upper()
                    _mapped = _lf_map.get(_fc)
                    if _mapped:
                        service_fluid = _mapped
                        break

        # ── Area / Unit — multi-strategy extraction ───────────────────────
        # Strategy 1: search a wider context (soft-coded area_context_chars).
        # Uses capture group(1) — returns just the code, not the whole keyword match.
        _a_start = max(0, m.start() - area_ctx_chars)
        _a_end   = min(len(text), m.end() + area_ctx_chars)
        area_m   = area_re.search(text[_a_start:_a_end])
        area     = area_m.group(1).strip() if area_m else ''

        # Strategy 2: derive from serial number digits (O&G tag-number convention).
        # V-101 → "100", P-2201 → "2200", E-10001 → "10000"
        if not area and area_from_tag_heuristic:
            _digits = re.sub(r'[^0-9]', '', m.group(2))
            if len(_digits) >= 3:
                area = _digits[0] + '0' * (len(_digits) - 1)

        # ── Nozzle connections — multi-strategy extraction ────────────────
        _nzl_start    = max(0, m.start() - nozzle_ctx_chars)
        _nzl_end      = min(len(text), m.end() + nozzle_ctx_chars)
        _nzl_ctx      = text[_nzl_start:_nzl_end]
        # Strategy 1: N1 / N-1 / N2A nozzle tag pattern
        nozzle_tokens = list(dict.fromkeys(nozzle_re.findall(_nzl_ctx)))
        # Strategy 2: size-prefixed nozzle labels  e.g. 4"-N1, 6"-N2A
        for _snm in re.finditer(
            r'\b\d{1,3}\s*["\']\s*-?\s*(N[-]?[0-9]{1,2}[A-Z]?)\b',
            _nzl_ctx, re.IGNORECASE
        ):
            _tok = _snm.group(1).upper()
            if _tok not in nozzle_tokens:
                nozzle_tokens.append(_tok)
        # Strategy 3: functional orientation labels as fallback when no N-tags found
        # e.g. INLET, OUTLET, SUCTION, DISCHARGE on equipment bubbles
        if not nozzle_tokens:
            _orient_hits = re.findall(
                r'\b(INLET|OUTLET|SUCT(?:ION)?|DISCH(?:ARGE)?|VENT|DRAIN|BYPASS'
                r'|OVERFLOW|RECYCLE|RETURN|FEED|PRODUCT|OVERHEAD|BOTTOM(?:S)?)\b',
                _nzl_ctx, re.IGNORECASE,
            )
            for _ow in dict.fromkeys(w.capitalize() for w in _orient_hits):
                nozzle_tokens.append(_ow)
        nozzle_tokens = nozzle_tokens[:8]

        # ── Material / piping spec — multi-strategy extraction ────────────
        _mat_start     = max(0, m.start() - mat_ctx_chars)
        _mat_end       = min(len(text), m.end() + mat_ctx_chars)
        _mat_ctx       = text[_mat_start:_mat_end]
        material_class = ''
        # Strategy 1: label-based — MATERIAL: CS/SS316, SHELL: DSS, MTL: INCONEL
        _mat_lm = material_label_re.search(_mat_ctx)
        if _mat_lm:
            _raw_mat = _mat_lm.group(1).split('\n')[0].strip().rstrip('.,;/ ')
            if len(_raw_mat) >= 2:
                material_class = _raw_mat[:25].upper()
        # Strategy 2: pattern scan in wider context
        if not material_class:
            mat_matches    = mat_re.findall(_mat_ctx)
            material_class = mat_matches[0].upper() if mat_matches else ''
        # Strategy 3: derive material hint from pipe-class prefix in line connection tags.
        # Line tag format: SIZE"-FLUID-SEQ-PIPECLASS[-SUFFIX]
        # First 2 digits of the pipe-class code encode material in most spec books
        # (soft-coded in pipe_class_prefix_map — add project-specific mappings there).
        if not material_class and lc_tokens:
            _pc_prefix_map = {k: v for k, v in config.get('pipe_class_prefix_map', {}).items()
                              if not str(k).startswith('_')}
            for _lc in lc_tokens:
                # Pipe class is typically 5-8 digits separated by '-' or '_'
                _pc_m = re.search(r'[\-_](\d{4,8})(?:[\-_][A-Z0-9]{0,5})?(?:\s|$)', _lc)
                if _pc_m:
                    _prefix = _pc_m.group(1)[:2]
                    _mat = _pc_prefix_map.get(_prefix)
                    if _mat:
                        material_class = _mat
                        break

        # ── Material OCR sanity filter (soft-coded) ───────────────────────
        # Reject MOC strings that contain mixed-case gibberish (OCR bleed from
        # annotations, e.g. "CS +CuNINoAN", "HALF OPEN PIPE") or invalid chars.
        # A valid MOC token is ALL UPPERCASE engineering code or contains only
        # uppercase + digits + + / - space. The config flag
        # 'material_reject_invalid_ocr' toggles the filter.
        if material_class and bool(ext_cfg.get('material_reject_invalid_ocr', True)):
            _mc_upper = material_class.upper().strip()
            # Count alpha chars that would be lowercase in the ORIGINAL (pre-upper)
            _had_lowercase = any(c.islower() for c in material_class)
            # Known-good token shape: 2-25 chars of [A-Z0-9 +/.-]
            _shape_ok = bool(re.fullmatch(r'[A-Z0-9 +/\.\-]{2,25}', _mc_upper))
            # Reject common OCR-garbage keywords that are NOT materials
            _garbage_kw = re.search(
                r'\b(?:HALF|OPEN|PIPE|CuNINoAN|LEVEL|ONLY|NOTE|SHALL|SEE)\b',
                material_class, re.IGNORECASE,
            )
            if (not _shape_ok) or _garbage_kw or (_had_lowercase and not _shape_ok):
                print(f'[EQ-DIAG][MOC] rejected OCR-garbage MOC for {tag}: {material_class!r}', flush=True)
                material_class = ''

        # ── Process note references — wider context scan ──────────────────
        _nt_start     = max(0, m.start() - note_ctx_chars)
        _nt_end       = min(len(text), m.end() + note_ctx_chars)
        _nt_ctx       = text[_nt_start:_nt_end]
        note_matches  = list(dict.fromkeys(
            n.strip() for n in note_re.findall(_nt_ctx)
        ))[:3]
        process_notes = ', '.join(note_matches) if note_matches else ''

        # ── Process parameter extraction (P&ID data-bubble / annotation) ──
        # Uses a narrower context window so values are specific to this tag.
        _pp_ctx_chars  = int(ext_cfg.get('process_param_context_chars', 300))
        _pp_start      = max(0, m.start() - _pp_ctx_chars)
        _pp_end        = min(len(text), m.end() + _pp_ctx_chars)
        _pp_ctx        = text[_pp_start:_pp_end].upper()

        # Soft-coded regex patterns (all read from config)
        _press_val_pat  = ext_cfg.get('pressure_value_pattern',
                           r'(-?\d+(?:\.\d+)?)\s*(PSIG|PSIA|PSI|barg|bara|kPag|MPa|bar)\b')
        _temp_val_pat   = ext_cfg.get('temperature_value_pattern',
                           r'(-?\d+(?:\.\d+)?)\s*(?:°\s*[FC]|DEG\.?\s*[FC]|DEGF|DEGC)')
        _flow_lbl_pat   = ext_cfg.get('flowrate_label_pattern',
                           r'(?:Q\s*[:=]|FLOW\s*RATE|FLOWRATE|CAPACITY|DESIGN\s*FLOW|DUTY)\s*[:=/(]?')
        _flow_val_pat   = ext_cfg.get('flowrate_value_pattern',
                           r'(\d+(?:[,.]\d+)?)\s*(M3/H|M3/HR|NM3/H|NM3/HR|SM3/D|MMSCFD|BBL/D|BBL/H|BPD|GPM|T/H|KG/H|KG/HR|MW|KW|MMBTU/H)')
        _flow_bare_pat  = ext_cfg.get('flowrate_bare_value_pattern',
                           r'(\d+(?:[,.]\d+)?)\s*(M3/H|M3/HR|NM3/H|NM3/HR|SM3/D|MMSCFD|BBL/D|GPM|T/H|KG/H|MW|KW|MMBTU/H)')
        _flow_ctx_chars = int(ext_cfg.get('flowrate_context_chars', 500))
        _op_press_lbl   = ext_cfg.get('oper_pressure_label_pattern',
                           r'(?:OPER(?:ATING)?|OP\.?)\s*PRESS(?:URE)?\.?\s*[-:=/(]')
        _des_press_lbl  = ext_cfg.get('design_pressure_label_pattern',
                           r'(?:DES(?:IGN)?\.?(?:\s*/\s*SET)?|SET)\s*PRESS(?:URE)?\.?\s*[-:=/(]')
        _op_temp_lbl    = ext_cfg.get('oper_temp_label_pattern',
                           r'(?:OPER(?:ATING)?|OP\.?)\s*TEMP(?:ERATURE)?\.?\s*[-:=/(]')
        _des_temp_lbl   = ext_cfg.get('design_temp_label_pattern',
                           r'DES(?:IGN)?\.?\s*TEMP(?:ERATURE)?\.?\s*[-:=/(]')
        # Soft-coded: slash-pair dual-value patterns for "LABEL (MIN/MAX) : V1/V2 UNIT"
        # format common on O&G P&ID data boxes (single cell stores both min + max).
        _dual_temp_pat  = ext_cfg.get('dual_value_temp_pattern',
                           r'(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)\s*(°\s*[FC]|DEG\.?\s*[FC]|DEGF|DEGC)?')
        _dual_press_pat = ext_cfg.get('dual_value_press_pattern',
                           r'(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)\s*(PSIG|PSIA|PSI|barg|bara|kPag|MPa|bar)\b')
        _ins_codes      = [c.upper() for c in ext_cfg.get('insulation_codes',
                           ['HOT', 'COLD', 'PERS', 'HT', 'CT', 'TRACED', 'EHT', 'BARE', 'ACOUSTIC'])]
        _dim_len_lbl    = ext_cfg.get('dimension_length_label_pattern',
                           r'(?:LENGTH|HEIGHT|TL[-/]TL|TAN[-/]TAN|LONG|T/T)\s*[:=]?')
        _dim_dia_lbl    = ext_cfg.get('dimension_diameter_label_pattern',
                           r'(?:DIA(?:METER)?|O\.?D\.?|BORE|I\.?D\.?|NB|DN)\s*[:=]?')
        _dim_val_pat    = ext_cfg.get('dimension_value_pattern', r'(\d+(?:\.\d+)?)\s*(mm|M)?')
        _mtr_pat        = ext_cfg.get('motor_rating_pattern',
                           r'(?:MOTOR|DRIVER|RATED\s*POWER|INSTALLED\s*POWER)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(kW|KW|HP|BHP|KVA)\b')
        _mtr_bare_pat   = ext_cfg.get('motor_rating_bare_pattern', r'(\d+(?:\.\d+)?)\s*(kW|KW)\b')
        _qual_pat       = ext_cfg.get('quality_required_pattern',
                           r'(?:QUALITY|QC|NDE|NDT|INSPECT(?:ION)?)\s*[:=]?\s*([A-D](?:\s+LEVEL)?|LEVEL\s*[A-D]|(?:100%\s*)?(?:RT|MT|UT|PT|VT)(?:[+&,/\s]+(?:RT|MT|UT|PT|VT))*)')

        # ── Oper. Pressure ────────────────────────────────────────────────
        oper_pressure = ''
        _op_lbl_m = re.search(_op_press_lbl, _pp_ctx, re.IGNORECASE)
        if _op_lbl_m:
            _after_lbl = _pp_ctx[_op_lbl_m.end():]
            _pv = re.search(_press_val_pat, _after_lbl[:80], re.IGNORECASE)
            if _pv:
                oper_pressure = f'{_pv.group(1)} {_pv.group(2)}'
            elif not _pv:
                # Fallback: label matched at "(MIN/MAX)" qualifier  →  after = "MIN/MAX) : 155 psig".
                # Skip over the qualifier and grab the first pressure value after ":".
                _after_colon = re.search(r'\)\s*[:,=]\s*(.{1,60})', _after_lbl[:80], re.IGNORECASE)
                if _after_colon:
                    _pv2 = re.search(_press_val_pat, _after_colon.group(1), re.IGNORECASE)
                    if _pv2:
                        oper_pressure = f'{_pv2.group(1)} {_pv2.group(2)}'

        # ── Design Pressure min / max ─────────────────────────────────────
        design_pressure_min = ''
        design_pressure_max = ''
        _dp_lbl_ms = list(re.finditer(_des_press_lbl, _pp_ctx, re.IGNORECASE))
        _dp_vals   = []   # list of (numeric_value, unit_string)
        for _dlm in _dp_lbl_ms:
            _win80 = _pp_ctx[_dlm.end():_dlm.end() + 80]
            # Strategy A: slash-pair with unit at end e.g. "195 / -13.2 psig"
            # (label matched at "(MIN/MAX)" qualifier → after = "MIN/MAX) : 195 / -13.2 psig")
            _dp_slash = re.search(_dual_press_pat, _win80, re.IGNORECASE)
            if _dp_slash:
                _pu = _dp_slash.group(3).upper()
                _dp_vals.append((float(_dp_slash.group(1)), _pu))
                _dp_vals.append((float(_dp_slash.group(2)), _pu))
            else:
                # Strategy B: standard "VALUE UNIT" or "VALUE UNIT / VALUE UNIT"
                for _pv in re.finditer(_press_val_pat, _win80, re.IGNORECASE):
                    _dp_vals.append((float(_pv.group(1)), _pv.group(2)))
            if not _dp_vals:
                # Strategy C: unit embedded in label parens e.g. "DES./SET PRESS (PSIG) : 195 / FV".
                # Locate (UNIT) just after the label match, collect numeric values that follow;
                # non-numeric tokens like FV (Full Vacuum) are silently skipped.
                _pu_m = re.search(
                    r'\(([A-Za-z]{2,5})\)\s*[:,=\s]+(-?\d+(?:\.\d+)?)', _win80, re.IGNORECASE
                )
                if _pu_m:
                    _pu = _pu_m.group(1).upper()
                    if _pu in ('PSIG', 'PSIA', 'PSI', 'BARG', 'BARA', 'BAR', 'KPAG', 'MPA'):
                        _dp_vals.append((float(_pu_m.group(2)), _pu))
                        # Also grab second value (MIN or MAX partner), ignoring non-numerics
                        _rest = _win80[_pu_m.end():]
                        _v2 = re.search(r'[/,\s]+(-?\d+(?:\.\d+)?)', _rest[:30])
                        if _v2:
                            try:
                                _dp_vals.append((float(_v2.group(1)), _pu))
                            except ValueError:
                                pass
        if _dp_vals:
            _dp_nums  = [v for v, _u in _dp_vals]
            _dp_units = _dp_vals[0][1]  # use unit from first match
            if len(_dp_nums) == 1:
                # Single value = maximum design pressure; MIN is unspecified
                design_pressure_max = f'{_dp_nums[0]} {_dp_units}'
            else:
                design_pressure_min = f'{min(_dp_nums)} {_dp_units}'
                design_pressure_max = f'{max(_dp_nums)} {_dp_units}'
        elif not oper_pressure:
            # Fallback: any bare pressure value in narrow context
            _bare_pv = re.search(_press_val_pat, _pp_ctx, re.IGNORECASE)
            if _bare_pv:
                design_pressure_max = f'{_bare_pv.group(1)} {_bare_pv.group(2)}'

        # ── Oper. Temperature ─────────────────────────────────────────────
        oper_temperature = ''
        _ot_lbl_m = re.search(_op_temp_lbl, _pp_ctx, re.IGNORECASE)
        if _ot_lbl_m:
            _ot_after = _pp_ctx[_ot_lbl_m.end():_ot_lbl_m.end() + 80]
            # Strategy 1: standard single value e.g. "175 °F"
            _tv = re.search(_temp_val_pat, _ot_after, re.IGNORECASE)
            if _tv:
                oper_temperature = f'{_tv.group(1)} °F'
            else:
                # Strategy 2: slash-pair dual value e.g. "105/60 °F" or "105 / 60 °F"
                # Matches "LABEL (MIN/MAX) : 105/60 °F" where unit is only at end.
                # The normalizer converts "105/60 °F" → "60 – 105 °F" (ascending range).
                _ot_dual = re.search(_dual_temp_pat, _ot_after, re.IGNORECASE)
                if _ot_dual and _ot_dual.group(1) != _ot_dual.group(2):
                    _raw_unit = (_ot_dual.group(3) or '°F').strip()
                    oper_temperature = f'{_ot_dual.group(1)}/{_ot_dual.group(2)} {_raw_unit}'
                else:
                    # Strategy 3: unit embedded in label parens e.g. "OPER TEMP (°F) : 175".
                    # The label matches at "(", leaving "°F) : 175" in the window.
                    # Detect °C vs °F from label context, then grab first bare number.
                    _ot_unit = 'C' if re.search(
                        r'\(°?\s*C\)', _pp_ctx[_ot_lbl_m.start():_ot_lbl_m.end() + 8], re.IGNORECASE
                    ) else 'F'
                    _bare_t = re.search(
                        r'[°FC)\s]+[:,=\s]+(-?\d+(?:\.\d+)?)', _ot_after, re.IGNORECASE
                    )
                    if _bare_t:
                        oper_temperature = f'{_bare_t.group(1)} °{_ot_unit}'
        # Normalise dual-temperature values e.g. "105/60 °F" → "60 – 105 °F"
        oper_temperature = _normalize_oper_temp(oper_temperature)

        # ── Design Temp min / max ─────────────────────────────────────────
        design_temp_min = ''
        design_temp_max = ''
        _dt_lbl_ms = list(re.finditer(_des_temp_lbl, _pp_ctx, re.IGNORECASE))
        _dt_vals   = []  # list of numeric values (°F)
        for _dtlm in _dt_lbl_ms:
            _win80 = _pp_ctx[_dtlm.end():_dtlm.end() + 80]
            # Strategy A: slash-pair with unit at end e.g. "185 / -13.2 °F"
            # (label matches at "(MIN/MAX)" qualifier → after = "MIN/MAX) : 185 / -13.2 °F").
            # _dual_temp_pat captures the two numbers (group 3 = optional unit).
            _dt_dual = re.search(_dual_temp_pat, _win80, re.IGNORECASE)
            if _dt_dual and _dt_dual.group(1) != _dt_dual.group(2):
                _dt_vals.extend([float(_dt_dual.group(1)), float(_dt_dual.group(2))])
            else:
                # Strategy B: standard "VALUE °F" or two "VALUE °F" entries
                for _tv in re.finditer(_temp_val_pat, _win80, re.IGNORECASE):
                    _dt_vals.append(float(_tv.group(1)))
        if _dt_vals:
            if len(_dt_vals) == 1:
                # Single design temperature = maximum; MIN is unspecified
                design_temp_max = f'{_dt_vals[0]} °F'
            else:
                design_temp_min = f'{min(_dt_vals)} °F'
                design_temp_max = f'{max(_dt_vals)} °F'

        # ── Design Flowrate ───────────────────────────────────────────────
        # Use a wider context window than the general _pp_ctx (soft-coded via
        # flowrate_context_chars) since flow annotations on P&IDs often sit in
        # connected line labels some distance from the equipment symbol.
        _fl_start  = max(0, m.start() - _flow_ctx_chars)
        _fl_end    = min(len(text), m.end() + _flow_ctx_chars)
        _fl_ctx    = text[_fl_start:_fl_end].upper()
        design_flowrate = ''

        # Strategy 1: label-based (Q= / FLOW RATE: / CAPACITY: / DUTY: …)
        _fl_lbl_m = re.search(_flow_lbl_pat, _fl_ctx, re.IGNORECASE)
        if _fl_lbl_m:
            _fv = re.search(_flow_val_pat, _fl_ctx[_fl_lbl_m.end():_fl_lbl_m.end() + 80], re.IGNORECASE)
            if _fv:
                _val = _fv.group(1).replace(',', '.')
                design_flowrate = f'{_val} {_fv.group(2).upper()}'

        # Strategy 2: bare unit scan — number immediately followed by a
        # recognised flow/duty unit, no label required.  Avoids picking up
        # tag serial numbers by requiring the number > 0 and the unit token
        # is word-bounded (e.g. '100 M3/H' but not '308-TF').
        if not design_flowrate:
            for _bm in re.finditer(_flow_bare_pat, _fl_ctx, re.IGNORECASE):
                _bval = float(_bm.group(1).replace(',', '.'))
                if _bval > 0:
                    design_flowrate = f'{_bm.group(1).replace(",", ".")} {_bm.group(2).upper()}'
                    break

        # Strategy 3: for heaters / heat-exchangers derive duty from any
        # kW or MW annotation.  Only applies when no flow-unit was found.
        if not design_flowrate and prefix in {'E', 'H', 'HT', 'AG', 'CL', 'VR'}:
            _duty_m = re.search(r'(\d+(?:[,.]\d+)?)\s*(MW|KW)\b', _fl_ctx, re.IGNORECASE)
            if _duty_m:
                _dval = float(_duty_m.group(1).replace(',', '.'))
                if _dval > 0:
                    design_flowrate = f'{_dval} {_duty_m.group(2).upper()} (Duty)'

        # ── Insulation ────────────────────────────────────────────────────
        insulation = ''
        _ins_lbl_m = re.search(
            r'(?:INSUL(?:ATION)?|INS|TRACE)\s*[:=/]?\s*([A-Z]{2,10})', _pp_ctx, re.IGNORECASE
        )
        if _ins_lbl_m and _ins_lbl_m.group(1).upper() in _ins_codes:
            insulation = _ins_lbl_m.group(1).upper()
        elif not insulation:
            for _ic in _ins_codes:
                if re.search(r'\b' + _ic + r'\b', _pp_ctx):
                    insulation = _ic
                    break

        # Strategy 2: infer BARE for static equipment with no insulation label.
        # Vessels, drums, separators, tanks, columns etc. on O&G P&IDs are
        # typically BARE unless specifically annotated.  Soft-coded via
        # 'insulation_bare_default_prefixes' in equipment_type_config.json.
        if not insulation:
            _bare_pfxs = {p.upper() for p in ext_cfg.get('insulation_bare_default_prefixes', [
                'V', 'T', 'D', 'S', 'TK', 'F', 'R', 'SC', 'AB', 'CY',
                'FX', 'SK', 'SX', 'VX', 'PF',
            ])}
            if prefix.upper() in _bare_pfxs:
                insulation = 'BARE'

        # ── Dimensions ────────────────────────────────────────────────────
        # Uses a wider context window than _pp_ctx because dimension data
        # boxes on P&IDs are often in a separate table whose OCR text can be
        # far from the equipment tag text.  Both values are soft-coded via
        # dimension_length_context_chars (default 600) and
        # dimension_value_window (default 60 chars after label end).
        _dim_ctx_chars = int(ext_cfg.get('dimension_length_context_chars', 600))
        _dim_val_win   = int(ext_cfg.get('dimension_value_window', 60))
        _dim_start     = max(0, m.start() - _dim_ctx_chars)
        _dim_end       = min(len(text), m.end() + _dim_ctx_chars)
        _dim_ctx       = text[_dim_start:_dim_end].upper()
        dimension_length   = ''
        dimension_diameter = ''

        # ── Soft-coded: reject dimension values that echo the tag serial ──
        # OCR on P&IDs sometimes bonds the numeric serial from the equipment
        # tag (e.g. "851" in "P-851A-TF") to a nearby "MM" / "M" literal,
        # producing spurious values like "851 MM" or "803 MM".  We build a
        # reject-set from the current tag's numeric serial (and close OCR
        # variants) so the dimension scan skips them. Controlled by
        # 'dimension_reject_tag_serial' (default true) in extraction config.
        _dim_reject_tag_serial = bool(ext_cfg.get('dimension_reject_tag_serial', True))
        _tag_serial_reject: set = set()
        if _dim_reject_tag_serial:
            _tag_ser_m = re.search(r'-([0-9]{3,5})', tag)
            if _tag_ser_m:
                _ts = _tag_ser_m.group(1)
                _tag_serial_reject.add(_ts)
                # Also reject "X<serial>" style OCR bleed where a leading digit
                # (e.g. 82106 = "8" merged with "2106"/"106" page-y-offset noise)
                # contains the tag serial as a substring. Anchored to exact
                # serial-length match only to avoid rejecting legitimate 500mm etc.
        # ── Length / Height ──────────────────────────────────────────────
        # OCR artefact: P&ID pipe annotations such as  2"-FL-ACIN-xxxx  are
        # often read as  "2 M"  or  "2 MM"  by Tesseract, and re.search would
        # pick that FIRST match before the real vessel length (e.g. 15.0 M).
        # Fix: collect ALL value matches in the scan window, apply soft-coded
        # minimum thresholds (DIMENSION_LENGTH_MIN_M / _MIN_MM) and take the
        # LARGEST passing value — the vessel length always wins over pipe sizes.
        _len_lbl_m = re.search(_dim_len_lbl, _dim_ctx, re.IGNORECASE)
        if _len_lbl_m:
            _win = _dim_ctx[_len_lbl_m.end():_len_lbl_m.end() + _dim_val_win]
            _best_len_val, _best_len_str = 0.0, ''
            for _dv in re.finditer(_dim_val_pat, _win, re.IGNORECASE):
                _raw_num = _dv.group(1)
                # Reject the current tag's own serial number (OCR bleed)
                if _raw_num in _tag_serial_reject:
                    continue
                _val  = float(_raw_num)
                _unit = (_dv.group(2) or '').upper()
                if _unit == 'M' and _val >= DIMENSION_LENGTH_MIN_M and _val > _best_len_val:
                    _best_len_val = _val
                    _best_len_str = f'{_raw_num} M'
                elif _unit in ('MM', '') and _val >= DIMENSION_LENGTH_MIN_MM and _val > _best_len_val:
                    _best_len_val = _val
                    _best_len_str = f'{_raw_num} {_unit or "MM"}'
            dimension_length = _best_len_str
        # ── Diameter / Width ─────────────────────────────────────────────
        _dia_lbl_m = re.search(_dim_dia_lbl, _dim_ctx, re.IGNORECASE)
        if _dia_lbl_m:
            _win = _dim_ctx[_dia_lbl_m.end():_dia_lbl_m.end() + _dim_val_win]
            _best_dia_val, _best_dia_str = 0.0, ''
            for _dv in re.finditer(_dim_val_pat, _win, re.IGNORECASE):
                _raw_num = _dv.group(1)
                if _raw_num in _tag_serial_reject:
                    continue
                _val  = float(_raw_num)
                _unit = (_dv.group(2) or '').upper()
                if _unit == 'M' and _val >= DIMENSION_DIAMETER_MIN_M and _val > _best_dia_val:
                    _best_dia_val = _val
                    _best_dia_str = f'{_raw_num} M'
                elif _unit in ('MM', '') and _val >= DIMENSION_DIAMETER_MIN_MM and _val > _best_dia_val:
                    _best_dia_val = _val
                    _best_dia_str = f'{_raw_num} {_unit or "MM"}'
            dimension_diameter = _best_dia_str

        # ── Motor Rating ─────────────────────────────────────────────────
        # Uses a WIDER context window than _pp_ctx so motor callouts attached
        # via lead lines (far from the tag in OCR text order) are still found.
        # Soft-coded via motor_rating_context_chars (default 800 chars each
        # side).  OCR at 0°/90°/180°/270° (ocr_rotation_angles) means vertical
        # and downward-oriented motor annotations are already in the text pool.
        _mtr_ctx_chars = int(ext_cfg.get('motor_rating_context_chars', 800))
        _mtr_start     = max(0, m.start() - _mtr_ctx_chars)
        _mtr_end       = min(len(text), m.end() + _mtr_ctx_chars)
        _mtr_ctx       = text[_mtr_start:_mtr_end].upper()
        motor_rating = ''
        _mr_m = re.search(_mtr_pat, _mtr_ctx, re.IGNORECASE)
        if _mr_m:
            motor_rating = f'{_mr_m.group(1)} {_mr_m.group(2).upper()}'
        elif not motor_rating:
            _mr_bare_m = re.search(_mtr_bare_pat, _mtr_ctx, re.IGNORECASE)
            if _mr_bare_m:
                motor_rating = f'{_mr_bare_m.group(1)} {_mr_bare_m.group(2).upper()}'

        # Soft-coded: non-rotating equipment has no motor.  Prefixes are
        # configurable via 'motor_na_prefixes' in equipment_type_config.json.
        # The display value ('No', '', 'N/A' …) is controlled by
        # 'motor_na_display_value' (default 'No') — change in config only.
        _motor_na_pfxs = {p.upper() for p in ext_cfg.get('motor_na_prefixes', [
            'V', 'T', 'D', 'F', 'R', 'E', 'S', 'TK', 'SC', 'AB', 'ST', 'FL',
            'CY', 'DR', 'FG', 'MS', 'SK', 'HX', 'SX', 'FX', 'VX', 'GX',
            'PF', 'DP', 'AN', 'EJ', 'MX',
        ])}
        _motor_na_val = str(ext_cfg.get('motor_na_display_value', 'No'))
        if not motor_rating and prefix.upper() in _motor_na_pfxs:
            motor_rating = _motor_na_val

        # ── Quality Specification (→ Remarks) ────────────────────────────
        # Core logic unchanged — still extracts quality class / NACE compliance.
        # Result is now stored in quality_spec and routed to the 'remarks' field
        # so that 'quality_required' (→ "Quantity Required" column) can hold the
        # numerical count.  Controlled by QUALITY_SPEC_* module-level constants.
        #
        # Soft-coded toggle: 'quality_spec_in_remarks' (default False) — when
        # False, quality/NACE references are NOT routed into the Remarks column.
        # Oil & Gas equipment lists typically keep Remarks for project notes only;
        # NACE applicability is derived from service fluid / drawing notes and is
        # NOT a per-row property. Set to True to restore legacy behaviour.
        _quality_in_remarks = bool(ext_cfg.get('quality_spec_in_remarks', False))
        quality_spec = ''
        if _quality_in_remarks:
            _qr_m = re.search(_qual_pat, _pp_ctx, re.IGNORECASE)
            if _qr_m:
                quality_spec = _qr_m.group(1).strip().upper()

        # Strategy 2: scan wider context for explicit NACE reference.
        # P&IDs for sour-service equipment (H2S/SOUR GAS) should comply with
        # NACE MR0175 / ISO 15156.  When the drawing has NACE in the notes or
        # title block, apply it.  Threshold is soft-coded via
        # 'quality_nace_context_chars' (default 1500).
        if _quality_in_remarks and not quality_spec:
            _qual_ctx_chars = int(ext_cfg.get('quality_nace_context_chars', 1500))
            _qual_ctx_start = max(0, m.start() - _qual_ctx_chars)
            _qual_ctx_end   = min(len(text), m.end() + _qual_ctx_chars)
            _qual_wide_ctx  = text[_qual_ctx_start:_qual_ctx_end]
            if re.search(r'\bNACE\s*MR\s*0175\b', _qual_wide_ctx, re.IGNORECASE):
                quality_spec = QUALITY_SPEC_NACE_FULL
            elif re.search(r'\bNACE\b', _qual_wide_ctx, re.IGNORECASE):
                quality_spec = QUALITY_SPEC_NACE_SHORT

        # Strategy 3: infer from sour-service context.  If sour gas / H2S /
        # HIC is detected anywhere in the drawing, all vessels must comply
        # with NACE MR0175.  Soft-coded disable via
        # 'quality_infer_from_sour_service' = false.
        _infer_sour = bool(ext_cfg.get('quality_infer_from_sour_service', True))
        if _quality_in_remarks and not quality_spec and _infer_sour:
            _sour_ctx = text[:min(len(text), 6000)]  # check first 6 k chars
            if re.search(r'\bSOUR\s+GAS\b|\bH2S\b|\bHIC\b|\bSSC\b',
                         _sour_ctx, re.IGNORECASE):
                quality_spec = QUALITY_SPEC_NACE_FULL
            elif service_fluid and re.search(r'\bsour\b', service_fluid, re.IGNORECASE):
                quality_spec = QUALITY_SPEC_NACE_FULL

        # ── Quantity Required (numerical count) ───────────────────────────
        # Extracts an explicit count callout (QTY: 2, NO. REQD 1, etc.) from
        # the narrow context window.  Defaults to QUANTITY_REQUIRED_DEFAULT
        # ('1') when no callout is found — the vast majority of P&ID equipment
        # items are single-unit.  Pattern is soft-coded via module-level
        # QUANTITY_REQUIRED_PATTERN.
        quantity_required = QUANTITY_REQUIRED_DEFAULT
        _qty_m = re.search(QUANTITY_REQUIRED_PATTERN, _pp_ctx, re.IGNORECASE)
        if _qty_m:
            quantity_required = _qty_m.group(1).strip()

        # ── Revision & SL No — pre-tag token scan ─────────────────────────
        # Tabular PDF text writes table cells as newline-separated tokens in
        # reading order.  The pattern is: ...\n[sl_no]\n[revision]\n[TAG]\n...
        # We scan _REV_PRE_TAG_TOKENS tokens immediately before the tag match.
        # Soft-coded via _REV_PRE_TAG_WIN_CHARS, _REV_PRE_TAG_TOKENS,
        # _clean_revision(), and the SL-no regex below.
        _pre_text = text[max(0, m.start() - _REV_PRE_TAG_WIN_CHARS):m.start()]
        _pre_toks = [t.strip() for t in _pre_text.split('\n') if t.strip()]
        revision  = ''
        sl_no     = ''
        if _pre_toks:
            # Closest token before the tag is most likely the revision cell
            _rev_candidate = _clean_revision(_pre_toks[-1])
            if _rev_candidate:
                revision = _rev_candidate
                # Token before the revision cell is likely the SL No
                if len(_pre_toks) >= 2 and re.match(r'^\d{1,3}$', _pre_toks[-2]):
                    sl_no = _pre_toks[-2]
            elif re.match(r'^\d{1,3}$', _pre_toks[-1]):
                # Last token is a number; could be sl_no with no revision column
                sl_no = _pre_toks[-1]

        results.append({
            'tag':                 tag,
            'type_label':          type_label,
            'description':         description,
            'revision':            revision,
            'sl_no':               sl_no,
            'area':                area,
            'drawing_ref':         drawing_ref,
            'line_connections':    lc_tokens,
            'nozzle_connections':  nozzle_tokens,
            'service_fluid':       service_fluid,
            'material_class':      material_class,
            'process_notes':       process_notes,
            # Process parameters extracted from data-bubble context
            'design_flowrate':     design_flowrate,
            'oper_pressure':       oper_pressure,
            'oper_temperature':    oper_temperature,
            'design_pressure_min': design_pressure_min,
            'design_pressure_max': design_pressure_max,
            'design_temp_min':     design_temp_min,
            'design_temp_max':     design_temp_max,
            'insulation':          insulation,
            'dimension_length':    dimension_length,
            'dimension_diameter':  dimension_diameter,
            'motor_rating':        motor_rating,
            'quality_required':    quantity_required,
            'remarks':             quality_spec,
        })

        # ── Merge data-box index values (fill empty fields only) ──────────
        # Values extracted by _build_databox_index from the full text are
        # merged in here.  Only fills columns that the narrow context-window
        # extraction above left empty — never overwrites a found value.
        _db_vals = _databox_idx.get(tag.upper(), {})
        if _db_vals:
            _last = results[-1]
            for _fk, _fv in _db_vals.items():
                if _fk in _last:
                    # For _min/_max fields the explicit databox label extraction
                    # (e.g. "DESIGN TEMP (MAX/MIN): 185 F / -13.2 F") is more
                    # reliable than the narrow context scan — always override
                    # so a wrong single-value pickup from context is corrected.
                    if _fk.endswith(('_min', '_max')):
                        if _fv:  # only write a non-empty databox value
                            _last[_fk] = _fv
                    elif not _last[_fk]:
                        _last[_fk] = _fv
            # Normalise oper_temperature if it was filled from the databox
            # (the narrow-context path already calls _normalize_oper_temp;
            #  the databox path does not, so we apply it here).
            if _last.get('oper_temperature') and '/' in _last['oper_temperature']:
                _last['oper_temperature'] = _normalize_oper_temp(_last['oper_temperature'])
            # ── Fallback: split raw "X / Y" values that landed in a single
            # _max or _min field (happens when the (MAX/MIN) label variant
            # did not match and the shorter fallback label fired instead,
            # e.g. "design temp" → design_temp_max = "185 F / -13.2 F").
            # We detect these, split numerically and re-assign correctly.
            for _single_fk in ('design_temp_max', 'design_pressure_max', 'design_temp_min', 'design_pressure_min'):
                _sv = _last.get(_single_fk, '')
                if _sv and '/' in _sv:
                    # First, strip non-numeric FV-type suffixes on pressure fields
                    if 'pressure' in _single_fk:
                        _sv_cleaned = _clean_pressure_value(_sv)
                        if _sv_cleaned != _sv:
                            _last[_single_fk] = _sv_cleaned
                            continue
                    _sparts = re.split(r'\s*/\s*', _sv, maxsplit=1)
                    if len(_sparts) == 2:
                        _sn0 = re.search(r'-?\d+(?:\.\d+)?', _sparts[0])
                        _sn1 = re.search(r'-?\d+(?:\.\d+)?', _sparts[1])
                        if _sn0 and _sn1:
                            _sf0, _sf1 = float(_sn0.group()), float(_sn1.group())
                            # Determine the _min/_max counterpart field name
                            if _single_fk.endswith('_max'):
                                _counterpart = _single_fk[:-4] + '_min'
                            else:
                                _counterpart = _single_fk[:-4] + '_max'
                            # Assign larger to _max, smaller to _min
                            _larger  = _sparts[0] if _sf0 >= _sf1 else _sparts[1]
                            _smaller = _sparts[0] if _sf0 <= _sf1 else _sparts[1]
                            if _single_fk.endswith('_max'):
                                _last[_single_fk] = _larger
                                if _counterpart in _last and not _last[_counterpart]:
                                    _last[_counterpart] = _smaller
                            else:
                                _last[_single_fk] = _smaller
                                if _counterpart in _last and not _last[_counterpart]:
                                    _last[_counterpart] = _larger
            print(
                f'[EQ-DIAG][DataBox] Merged into {tag}: '
                + str({k: v for k, v in _db_vals.items() if _last.get(k)}),
                flush=True,
            )

    results.sort(key=lambda x: x['tag'])

    # ── Gate 2: data-box presence post-filter ────────────────────────────
    # Soft-coded via 'require_at_least_one_param' and
    # 'prefer_databox_index_filter' in equipment_type_config.json.
    #
    # Two-tier strategy:
    #
    # Tier A — _databox_idx (preferred, more reliable):
    #   _build_databox_index scans label:value pairs across the FULL OCR
    #   text and associates each value with the NEAREST preceding tag. This
    #   proximity-based attribution is immune to context-window bleed: even
    #   if V-804-TF and V-308-TF both appear in a 400-char window around a
    #   "OPER PRESS" label, the LAST tag before the label in OCR text order
    #   is V-308-TF (the actual data-box owner). So _databox_idx accurately
    #   contains {V-308-TF: {oper_pressure: ..., ...}} and nothing for
    #   referenced tags like V-804-TF.
    #   When _databox_idx has any entries, use it as the SOLE authoritative
    #   list of primary equipment on this drawing.
    #   Controlled by 'prefer_databox_index_filter' (default true).
    #
    # Tier B — context-window param-field check (fallback):
    #   Used when _databox_idx is empty (label format not in databox_label_map,
    #   drawing has no data boxes, etc.). Any item with at least one non-empty
    #   core process parameter from the regex context-window pass is kept.
    #   Controlled by 'require_at_least_one_param' (default true).
    if bool(ext_cfg.get('require_at_least_one_param', True)):
        _primary_fields = ext_cfg.get(
            'param_fields_for_primary_check',
            ['oper_pressure', 'oper_temperature',
             'design_pressure_min', 'design_pressure_max',
             'design_temp_min', 'design_temp_max',
             'dimension_length', 'dimension_diameter',
             'design_flowrate'],
        )
        _prefer_db_idx = bool(ext_cfg.get('prefer_databox_index_filter', True))
        _kept, _dropped = [], []

        if _prefer_db_idx and _databox_idx:
            # ── Tier A: databox-index is authoritative ───────────────────
            # Only tags confirmed by the proximity-based label scan are primary.
            for _item in results:
                if _item['tag'].upper() in _databox_idx:
                    _kept.append(_item)
                else:
                    _dropped.append(_item['tag'])
            print(
                f'[EQ-DIAG] Gate2-TierA: kept {len(_kept)} databox-indexed tag(s), '
                f'removed {len(_dropped)} referenced tag(s): {_dropped}',
                flush=True,
            )
        else:
            # ── Tier B: context-window param check (fallback) ────────────
            for _item in results:
                _has_param = any(_item.get(f) for f in _primary_fields)
                if _has_param:
                    _kept.append(_item)
                else:
                    _dropped.append(_item['tag'])
            if _dropped:
                print(
                    f'[EQ-DIAG] Gate2-TierB: removed {len(_dropped)} no-param '
                    f'referenced tag(s): {_dropped}',
                    flush=True,
                )

        results = _kept if _kept else results  # safety: never return empty when all pass

    # ── Description reject filter ───────────────────────────────────
    # Soft-coded via 'description_reject_patterns' in equipment_type_config.json.
    # Any item whose description field matches ANY pattern is excluded.
    # Use-case: OCR picks up revision-cloud annotation text ("Revision Cloud")
    # or description-only rows that have no real equipment tag meaning.
    _desc_reject_pats = ext_cfg.get('description_reject_patterns', [])
    if _desc_reject_pats:
        _pre_dr = len(results)
        results = [
            _item for _item in results
            if not any(
                re.search(_drpat, _item.get('description', '') or '')
                for _drpat in _desc_reject_pats
            )
        ]
        _removed_dr = _pre_dr - len(results)
        if _removed_dr:
            print(
                f'[EQ-DIAG] Description-reject filter removed {_removed_dr} item(s)',
                flush=True,
            )

    # ── Description CLEAR filter (keeps the row, blanks the description) ──
    # Soft-coded via 'description_clear_patterns'. Matches cases where the
    # regex fallback picked up notes-bleed text as the description (e.g.
    # "Rated Flow Hr Pump Supplied With", "All Instrument Tag Numbers Are
    # Suffixed", single-word "Pump"). Clearing the description here means the
    # AI gap-fill pass (ai_authoritative_fields='description') will re-extract
    # the correct data-box title. Unlike the reject filter, the equipment row
    # itself is kept.
    _desc_clear_pats = ext_cfg.get('description_clear_patterns', [])
    if _desc_clear_pats:
        _cleared = 0
        for _item in results:
            _desc = _item.get('description', '') or ''
            if not _desc:
                continue
            if any(re.search(_dcpat, _desc) for _dcpat in _desc_clear_pats):
                _item['description'] = ''
                _cleared += 1
        if _cleared:
            print(
                f'[EQ-DIAG] Description-clear filter blanked {_cleared} description(s)'
                ' for AI re-extraction',
                flush=True,
            )
    # ── Dedup-by-description gate ──────────────────────────────────────
    # Soft-coded via 'dedup_by_description_enabled' in equipment_type_config.json.
    # When multiple extracted items share the same description (case-insensitive),
    # keep only the one with the most populated fields (highest richness score).
    # This eliminates OCR-variant duplicates such as V-803-TF vs V-803-TEF when
    # both are matched to the same physical equipment description on the drawing.
    if bool(ext_cfg.get('dedup_by_description_enabled', True)):
        def _richness(item):
            return sum(1 for v in item.values() if v and str(v).strip() not in ('', '—', 'None'))
        _desc_groups: dict = {}
        for _item in results:
            _dkey = ((_item.get('description') or '').strip().lower()) or f'__notag_{_item["tag"]}'
            if _dkey not in _desc_groups:
                _desc_groups[_dkey] = []
            _desc_groups[_dkey].append(_item)
        _before_dedup = len(results)
        results = [
            max(_grp, key=_richness)
            for _grp in _desc_groups.values()
        ]
        _removed_dd = _before_dedup - len(results)
        if _removed_dd:
            print(
                f'[EQ-DIAG] Dedup-by-description removed {_removed_dd} duplicate(s)',
                flush=True,
            )

    # ── Uppercase normalization for material/abbreviation fields ────────────
    # Soft-coded via 'uppercase_fields' in equipment_type_config.json. Material
    # codes ("cs", "cS + LINING") must render as canonical uppercase ("CS",
    # "CS + LINING"). Uses casefold-aware upper() and preserves internal
    # punctuation/spaces. Only applied to fields listed in config.
    _upper_fields = ext_cfg.get('uppercase_fields', ['moc', 'insulation'])
    if _upper_fields:
        for _item in results:
            for _uf in _upper_fields:
                _uv = _item.get(_uf, '')
                if isinstance(_uv, str) and _uv and _uv.strip() not in ('—', 'N/A'):
                    _item[_uf] = _uv.upper()

    # Soft-coded via config key 'minmax_correction_pairs' in the 'extraction'
    # section. Ensures that design_temp_max is always numerically >= design_temp_min
    # regardless of which code path populated them.
    _minmax_pairs = ext_cfg.get('minmax_correction_pairs', [
        ['design_temp_min',      'design_temp_max'],
        ['design_pressure_min',  'design_pressure_max'],
    ])
    for _item in results:
        for _fmin, _fmax in _minmax_pairs:
            _vmin = _item.get(_fmin, '')
            _vmax = _item.get(_fmax, '')
            if _vmin and _vmax:
                _nmin = re.search(r'-?\d+(?:\.\d+)?', _vmin)
                _nmax = re.search(r'-?\d+(?:\.\d+)?', _vmax)
                if _nmin and _nmax and float(_nmin.group()) > float(_nmax.group()):
                    # Values are inverted — swap
                    _item[_fmin], _item[_fmax] = _vmax, _vmin

    # ── Post-deduplication: canonicalize to the most-specific tag form ────
    # When the same physical equipment appears in two forms (e.g. both "V-308"
    # and "V-308-TF"), keep only the longer (more specific) form.
    # Two different equipment with the same base but different suffixes
    # (e.g. V-805-TF vs V-805-1F) are intentionally kept as separate rows.
    _suffix_re = re.compile(r'^([A-Z]{1,2}-[0-9]{3,5}[A-Z]?)-[A-Z0-9]{1,4}$')
    _full_tag_bases: set = set()
    for _item in results:
        _m = _suffix_re.match(_item['tag'])
        if _m:
            _full_tag_bases.add(_m.group(1))  # e.g. "V-308" from "V-308-TF"
    # Remove bare-base entries that have a suffixed sibling
    results = [
        _item for _item in results
        if _item['tag'] not in _full_tag_bases
    ]

    # ── Sibling-unit merge (soft-coded, additive post-processing) ──────────
    # Collapses multi-unit equipment that share one data box (e.g. triple-50%
    # pump train "P-851A/B/C-TF") into a single row whose tag lists every
    # variant letter joined by 'sibling_merge_separator'. All other fields
    # (description, process params) inherit from the first variant — those
    # values come from the shared data box on the drawing so they apply
    # identically to every unit.
    #
    # Controlled by:
    #   merge_sibling_unit_variants      (bool, default True)
    #   sibling_merge_separator          (str,  default '/')
    #   sibling_merge_min_group_size     (int,  default 2)
    #
    # No extraction logic above this point is altered; this step only
    # rewrites the final results list.
    if bool(ext_cfg.get('merge_sibling_unit_variants', True)):
        _sep       = str(ext_cfg.get('sibling_merge_separator', '/'))
        _min_group = int(ext_cfg.get('sibling_merge_min_group_size', 2))
        # Matches  PREFIX-NNN + single alpha variant + optional -SUFFIX
        # e.g. P-851A-TF  →  base='P-851', variant='A', suffix='TF'
        _sibling_split_re = re.compile(
            r'^([A-Z]{1,2}-[0-9]{3,5})([A-Z])(?:-([A-Z0-9]{1,4}))?$'
        )
        _groups: dict = {}          # group_key → [indices in results]
        _group_order: list = []      # first-seen order of group keys
        for _idx, _item in enumerate(results):
            _sm = _sibling_split_re.match(_item['tag'])
            if not _sm:
                continue
            _base, _variant, _sfx = _sm.group(1), _sm.group(2), _sm.group(3) or ''
            _key = (_base, _sfx)
            if _key not in _groups:
                _groups[_key] = []
                _group_order.append(_key)
            _groups[_key].append((_idx, _variant))

        _to_remove: set = set()
        _merged_updates: dict = {}   # idx → new tag / qty
        for _key in _group_order:
            _members = _groups[_key]
            if len(_members) < _min_group:
                continue
            # Preserve first-seen variant order (keeps A/B/C ordering natural)
            _base, _sfx = _key
            _variants_seen: list = []
            for _idx, _variant in _members:
                if _variant not in _variants_seen:
                    _variants_seen.append(_variant)
            _merged_tag = f"{_base}{_sep.join(_variants_seen)}"
            if _sfx:
                _merged_tag = f"{_merged_tag}-{_sfx}"
            _keep_idx = _members[0][0]
            _merged_updates[_keep_idx] = {
                'tag':              _merged_tag,
                'quality_required': str(len(_variants_seen)),
            }
            for _idx, _ in _members[1:]:
                _to_remove.add(_idx)
            print(
                f'[EQ-DIAG] Sibling-merge: {[results[i]["tag"] for i, _ in _members]} '
                f'→ {_merged_tag!r} (qty={len(_variants_seen)})',
                flush=True,
            )

        if _to_remove or _merged_updates:
            _new_results = []
            for _idx, _item in enumerate(results):
                if _idx in _to_remove:
                    continue
                if _idx in _merged_updates:
                    _item = dict(_item)
                    _item.update(_merged_updates[_idx])
                _new_results.append(_item)
            results = _new_results

    return results


def _dedup_equipment_by_tag(items: list) -> list:
    """
    Deduplicate a merged multi-page equipment list by tag.

    When the same tag appears on more than one page (e.g. it was referenced on
    page 1 AND has its own data box on page 2), keep only the entry with the
    most populated fields.  This ensures the richest extraction result wins.
    """
    by_tag: dict = {}
    for item in items:
        tag = (item.get('tag') or '').upper()
        if not tag:
            continue
        if tag not in by_tag:
            by_tag[tag] = item
        else:
            _skip_keys = {'sl_no', 'tag', 'type_label', 'area', 'drawing_ref',
                          'line_connections', 'nozzle_connections'}
            _pop_new = sum(1 for k, v in item.items()
                           if k not in _skip_keys and v and v not in ('', 'No', [], 'N/A'))
            _pop_old = sum(1 for k, v in by_tag[tag].items()
                           if k not in _skip_keys and v and v not in ('', 'No', [], 'N/A'))
            if _pop_new > _pop_old:
                by_tag[tag] = item
    result = list(by_tag.values())
    print(f'[EQ-DIAG] _dedup_equipment_by_tag: {len(items)} in → {len(result)} unique tags out', flush=True)

    # ── Cross-page sibling-unit merge (soft-coded) ──────────────────────────
    # After multi-page dedup, sibling variants (P-851A / P-851B / P-851C)
    # extracted from DIFFERENT pages are now all present in `result`. The
    # per-page sibling merge inside _extract_equipment_items cannot collapse
    # them because they never lived in the same page's results. Run one more
    # pass here. Toggle via 'merge_sibling_unit_variants_cross_page'
    # (default True, inherits same merge settings).
    try:
        cfg     = _load_config()
        ext_cfg = cfg.get('extraction', {})
        if bool(ext_cfg.get('merge_sibling_unit_variants_cross_page', True)) \
                and bool(ext_cfg.get('merge_sibling_unit_variants', True)):
            _sep       = str(ext_cfg.get('sibling_merge_separator', '/'))
            _min_group = int(ext_cfg.get('sibling_merge_min_group_size', 2))
            _sibling_split_re = re.compile(
                r'^([A-Z]{1,2}-[0-9]{3,5})([A-Z])(?:-([A-Z0-9]{1,4}))?$'
            )
            _groups: dict = {}
            _order:  list = []
            for _i, _it in enumerate(result):
                _sm = _sibling_split_re.match((_it.get('tag') or '').upper())
                if not _sm:
                    continue
                _key = (_sm.group(1), _sm.group(3) or '')
                if _key not in _groups:
                    _groups[_key] = []
                    _order.append(_key)
                _groups[_key].append((_i, _sm.group(2)))

            _drop: set = set()
            _upd:  dict = {}
            for _key in _order:
                _members = _groups[_key]
                if len(_members) < _min_group:
                    continue
                _variants_seen: list = []
                for _i, _v in _members:
                    if _v not in _variants_seen:
                        _variants_seen.append(_v)
                _base, _sfx = _key
                _merged_tag = f"{_base}{_sep.join(_variants_seen)}"
                if _sfx:
                    _merged_tag = f"{_merged_tag}-{_sfx}"
                # Keep the richest member as the carrier row
                def _r(it):
                    return sum(1 for v in it.values() if v and str(v).strip() not in ('', '—', 'None', 'No'))
                _best = max(_members, key=lambda m: _r(result[m[0]]))[0]
                _upd[_best] = {
                    'tag':              _merged_tag,
                    'quality_required': str(len(_variants_seen)),
                }
                for _i, _ in _members:
                    if _i != _best:
                        _drop.add(_i)
                print(
                    f'[EQ-DIAG] Cross-page sibling-merge: '
                    f'{[result[i]["tag"] for i, _ in _members]} → {_merged_tag!r}',
                    flush=True,
                )
            if _drop or _upd:
                _new: list = []
                for _i, _it in enumerate(result):
                    if _i in _drop:
                        continue
                    if _i in _upd:
                        _it = dict(_it)
                        _it.update(_upd[_i])
                    _new.append(_it)
                result = _new
    except Exception as _e:
        print(f'[EQ-DIAG] Cross-page sibling-merge skipped: {_e}', flush=True)

    return result


_result_store: dict = {}

# ── Soft-coded AI gap-fill constants ─────────────────────────────────────────
# Default list of fields the AI gap-fill pass will attempt.
# Controlled per-deploy via equipment_type_config.json extraction.ai_gap_fill_fields.
# Fields already populated by the regex pass are NEVER overwritten.
_AI_GAP_FILL_DEFAULT_FIELDS = [
    'oper_pressure',
    'oper_temperature',
    'design_pressure_min',
    'design_pressure_max',
    'design_temp_min',
    'design_temp_max',
    'design_flowrate',
    'moc',
    'insulation',
    'description',
]

# JSON response validation: AI must return keys matching these field names.
# Values must be strings (or null). Any other structure is rejected.
_AI_FILL_FIELD_SET = set(_AI_GAP_FILL_DEFAULT_FIELDS)


def _ai_gap_fill_pid_items(items: list, text: str, config: dict) -> list:
    """
    Multi-model AI gap-fill pass for P&ID drawing mode.

    For each equipment item with one or more empty target fields, re-extracts
    a wider text context window around the tag and sends it to:
      1. OpenAI GPT-4o (primary)  — with automatic Gemini fallback inside
         MultiModelAIService if the OpenAI quota is exceeded.
      2. Gemini Flash             — as an independent second-opinion when
         ai_gap_fill_provider == 'both', to fill any fields GPT-4o left null.

    The AI is prompted to return ONLY a flat JSON object.  Values are merged
    into the item ONLY when the field is still empty after regex extraction.

    Soft-coded via equipment_type_config.json extraction section:
      ai_gap_fill_enabled       : true/false  (default true)
      ai_gap_fill_provider      : "both" | "openai" | "gemini"
      ai_gap_fill_fields        : list of field keys to attempt
      ai_gap_fill_context_chars : chars each side of the tag (default 800)
      ai_gap_fill_max_tokens    : max tokens for AI response (default 350)
      ai_gap_fill_temperature   : sampling temperature (default 0)
      ai_gap_fill_min_empty_fields : minimum empty fields before AI is called
    """
    import json as _json

    ext_cfg     = config.get('extraction', {})
    enabled     = bool(ext_cfg.get('ai_gap_fill_enabled', True))
    if not enabled:
        return items

    fill_fields = list(ext_cfg.get('ai_gap_fill_fields', _AI_GAP_FILL_DEFAULT_FIELDS))
    ctx_chars   = int(ext_cfg.get('ai_gap_fill_context_chars', 800))
    max_tokens  = int(ext_cfg.get('ai_gap_fill_max_tokens', 350))
    temperature = float(ext_cfg.get('ai_gap_fill_temperature', 0))
    min_empty   = int(ext_cfg.get('ai_gap_fill_min_empty_fields', 1))
    provider    = str(ext_cfg.get('ai_gap_fill_provider', 'both')).lower()
    # Fields where AI result ALWAYS overrides regex — ensures deterministic,
    # consistent output across users regardless of OCR noise (e.g. description).
    authoritative_fields = set(ext_cfg.get('ai_authoritative_fields', []))

    # Lazy-import to avoid circular imports and keep startup fast
    try:
        from apps.pid_analysis.multi_model_service import MultiModelAIService
        ai = MultiModelAIService()
    except Exception as _e:
        print(f'[EQ-DIAG][AI-FILL] Service init failed: {_e}', flush=True)
        return items

    text_upper = text.upper()

    # ── Shared prompt builder ────────────────────────────────────────────────
    _FIELD_HINTS = {
        'oper_pressure':      'Operating pressure with unit (e.g. "155 PSIG" or "10 barg")',
        'oper_temperature':   'Operating temperature with unit (e.g. "105 °F" or "40 °C")',
        'design_pressure_min': 'Minimum design/set pressure with unit (e.g. "-13.2 PSIG")',
        'design_pressure_max': 'Maximum design/set pressure with unit (e.g. "195 PSIG")',
        'design_temp_min':    'Minimum design temperature with unit (e.g. "-13 °F")',
        'design_temp_max':    'Maximum design temperature with unit (e.g. "185 °F")',
        'design_flowrate':    'Design flowrate or capacity with unit (e.g. "327 M3" or "100 M3/H")',
        'moc':                'Material of construction abbreviation (e.g. "CS", "SS316L", "DUPLEX")',
        'insulation':         'Insulation type code (e.g. "PERS", "HOT", "BARE", "TRACED")',
        'description':        (
            'Equipment service description — the exact 2-8 word ALL-CAPS title '
            'printed in the equipment data box directly below or next to the tag '
            '(e.g. "MRD THREE PHASE SEPARATOR", "MRD PRODUCED WATER DISPOSAL PUMP", '
            '"MRD MEA INLET SCRUBBER"). Prefer the literal words on the drawing. '
            'Do NOT use connector/routing text like "SOUR GAS TO ..." or "FROM ..." '
            'or "TO ...". Do NOT include pipe tags, line numbers, or drawing numbers. '
            'Return the title as ALL UPPERCASE.'
        ),
    }

    def _build_prompt(tag: str, empty: list, ctx: str) -> str:
        field_list = '\n'.join(
            f'  "{f}": {_FIELD_HINTS.get(f, f)}'
            for f in empty
        )
        return (
            f'You are an Oil & Gas P&ID data extraction assistant.\n'
            f'Extract the following fields for equipment tag {tag} from the text excerpt below.\n'
            f'Return ONLY a valid JSON object with these exact keys. '
            f'Set the value to null if the field cannot be found.\n'
            f'Do NOT include markdown fences, explanations, or keys not listed.\n\n'
            f'Fields to extract:\n{field_list}\n\n'
            f'TEXT EXCERPT:\n{ctx}'
        )

    def _call_ai(prompt: str, model_hint: str) -> dict:
        """Call the AI and return parsed dict, or {} on failure."""
        try:
            raw = ai.chat_completion(
                messages=[{'role': 'user', 'content': prompt}],
                model=model_hint,
                max_tokens=max_tokens,
                temperature=temperature,
            )
            # Strip markdown code fences if present
            cleaned = re.sub(r'^```[a-z]*\s*', '', raw.strip(), flags=re.IGNORECASE)
            cleaned = re.sub(r'\s*```$', '', cleaned.strip())
            parsed  = _json.loads(cleaned)
            if isinstance(parsed, dict):
                return parsed
        except Exception as _e:
            print(f'[EQ-DIAG][AI-FILL] parse/call error ({model_hint}): {_e}', flush=True)
        return {}

    # ── Per-item gap fill ────────────────────────────────────────────────────
    for item in items:
        tag = item.get('tag', '')
        if not tag:
            continue

        empty_fields = [f for f in fill_fields if not item.get(f)]
        # Authoritative fields are ALWAYS queried — even when regex populated
        # them — so the AI result overrides inconsistent OCR-derived values.
        forced_fields = [
            f for f in fill_fields
            if f in authoritative_fields and f not in empty_fields
        ]
        query_fields = empty_fields + forced_fields
        if len(query_fields) < min_empty:
            continue

        # Re-locate tag in text for context window
        idx = text_upper.find(tag.upper())
        if idx == -1:
            continue
        ctx_start = max(0, idx - ctx_chars // 2)
        ctx_end   = min(len(text), idx + ctx_chars // 2)
        ctx       = text[ctx_start:ctx_end]

        prompt = _build_prompt(tag, query_fields, ctx)

        # ── Pass 1: primary provider (GPT-4o, with auto Gemini fallback) ──
        gpt_result = {}
        if provider in ('openai', 'both'):
            gpt_result = _call_ai(prompt, 'openai')

        # ── Pass 2: Gemini second-opinion (fills any fields GPT-4o left null) ──
        gem_result = {}
        if provider in ('gemini', 'both'):
            still_empty = [f for f in query_fields if not gpt_result.get(f)]
            if still_empty:
                gem_result = _call_ai(_build_prompt(tag, still_empty, ctx), 'gemini')

        # ── Merge: OpenAI wins over Gemini.
        # Authoritative fields: AI always overrides regex.
        # Non-authoritative fields: AI fills only when regex left it empty.
        filled = []
        for f in query_fields:
            val = gpt_result.get(f) or gem_result.get(f)
            if not val or str(val).strip().lower() in ('null', 'none', ''):
                continue
            is_authoritative = f in authoritative_fields
            if item.get(f) and not is_authoritative:
                continue                       # already filled by regex — skip
            new_val = str(val).strip()
            if is_authoritative and item.get(f) == new_val:
                continue                       # no change
            item[f] = new_val
            filled.append(f)

        if filled:
            print(f'[EQ-DIAG][AI-FILL] {tag}: AI filled {filled}', flush=True)

    # ── Uppercase normalization for material/abbreviation fields ─────────
    # Applied again here so AI-filled MOC / insulation values also render in
    # canonical uppercase. Soft-coded via 'uppercase_fields' in config.
    _upper_fields_ai = ext_cfg.get('uppercase_fields', ['moc', 'insulation'])
    if _upper_fields_ai:
        for _it in items:
            for _uf in _upper_fields_ai:
                _uv = _it.get(_uf, '')
                if isinstance(_uv, str) and _uv and _uv.strip() not in ('—', 'N/A'):
                    _it[_uf] = _uv.upper()

    return items


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyze_pid_equipment(request):
    """POST /api/v1/pid/equipment/analyze/
    Accepts a single P&ID PDF, dispatches extraction to a Celery background task,
    and returns HTTP 202 with upload_id immediately.
    The frontend polls /status/<upload_id>/ every few seconds until 'completed'.
    """
    from apps.pid_analysis.tasks import run_equipment_analysis_task   # lazy import avoids circular

    config  = _load_config()
    ext_cfg = config.get('extraction', {})
    allowed = [e.lower() for e in ext_cfg.get('allowed_extensions', ['pdf'])]
    max_mb  = float(ext_cfg.get('max_file_size_mb', 50))

    pid_file = request.FILES.get('file') or (list(request.FILES.values())[0] if request.FILES else None)
    if not pid_file:
        return Response({'error': 'No file provided', 'success': False},
                        status=drf_status.HTTP_400_BAD_REQUEST)

    ext = pid_file.name.rsplit('.', 1)[-1].lower()
    if ext not in allowed:
        return Response({'error': f'Unsupported format: .{ext}. Allowed: {", ".join(allowed)}', 'success': False},
                        status=drf_status.HTTP_400_BAD_REQUEST)

    if pid_file.size > max_mb * 1024 * 1024:
        return Response({'error': f'File exceeds {max_mb} MB limit', 'success': False},
                        status=drf_status.HTTP_400_BAD_REQUEST)

    upload_id = f'EQ-{uuid.uuid4().hex[:12].upper()}'

    # Encode file bytes as base64 so they can be passed as a JSON-serialisable
    # task argument (Celery serialises args to JSON by default).
    file_b64 = base64.b64encode(pid_file.read()).decode('ascii')

    # Mark as 'processing' in cache immediately so status polls never return 404.
    cache.set(
        EQ_RESULT_CACHE_KEY_FMT.format(upload_id=upload_id),
        {'status': 'processing', 'progress': 0, 'message': 'Queued for extraction…'},
        EQ_RESULT_CACHE_TTL_S,
    )

    _dispatch_eq_task(run_equipment_analysis_task, upload_id, config,
                      upload_id, file_b64, pid_file.name)
    logger.info('[EquipmentList] Task dispatched  upload_id=%s  file=%s', upload_id, pid_file.name)

    return Response(
        {'upload_id': upload_id, 'status': 'processing', 'message': 'Extraction queued'},
        status=drf_status.HTTP_202_ACCEPTED,
    )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_equipment_analysis_status(request, upload_id):
    """GET /api/v1/pid/equipment/status/<upload_id>/"""
    entry = _eq_get_result_entry(upload_id)
    if not entry:
        return Response({'upload_id': upload_id, 'status': 'not_found', 'progress': 0},
                        status=drf_status.HTTP_404_NOT_FOUND)
    s = entry.get('status', 'processing')
    return Response({
        'upload_id': upload_id,
        'status':    s,
        'progress':  entry.get('progress', 100 if s == 'completed' else 50),
        'message':   entry.get('message', entry.get('error',
                         'Extraction complete' if s == 'completed' else 'Processing…')),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_equipment_analysis_results(request, upload_id):
    """GET /api/v1/pid/equipment/results/<upload_id>/"""
    entry = _eq_get_result_entry(upload_id)
    if not entry:
        return Response({'error': 'Results not found — re-upload the file', 'upload_id': upload_id},
                        status=drf_status.HTTP_404_NOT_FOUND)
    if entry.get('status') == 'failed':
        return Response({'error': entry.get('error', 'Extraction failed'), 'upload_id': upload_id},
                        status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)

    config = _load_config()
    return Response({
        'success':     True,
        'upload_id':   upload_id,
        'equipment':   entry.get('equipment', []),
        'total':       entry.get('total', 0),
        'drawing_ref': entry.get('drawing_ref', ''),
        'columns':     [c['label'] for c in config.get('excel_columns', []) if c['key'] != 'sl_no'],
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_equipment_excel(request, upload_id):
    """GET /api/v1/pid/equipment/download-excel/<upload_id>/"""
    entry = _eq_get_result_entry(upload_id)
    if not entry or entry.get('status') != 'completed':
        return Response({'error': 'Results not available - re-upload the file'},
                        status=drf_status.HTTP_404_NOT_FOUND)

    config    = _load_config()
    col_defs  = config.get('excel_columns', [])
    equipment = entry.get('equipment', [])
    drawing   = entry.get('drawing_ref', 'equipment_list')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Equipment List'

        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border  = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )
        alt_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

        headers = [c['label'] for c in col_defs]
        for col_idx, label in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border

        ws.row_dimensions[1].height = 30

        for row_idx, item in enumerate(equipment, 2):
            row_fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col_def in enumerate(col_defs, 1):
                key   = col_def['key']
                value = item.get(key, '')
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value) if value else '-'
                elif value == '' or value is None:
                    value = '-'
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center', wrap_text=(key in ('line_connections', 'description')))
                cell.border    = thin_border
                if row_fill:
                    cell.fill = row_fill

        for col_idx, col_def in enumerate(col_defs, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 18)

        ws.freeze_panes = 'A2'

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = re.sub(r'[^\w\-]', '_', drawing)
        response  = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{EQUIPMENT_EXCEL_FILENAME}"'
        return response

    except ImportError:
        return Response({'error': 'openpyxl is not installed on the server'},
                        status=drf_status.HTTP_501_NOT_IMPLEMENTED)
    except Exception as exc:
        logger.error('[EquipmentList] Excel error: %s', exc, exc_info=True)
        return Response({'error': f'Excel generation failed: {exc}'},
                        status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyze_pid_equipment_batch(request):
    """POST /api/v1/pid/equipment/analyze-batch/
    Accepts multiple files, dispatches extraction to a Celery background task,
    and returns HTTP 202 with upload_id immediately.
    """
    from apps.pid_analysis.tasks import run_equipment_batch_analysis_task   # lazy import

    config  = _load_config()
    ext_cfg = config.get('extraction', {})
    allowed = [e.lower() for e in ext_cfg.get('allowed_extensions', ['pdf'])]
    max_mb  = float(ext_cfg.get('max_file_size_mb', 50))

    files = list(request.FILES.values())
    if not files:
        return Response({'error': 'No files provided', 'success': False},
                        status=drf_status.HTTP_400_BAD_REQUEST)

    # Validate all files up-front before dispatching the task
    for pid_file in files:
        ext = pid_file.name.rsplit('.', 1)[-1].lower()
        if ext not in allowed:
            return Response(
                {'error': f'Unsupported format: .{ext}. Allowed: {", ".join(allowed)}', 'success': False},
                status=drf_status.HTTP_400_BAD_REQUEST,
            )
        if pid_file.size > max_mb * 1024 * 1024:
            return Response(
                {'error': f'{pid_file.name} exceeds {max_mb} MB limit', 'success': False},
                status=drf_status.HTTP_400_BAD_REQUEST,
            )

    upload_id  = f'EQB-{uuid.uuid4().hex[:12].upper()}'

    # Encode each file as base64 for Celery JSON serialisation
    files_data = [
        {'b64': base64.b64encode(pid_file.read()).decode('ascii'), 'filename': pid_file.name}
        for pid_file in files
    ]

    cache.set(
        EQ_RESULT_CACHE_KEY_FMT.format(upload_id=upload_id),
        {'status': 'processing', 'progress': 0, 'message': f'Queued: 0 / {len(files)} files…'},
        EQ_RESULT_CACHE_TTL_S,
    )

    _dispatch_eq_task(run_equipment_batch_analysis_task, upload_id, config,
                      upload_id, files_data)
    logger.info('[EquipmentList Batch] Task dispatched  upload_id=%s  files=%d', upload_id, len(files))

    return Response(
        {'upload_id': upload_id, 'status': 'processing', 'message': f'{len(files)} file(s) queued'},
        status=drf_status.HTTP_202_ACCEPTED,
    )
