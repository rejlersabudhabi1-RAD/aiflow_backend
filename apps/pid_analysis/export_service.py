"""
P&ID Report Export Service
Generate professional reports in PDF, Excel, and CSV formats with customizable branding
"""
from django.http import HttpResponse
from django.conf import settings
from datetime import datetime
import csv
import io


class PIDReportExportService:
    """Service for exporting P&ID reports in various formats"""
    
    def __init__(self):
        """Initialize with soft-coded configuration from settings"""
        # Company Branding (soft-coded)
        self.company_name = getattr(settings, 'REPORT_COMPANY_NAME', 'REJLERS ABU DHABI')
        self.company_subtitle = getattr(settings, 'REPORT_COMPANY_SUBTITLE', 'Engineering & Design Consultancy')
        self.company_website = getattr(settings, 'REPORT_COMPANY_WEBSITE', 'www.rejlers.com/ae')
        
        # Report Colors (soft-coded, add # prefix for hex colors)
        primary = getattr(settings, 'REPORT_PRIMARY_COLOR', '003366')
        secondary = getattr(settings, 'REPORT_SECONDARY_COLOR', 'FFA500')
        text = getattr(settings, 'REPORT_TEXT_COLOR', '333333')
        border = getattr(settings, 'REPORT_BORDER_COLOR', 'CCCCCC')
        
        self.colors = {
            'primary': f'#{primary}' if not primary.startswith('#') else primary,
            'secondary': f'#{secondary}' if not secondary.startswith('#') else secondary,
            'text': f'#{text}' if not text.startswith('#') else text,
            'border': f'#{border}' if not border.startswith('#') else border
        }
        
        # Report Template Settings (soft-coded)
        self.report_title = getattr(settings, 'REPORT_TITLE', 'P&ID DESIGN VERIFICATION REPORT')
        self.footer_text = getattr(settings, 'REPORT_FOOTER_TEXT', 'CONFIDENTIAL ENGINEERING DOCUMENT')
        self.footer_note = getattr(settings, 'REPORT_FOOTER_NOTE_FORMATTED', f'This document is the property of {self.company_name}. Unauthorized distribution is prohibited.')
    
    def export_pdf(self, drawing):
        """Export report as PDF with customizable branding"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        
        buffer = io.BytesIO()
        
        # Create PDF document (landscape for better table display)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2.5*cm,
            bottomMargin=2.5*cm,
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(self.colors['primary']),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor(self.colors['primary']),
            spaceAfter=12,
            alignment=TA_LEFT,
            fontName='Helvetica-Bold'
        )
        
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor(self.colors['primary']),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Header
        header_data = [
            [Paragraph(f'<b>{self.company_name}</b>', header_style)],
            [Paragraph(self.company_subtitle, styles['Normal'])],
            [Paragraph(self.company_website, styles['Normal'])]
        ]
        header_table = Table(header_data, colWidths=[landscape(A4)[0] - 4*cm])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 16),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(self.colors['primary'])),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Title
        elements.append(Paragraph(self.report_title, title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Drawing Information
        elements.append(Paragraph('DRAWING INFORMATION', subtitle_style))
        
        drawing_info = [
            ['Drawing Number:', drawing.drawing_number or 'N/A'],
            ['Drawing Title:', drawing.drawing_title or 'N/A'],
            ['Revision:', drawing.revision or 'N/A'],
            ['Project Name:', drawing.project_name or 'N/A'],
            ['Analysis Date:', drawing.analysis_completed_at.strftime('%d-%b-%Y %H:%M') if drawing.analysis_completed_at else 'N/A'],
            ['Report Generated:', datetime.now().strftime('%d-%b-%Y %H:%M')],
        ]
        
        info_table = Table(drawing_info, colWidths=[5*cm, 15*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(self.colors['border'])),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor(self.colors['primary'])),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.8*cm))
        
        # Summary Statistics
        report = drawing.analysis_report
        elements.append(Paragraph('ANALYSIS SUMMARY', subtitle_style))
        
        summary_data = [
            ['Total Issues', 'Pending', 'Approved', 'Ignored'],
            [str(report.total_issues), str(report.pending_count), str(report.approved_count), str(report.ignored_count)]
        ]
        
        summary_table = Table(summary_data, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(self.colors['border'])),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 1*cm))
        
        # Page break before issues table
        elements.append(PageBreak())
        
        # Issues Table
        elements.append(Paragraph('DETAILED ISSUES & OBSERVATIONS', subtitle_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Table headers
        issues_data = [
            ['#', 'P&ID Ref', 'Issue Observed', 'Action Required', 'Severity', 'Status']
        ]
        
        # Add issues with comprehensive fallback strategy
        issues_found = False
        
        # Method 1: Try database issues first
        db_issues = list(report.issues.all().order_by('serial_number'))
        if db_issues:
            for issue in db_issues:
                issues_data.append([
                    str(issue.serial_number),
                    issue.pid_reference[:30],
                    issue.issue_observed[:80],
                    issue.action_required[:80],
                    issue.severity.upper(),
                    issue.status.upper()
                ])
            issues_found = True
        
        # Method 2: Try report_data JSON if no database issues
        if not issues_found and hasattr(report, 'report_data') and isinstance(report.report_data, dict):
            # Try multiple possible keys
            possible_keys = ['issues', 'identified_issues', 'analysis_issues', 'pid_issues']
            for key in possible_keys:
                if key in report.report_data and isinstance(report.report_data[key], list):
                    for issue in report.report_data[key]:
                        if isinstance(issue, dict):
                            issues_data.append([
                                str(issue.get('serial_number', '')),
                                str(issue.get('pid_reference', ''))[:30],
                                str(issue.get('issue_observed', ''))[:80],
                                str(issue.get('action_required', ''))[:80],
                                str(issue.get('severity', '')).upper(),
                                str(issue.get('status', '')).upper()
                            ])
                    issues_found = True
                    break
            
            # Try nested structure
            if not issues_found and 'analysis_result' in report.report_data:
                analysis_result = report.report_data['analysis_result']
                if isinstance(analysis_result, dict) and 'issues' in analysis_result:
                    for issue in analysis_result['issues']:
                        if isinstance(issue, dict):
                            issues_data.append([
                                str(issue.get('serial_number', '')),
                                str(issue.get('pid_reference', ''))[:30],
                                str(issue.get('issue_observed', ''))[:80],
                                str(issue.get('action_required', ''))[:80],
                                str(issue.get('severity', '')).upper(),
                                str(issue.get('status', '')).upper()
                            ])
                    issues_found = True
        
        # Method 3: Generate placeholder if report claims issues exist
        if not issues_found and report.total_issues > 0:
            for i in range(min(report.total_issues, 10)):
                issues_data.append([
                    str(i + 1),
                    f'REF-{i+1:03d}',
                    'Issue data not found in serialization',
                    'Investigate backend data pipeline',
                    'OBSERVATION',
                    'PENDING'
                ])
        
        issues_table = Table(issues_data, colWidths=[1.5*cm, 4*cm, 8*cm, 8*cm, 2.5*cm, 2.5*cm])
        issues_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(self.colors['border'])),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (4, 1), (5, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
        ]))
        elements.append(issues_table)
        
        # Specification Breaks Section
        if hasattr(report, 'report_data') and isinstance(report.report_data, dict):
            spec_breaks = report.report_data.get('specification_breaks', [])
            if spec_breaks:
                elements.append(PageBreak())
                elements.append(Paragraph('SPECIFICATION BREAKS', subtitle_style))
                elements.append(Spacer(1, 0.3*cm))
                
                spec_breaks_data = [
                    ['ID', 'Location', 'Upstream Spec', 'Downstream Spec', 'Reason', 'Marked']
                ]
                
                for sb in spec_breaks:
                    upstream = f"{sb.get('upstream_spec', {}).get('material_spec', 'N/A')}\n{sb.get('upstream_spec', {}).get('pressure_class', 'N/A')}"
                    downstream = f"{sb.get('downstream_spec', {}).get('material_spec', 'N/A')}\n{sb.get('downstream_spec', {}).get('pressure_class', 'N/A')}"
                    
                    spec_breaks_data.append([
                        sb.get('spec_break_id', '')[:10],
                        sb.get('location', '')[:40],
                        upstream,
                        downstream,
                        sb.get('reason_for_break', '')[:50],
                        '✓' if sb.get('break_properly_marked') == 'Yes' else '✗'
                    ])
                
                spec_table = Table(spec_breaks_data, colWidths=[2*cm, 6*cm, 4*cm, 4*cm, 6*cm, 1.5*cm])
                spec_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(self.colors['border'])),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (5, 1), (5, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAF5FF')]),
                ]))
                elements.append(spec_table)
        
        # Footer
        elements.append(Spacer(1, 1*cm))
        footer_text = f"<b>{self.footer_text}</b><br/>{self.footer_note}<br/><i>Generated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}</i>"
        elements.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#666666'))))
        
        # Build PDF
        doc.build(elements)
        
        # FileResponse
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="PID_Analysis_{drawing.drawing_number or "Report"}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        return response
    
    def export_excel(self, drawing):
        """Export report as Excel with customizable branding"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse(
                "openpyxl not installed. Please install it to use Excel export.",
                status=500
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&ID Analysis Report"
        
        # Styles
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        
        title_font = Font(name='Arial', size=20, bold=True, color='003366')
        subtitle_font = Font(name='Arial', size=12, bold=True, color='003366')
        normal_font = Font(name='Arial', size=10)
        bold_font = Font(name='Arial', size=10, bold=True)
        
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Company Header
        ws.merge_cells('A1:F1')
        ws['A1'] = self.company_name
        ws['A1'].font = Font(name='Arial', size=18, bold=True, color='003366')
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:F2')
        ws['A2'] = self.company_subtitle
        ws['A2'].font = Font(name='Arial', size=10, color='666666')
        ws['A2'].alignment = center_alignment
        
        ws.merge_cells('A3:F3')
        ws['A3'] = self.company_website
        ws['A3'].font = Font(name='Arial', size=9, color='0066CC')
        ws['A3'].alignment = center_alignment
        
        # Title
        ws.merge_cells('A5:F5')
        ws['A5'] = self.report_title
        ws['A5'].font = title_font
        ws['A5'].alignment = center_alignment
        
        # Drawing Information
        row = 7
        ws[f'A{row}'] = 'DRAWING INFORMATION'
        ws[f'A{row}'].font = subtitle_font
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        info_data = [
            ['Drawing Number:', drawing.drawing_number or 'N/A'],
            ['Drawing Title:', drawing.drawing_title or 'N/A'],
            ['Revision:', drawing.revision or 'N/A'],
            ['Project Name:', drawing.project_name or 'N/A'],
            ['Analysis Date:', drawing.analysis_completed_at.strftime('%d-%b-%Y %H:%M') if drawing.analysis_completed_at else 'N/A'],
            ['Report Generated:', datetime.now().strftime('%d-%b-%Y %H:%M')],
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = bold_font
            ws[f'A{row}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = normal_font
            ws.merge_cells(f'B{row}:F{row}')
            row += 1
        
        # Summary
        row += 2
        ws[f'A{row}'] = 'ANALYSIS SUMMARY'
        ws[f'A{row}'].font = subtitle_font
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        report = drawing.analysis_report
        summary_headers = ['Total Issues', 'Pending', 'Approved', 'Ignored']
        summary_values = [report.total_issues, report.pending_count, report.approved_count, report.ignored_count]
        
        for col, (header, value) in enumerate(zip(summary_headers, summary_values), start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            
            value_cell = ws.cell(row=row+1, column=col)
            value_cell.value = value
            value_cell.font = Font(name='Arial', size=12, bold=True)
            value_cell.alignment = center_alignment
            value_cell.border = border
        
        # Issues Table
        row += 4
        ws[f'A{row}'] = 'DETAILED ISSUES & OBSERVATIONS'
        ws[f'A{row}'].font = subtitle_font
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        headers = ['#', 'P&ID Reference', 'Issue Observed', 'Action Required', 'Severity', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Add issues with comprehensive fallback strategy
        issues_found = False
        
        # Method 1: Try database issues first
        db_issues = list(report.issues.all().order_by('serial_number'))
        if db_issues:
            for issue in db_issues:
                row += 1
                data = [
                    issue.serial_number,
                    issue.pid_reference,
                    issue.issue_observed,
                    issue.action_required,
                    issue.severity.upper(),
                    issue.status.upper()
                ]
                
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.font = normal_font
                    cell.alignment = left_alignment if col in [2, 3, 4] else center_alignment
                    cell.border = border
            issues_found = True
        
        # Method 2: Try report_data JSON if no database issues
        if not issues_found and hasattr(report, 'report_data') and isinstance(report.report_data, dict):
            # Try multiple possible keys
            possible_keys = ['issues', 'identified_issues', 'analysis_issues', 'pid_issues']
            for key in possible_keys:
                if key in report.report_data and isinstance(report.report_data[key], list):
                    for issue in report.report_data[key]:
                        if isinstance(issue, dict):
                            row += 1
                            data = [
                                issue.get('serial_number', ''),
                                issue.get('pid_reference', ''),
                                issue.get('issue_observed', ''),
                                issue.get('action_required', ''),
                                str(issue.get('severity', '')).upper(),
                                str(issue.get('status', '')).upper()
                            ]
                            
                            for col, value in enumerate(data, start=1):
                                cell = ws.cell(row=row, column=col)
                                cell.value = value
                                cell.font = normal_font
                                cell.alignment = left_alignment if col in [2, 3, 4] else center_alignment
                                cell.border = border
                    issues_found = True
                    break
        
        # Method 3: Generate placeholder if report claims issues exist
        if not issues_found and report.total_issues > 0:
            for i in range(min(report.total_issues, 10)):
                row += 1
                data = [
                    i + 1,
                    f'REF-{i+1:03d}',
                    'Issue data not found in serialization',
                    'Investigate backend data pipeline',
                    'OBSERVATION',
                    'PENDING'
                ]
                
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.font = normal_font
                    cell.alignment = left_alignment if col in [2, 3, 4] else center_alignment
                    cell.border = border
        
        # Specification Breaks Section
        if hasattr(report, 'report_data') and isinstance(report.report_data, dict):
            spec_breaks = report.report_data.get('specification_breaks', [])
            if spec_breaks:
                row += 4
                ws[f'A{row}'] = 'SPECIFICATION BREAKS'
                ws[f'A{row}'].font = subtitle_font
                ws.merge_cells(f'A{row}:F{row}')
                
                row += 1
                spec_headers = ['ID', 'Location', 'Upstream Spec', 'Downstream Spec', 'Reason', 'Marked']
                for col, header in enumerate(spec_headers, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
                    cell.alignment = center_alignment
                    cell.border = border
                
                for sb in spec_breaks:
                    row += 1
                    upstream = f"{sb.get('upstream_spec', {}).get('material_spec', 'N/A')} {sb.get('upstream_spec', {}).get('pressure_class', '')}"
                    downstream = f"{sb.get('downstream_spec', {}).get('material_spec', 'N/A')} {sb.get('downstream_spec', {}).get('pressure_class', '')}"
                    
                    spec_data = [
                        sb.get('spec_break_id', ''),
                        sb.get('location', ''),
                        upstream,
                        downstream,
                        sb.get('reason_for_break', ''),
                        '✓' if sb.get('break_properly_marked') == 'Yes' else '✗'
                    ]
                    
                    for col, value in enumerate(spec_data, start=1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        cell.font = normal_font
                        cell.alignment = left_alignment if col in [2, 3, 4, 5] else center_alignment
                        cell.border = border
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Footer
        row += 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f'{self.footer_text} - Generated: {datetime.now().strftime("%d-%b-%Y %H:%M:%S")}'
        ws[f'A{row}'].font = Font(name='Arial', size=8, italic=True, color='666666')
        ws[f'A{row}'].alignment = center_alignment
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="PID_Analysis_{drawing.drawing_number or "Report"}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    def export_csv(self, drawing):
        """Export report as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="PID_Analysis_{drawing.drawing_number or "Report"}_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        
        # Header
        writer.writerow([f'{self.company_name} - {self.report_title}'])
        writer.writerow([self.company_website])
        writer.writerow([])
        
        # Drawing Information
        writer.writerow(['DRAWING INFORMATION'])
        writer.writerow(['Drawing Number', drawing.drawing_number or 'N/A'])
        writer.writerow(['Drawing Title', drawing.drawing_title or 'N/A'])
        writer.writerow(['Revision', drawing.revision or 'N/A'])
        writer.writerow(['Project Name', drawing.project_name or 'N/A'])
        writer.writerow(['Analysis Date', drawing.analysis_completed_at.strftime('%d-%b-%Y %H:%M') if drawing.analysis_completed_at else 'N/A'])
        writer.writerow(['Report Generated', datetime.now().strftime('%d-%b-%Y %H:%M')])
        writer.writerow([])
        
        # Summary
        report = drawing.analysis_report
        writer.writerow(['ANALYSIS SUMMARY'])
        writer.writerow(['Total Issues', 'Pending', 'Approved', 'Ignored'])
        writer.writerow([report.total_issues, report.pending_count, report.approved_count, report.ignored_count])
        writer.writerow([])
        
        # Issues
        writer.writerow(['DETAILED ISSUES & OBSERVATIONS'])
        writer.writerow(['#', 'P&ID Reference', 'Category', 'Issue Observed', 'Action Required', 'Severity', 'Status', 'Approval', 'Remark'])
        
        # Add issues with comprehensive fallback strategy
        issues_found = False
        
        # Method 1: Try database issues first
        db_issues = list(report.issues.all().order_by('serial_number'))
        if db_issues:
            for issue in db_issues:
                writer.writerow([
                    issue.serial_number,
                    issue.pid_reference,
                    issue.category,
                    issue.issue_observed,
                    issue.action_required,
                    issue.severity.upper(),
                    issue.status.upper(),
                    issue.approval,
                    issue.remark
                ])
            issues_found = True
        
        # Method 2: Try report_data JSON if no database issues
        if not issues_found and hasattr(report, 'report_data') and isinstance(report.report_data, dict):
            # Try multiple possible keys
            possible_keys = ['issues', 'identified_issues', 'analysis_issues', 'pid_issues']
            for key in possible_keys:
                if key in report.report_data and isinstance(report.report_data[key], list):
                    for issue in report.report_data[key]:
                        if isinstance(issue, dict):
                            writer.writerow([
                                issue.get('serial_number', ''),
                                issue.get('pid_reference', ''),
                                issue.get('category', ''),
                                issue.get('issue_observed', ''),
                                issue.get('action_required', ''),
                                str(issue.get('severity', '')).upper(),
                                str(issue.get('status', '')).upper(),
                                issue.get('approval', ''),
                                issue.get('remark', '')
                            ])
                    issues_found = True
                    break
        
        # Method 3: Generate placeholder if report claims issues exist
        if not issues_found and report.total_issues > 0:
            for i in range(min(report.total_issues, 10)):
                writer.writerow([
                    i + 1,
                    f'REF-{i+1:03d}',
                    'OBSERVATION',
                    'Issue data not found in serialization',
                    'Investigate backend data pipeline',
                    'OBSERVATION',
                    'PENDING',
                    'N/A',
                    'Data retrieval issue'
                ])
        
        writer.writerow([])
        
        # Specification Breaks
        if hasattr(report, 'report_data') and isinstance(report.report_data, dict):
            spec_breaks = report.report_data.get('specification_breaks', [])
            if spec_breaks:
                writer.writerow(['SPECIFICATION BREAKS'])
                writer.writerow(['ID', 'Location', 'Upstream Spec', 'Downstream Spec', 'Reason', 'Properly Marked', 'Transition Required', 'Cost Impact'])
                
                for sb in spec_breaks:
                    upstream = f"{sb.get('upstream_spec', {}).get('material_spec', 'N/A')} {sb.get('upstream_spec', {}).get('pressure_class', '')}"
                    downstream = f"{sb.get('downstream_spec', {}).get('material_spec', 'N/A')} {sb.get('downstream_spec', {}).get('pressure_class', '')}"
                    
                    writer.writerow([
                        sb.get('spec_break_id', ''),
                        sb.get('location', ''),
                        upstream,
                        downstream,
                        sb.get('reason_for_break', ''),
                        sb.get('break_properly_marked', 'N/A'),
                        sb.get('transition_piece_required', 'N/A'),
                        sb.get('cost_impact', 'N/A')
                    ])
                
                writer.writerow([])
        
        writer.writerow([f'{self.footer_text} - Generated: {datetime.now().strftime("%d-%b-%Y %H:%M:%S")}'])
        
        return response
