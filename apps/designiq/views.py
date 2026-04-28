"""
DesignIQ Views - AI-Powered Design Analysis API
Intelligent design verification, optimization, and recommendations
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Q, Count
from django.utils import timezone
from django.http import HttpResponse
from django.conf import settings
import logging
import json
import os
import threading
import uuid
import concurrent.futures

from .models import DesignProject, DesignAnalysis, DesignOptimization, DesignTemplate, EngineeringListItem, LIST_TYPES
from .s3_utils import s3_storage  # S3 document storage
from .serializers import (
    DesignProjectListSerializer, DesignProjectDetailSerializer,
    DesignProjectCreateSerializer, DesignAnalysisSerializer,
    DesignOptimizationSerializer, DesignTemplateSerializer,
    DesignAnalysisCreateSerializer, EngineeringListItemSerializer,
    EngineeringListItemListSerializer, ListTypeConfigSerializer
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Thread-based extraction fallback (used when Celery broker is unavailable)
# Progress is written to /tmp/base_extraction_{task_id}.json so every
# Gunicorn worker on the same container can read it during polling.
# ---------------------------------------------------------------------------
def _extract_pid_no_per_page(file_path: str) -> dict:
    """
    Build a {page_index (1-based): dwg_no} map for a P&ID PDF.

    Reuses the title-block drawing-number extractor from apps.pid_analysis
    (text-based fallback + PyMuPDF coord-based), called per page so each line
    row can be stamped with the P&ID drawing number it belongs to. Returns
    an empty dict on any failure — caller falls back gracefully.

    Soft-coded: no new knobs here; all behaviour flows from the existing
    title-block constants in equipment_analysis_views.
    """
    page_map: dict = {}
    try:
        import fitz as _fitz  # PyMuPDF — already a project dependency
        from apps.pid_analysis.equipment_analysis_views import (
            _extract_titleblock_dwg_no,
        )
        doc = _fitz.open(file_path)
        try:
            for _pg_idx in range(len(doc)):
                try:
                    _text = doc[_pg_idx].get_text() or ''
                    _dwg = _extract_titleblock_dwg_no(_text) or ''
                except Exception:
                    _dwg = ''
                page_map[_pg_idx + 1] = _dwg  # 1-based to match pid_ocr_extractor
        finally:
            doc.close()
    except Exception as _e:
        logger.warning(f'[base_extract_thread] P&ID No. per-page extraction failed: {_e}')
    return page_map


def _run_base_extraction_in_thread(task_id, file_path, filename, include_area, format_type, legend_file_path=None):
    """Spawn a daemon thread that runs P&ID OCR and writes progress to /tmp/."""
    progress_file = f'/tmp/base_extraction_{task_id}.json'

    def _write(state, percent, status_msg, result=None, error=None, **extra):
        payload = {'task_id': task_id, 'state': state, 'percent': percent, 'status': status_msg}
        if result is not None:
            payload['result'] = result
        if error is not None:
            payload['error'] = error
        if extra:
            payload.update(extra)
        try:
            with open(progress_file, 'w') as fh:
                json.dump(payload, fh)
        except Exception as we:
            logger.warning(f'[base_extract_thread] progress write failed: {we}')

    # Soft-coded progress band for the per-page loop (same shape as Celery task).
    _PAGE_PROG_START = 20
    _PAGE_PROG_END   = 80

    def _page_progress(page_num, total_pages, lines_so_far, phase):
        span = _PAGE_PROG_END - _PAGE_PROG_START
        frac = max(0.0, min(1.0, (page_num - 1) / max(1, total_pages)))
        percent = int(_PAGE_PROG_START + span * frac)
        _write(
            'PROGRESS', percent,
            f'Page {page_num}/{total_pages} — {lines_so_far} lines extracted so far…',
            current_page=page_num, total_pages=total_pages,
            lines_so_far=lines_so_far, phase=phase,
        )

    def _run():
        try:
            logger.info(f'[base_extract_thread] START task_id={task_id} file={filename}')
            from apps.designiq.pid_ocr_extractor_v2 import PIDLineExtractorV2
            _write('PROGRESS', 5, 'Initializing OCR engine…')
            extractor = PIDLineExtractorV2()

            # Load legend knowledge if a legend file was uploaded
            service_codes = {}
            insulation_codes = {}
            if legend_file_path and os.path.exists(legend_file_path):
                try:
                    from apps.pid_verification.services.legend_knowledge import (
                        extract_text_from_pdf, parse_legend_knowledge,
                    )
                    _write('PROGRESS', 10, 'Parsing legend sheet…')
                    legend_text = extract_text_from_pdf(legend_file_path)
                    legend_data = parse_legend_knowledge(legend_text)
                    service_codes = legend_data.get('service_codes', {})
                    insulation_codes = legend_data.get('insulation_codes', {})
                    logger.info(
                        f'[base_extract_thread] legend: {len(service_codes)} service codes, '
                        f'{len(insulation_codes)} insulation codes'
                    )
                except Exception as le:
                    logger.warning(f'[base_extract_thread] legend parse failed: {le}')
                finally:
                    try:
                        os.unlink(legend_file_path)
                    except Exception:
                        pass

            _write('PROGRESS', 15, f'Running extraction on {filename}…')
            extracted_lines = extractor.extract_from_pdf(
                file_path,
                include_area=include_area,
                format_type=format_type,
                progress_callback=_page_progress,
            )
            _write('PROGRESS', 85, f'OCR complete: {len(extracted_lines)} lines found. Formatting…')
            # ── Per-page drawing-number (P&ID No.) extraction ────────────────
            # Soft-coded: reuses pid_analysis title-block logic unchanged.
            # Builds a {page_index_1based: dwg_no} dict used below. Any failure
            # is non-fatal — the pid_no field simply remains empty.
            _page_to_dwg = _extract_pid_no_per_page(file_path)
            _doc_dwg = next((v for v in _page_to_dwg.values() if v), '')
            logger.info(
                f'[base_extract_thread] P&ID No. map: {_page_to_dwg} (doc fallback={_doc_dwg!r})'
            )

            # ── Soft-coded breaker / page-connector inference ────────────────
            # Fills empty from_line / to_line by spatial proximity to breaker
            # tags (E3-SP-XXXX, TP-..., IP-..., etc.). Patterns and thresholds
            # live in apps/designiq/breaker_inference.py — edit there freely.
            # Pure post-processing; never overwrites existing values; runs only
            # while the temp PDF is still on disk.
            try:
                from apps.designiq.breaker_inference import infer_breakers_for_lines
                _write('PROGRESS', 90, 'Inferring From/To from page-connector breakers…')
                infer_breakers_for_lines(extracted_lines, file_path)
            except Exception as _be:
                logger.warning(f'[base_extract_thread] breaker inference skipped: {_be}')

            # Build 10-column output structure with EXPLICIT field mapping
            # Columns: Original Detection, Size, Fluid Code, Fluid Description,
            #          Sequence No, Piping Spec, Piping Spec Desc, Dept Deviation,
            #          Insulation, Insulation Description, P&ID No.
            base_data = []
            for line in extracted_lines:
                fluid = line.get('fluid_code', '')
                insul = line.get('insulation', '')
                piping_spec = line.get('piping_spec', line.get('pipr_class', ''))
                _pg = line.get('page')
                _pid_no = _page_to_dwg.get(_pg, '') or _doc_dwg
                base_data.append({
                    'original_detection':  line.get('original_detection', line.get('line_number', '')),
                    'size':                line.get('size', ''),
                    'fluid_code':          fluid,
                    'fluid_description':   service_codes.get(fluid.upper(), ''),
                    'sequence_no':         line.get('sequence_no', ''),
                    'piping_spec':         piping_spec,
                    'dept_deviation':      line.get('dept_deviation', ''),
                    'insulation':          insul,
                    'insulation_desc':     insulation_codes.get(insul.upper(), ''),
                    # From / To — produced by spatial / vision / geometric /
                    # breaker-inference passes upstream. Frontend column resolver
                    # has fallback keys so any of these surfaces correctly.
                    'from_line':           line.get('from_line', ''),
                    'to_line':             line.get('to_line', ''),
                    'from_equipment':      line.get('from_equipment', ''),
                    'to_equipment':        line.get('to_equipment', ''),
                    'pid_no':              _pid_no,
                })
            logger.info(f'[base_extract_thread] Formatted {len(base_data)} rows with 10 columns (pid_no enriched)')
            if os.path.exists(file_path):
                try:
                    os.unlink(file_path)
                except Exception as de:
                    logger.warning(f'[base_extract_thread] temp file delete failed: {de}')
            result = {
                'success': True,
                'total_lines': len(base_data),
                'data': base_data,
                'columns': 11,
                'message': f'Successfully extracted {len(base_data)} lines from {filename}',
            }
            _write('SUCCESS', 100, 'Extraction complete!', result=result)
            logger.info(f'[base_extract_thread] DONE task_id={task_id} lines={len(base_data)}')
        except Exception as exc:
            logger.error(f'[base_extract_thread] FAILED task_id={task_id}: {exc}', exc_info=True)
            if os.path.exists(file_path):
                try:
                    os.unlink(file_path)
                except Exception:
                    pass
            _write('FAILURE', 0, 'Extraction failed', error=str(exc))

    t = threading.Thread(target=_run, name=f'base_extract_{task_id}', daemon=True)
    t.start()

# Global singleton for PIDLineExtractorV2 to avoid reinitialization
_pid_extractor_instance = None
_pid_extractor_lock = None

def get_pid_extractor():
    """Get or create the global PIDLineExtractorV2 instance"""
    global _pid_extractor_instance, _pid_extractor_lock
    
    if _pid_extractor_lock is None:
        import threading
        _pid_extractor_lock = threading.Lock()
    
    if _pid_extractor_instance is None:
        with _pid_extractor_lock:
            # Double-check pattern
            if _pid_extractor_instance is None:
                logger.info("🚀 Initializing PIDLineExtractorV2 (one-time initialization)...")
                from apps.designiq.pid_ocr_extractor_v2 import PIDLineExtractorV2
                _pid_extractor_instance = PIDLineExtractorV2()
                logger.info("✅ PIDLineExtractorV2 ready for all requests")
    
    return _pid_extractor_instance


def load_module_config():
    """Load DesignIQ module configuration from JSON file"""
    config_path = os.path.join(
        settings.BASE_DIR,
        'apps', 'designiq', 'config', 'module_config.json'
    )
    try:
        with open(config_path, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Failed to load module config: {e}")
        # Return default configuration if file not found
        return {
            "design_modules": {"enabled_modules": [], "disabled_modules": []},
            "list_types": {"enabled": []},
            "features": {
                "show_design_type_cards": False,
                "show_engineering_lists": True
            }
        }


class DesignProjectViewSet(viewsets.ModelViewSet):
    """
    ViewSet for DesignIQ Projects
    Handles design project creation, analysis, and AI-powered insights
    """
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def get_queryset(self):
        """Get projects for current user with optional filtering"""
        queryset = DesignProject.objects.all()
        
        # Filter by user unless staff
        if not self.request.user.is_staff:
            queryset = queryset.filter(created_by=self.request.user)
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by design type
        design_type = self.request.query_params.get('design_type')
        if design_type:
            queryset = queryset.filter(design_type=design_type)
        
        # Search
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(project_name__icontains=search) |
                Q(description__icontains=search) |
                Q(organization__icontains=search)
            )
        
        return queryset.select_related('created_by').prefetch_related('analyses', 'optimizations')
    
    def get_serializer_class(self):
        """Use different serializers for different actions"""
        if self.action == 'retrieve':
            return DesignProjectDetailSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return DesignProjectCreateSerializer
        return DesignProjectListSerializer
    
    def perform_create(self, serializer):
        """Create project and set user"""
        project = serializer.save(created_by=self.request.user)
        logger.info(f"[DesignIQ] Project created: {project.id} by {self.request.user.email}")
    
    @action(detail=True, methods=['post'])
    def analyze(self, request, pk=None):
        """
        Trigger AI analysis for a project
        POST /api/v1/designiq/projects/{id}/analyze/
        Body: {
            "parameters": {...},  // Optional analysis parameters
            "force_reanalysis": false  // Re-analyze even if already completed
        }
        """
        project = self.get_object()
        
        if project.status == 'analyzing':
            return Response(
                {"error": "Analysis already in progress"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Update status
            project.status = 'analyzing'
            project.save()
            
            # Here you would trigger your AI analysis
            # For now, we'll return a placeholder response
            # TODO: Integrate with actual AI service
            
            logger.info(f"[DesignIQ] Analysis triggered for project: {project.id}")
            
            return Response({
                "message": "Analysis started successfully",
                "project_id": str(project.id),
                "status": "analyzing"
            })
            
        except Exception as e:
            logger.error(f"[DesignIQ] Analysis error: {str(e)}")
            project.status = 'failed'
            project.error_message = str(e)
            project.save()
            
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """
        Get project summary with statistics
        GET /api/v1/designiq/projects/{id}/summary/
        """
        project = self.get_object()
        
        analyses_stats = project.analyses.aggregate(
            total=Count('id'),
            critical=Count('id', filter=Q(severity='critical')),
            high=Count('id', filter=Q(severity='high')),
            resolved=Count('id', filter=Q(is_resolved=True))
        )
        
        optimizations_stats = project.optimizations.aggregate(
            total=Count('id'),
            high_impact=Count('id', filter=Q(impact='high')),
            implemented=Count('id', filter=Q(is_implemented=True))
        )
        
        return Response({
            "project": DesignProjectDetailSerializer(project).data,
            "analyses": analyses_stats,
            "optimizations": optimizations_stats,
            "summary": {
                "total_findings": analyses_stats['total'],
                "critical_issues": analyses_stats['critical'],
                "high_priority_issues": analyses_stats['high'],
                "resolution_rate": (analyses_stats['resolved'] / analyses_stats['total'] * 100) if analyses_stats['total'] > 0 else 0,
                "optimization_count": optimizations_stats['total'],
                "implemented_optimizations": optimizations_stats['implemented']
            }
        })
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """
        Get dashboard statistics for all user projects
        GET /api/v1/designiq/projects/dashboard/
        """
        queryset = self.get_queryset()
        
        stats = {
            "total_projects": queryset.count(),
            "by_status": {
                "draft": queryset.filter(status='draft').count(),
                "analyzing": queryset.filter(status='analyzing').count(),
                "completed": queryset.filter(status='completed').count(),
                "failed": queryset.filter(status='failed').count(),
            },
            "by_design_type": {},
            "recent_projects": DesignProjectListSerializer(
                queryset.order_by('-created_at')[:5],
                many=True
            ).data
        }
        
        # Get counts by design type
        for choice_value, choice_label in DesignProject.DESIGN_TYPE_CHOICES:
            stats['by_design_type'][choice_value] = queryset.filter(design_type=choice_value).count()
        
        return Response(stats)
    
    @action(detail=False, methods=['get'], url_path='module-config')
    def module_config(self, request):
        """
        Get DesignIQ module configuration (soft-coded)
        GET /api/v1/designiq/projects/module-config/
        
        Returns configuration for which modules and features are enabled/disabled
        """
        config = load_module_config()
        return Response(config)


class DesignAnalysisViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Design Analyses
    """
    
    permission_classes = [IsAuthenticated]
    serializer_class = DesignAnalysisSerializer
    
    def get_queryset(self):
        """Get analyses for user's projects"""
        if self.request.user.is_staff:
            return DesignAnalysis.objects.all().select_related('project', 'resolved_by')
        
        return DesignAnalysis.objects.filter(
            project__created_by=self.request.user
        ).select_related('project', 'resolved_by')
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return DesignAnalysisCreateSerializer
        return DesignAnalysisSerializer
    
    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """
        Mark analysis as resolved
        POST /api/v1/designiq/analyses/{id}/resolve/
        Body: {"resolution_notes": "Fixed by..."}
        """
        analysis = self.get_object()
        
        analysis.is_resolved = True
        analysis.resolved_by = request.user
        analysis.resolved_at = timezone.now()
        analysis.resolution_notes = request.data.get('resolution_notes', '')
        analysis.save()
        
        return Response(DesignAnalysisSerializer(analysis).data)


class DesignOptimizationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Design Optimizations
    """
    
    permission_classes = [IsAuthenticated]
    serializer_class = DesignOptimizationSerializer
    
    def get_queryset(self):
        """Get optimizations for user's projects"""
        if self.request.user.is_staff:
            return DesignOptimization.objects.all().select_related('project', 'implemented_by')
        
        return DesignOptimization.objects.filter(
            project__created_by=self.request.user
        ).select_related('project', 'implemented_by')
    
    @action(detail=True, methods=['post'])
    def implement(self, request, pk=None):
        """
        Mark optimization as implemented
        POST /api/v1/designiq/optimizations/{id}/implement/
        Body: {"implementation_notes": "..."}
        """
        optimization = self.get_object()
        
        optimization.is_implemented = True
        optimization.implemented_by = request.user
        optimization.implemented_at = timezone.now()
        
        if 'implementation_notes' in request.data:
            optimization.implementation_notes = request.data['implementation_notes']
        
        optimization.save()
        
        return Response(DesignOptimizationSerializer(optimization).data)


class DesignTemplateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Design Templates
    """
    
    permission_classes = [IsAuthenticated]
    serializer_class = DesignTemplateSerializer
    
    def get_queryset(self):
        """Get public templates and user's private templates"""
        queryset = DesignTemplate.objects.filter(
            Q(is_public=True) | Q(created_by=self.request.user)
        ).select_related('created_by')
        
        design_type = self.request.query_params.get('design_type')
        if design_type:
            queryset = queryset.filter(design_type=design_type)
        
        return queryset.order_by('-usage_count', 'name')
    
    @action(detail=True, methods=['post'])
    def use_template(self, request, pk=None):
        """
        Create a new project from template
        POST /api/v1/designiq/templates/{id}/use_template/
        Body: {
            "project_name": "...",
            "parameters": {...}
        }
        """
        template = self.get_object()
        
        # Increment usage count
        template.usage_count += 1
        template.save()
        
        # Create project from template
        project_data = {
            "project_name": request.data.get('project_name', f"Project from {template.name}"),
            "design_type": template.design_type,
            "description": f"Created from template: {template.name}",
            "design_parameters": request.data.get('parameters', template.template_data),
            "created_by": request.user,
        }
        
        project = DesignProject.objects.create(**project_data)
        
        return Response({
            "message": "Project created from template",
            "project": DesignProjectDetailSerializer(project).data
        }, status=status.HTTP_201_CREATED)


class EngineeringListItemViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Engineering List Items
    Handles Line List, Equipment List, Tie-In List, and Alarm/Trip List
    """
    
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    
    def get_queryset(self):
        """Get list items with filtering"""
        queryset = EngineeringListItem.objects.select_related('project', 'created_by')
        
        # Filter by list type
        list_type = self.request.query_params.get('list_type')
        if list_type:
            queryset = queryset.filter(list_type=list_type)
        
        # Filter by project
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by validation status
        is_validated = self.request.query_params.get('is_validated')
        if is_validated is not None:
            queryset = queryset.filter(is_validated=is_validated.lower() == 'true')
        
        # Search by item tag or description
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(item_tag__icontains=search) | Q(description__icontains=search)
            )
        
        return queryset.order_by('-created_at')
    
    def get_serializer_class(self):
        """Use different serializers for list vs detail"""
        if self.action == 'list':
            return EngineeringListItemListSerializer
        return EngineeringListItemSerializer
    
    @action(detail=False, methods=['get'])
    def list_types(self, request):
        """Get available list types configuration"""
        list_types_data = [
            {
                'code': code,
                'name': config['name'],
                'icon': config['icon'],
                'description': config['description'],
                'default_fields': config['default_fields']
            }
            for code, config in LIST_TYPES.items()
        ]
        
        serializer = ListTypeConfigSerializer(list_types_data, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get statistics for list items"""
        list_type = request.query_params.get('list_type')
        project_id = request.query_params.get('project')
        
        queryset = self.get_queryset()
        if list_type:
            queryset = queryset.filter(list_type=list_type)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        stats = {
            'total': queryset.count(),
            'by_status': {
                'active': queryset.filter(status='active').count(),
                'pending': queryset.filter(status='pending').count(),
                'approved': queryset.filter(status='approved').count(),
                'rejected': queryset.filter(status='rejected').count(),
                'inactive': queryset.filter(status='inactive').count(),
            },
            'validated': queryset.filter(is_validated=True).count(),
            'not_validated': queryset.filter(is_validated=False).count(),
            'by_list_type': {}
        }
        
        # Count by list type
        for code, config in LIST_TYPES.items():
            count = queryset.filter(list_type=code).count()
            if count > 0:
                stats['by_list_type'][code] = {
                    'name': config['name'],
                    'count': count
                }
        
        return Response(stats)
    
    @action(detail=True, methods=['post'])
    def validate_item(self, request, pk=None):
        """Validate a list item"""
        item = self.get_object()
        
        item.is_validated = True
        item.validated_by = request.user
        item.validated_at = timezone.now()
        item.validation_notes = request.data.get('notes', '')
        item.save()
        
        return Response({
            "message": "Item validated successfully",
            "item": EngineeringListItemSerializer(item, context={'request': request}).data
        })
    
    @action(detail=True, methods=['post'])
    def bulk_import(self, request):
        """Bulk import items from CSV/Excel data"""
        items_data = request.data.get('items', [])
        list_type = request.data.get('list_type')
        project_id = request.data.get('project')
        
        if not list_type or list_type not in LIST_TYPES:
            return Response(
                {"error": "Valid list_type is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created_items = []
        errors = []
        
        for idx, item_data in enumerate(items_data):
            try:
                item = EngineeringListItem.objects.create(
                    list_type=list_type,
                    project_id=project_id if project_id else None,
                    item_tag=item_data.get('item_tag', f'ITEM-{idx+1}'),
                    description=item_data.get('description', ''),
                    data=item_data.get('data', {}),
                    status=item_data.get('status', 'active'),
                    created_by=request.user
                )
                created_items.append(item)
            except Exception as e:
                errors.append({
                    'row': idx + 1,
                    'error': str(e),
                    'data': item_data
                })
        
        return Response({
            "message": f"Imported {len(created_items)} items",
            "created": len(created_items),
            "errors": len(errors),
            "error_details": errors if errors else None
        }, status=status.HTTP_201_CREATED if created_items else status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def export(self, request):
        """Export list items to structured format"""
        queryset = self.get_queryset()
        list_type = request.query_params.get('list_type')
        
        if list_type:
            queryset = queryset.filter(list_type=list_type)
        
        serializer = EngineeringListItemSerializer(queryset, many=True, context={'request': request})
        
        return Response({
            "list_type": list_type,
            "count": queryset.count(),
            "items": serializer.data
        })
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_pid(self, request):
        """
        Upload P&ID PDF and queue async OCR processing (OPTIMIZED - NO TIMEOUT)
        
        ENRICHMENT LAYER: Optionally accepts HMB/PMS/NACE documents for smart enrichment
        Base extraction runs first, enrichment layer adds columns if docs provided
        
        This endpoint immediately returns after uploading the file and queuing a Celery task.
        The actual OCR processing happens asynchronously in the background.
        
        Frontend should poll /upload_pid_status/{task_id}/ for progress and results.
        
        Response includes:
        - task_id: Use this to check processing status
        - message: Instructions for checking status
        - estimated_time: Rough estimate based on file size
        """
        pid_file = request.FILES.get('pid_file')
        list_type = request.data.get('list_type', 'line_list')
        
        # ENRICHMENT LAYER: Optional documents (do NOT block processing)
        hmb_file = request.FILES.get('hmb_file')
        pms_file = request.FILES.get('pms_file')
        nace_file = request.FILES.get('nace_file')
        stress_criticality_file = request.FILES.get('stress_criticality_file')  # NEW: 5th document for stress criticality
        
        if not pid_file:
            return Response({
                "error": "No P&ID file provided"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not pid_file.name.endswith('.pdf'):
            return Response({
                "error": "Only PDF files are supported"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if list_type not in LIST_TYPES:
            return Response({
                "error": f"Invalid list_type. Must be one of: {', '.join(LIST_TYPES.keys())}"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            from django.core.files.storage import default_storage
            from .models import DesignProject
            from .tasks import process_pid_upload_async
            from io import BytesIO
            
            # Get or create project
            project, _ = DesignProject.objects.get_or_create(
                project_name="P&ID Upload Project",
                defaults={
                    'created_by': request.user,
                    'design_type': 'pid',
                    'status': 'active'
                }
            )
            
            # Generate unique document ID
            last_doc_id = 0
            existing_items = EngineeringListItem.objects.filter(
                data__has_key='document_id'
            ).order_by('-created_at').first()
            
            if existing_items and existing_items.data.get('document_id'):
                try:
                    doc_id_str = existing_items.data['document_id'].split('-')[0]
                    last_doc_id = int(doc_id_str)
                except (ValueError, IndexError):
                    pass
            
            new_doc_id = last_doc_id + 1
            document_id = f"{new_doc_id:04d}-{pid_file.name}"
            
            logger.info(f"🆔 Generated Document ID: {document_id} (size: {pid_file.size / 1024 / 1024:.2f} MB)")
            
            # Read file content once
            pid_file.seek(0)
            file_content = pid_file.read()
            
            # Upload to S3 (if configured)
            s3_file = BytesIO(file_content)
            s3_result = s3_storage.upload_document(
                file_obj=s3_file,
                document_id=document_id,
                original_filename=pid_file.name
            )
            
            if s3_result['success']:
                logger.info(f"☁️ Uploaded to S3: {s3_result['s3_key']}")
                saved_path = s3_result['s3_key']
                storage_type = 's3'
                s3_url = s3_result['s3_url']
            else:
                logger.warning(f"⚠️ S3 upload failed, using local storage")
                local_file = BytesIO(file_content)
                local_file.seek(0)
                file_path = f"designiq/pid_uploads/{timezone.now().strftime('%Y/%m/%d')}/{document_id}"
                saved_path = default_storage.save(file_path, local_file)
                storage_type = 'local'
                s3_url = None
            
            # Get processing options
            include_area = request.POST.get('include_area', 'false').lower() == 'true'
            format_type = request.POST.get('format_type', 'onshore').lower()
            
            # ENRICHMENT LAYER: Process optional documents
            enrichment_files = {}
            if hmb_file:
                logger.info(f"📊 HMB document attached: {hmb_file.name}")
                hmb_content = hmb_file.read()
                enrichment_files['hmb'] = {
                    'filename': hmb_file.name,
                    'content': hmb_content,
                    'size': len(hmb_content)
                }
            if pms_file:
                logger.info(f"🔧 PMS document attached: {pms_file.name}")
                pms_content = pms_file.read()
                enrichment_files['pms'] = {
                    'filename': pms_file.name,
                    'content': pms_content,
                    'size': len(pms_content)
                }
            if nace_file:
                logger.info(f"⚗️ NACE document attached: {nace_file.name}")
                nace_content = nace_file.read()
                enrichment_files['nace'] = {
                    'filename': nace_file.name,
                    'content': nace_content,
                    'size': len(nace_content)
                }
            if stress_criticality_file:
                logger.info(f"⚡ Stress Criticality document attached: {stress_criticality_file.name}")
                stress_content = stress_criticality_file.read()
                enrichment_files['stress_criticality'] = {
                    'filename': stress_criticality_file.name,
                    'content': stress_content,
                    'size': len(stress_content)
                }
            
            # Queue Celery task for async processing (or execute immediately if EAGER mode)
            # PRODUCTION FIX: Catch ALL connection errors and fallback to direct execution
            try:
                task = process_pid_upload_async.delay(
                    file_path=saved_path,
                    filename=pid_file.name,
                    list_type=list_type,
                    user_id=request.user.id,
                    project_id=project.id if project else None,
                    document_id=document_id,
                    storage_type=storage_type,
                    s3_url=s3_url,
                    include_area=include_area,
                    format_type=format_type,
                    enrichment_files=enrichment_files if enrichment_files else None
                )
            except Exception as task_err:
                # ANY error calling task (Redis, Celery, Connection, etc.) - fallback to direct execution
                logger.warning(f"⚠️ Task dispatch failed: {type(task_err).__name__}: {task_err}")
                logger.info(f"🔄 Falling back to DIRECT synchronous execution (no broker required)")
                
                # Import and call the task function directly (bypasses Celery/Redis entirely)
                from apps.designiq.tasks import process_pid_upload_async
                
                try:
                    result = process_pid_upload_async(
                        file_path=saved_path,
                        filename=pid_file.name,
                        list_type=list_type,
                        user_id=request.user.id,
                        project_id=project.id if project else None,
                        document_id=document_id,
                        storage_type=storage_type,
                        s3_url=s3_url,
                        include_area=include_area,
                        format_type=format_type,
                        enrichment_files=enrichment_files
                    )
                    
                    logger.info(f"✅ Direct execution complete: {result.get('total_items', 0)} items")
                    
                    return Response({
                        "success": True,
                        "message": result.get('message', 'P&ID processed successfully'),
                        "document_id": document_id,
                        "filename": pid_file.name,
                        "extracted_lines": result.get('extracted_lines', []),
                        "items_created": result.get('items_created', 0),
                        "items_updated": result.get('items_updated', 0),
                        "total_items": result.get('total_items', 0),
                        "execution_mode": "direct_sync_fallback"
                    }, status=200)
                    
                except Exception as exec_err:
                    # Even direct execution failed - return detailed error
                    logger.error(f"❌ Direct execution ALSO failed: {exec_err}")
                    logger.exception("Full traceback:")
                    
                    return Response({
                        "success": False,
                        "error": f"Processing failed: {str(exec_err)}",
                        "error_type": type(exec_err).__name__,
                        "document_id": document_id,
                        "filename": pid_file.name
                    }, status=500)
            
            # Check if running in EAGER mode (synchronous execution)
            from django.conf import settings
            is_eager = getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
            
            if is_eager:
                # EAGER mode: Task completed immediately, return result now
                logger.info(f"⚡ EAGER mode: Task {task.id} completed synchronously")
                
                # Get result from task
                if task.successful():
                    result = task.result
                    logger.info(f"✅ Returning {result.get('total_items', 0)} items directly to frontend")
                    
                    return Response({
                        "success": True,
                        "message": result.get('message', 'P&ID processed successfully'),
                        "document_id": document_id,
                        "filename": pid_file.name,
                        "extracted_lines": result.get('extracted_lines', []),
                        "enriched_data": result.get('enriched_data', []),  # ENRICHMENT LAYER
                        "items_created": result.get('items_created', 0),
                        "items_updated": result.get('items_updated', 0),
                        "total_items": result.get('total_items', 0),
                        "format_type": result.get('format_type', format_type),
                        "include_area": result.get('include_area', include_area),
                        "s3_url": s3_url
                    }, status=status.HTTP_200_OK)
                else:
                    # Task failed in EAGER mode
                    error = str(task.result) if task.result else "Task failed"
                    logger.error(f"❌ EAGER mode task failed: {error}")
                    return Response({
                        "success": False,
                        "error": f"Processing failed: {error}"
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                # Async mode: Return task ID for polling
                file_size_mb = pid_file.size / 1024 / 1024
                estimated_seconds = int(file_size_mb * 45)  # 45 seconds per MB average
                
                logger.info(f"🔄 Async mode: Queued task {task.id} (estimated {estimated_seconds}s)")
                
                return Response({
                    "success": True,
                    "task_id": task.id,
                    "message": "PDF uploaded successfully. Processing in background...",
                    "document_id": document_id,
                    "filename": pid_file.name,
                    "file_size_mb": round(file_size_mb, 2),
                    "estimated_time_seconds": estimated_seconds,
                    "status_endpoint": f"/api/v1/designiq/lists/upload_pid_status/{task.id}/",
                    "instructions": "Poll the status_endpoint to check progress and get results when complete"
                }, status=status.HTTP_202_ACCEPTED)
            
        except Exception as e:
            logger.error(f"❌ Error uploading P&ID: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to upload P&ID: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_enriched_pid(self, request):
        """
        🧠 4-Document Smart Enrichment Upload
        
        Upload P&ID + HMB + PMS + NACE for 34-column line list extraction
        
        Required:
        - pid_file: P&ID PDF (mandatory)
        - hmb_file: HMB/PFD PDF (mandatory)
        - pms_file: PMS PDF (mandatory) 
        - nace_file: Material Selection/NACE PDF (mandatory)
        
        Returns: Enriched line data with up to 34 columns
        """
        # Get all 4 documents
        pid_file = request.FILES.get('pid_file')
        hmb_file = request.FILES.get('hmb_file')
        pms_file = request.FILES.get('pms_file')
        nace_file = request.FILES.get('nace_file')
        
        list_type = request.data.get('list_type', 'line_list')
        
        # Validate all 4 documents are present
        if not pid_file:
            return Response({
                "error": "P&ID file is mandatory"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not hmb_file:
            return Response({
                "error": "HMB file is mandatory"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not pms_file:
            return Response({
                "error": "PMS file is mandatory"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not nace_file:
            return Response({
                "error": "NACE file is mandatory"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            from django.core.files.storage import default_storage
            from .models import DesignProject
            from .tasks_enriched import process_enriched_pid_upload
            from io import BytesIO
            
            # Get or create project
            project, _ = DesignProject.objects.get_or_create(
                project_name="P&ID Enriched Upload Project",
                defaults={
                    'created_by': request.user,
                    'design_type': 'pid',
                    'status': 'active'
                }
            )
            
            # Generate document ID
            last_doc_id = 0
            existing_items = EngineeringListItem.objects.filter(
                data__has_key='document_id'
            ).order_by('-created_at').first()
            
            if existing_items and existing_items.data.get('document_id'):
                try:
                    doc_id_str = existing_items.data['document_id'].split('-')[0]
                    last_doc_id = int(doc_id_str)
                except (ValueError, IndexError):
                    pass
            
            new_doc_id = last_doc_id + 1
            document_id = f"{new_doc_id:04d}-{pid_file.name}"
            
            logger.info(f"🧠 Starting 4-document enrichment: {document_id}")
            
            # Upload all 4 documents to storage
            storage_type = 'local'
            
            # Save P&ID
            pid_file.seek(0)
            pid_path = default_storage.save(
                f"designiq/enriched_uploads/{timezone.now().strftime('%Y/%m/%d')}/{document_id}",
                pid_file
            )
            
            # Save HMB
            hmb_file.seek(0)
            hmb_path = default_storage.save(
                f"designiq/enriched_uploads/{timezone.now().strftime('%Y/%m/%d')}/HMB-{document_id}",
                hmb_file
            )
            
            # Save PMS
            pms_file.seek(0)
            pms_path = default_storage.save(
                f"designiq/enriched_uploads/{timezone.now().strftime('%Y/%m/%d')}/PMS-{document_id}",
                pms_file
            )
            
            # Save NACE
            nace_file.seek(0)
            nace_path = default_storage.save(
                f"designiq/enriched_uploads/{timezone.now().strftime('%Y/%m/%d')}/NACE-{document_id}",
                nace_file
            )
            
            # Get format options
            include_area = request.POST.get('include_area', 'false').lower() == 'true'
            format_type = request.POST.get('format_type', 'onshore').lower()
            
            # Convert relative paths to absolute paths for task processing
            pid_absolute_path = default_storage.path(pid_path)
            hmb_absolute_path = default_storage.path(hmb_path)
            pms_absolute_path = default_storage.path(pms_path)
            nace_absolute_path = default_storage.path(nace_path)
            
            # Queue enrichment task
            task = process_enriched_pid_upload.delay(
                pid_file_path=pid_absolute_path,
                pid_filename=pid_file.name,
                hmb_file_path=hmb_absolute_path,
                pms_file_path=pms_absolute_path,
                nace_file_path=nace_absolute_path,
                list_type=list_type,
                user_id=request.user.id,
                project_id=project.id,
                document_id=document_id,
                storage_type=storage_type,
                include_area=include_area,
                format_type=format_type
            )
            
            # Check if EAGER mode
            from django.conf import settings
            is_eager = getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
            
            if is_eager:
                logger.info(f"⚡ EAGER mode: Task {task.id} completed synchronously")
                
                if task.successful():
                    result = task.result
                    logger.info(f"✅ Enrichment complete: {result.get('total_items', 0)} items with 34 columns")
                    
                    return Response({
                        "success": True,
                        "message": result.get('message', '4-document enrichment complete'),
                        "document_id": document_id,
                        "enriched_lines": result.get('extracted_lines', []),
                        "items_created": result.get('items_created', 0),
                        "items_updated": result.get('items_updated', 0),
                        "total_items": result.get('total_items', 0),
                        "has_hmb_enrichment": result.get('has_hmb_enrichment', False),
                        "has_pms_enrichment": result.get('has_pms_enrichment', False),
                        "has_nace_enrichment": result.get('has_nace_enrichment', False),
                        "format_type": result.get('format_type', format_type),
                        "include_area": result.get('include_area', include_area)
                    }, status=status.HTTP_200_OK)
                else:
                    error = str(task.result) if task.result else "Unknown error"
                    return Response({
                        "error": f"Enrichment failed: {error}"
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            else:
                # Async mode
                logger.info(f"🔄 Async mode: Queued enrichment task {task.id}")
                
                return Response({
                    "success": True,
                    "task_id": task.id,
                    "message": "All 4 documents uploaded successfully. Processing enrichment...",
                    "document_id": document_id,
                    "estimated_time_minutes": 5,
                    "status_endpoint": f"/api/v1/designiq/lists/enriched_upload_status/{task.id}/",
                    "instructions": "Poll the status_endpoint to check progress"
                }, status=status.HTTP_202_ACCEPTED)
        
        except Exception as e:
            logger.error(f"❌ Error in enriched upload: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to process enriched upload: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='upload_pid_status/(?P<task_id>[^/.]+)')
    def upload_pid_status(self, request, task_id=None):
        """
        Check status of async P&ID upload processing
        
        GET /api/v1/designiq/lists/upload_pid_status/{task_id}/
        
        Returns:
        - state: PENDING, PROCESSING, SUCCESS, or FAILURE
        - progress: 0-100% (if PROCESSING)
        - status: Human-readable status message
        - result: Processing results (if SUCCESS)
        - error: Error message (if FAILURE)
        """
        from celery.result import AsyncResult
        from django.core.cache import cache
        
        if not task_id:
            return Response({
                "error": "Task ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Check cache first for fast response
            cache_key = f'pid_upload_progress_{task_id}'
            cached_data = cache.get(cache_key)
            
            if cached_data:
                return Response(cached_data)
            
            # Fall back to Celery task result
            task = AsyncResult(task_id)
            
            response_data = {
                'task_id': task_id,
                'state': task.state,
            }
            
            if task.state == 'PENDING':
                response_data.update({
                    'status': 'Task is queued, waiting to start...',
                    'percent': 0
                })
            elif task.state == 'PROGRESS':
                response_data.update(task.info)
            elif task.state == 'SUCCESS':
                response_data.update({
                    'status': 'Processing complete!',
                    'percent': 100,
                    'result': task.result
                })
            elif task.state == 'FAILURE':
                response_data.update({
                    'status': 'Processing failed',
                    'percent': 0,
                    'error': str(task.info) if task.info else 'Unknown error'
                })
            else:
                response_data.update({
                    'status': f'Task state: {task.state}',
                    'percent': 0
                })
            
            return Response(response_data)
            
        except Exception as e:
            logger.error(f"❌ Error checking task status: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to check task status: {str(e)}",
                "task_id": task_id
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='enriched_upload_status/(?P<task_id>[^/.]+)')
    def enriched_upload_status(self, request, task_id=None):
        """
        Check status of 4-document enriched upload processing
        
        GET /api/v1/designiq/lists/enriched_upload_status/{task_id}/
        
        Returns:
        - state: PENDING, PROCESSING, SUCCESS, or FAILURE
        - progress: 0-100% (if PROCESSING)
        - current_phase: P&ID, HMB, PMS, or NACE
        - status: Human-readable status message
        - result: Processing results with enrichment flags (if SUCCESS)
        - error: Error message (if FAILURE)
        """
        from celery.result import AsyncResult
        from django.core.cache import cache
        
        if not task_id:
            return Response({
                "error": "Task ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Check cache first for fast response
            cache_key = f'enriched_upload_progress_{task_id}'
            cached_data = cache.get(cache_key)
            
            if cached_data:
                return Response(cached_data)
            
            # Fall back to Celery task result
            task = AsyncResult(task_id)
            
            response_data = {
                'task_id': task_id,
                'state': task.state,
            }
            
            if task.state == 'PENDING':
                response_data.update({
                    'status': '⏳ Enrichment queued, waiting to start...',
                    'percent': 0,
                    'current_phase': 'queued'
                })
            elif task.state == 'PROGRESS':
                info = task.info or {}
                response_data.update({
                    'status': info.get('status', 'Processing...'),
                    'percent': info.get('percent', 0),
                    'current_phase': info.get('current_phase', 'unknown'),
                    'items_processed': info.get('items_processed', 0),
                    'has_hmb_enrichment': info.get('has_hmb_enrichment', False),
                    'has_pms_enrichment': info.get('has_pms_enrichment', False),
                    'has_nace_enrichment': info.get('has_nace_enrichment', False)
                })
            elif task.state == 'SUCCESS':
                result = task.result or {}
                response_data.update({
                    'status': '✅ 4-document enrichment complete!',
                    'percent': 100,
                    'current_phase': 'complete',
                    'result': {
                        'success': result.get('success', True),
                        'total_items': result.get('total_items', 0),
                        'items_created': result.get('items_created', 0),
                        'items_updated': result.get('items_updated', 0),
                        'has_hmb_enrichment': result.get('has_hmb_enrichment', False),
                        'has_pms_enrichment': result.get('has_pms_enrichment', False),
                        'has_nace_enrichment': result.get('has_nace_enrichment', False),
                        'enriched_lines': result.get('extracted_lines', []),
                        'message': result.get('message', '')
                    }
                })
            elif task.state == 'FAILURE':
                response_data.update({
                    'status': '❌ Enrichment failed',
                    'percent': 0,
                    'current_phase': 'error',
                    'error': str(task.info) if task.info else 'Unknown error'
                })
            else:
                response_data.update({
                    'status': f'Task state: {task.state}',
                    'percent': 0,
                    'current_phase': 'unknown'
                })
            
            return Response(response_data)
            
        except Exception as e:
            logger.error(f"❌ Error checking enriched task status: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to check enriched task status: {str(e)}",
                "task_id": task_id
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='documents')
    def list_documents(self, request):
        """
        List all uploaded P&ID documents with their unique IDs
        Groups line items by document_id
        """
        try:
            list_type = request.query_params.get('list_type', 'line_list')
            
            # Get all items with document IDs
            items = EngineeringListItem.objects.filter(
                list_type=list_type,
                data__has_key='document_id'
            ).order_by('-created_at')
            
            # Group by document_id
            documents_map = {}
            for item in items:
                doc_id = item.data.get('document_id')
                if not doc_id:
                    continue
                
                if doc_id not in documents_map:
                    documents_map[doc_id] = {
                        'document_id': doc_id,
                        'filename': item.data.get('filename', 'Unknown'),
                        'original_filename': item.data.get('filename', 'Unknown'),
                        'document_path': item.data.get('document_path', ''),
                        'storage_type': item.data.get('storage_type', 'local'),
                        's3_url': item.data.get('s3_url'),
                        'upload_date': item.data.get('upload_timestamp', item.created_at.isoformat()),
                        'uploaded_by': item.created_by.get_full_name() if item.created_by else 'Unknown',
                        'format_type': item.data.get('format_type', 'general'),
                        'line_count': 0,
                        'item_ids': []
                    }
                
                documents_map[doc_id]['line_count'] += 1
                documents_map[doc_id]['item_ids'].append(item.id)
            
            documents_list = list(documents_map.values())
            
            return Response({
                'documents': documents_list,
                'total_documents': len(documents_list)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error listing documents: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to list documents: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='documents/(?P<document_id>.+)/download')
    def download_document(self, request, document_id=None):
        """
        Download P&ID document by document_id
        Supports both S3 and local storage
        """
        try:
            logger.info(f"📥 Download request for document_id: {document_id}")
            
            # Find an item with this document_id to get the storage info
            item = EngineeringListItem.objects.filter(
                data__document_id=document_id
            ).first()
            
            if not item:
                logger.warning(f"❌ Document not found in database: {document_id}")
                return Response({
                    "error": f"Document not found: {document_id}"
                }, status=status.HTTP_404_NOT_FOUND)
            
            storage_type = item.data.get('storage_type', 'local')
            document_path = item.data.get('document_path', '')
            filename = item.data.get('filename', document_id)
            
            if storage_type == 's3':
                # Generate presigned URL for S3 (1 hour expiration)
                presigned_url = s3_storage.generate_presigned_url(
                    s3_key=document_path,
                    expiration=3600
                )
                
                if presigned_url:
                    # Return presigned URL
                    return Response({
                        'url': presigned_url,
                        'filename': filename,
                        'storage_type': 's3',
                        'expires_in': 3600
                    }, status=status.HTTP_200_OK)
                else:
                    return Response({
                        "error": "Failed to generate download URL"
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            else:
                # Local storage - serve file directly
                from django.core.files.storage import default_storage
                
                if not default_storage.exists(document_path):
                    return Response({
                        "error": f"Document file not found: {document_path}"
                    }, status=status.HTTP_404_NOT_FOUND)
                
                file = default_storage.open(document_path, 'rb')
                response = HttpResponse(file.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'inline; filename="{filename}"'
                file.close()
                
                return response
            
        except Exception as e:
            logger.error(f"Error downloading document: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to download document: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['delete'], url_path='documents/(?P<document_id>.+)')
    def delete_document(self, request, document_id=None):
        """
        Delete all line items associated with a document ID
        """
        try:
            logger.info(f"🗑️ Delete request for document_id: {document_id}")
            
            # Find all items with this document_id
            items = EngineeringListItem.objects.filter(
                data__document_id=document_id
            )
            
            count = items.count()
            if count == 0:
                logger.warning(f"❌ No items found for document: {document_id}")
                return Response({
                    "error": f"No items found for document ID: {document_id}"
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Delete all items
            items.delete()
            
            logger.info(f"🗑️ Deleted document {document_id} with {count} line items")
            
            return Response({
                "message": f"Successfully deleted document {document_id}",
                "items_deleted": count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error deleting document: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to delete document: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='export-document-excel')
    def export_document_excel(self, request):
        """
        Export document line items to Excel (CRS multi-revision pattern)
        GET /api/v1/designiq/lists/export-document-excel/?document_id={document_id}
        
        Query Parameters:
            document_id: The document ID to export (required)
        
        Returns Excel file with line items
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            # Get document_id from query parameters
            document_id = request.query_params.get('document_id')
            if not document_id:
                return Response({
                    "error": "document_id query parameter is required"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"📊 Excel export request for document: {document_id}")
            
            # Find all items with this document_id
            items = EngineeringListItem.objects.filter(
                list_type='line_list',
                data__document_id=document_id
            ).order_by('item_tag')
            
            if not items.exists():
                logger.warning(f"❌ No items found for document: {document_id}")
                return Response({
                    "error": f"No line items found for document ID: {document_id}"
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Log sample data to debug
            first_item = items.first()
            logger.info(f"📋 Sample item data keys: {list(first_item.data.keys())}")
            logger.info(f"📋 Has criticality_stress? {'criticality_stress' in first_item.data}")
            if 'criticality_stress' in first_item.data:
                logger.info(f"📋 Sample criticality_stress value: {first_item.data['criticality_stress']}")
            else:
                logger.warning(f"⚠️ criticality_stress NOT FOUND in database! Old data needs re-upload.")
                logger.info(f"📋 All available keys: {list(first_item.data.keys())}")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Line List"
            
            # Header style
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            
            # Define headers - Include ALL enriched columns (17 base + 26 enriched + criticality_stress)
            headers = [
                # Base columns (17)
                'Line Number', 'Size', 'Fluid Code', 'Fluid Description',
                'Sequence No', 'Pipe Class', 'Insulation', 'Area',
                'FROM', 'TO', 'Flow Detection Method', 'Flow Confidence',
                'Page', 'Confidence', 'Document ID', 'Filename', 'Upload Date',
                # Enriched columns (26)
                'Flow Medium', 'Two Phase', 'Surge Flow', 'Flow Max', 'Density',
                'Normal Pressure', 'Normal Temp', 'Design Pressure', 'Minimax Design Temp',
                'Design Code', 'Category M Fluid', 'Schedule / Wall THK', 'Stress Relief',
                'PWHT', 'RT', 'MT/PT', 'Hardness', 'Visual', 'NACE MR 0175',
                'Piping Rated Pressure', 'Test Pressure', 'Test Medium',
                'PID No', 'PID Rev', 'Date', 'Criticality Code',
                # Stress criticality column (1)
                'Criticality Stress',
                # Status columns (2)
                'Status', 'Validated'
            ]
            
            ws.append(headers)
            
            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data rows
            for item in items:
                ws.append([
                    # Base columns (17)
                    item.item_tag,
                    item.data.get('size', ''),
                    item.data.get('fluid_code', ''),
                    item.data.get('fluid_description', ''),
                    item.data.get('sequence_no', ''),
                    item.data.get('pipr_class', ''),
                    item.data.get('insulation', ''),
                    item.data.get('area', ''),
                    item.data.get('from_line', ''),
                    item.data.get('to_line', ''),
                    item.data.get('flow_detection_method', ''),
                    item.data.get('flow_confidence', ''),
                    item.data.get('page', ''),
                    item.data.get('confidence', ''),
                    item.data.get('document_id', ''),
                    item.data.get('filename', ''),
                    item.data.get('upload_date', ''),
                    # Enriched columns (26)
                    item.data.get('flow_medium', ''),
                    item.data.get('two_phase', ''),
                    item.data.get('surge_flow', ''),
                    item.data.get('flow_max', ''),
                    item.data.get('density', ''),
                    item.data.get('normal_pressure', ''),
                    item.data.get('normal_temp', ''),
                    item.data.get('design_pressure', ''),
                    item.data.get('minimax_design_temp', ''),
                    item.data.get('design_code', ''),
                    item.data.get('category_m_fluid', ''),
                    item.data.get('schedule_wall_thk', ''),
                    item.data.get('stress_relief', ''),
                    item.data.get('pwht', ''),
                    item.data.get('rt', ''),
                    item.data.get('mt_pt', ''),
                    item.data.get('hardness', ''),
                    item.data.get('visual', ''),
                    item.data.get('nace_mr_0175', ''),
                    item.data.get('piping_rated_pressure', ''),
                    item.data.get('test_pressure', ''),
                    item.data.get('test_medium', ''),
                    item.data.get('pid_no', ''),
                    item.data.get('pid_rev', ''),
                    item.data.get('date', ''),
                    item.data.get('criticality_code', ''),
                    # Stress criticality column (1)
                    item.data.get('criticality_stress', 'N/A'),
                    # Status columns (2)
                    item.status,
                    'Yes' if item.is_validated else 'No'
                ])
            
            # Auto-size columns
            for col_num in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_num)
                max_length = len(headers[col_num - 1])
                for cell in ws[col_letter]:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = min(cell_length, 50)
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Get filename from first item
            filename = items.first().data.get('filename', document_id)
            safe_filename = "".join(c for c in filename if c.isalnum() or c in "-_.").strip() or "line_list"
            excel_filename = f"{safe_filename}_line_list.xlsx"
            
            logger.info(f"✅ Generated Excel with {items.count()} line items")
            
            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
            response['X-Item-Count'] = str(items.count())
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting Excel: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to export Excel: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='previous_outputs')
    def previous_outputs(self, request):
        """
        List all previously processed P&ID outputs for download
        
        GET /api/v1/designiq/lists/previous_outputs/
        
        Query Parameters:
        - list_type: Filter by list type (default: line_list)
        - limit: Number of results (default: 50)
        
        Returns list of historical P&ID processing outputs with download info
        """
        try:
            from .models import ProcessedPIDOutput
            
            list_type = request.query_params.get('list_type', 'line_list')
            limit = int(request.query_params.get('limit', 50))
            
            # Check if table exists by attempting a simple query
            try:
                outputs = ProcessedPIDOutput.objects.filter(
                    list_type=list_type
                ).order_by('-processing_date')[:limit]
                
                results = []
                for output in outputs:
                    try:
                        results.append({
                            'id': str(output.id),
                            'pid_number': output.pid_number or '',
                            'pid_revision': output.pid_revision or '',
                            'processing_date': output.processing_date.strftime('%Y-%m-%d %H:%M') if output.processing_date else '',
                            'processed_by': output.processed_by.email if output.processed_by else 'Unknown',
                            'total_lines': output.total_lines or 0,
                            'total_columns': output.total_columns or 0,
                            'excel_filename': output.excel_filename or '',
                            'file_size_mb': round(output.file_size / (1024 * 1024), 2) if output.file_size else 0,
                            'enrichment_enabled': output.enrichment_enabled or False,
                            'format_type': output.format_type or 'general',
                            'has_file': bool(output.excel_file) if hasattr(output, 'excel_file') else False
                        })
                    except Exception as e:
                        logger.warning(f"Error processing output {output.id}: {e}")
                        continue
                
                logger.info(f"📂 Retrieved {len(results)} previous outputs for {list_type}")
                
                return Response({
                    'success': True,
                    'count': len(results),
                    'outputs': results
                })
            except Exception as db_error:
                logger.warning(f"Database query error in previous_outputs: {db_error}")
                # Return empty list if table doesn't exist or query fails
                return Response({
                    'success': True,
                    'count': 0,
                    'outputs': [],
                    'message': 'No previous outputs available'
                })
        except Exception as e:
            logger.error(f"Error in previous_outputs endpoint: {e}")
            return Response({
                'success': False,
                'error': 'Failed to fetch previous outputs',
                'outputs': []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        except Exception as e:
            logger.error(f"Error fetching previous outputs: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to fetch previous outputs: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='download_output/(?P<output_id>[^/.]+)')
    def download_output(self, request, output_id=None):
        """
        Download a previously generated Excel file
        
        GET /api/v1/designiq/lists/download_output/{output_id}/
        
        Returns the Excel file for download without reprocessing
        """
        from .models import ProcessedPIDOutput
        from django.http import FileResponse
        import os
        
        if not output_id:
            return Response({
                "error": "Output ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            output = ProcessedPIDOutput.objects.get(id=output_id)
            
            if not output.excel_file:
                return Response({
                    "error": "Excel file not found for this output"
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Check if file exists
            if not output.excel_file.storage.exists(output.excel_file.name):
                return Response({
                    "error": "Excel file no longer exists on server"
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Open file and create response
            file_handle = output.excel_file.open('rb')
            response = FileResponse(
                file_handle,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{output.excel_filename}"'
            response['X-PID-Number'] = output.pid_number
            response['X-PID-Revision'] = output.pid_revision
            response['X-Processing-Date'] = output.processing_date.strftime('%Y-%m-%d')
            
            logger.info(f"📥 Downloaded: {output.excel_filename} (ID: {output_id})")
            
            return response
            
        except ProcessedPIDOutput.DoesNotExist:
            return Response({
                "error": "Output not found"
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error downloading output: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to download output: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='base_extraction')
    def base_extraction(self, request):
        """
        🎯 BASE EXTRACTION ENDPOINT - P&ID Only (Async, production-safe)

        POST /api/v1/designiq/lists/base_extraction/

        Accepts: P&ID file only (PDF)
        Returns (HTTP 202): { task_id, status_endpoint } — poll for progress/results
        Returns (HTTP 200): { success, data, ... } — only when Celery is in EAGER mode (local dev)

        Columns: Original Detection, Fluid Code, Size, Sequence No,
                 PIPR Class, Insulation, From, To

        Why async? Railway's reverse proxy times out HTTP requests after ~60 s.
        OCR extraction takes several minutes, so we offload it to a Celery worker
        and let the frontend poll /base_extraction_status/{task_id}/ for results.
        """
        from apps.designiq.tasks import base_extract_lines_async

        logger.info("=" * 80)
        logger.info("🎯 BASE EXTRACTION REQUEST (async) – P&ID Only")
        logger.info("=" * 80)

        try:
            # ------------------------------------------------------------------
            # 1. Validate input
            # ------------------------------------------------------------------
            pid_file = request.FILES.get('pid_file')
            if not pid_file:
                return Response({'error': 'P&ID file is required'}, status=status.HTTP_400_BAD_REQUEST)

            if not pid_file.name.lower().endswith('.pdf'):
                return Response({'error': 'Only PDF files are supported'}, status=status.HTTP_400_BAD_REQUEST)

            include_area = request.POST.get('include_area', 'false').lower() == 'true'
            format_type = request.POST.get('format_type', 'onshore').lower()

            logger.info(f"📄 File: {pid_file.name} ({pid_file.size / 1024 / 1024:.2f} MB)")
            logger.info(f"📍 Format: {format_type}, Include Area: {include_area}")

            # ------------------------------------------------------------------
            # 2. Save P&ID to a temporary file (Celery worker needs a path on disk)
            # ------------------------------------------------------------------
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                for chunk in pid_file.chunks():
                    tmp_file.write(chunk)
                tmp_path = tmp_file.name

            # Save legend file if provided
            legend_tmp_path = None
            legend_file = request.FILES.get('legend_file')
            if legend_file:
                if not legend_file.name.lower().endswith('.pdf'):
                    return Response({'error': 'Legend file must be a PDF'}, status=status.HTTP_400_BAD_REQUEST)
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as leg_tmp:
                    for chunk in legend_file.chunks():
                        leg_tmp.write(chunk)
                    legend_tmp_path = leg_tmp.name
                logger.info(f"📋 Legend file: {legend_file.name}")

            logger.info(f"💾 Saved to temp file: {tmp_path}")

            # ------------------------------------------------------------------
            # 3. Dispatch — two modes, chosen by EAGER flag:
            #
            #  A. EAGER (CELERY_TASK_ALWAYS_EAGER=True, local dev):
            #     Celery runs the task inline — no broker needed.
            #     Returns HTTP 200 immediately with extraction results.
            #
            #  B. Production (default):
            #     ALWAYS use thread-based extraction.  The thread writes
            #     progress to /tmp/base_extraction_{task_id}.json so that
            #     any Gunicorn worker on the same container can serve the
            #     polling endpoint.  The POST handler returns HTTP 202
            #     in < 1 second regardless of Redis/Celery availability.
            #
            #  SOFT-CODED opt-in: set env var
            #     CELERY_BASE_EXTRACTION_PREFER_CELERY=true
            #  to use Celery when a live broker is configured.  The
            #  CELERY_BROKER_CONNECTION_TIMEOUT setting controls how long
            #  we wait before falling back to thread (max 8 s, non-blocking).
            # ------------------------------------------------------------------
            is_eager = getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)

            # --- path A: EAGER --------------------------------------------------
            if is_eager:
                logger.info('⚡ EAGER mode — running task synchronously')
                task = base_extract_lines_async.delay(
                    file_path=tmp_path,
                    filename=pid_file.name,
                    include_area=include_area,
                    format_type=format_type,
                )
                if task.successful():
                    return Response(task.result, status=status.HTTP_200_OK)
                error = str(task.result) if task.result else 'Task failed in EAGER mode'
                return Response({'error': error}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            # --- path B: async (production) ------------------------------------
            # SOFT-CODED: prefer Celery only when explicitly opted in AND broker
            # URL is configured.  Otherwise always use the thread path which
            # is guaranteed non-blocking.
            _prefer_celery = getattr(settings, 'CELERY_BASE_EXTRACTION_PREFER_CELERY', False)
            _broker_url    = getattr(settings, 'CELERY_BROKER_URL', None)
            _task_id_str   = None
            _mode          = 'thread'

            if _prefer_celery and _broker_url:
                # Non-blocking Celery dispatch: create TPE without context manager
                # so we can call shutdown(wait=False) and never block the worker.
                _broker_timeout = getattr(
                    settings, 'CELERY_BROKER_CONNECTION_TIMEOUT', 5
                ) + 3
                _tpe    = concurrent.futures.ThreadPoolExecutor(max_workers=1)
                _future = _tpe.submit(
                    base_extract_lines_async.delay,
                    file_path=tmp_path,
                    filename=pid_file.name,
                    include_area=include_area,
                    format_type=format_type,
                )
                # CRITICAL: shutdown(wait=False) — never block even if Redis hangs
                _tpe.shutdown(wait=False)
                try:
                    task         = _future.result(timeout=_broker_timeout)
                    _task_id_str = task.id
                    _mode        = 'celery'
                    logger.info(f'✅ Celery task submitted: {_task_id_str}')
                except Exception as _celery_err:
                    logger.warning(
                        f'⚠️ Celery dispatch failed ({_celery_err.__class__.__name__}: '
                        f'{_celery_err}). Falling back to thread.'
                    )

            # Thread fallback (or always-thread when _prefer_celery is False)
            if _task_id_str is None:
                _task_id_str = str(uuid.uuid4())
                _mode        = 'thread'
                _run_base_extraction_in_thread(
                    _task_id_str, tmp_path, pid_file.name,
                    include_area, format_type, legend_tmp_path,
                )

            # ------------------------------------------------------------------
            # 4. Return 202 immediately — frontend polls for progress/results
            # ------------------------------------------------------------------
            file_size_mb      = pid_file.size / 1024 / 1024
            estimated_seconds = max(60, int(file_size_mb * 45))  # soft heuristic

            logger.info(
                f'🔄 Async mode ({_mode}): task {_task_id_str} '
                f'(~{estimated_seconds}s estimated)'
            )

            return Response({
                'success':           True,
                'task_id':           _task_id_str,
                'message':           'P&ID uploaded — processing in background. Poll the status endpoint.',
                'filename':          pid_file.name,
                'file_size_mb':      round(file_size_mb, 2),
                'estimated_time_seconds': estimated_seconds,
                'dispatch_mode':     _mode,
                'status_endpoint':   f'/api/v1/designiq/lists/base_extraction_status/{_task_id_str}/',
            }, status=status.HTTP_202_ACCEPTED)

        except Exception as e:
            logger.error(f"❌ base_extraction failed: {e}", exc_info=True)
            return Response({
                'error': str(e),
                'message': 'Base extraction failed',
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='base_extraction_status/(?P<task_id>[^/.]+)')
    def base_extraction_status(self, request, task_id=None):
        """
        📊 Poll the status of an async base extraction job

        GET /api/v1/designiq/lists/base_extraction_status/{task_id}/

        Returns:
        - state: PENDING | PROGRESS | SUCCESS | FAILURE
        - percent: 0-100
        - status: human-readable message
        - result: extraction data (only when state == SUCCESS)
        - error: error message (only when state == FAILURE)
        """
        from celery.result import AsyncResult
        from django.core.cache import cache

        if not task_id:
            return Response({'error': 'task_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # ------------------------------------------------------------------
            # A. Thread-based fallback: check /tmp/ progress file first.
            #    This is written by _run_base_extraction_in_thread() when the
            #    Celery broker is unavailable. All Gunicorn workers share /tmp/.
            # ------------------------------------------------------------------
            progress_file = f'/tmp/base_extraction_{task_id}.json'
            if os.path.exists(progress_file):
                try:
                    with open(progress_file, 'r') as fh:
                        return Response(json.load(fh))
                except Exception as read_err:
                    logger.warning(f'Could not read thread progress file: {read_err}')

            # ------------------------------------------------------------------
            # B. Celery path: check Redis cache then AsyncResult
            # ------------------------------------------------------------------
            cache_key = f'base_extraction_progress_{task_id}'
            try:
                cached = cache.get(cache_key)
            except Exception:
                cached = None  # Cache backend unreachable — skip gracefully
            if cached:
                return Response(cached)

            # C. Fall back to Celery AsyncResult (works when Redis/broker is live)
            try:
                task = AsyncResult(task_id)
                response_data = {'task_id': task_id, 'state': task.state}

                if task.state == 'PENDING':
                    response_data.update({'status': 'Queued, waiting to start…', 'percent': 0})
                elif task.state == 'PROGRESS':
                    response_data.update(task.info or {})
                elif task.state == 'SUCCESS':
                    response_data.update({'status': 'Extraction complete!', 'percent': 100, 'result': task.result})
                elif task.state == 'FAILURE':
                    response_data.update({'status': 'Extraction failed', 'percent': 0,
                                           'error': str(task.info) if task.info else 'Unknown error'})
                else:
                    response_data.update({'status': f'State: {task.state}', 'percent': 0})

                return Response(response_data)
            except Exception as celery_poll_err:
                # Broker unreachable — task not found in any store
                _err_low = str(celery_poll_err).lower()
                if any(x in _err_low for x in ('connection refused', 'errno 111', 'transport error')):
                    logger.warning(f'Celery broker unreachable during status poll: {celery_poll_err}')
                    return Response({
                        'task_id': task_id,
                        'state': 'PENDING',
                        'percent': 0,
                        'status': 'Task queued — broker temporarily unreachable, will retry…',
                    })
                raise

        except Exception as e:
            logger.error(f"❌ base_extraction_status error: {e}", exc_info=True)
            return Response({'error': str(e), 'task_id': task_id}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


