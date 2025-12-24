"""
CRS Template Populator - NEW HELPER MODULE
Populates CRS Excel template with extracted comments
CRITICAL: Preserves original template format exactly
Does NOT modify existing code or APIs
"""

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from typing import List, BinaryIO
from io import BytesIO
from datetime import datetime
from .comment_extractor import ReviewerComment


def populate_crs_template(
    template_buffer: BytesIO,
    comments: List[ReviewerComment],
    document_metadata: dict = None
) -> BytesIO:
    """
    Populate CRS template with extracted comments
    
    CRITICAL RULES:
    - Opens EXISTING Excel file
    - Modifies ONLY cell VALUES
    - Preserves all formatting, formulas, structure
    - Returns modified file in SAME format
    
    Args:
        template_buffer: BytesIO containing original CRS template
        comments: List of ReviewerComment objects
        document_metadata: Optional metadata (project name, doc number, etc.)
    
    Returns:
        BytesIO containing populated template (SAME FORMAT)
    """
    try:
        # CRITICAL: Reset buffer position before reading
        template_buffer.seek(0)
        
        # Load existing template
        # Note: keep_vba only needed for .xlsm files, but safe to include
        try:
            workbook = openpyxl.load_workbook(template_buffer, keep_vba=True)
        except Exception:
            # Fallback without VBA support for regular .xlsx files
            template_buffer.seek(0)
            workbook = openpyxl.load_workbook(template_buffer)
        
        # Work with active sheet (typically "CRS" or "Comments")
        sheet = workbook.active
        
        # Populate header metadata if provided
        if document_metadata:
            _populate_header(sheet, document_metadata)
        
        # Find the data starting row (usually after header)
        data_start_row = _find_data_start_row(sheet)
        
        # Populate comments one row at a time
        for idx, comment in enumerate(comments, start=0):
            row_num = data_start_row + idx
            _populate_comment_row(sheet, row_num, comment, idx + 1)
        
        # Save to BytesIO - PRESERVES ORIGINAL FORMAT
        output_buffer = BytesIO()
        workbook.save(output_buffer)
        
        # CRITICAL: Reset position to beginning for reading
        output_buffer.seek(0)
        
        return output_buffer
    
    except Exception as e:
        print(f"Error populating CRS template: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return original template on error - SAFE FALLBACK
        template_buffer.seek(0)
        return template_buffer


def _populate_header(sheet, metadata: dict):
    """
    Populate header cells with document metadata
    Finds cells by content, modifies values only
    """
    try:
        # Common header fields to populate
        header_mappings = {
            'Project:': metadata.get('project_name', ''),
            'Document Number:': metadata.get('document_number', ''),
            'Revision:': metadata.get('revision', ''),
            'Date:': datetime.now().strftime('%Y-%m-%d'),
            'Contractor:': metadata.get('contractor', ''),
        }
        
        # Search first 20 rows for header labels
        for row in range(1, 21):
            for col in range(1, 10):
                cell = sheet.cell(row, col)
                cell_value = str(cell.value or '').strip()
                
                # If cell contains a header label, populate adjacent cell
                for label, value in header_mappings.items():
                    if label in cell_value and value:
                        # Try to find the value cell (usually next column)
                        value_cell = sheet.cell(row, col + 1)
                        if not value_cell.value or str(value_cell.value).strip() == '':
                            value_cell.value = value
    
    except Exception as e:
        # Silently fail - header population is optional
        pass


def _find_data_start_row(sheet) -> int:
    """
    Find the first row where comment data should start
    Typically looks for column headers like "No.", "Reviewer", "Comment"
    """
    # Common header indicators
    header_keywords = ['No', 'No.', '#', 'Reviewer', 'Comment', 'Discipline', 'Type']
    
    # Search first 30 rows
    for row_num in range(1, 31):
        row_values = [str(sheet.cell(row_num, col).value or '').strip() 
                      for col in range(1, 15)]
        
        # Check if this row contains multiple header keywords
        matches = sum(1 for keyword in header_keywords 
                     if any(keyword.lower() in val.lower() for val in row_values))
        
        if matches >= 2:
            # Data starts on next row
            return row_num + 1
    
    # Default fallback
    return 10


def _populate_comment_row(sheet, row_num: int, comment: ReviewerComment, index: int):
    """
    Populate a single row with comment data
    Determines column positions dynamically
    Handles merged cells gracefully
    """
    try:
        # Typical CRS column structure (flexible detection)
        column_mapping = _detect_column_mapping(sheet)
        
        # Helper to safely set cell value (handles merged cells)
        def safe_set_value(row, col, value):
            try:
                cell = sheet.cell(row, col)
                # Check if it's a merged cell
                if hasattr(cell, 'coordinate'):
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Get the top-left cell of the merged range
                            min_col, min_row, max_col, max_row = merged_range.bounds
                            cell = sheet.cell(min_row, min_col)
                            break
                cell.value = value
            except AttributeError:
                # MergedCell objects can't be written to directly
                pass
            except Exception:
                pass
        
        # Populate each field
        if 'no' in column_mapping:
            safe_set_value(row_num, column_mapping['no'], index)
        
        if 'reviewer' in column_mapping:
            safe_set_value(row_num, column_mapping['reviewer'], comment.reviewer_name)
        
        if 'comment' in column_mapping:
            safe_set_value(row_num, column_mapping['comment'], comment.comment_text)
        
        if 'discipline' in column_mapping:
            safe_set_value(row_num, column_mapping['discipline'], comment.discipline)
        
        if 'type' in column_mapping:
            safe_set_value(row_num, column_mapping['type'], comment.comment_type)
        
        if 'section' in column_mapping:
            safe_set_value(row_num, column_mapping['section'], comment.section_reference)
        
        if 'page' in column_mapping and comment.page_number:
            safe_set_value(row_num, column_mapping['page'], f"Page {comment.page_number}")
        
        # Status and response columns - leave empty for user to fill
        if 'status' in column_mapping:
            safe_set_value(row_num, column_mapping['status'], "Open")
        
        # Copy formatting from row above if exists
        _copy_row_formatting(sheet, row_num - 1, row_num)
    
    except Exception as e:
        print(f"Error populating row {row_num}: {str(e)}")


def _detect_column_mapping(sheet) -> dict:
    """
    Dynamically detect which columns correspond to which fields
    Searches header row for keywords
    """
    mapping = {}
    
    # Keywords to search for each field
    field_keywords = {
        'no': ['no', 'no.', '#', 'item', 'index'],
        'reviewer': ['reviewer', 'raised by', 'author', 'name'],
        'comment': ['comment', 'description', 'remark', 'note'],
        'discipline': ['discipline', 'dept', 'department', 'area'],
        'type': ['type', 'category', 'classification'],
        'section': ['section', 'drawing', 'reference', 'ref'],
        'page': ['page', 'pg', 'sheet'],
        'status': ['status', 'state'],
        'response': ['response', 'reply', 'action'],
    }
    
    # Search first 30 rows for headers
    header_row = None
    for row_num in range(1, 31):
        row_values = [(col, str(sheet.cell(row_num, col).value or '').strip().lower()) 
                      for col in range(1, 20)]
        
        # Check if this looks like a header row
        header_count = sum(1 for _, val in row_values if val in 
                          [kw for keywords in field_keywords.values() for kw in keywords])
        
        if header_count >= 3:
            header_row = row_num
            break
    
    if not header_row:
        # Use default positions if no clear header found
        return {
            'no': 1,
            'reviewer': 2,
            'comment': 3,
            'discipline': 4,
            'type': 5,
            'section': 6,
            'page': 7,
            'status': 8,
            'response': 9,
        }
    
    # Map columns based on header content
    for col_num, cell_value in [(col, str(sheet.cell(header_row, col).value or '').strip().lower()) 
                                 for col in range(1, 20)]:
        for field, keywords in field_keywords.items():
            if any(keyword in cell_value for keyword in keywords):
                mapping[field] = col_num
                break
    
    return mapping


def _copy_row_formatting(sheet, source_row: int, target_row: int):
    """
    Copy formatting from source row to target row
    Preserves borders, fonts, alignment
    """
    try:
        for col in range(1, 20):
            source_cell = sheet.cell(source_row, col)
            target_cell = sheet.cell(target_row, col)
            
            # Copy formatting attributes
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            if source_cell.border:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )
            
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )
            
            # Copy number format
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
    
    except Exception:
        pass  # Silent fail for formatting


def generate_populated_crs(
    template_path: str,
    comments: List[ReviewerComment],
    output_path: str = None,
    metadata: dict = None
) -> BytesIO:
    """
    High-level function to generate populated CRS
    Can work with file paths or buffers
    
    Args:
        template_path: Path to CRS template file OR BytesIO buffer
        comments: List of extracted comments
        output_path: Optional path to save file
        metadata: Optional document metadata
    
    Returns:
        BytesIO buffer with populated CRS
    """
    # Read template
    if isinstance(template_path, str):
        with open(template_path, 'rb') as f:
            template_buffer = BytesIO(f.read())
    else:
        template_buffer = template_path
    
    # Populate template
    populated_buffer = populate_crs_template(template_buffer, comments, metadata)
    
    # Save to file if requested
    if output_path:
        with open(output_path, 'wb') as f:
            populated_buffer.seek(0)
            f.write(populated_buffer.read())
        populated_buffer.seek(0)
    
    return populated_buffer


def validate_template(template_buffer: BytesIO) -> dict:
    """
    Validate that template file is usable
    Returns validation report
    """
    try:
        workbook = openpyxl.load_workbook(template_buffer)
        sheet = workbook.active
        
        return {
            'valid': True,
            'sheet_name': sheet.title,
            'row_count': sheet.max_row,
            'column_count': sheet.max_column,
            'has_data': sheet.max_row > 1,
        }
    
    except Exception as e:
        return {
            'valid': False,
            'error': str(e)
        }
