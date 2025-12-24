#!/usr/bin/env python
"""
CRS Document Processing - System Test Script
Tests all components of the CRS document processing system
"""
import os
import sys
import django

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

def run_tests():
    print('')
    print('='*70)
    print('   CRS DOCUMENT PROCESSING - TEST SUMMARY')
    print('='*70)
    print('')

    all_passed = True

    # Test 1: Imports
    print('[1] Module Imports')
    try:
        from apps.crs_documents.helpers import (
            extract_reviewer_comments, get_crs_template, 
            populate_crs_template, get_template_info, CLEANER_AVAILABLE
        )
        from apps.crs_documents.helpers.comment_cleaner import get_comment_cleaner
        print('    Status: PASSED')
    except Exception as e:
        print(f'    Status: FAILED - {e}')
        all_passed = False

    # Test 2: Comment Cleaner
    print('')
    print('[2] Comment Cleaner')
    try:
        from apps.crs_documents.helpers.comment_cleaner import get_comment_cleaner
        cleaner = get_comment_cleaner()
        tests_passed = 0
        tests_total = 6
        
        test_cases = [
            ('Sreejith Rajeev Typewriter 166 Update design', False, 'Update design'),
            ('Typewriter 166', True, ''),
            ('John Smith Text Box Check valve sizing', False, 'Check valve sizing'),
            ('Update the design as discussed with Ahmed', False, 'Update the design as discussed with Ahmed'),
            ('AutoCAD SHX Text', True, ''),
            ('166', True, ''),
        ]
        
        for text, should_skip, expected in test_cases:
            result = cleaner.clean_comment(text)
            if result.should_skip == should_skip:
                tests_passed += 1
        
        status = "PASSED" if tests_passed == tests_total else "PARTIAL"
        print(f'    Tests: {tests_passed}/{tests_total} passed')
        print(f'    Status: {status}')
        if tests_passed != tests_total:
            all_passed = False
    except Exception as e:
        print(f'    Status: FAILED - {e}')
        all_passed = False

    # Test 3: S3 Template Access
    print('')
    print('[3] S3 Template Access')
    try:
        from apps.crs_documents.helpers import get_template_info
        info = get_template_info()
        s3_ok = info.get('s3_accessible', False) or len(info.get('local_templates', [])) > 0
        s3_svc = "Available" if info.get('s3_service_available') else "Not available"
        print(f'    S3 Service: {s3_svc}')
        print(f'    S3 Accessible: {info.get("s3_accessible", False)}')
        print(f'    Local Templates: {len(info.get("local_templates", []))}')
        if s3_ok:
            print('    Status: PASSED')
        else:
            print('    Status: FAILED')
            all_passed = False
    except Exception as e:
        print(f'    Status: FAILED - {e}')
        all_passed = False

    # Test 4: Template Fetch
    print('')
    print('[4] Template Fetch & Validate')
    try:
        import openpyxl
        from apps.crs_documents.helpers import get_crs_template
        template = get_crs_template()
        template.seek(0)
        wb = openpyxl.load_workbook(template)
        print(f'    Size: {len(template.getvalue())} bytes')
        print(f'    Sheets: {wb.sheetnames}')
        print('    Status: PASSED')
    except Exception as e:
        print(f'    Status: FAILED - {e}')
        all_passed = False

    # Test 5: Template Population
    print('')
    print('[5] Template Population')
    try:
        import openpyxl
        from apps.crs_documents.helpers import get_crs_template, populate_crs_template
        from apps.crs_documents.helpers.comment_extractor import ReviewerComment
        
        comments = []
        for i in range(3):
            c = ReviewerComment()
            c.reviewer_name = f'Reviewer {i+1}'
            c.comment_text = f'Test comment {i+1}'
            c.page_number = i + 1
            comments.append(c)
        
        template = get_crs_template()
        output = populate_crs_template(template, comments, {'project_name': 'Test'})
        
        output.seek(0)
        wb = openpyxl.load_workbook(output)
        print(f'    Output: {len(output.getvalue())} bytes')
        print('    Status: PASSED')
    except Exception as e:
        print(f'    Status: FAILED - {e}')
        all_passed = False

    print('')
    print('='*70)
    if all_passed:
        print('   ALL TESTS PASSED!')
    else:
        print('   SOME TESTS FAILED')
    print('='*70)
    print('')
    
    return 0 if all_passed else 1

if __name__ == '__main__':
    sys.exit(run_tests())
